# -*- coding: utf-8 -*-
"""Phase 97 D-NEW-5: chunk_locator strings per format.

Tests verify that each extractor call site passes the correct chunk_locator
string to _write_page_doc, and that the locators follow the documented patterns:
  - PDF: 'p. N' (e.g. 'p. 1', 'p. 2')
  - DOCX: 'paragraphs N-M' (e.g. 'paragraphs 1-20')
  - HTML (semantic): '§ <heading>' (e.g. '§ פרק א')
  - HTML (fallback): '¶ N-M' (e.g. '¶ 1-5')
  - XLSX: '<sheet>!R<n>:R<m>' (e.g. 'Synopsis!R1:R500')
  - CSV: 'rows N-M' (e.g. 'rows 1-200')
"""
import os
from pathlib import Path

import pytest


TESTS_DIR = Path(__file__).parent
REPO_ROOT = TESTS_DIR.parent


# ---------------------------------------------------------------------------
# Helpers — create minimal in-memory files for testing
# ---------------------------------------------------------------------------

def _write_txt_file(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


# ---------------------------------------------------------------------------
# Test: PDF locator 'p. N'
# ---------------------------------------------------------------------------

def test_pdf_locator_format(tmp_path):
    """D-NEW-5: PDF extractor call site passes chunk_locator='p. N' for each page."""
    from shared.local_indexer import LocalIndexer

    index_dir = str(tmp_path / "idx")
    lab_dir = str(tmp_path / "lab")
    db_path = str(tmp_path / "test.sqlite3")
    os.makedirs(index_dir)
    os.makedirs(lab_dir)

    folder = str(tmp_path / "docs")
    os.makedirs(folder)

    # Create a 3-page synthetic PDF using reportlab if available, else skip
    pdf_path = os.path.join(folder, "test_3pages.pdf")
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import letter
        c = rl_canvas.Canvas(pdf_path, pagesize=letter)
        for page_num in range(1, 4):
            c.drawString(100, 700, f"Page {page_num} content for testing chunk locator.")
            c.showPage()
        c.save()
    except ImportError:
        # If reportlab not available, use a pre-built minimal PDF
        # Minimal 3-page PDF (minimal structure for PyMuPDF to parse)
        pdf_content = b"""%PDF-1.4
1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj
2 0 obj<</Type/Pages/Kids[3 0 R 5 0 R 7 0 R]/Count 3>>endobj
3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R/Resources<</Font<</F1 9 0 R>>>>>>endobj
4 0 obj<</Length 44>>stream
BT /F1 12 Tf 100 700 Td (Page 1 content text.) Tj ET
endstream
endobj
5 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 6 0 R/Resources<</Font<</F1 9 0 R>>>>>>endobj
6 0 obj<</Length 44>>stream
BT /F1 12 Tf 100 700 Td (Page 2 content text.) Tj ET
endstream
endobj
7 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 8 0 R/Resources<</Font<</F1 9 0 R>>>>>>endobj
8 0 obj<</Length 44>>stream
BT /F1 12 Tf 100 700 Td (Page 3 content text.) Tj ET
endstream
endobj
9 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj
xref
0 10
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000274 00000 n
0000000368 00000 n
0000000527 00000 n
0000000621 00000 n
0000000780 00000 n
0000000874 00000 n
trailer<</Size 10/Root 1 0 R>>
startxref
944
%%EOF"""
        with open(pdf_path, "wb") as f:
            f.write(pdf_content)

    # Capture what chunk_locators are written
    locators_written = []
    indexer = LocalIndexer(index_dir, lab_dir, db_path)
    original_write = indexer._write_page_doc

    def capturing_write(sys_id, page_num, text, title, folder_id, chunk_locator=""):
        locators_written.append((page_num, chunk_locator))
        return original_write(sys_id, page_num, text, title, folder_id,
                              chunk_locator=chunk_locator)

    indexer._write_page_doc = capturing_write

    try:
        indexer.add_folder(folder)
        result = indexer.scan_all()
    finally:
        indexer.close()

    # Filter only PDF page locators (non-empty)
    pdf_locators = [(pn, loc) for pn, loc in locators_written if loc.startswith("p. ")]
    assert len(pdf_locators) >= 1, (
        f"Expected at least 1 PDF page with 'p. N' locator, got {locators_written}"
    )
    for page_num, locator in pdf_locators:
        assert locator == f"p. {page_num}", (
            f"PDF locator should be 'p. {page_num}', got '{locator}'"
        )


# ---------------------------------------------------------------------------
# Test: DOCX locator 'paragraphs N-M'
# ---------------------------------------------------------------------------

def test_docx_locator_format(tmp_path):
    """D-NEW-5: DOCX extractor passes chunk_locator='paragraphs N-M' for each chunk."""
    try:
        import docx
    except ImportError:
        pytest.skip("python-docx not installed")

    from shared.local_indexer import LocalIndexer

    index_dir = str(tmp_path / "idx")
    lab_dir = str(tmp_path / "lab")
    db_path = str(tmp_path / "test.sqlite3")
    os.makedirs(index_dir)
    os.makedirs(lab_dir)

    folder = str(tmp_path / "docs")
    os.makedirs(folder)

    # Create DOCX with 40 paragraphs -> 2 chunks at chunk_size=20
    doc = docx.Document()
    for i in range(1, 41):
        doc.add_paragraph(f"Paragraph {i} — test content for chunk locator validation.")
    docx_path = os.path.join(folder, "test_40para.docx")
    doc.save(docx_path)

    locators_written = []
    indexer = LocalIndexer(index_dir, lab_dir, db_path)
    original_write = indexer._write_page_doc

    def capturing_write(sys_id, page_num, text, title, folder_id, chunk_locator=""):
        if chunk_locator.startswith("paragraphs "):
            locators_written.append(chunk_locator)
        return original_write(sys_id, page_num, text, title, folder_id,
                              chunk_locator=chunk_locator)

    indexer._write_page_doc = capturing_write

    try:
        indexer.add_folder(folder)
        indexer.scan_all()
    finally:
        indexer.close()

    assert len(locators_written) >= 2, (
        f"Expected at least 2 DOCX chunk locators (40 paras / 20 per chunk), got {locators_written}"
    )
    # First chunk: paragraphs 1-20
    assert locators_written[0] == "paragraphs 1-20", (
        f"First chunk locator should be 'paragraphs 1-20', got '{locators_written[0]}'"
    )
    # Second chunk: paragraphs 21-40
    assert locators_written[1] == "paragraphs 21-40", (
        f"Second chunk locator should be 'paragraphs 21-40', got '{locators_written[1]}'"
    )


# ---------------------------------------------------------------------------
# Test: HTML semantic locator '§ <heading>'
# ---------------------------------------------------------------------------

def test_html_semantic_locator_format(tmp_path):
    """D-NEW-5 / Wave C contract: HTML semantic chunk under 'פרק א' yields locator '§ פרק א'.

    Semantic mode triggers when len(headings) >= 3 AND avg paragraphs-per-heading >= 5.
    Test uses 3 headings x 5 paragraphs each = avg 5 (at boundary for trigger).
    """
    from shared.local_indexer import extract_html_pages

    # Build paragraphs per section: 5 per heading to meet avg_inter >= 5 threshold
    def section(heading, n):
        paras = "\n".join(
            f"<p>תוכן {i} תחת {heading}. Paragraph {i} content.</p>"
            for i in range(1, n + 1)
        )
        return f"<h1>{heading}</h1>\n{paras}"

    html_content = (
        "<!DOCTYPE html>\n<html><head><title>Test Document</title></head>\n<body>\n"
        + section("פרק א", 5)
        + "\n"
        + section("פרק ב", 5)
        + "\n"
        + section("פרק ג", 5)
        + "\n</body></html>"
    )

    html_path = tmp_path / "test_semantic.html"
    html_path.write_text(html_content, encoding="utf-8")

    chunks = list(extract_html_pages(str(html_path)))
    assert len(chunks) >= 1, "Expected at least 1 HTML chunk"

    # Find the chunk for פרק א
    perek_alef_chunk = next(
        (c for c in chunks if "§ פרק א" in c[3]),
        None
    )
    assert perek_alef_chunk is not None, (
        f"Expected a chunk with locator '§ פרק א', got locators: {[c[3] for c in chunks]}"
    )


# ---------------------------------------------------------------------------
# Test: HTML fallback paragraph locator '¶ N-M'
# ---------------------------------------------------------------------------

def test_html_fallback_locator_format(tmp_path):
    """D-NEW-5 / Wave C contract: HTML fallback (no headings) yields '¶ N-M' locators."""
    from shared.local_indexer import extract_html_pages

    # HTML with no h1/h2 headings -> fallback paragraph chunking
    html_content = """<!DOCTYPE html>
<html><head><title>No Headings Document</title></head>
<body>
""" + "\n".join(f"<p>Paragraph {i} content for fallback chunking test.</p>" for i in range(1, 25)) + """
</body></html>"""

    html_path = tmp_path / "test_fallback.html"
    html_path.write_text(html_content, encoding="utf-8")

    chunks = list(extract_html_pages(str(html_path)))
    assert len(chunks) >= 1, "Expected at least 1 fallback HTML chunk"

    # All locators should follow '¶ N-M' pattern
    fallback_chunks = [c for c in chunks if c[3].startswith("¶ ")]
    assert len(fallback_chunks) >= 1, (
        f"Expected fallback locators starting with '¶ ', got: {[c[3] for c in chunks]}"
    )
    for chunk in fallback_chunks:
        locator = chunk[3]
        # Format: '¶ N-M'
        assert locator.startswith("¶ "), f"Fallback locator should start with '¶ ', got '{locator}'"
        parts = locator[2:].split("-")
        assert len(parts) == 2, f"Fallback locator should have N-M format, got '{locator}'"
        assert parts[0].isdigit() and parts[1].isdigit(), (
            f"Fallback locator N-M should be digits, got '{locator}'"
        )


# ---------------------------------------------------------------------------
# Test: XLSX locator '<sheet>!R<n>:R<m>'
# ---------------------------------------------------------------------------

def test_xlsx_locator_format(tmp_path):
    """D-NEW-5 / Wave C contract: XLSX Synopsis sheet rows 1-500 yields 'Synopsis!R1:R500'."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    from shared.local_indexer import extract_xlsx_pages

    xlsx_path = tmp_path / "test_synopsis.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synopsis"
    for row in range(1, 502):  # 501 rows -> 2 chunks (500 + 1)
        ws.cell(row=row, column=1, value=f"Row {row}")
        ws.cell(row=row, column=2, value=f"Content {row}")
    wb.save(str(xlsx_path))

    chunks = list(extract_xlsx_pages(str(xlsx_path)))
    assert len(chunks) >= 1, "Expected at least 1 XLSX chunk"

    # First chunk: Synopsis!R1:R500
    first_locator = chunks[0][3]
    assert first_locator == "Synopsis!R1:R500", (
        f"First XLSX chunk locator should be 'Synopsis!R1:R500', got '{first_locator}'"
    )


# ---------------------------------------------------------------------------
# Test: CSV locator 'rows N-M'
# ---------------------------------------------------------------------------

def test_csv_locator_format(tmp_path):
    """D-NEW-5 / Wave C contract: CSV rows 201-400 yields locator 'rows 201-400'."""
    from shared.local_indexer import extract_csv_pages

    csv_path = tmp_path / "test_rows.csv"
    # Write 400 rows to get 2 chunks (200 per chunk)
    with open(str(csv_path), "w", encoding="utf-8") as f:
        for i in range(1, 401):
            f.write(f"col1_{i},col2_{i},col3_{i}\n")

    chunks = list(extract_csv_pages(str(csv_path)))
    assert len(chunks) >= 2, f"Expected at least 2 CSV chunks for 400 rows, got {len(chunks)}"

    # Second chunk: rows 201-400
    second_locator = chunks[1][3]
    assert second_locator == "rows 201-400", (
        f"Second CSV chunk locator should be 'rows 201-400', got '{second_locator}'"
    )

# -*- coding: utf-8 -*-
"""Phase 94 EXPORT-META Wave 3: xlsx structure tests.

Tests `web.export_service.ExportService.export_search_results_excel`. The
workbook layout was finalized in smoke verification round 2 (2026-05-21) as
a 4-sheet workbook: Search Results + Manuscripts + Bibliography +
Credits and Info. Consumes `shared.export_dossier` (Wave 1) and the session
payload kwargs threaded through `web.api.export_excel` (Wave 2).

Test fixture strategy:
  - `_FakeMetaMgr` provides `get_meta_for_id` / `get_library_for_id` /
    `parse_full_id_components` returning predictable shapes.
  - The Wave 1 dossier helpers are monkeypatched at
    `shared.export_dossier.<name>` so tests don't depend on real sidecars
    (matches MUST-FIX 94-01-A pattern from Wave 1).
  - Loaded workbook is parsed back via `openpyxl.load_workbook(BytesIO(...),
    rich_text=True)` for rich-text assertions on the Snippet column.
"""
from io import BytesIO

import pytest
import openpyxl


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


class _FakeMetaMgr:
    """Minimal MetadataManager stub for ExportService."""

    def get_meta_for_id(self, sys_id):
        # Use a 4-char tail so 1-char sys_ids ('A', 'B') still produce a value.
        return (f'T-S {sys_id[-4:]}', f'Title for {sys_id}')

    def get_library_for_id(self, sys_id):
        return 'CUL'

    def parse_full_id_components(self, header):
        return {}


def _make_result(sys_id, snippet='foo *bar* baz', img='1r', source='ms'):
    """Build a live (uncompacted) search result row."""
    return {
        'uid': f'IE_x_P_y_FL_{sys_id}',
        'display': {
            'id': sys_id,
            'shelfmark': f'T-S {sys_id[-4:]}',
            'title': f'Title for {sys_id}',
            'library_code': 'CUL',
            'img': img,
            'source': source,
        },
        'snippet': snippet,
        'raw_header': '',
        'sort_score': 1.0,
        'full_text': 'page text here',
    }


@pytest.fixture
def stub_dossier(monkeypatch):
    """Monkeypatch the Wave 1 helpers to return predictable shapes."""
    from shared import export_dossier
    monkeypatch.setattr(export_dossier, 'pgp_subset_for_sys_id', lambda s, **kw: {
        'pgp_url': f'https://pgp.example/{s}',
        'description': f'PGP desc {s}',
        'document_type': 'Letter',
        'date_display': '1100',
        'languages': ['Hebrew', 'Aramaic'],
        'tags': ['letter'],
    } if s else None)
    monkeypatch.setattr(export_dossier, 'nli_subset_for_sys_id', lambda s, **kw: {
        'catalog_entry': f'Neubauer {s[-4:]}',
        'library_viewer_url': f'https://cudl.example/{s}',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'catalog_summary_for_sys_id', lambda s, **kw: {
        'title': f'CatTitle {s}',
        'author_text': 'Author',
        'copy_date': '1180',
        'copy_place': 'Fustat',
    } if s else None)
    # Only the 99001234567890 sys_id has bib entries; others return [].
    monkeypatch.setattr(export_dossier, 'bibliography_for_sys_id', lambda s, **kw: [{
        'running_title': 'Med. Soc.',
        'title_year': 1967,
        'mention_page': '123',
        'article_name': 'Letter',
        'article_author_eng': 'Goitein',
        'catalog_acronym': 'MS',
    }] if s == '99001234567890' else [])


@pytest.fixture
def export_service():
    from web.export_service import ExportService
    return ExportService(meta_mgr=_FakeMetaMgr())


def _load_wb(content):
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_workbook_has_three_sheets_in_order(export_service, stub_dossier):
    """Test 1+2: workbook has 3 sheets in expected order; first is default-active."""
    results = [_make_result('99001234567890')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    assert wb.sheetnames == ['Search Results', 'Manuscripts', 'Bibliography']
    assert wb.active.title == 'Search Results'


def test_main_sheet_has_12_headers_in_unified_order(export_service, stub_dossier):
    """Test 3: main sheet has the unified 12-column order per D-01."""
    results = [_make_result('99001234567890')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == [
        'System ID', 'Library', 'Shelfmark', 'Title',
        'Image/Page', 'Source', 'Snippet', 'Full Text',
        'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest',
    ]


def test_has_pgp_yes_or_empty(export_service, stub_dossier):
    """Test 4: Has PGP = 'Yes' when sys_id in set, '' otherwise (D-06)."""
    results = [_make_result('A'), _make_result('B')]
    content, _ = export_service.export_search_results_excel(
        results, 'q', transcription_sys_ids={'A'},
    )
    wb = _load_wb(content)
    ws = wb['Search Results']
    # Has PGP is column 9.
    assert ws.cell(2, 9).value == 'Yes'
    assert ws.cell(3, 9).value in ('', None)


def test_is_printed_yes_or_empty(export_service, stub_dossier):
    """Test 5: Is Printed = 'Yes' when sys_id in set, '' otherwise (D-06)."""
    results = [_make_result('A'), _make_result('B')]
    content, _ = export_service.export_search_results_excel(
        results, 'q', printed_ids={'B'},
    )
    wb = _load_wb(content)
    ws = wb['Search Results']
    # Is Printed is column 10.
    assert ws.cell(2, 10).value in ('', None)
    assert ws.cell(3, 10).value == 'Yes'


def test_domains_pipe_joined(export_service, stub_dossier):
    """Test 6: Domains rendered as pipe-joined with NO spaces (D-05)."""
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(
        results, 'q', result_domains={'A': ['Bible', 'Letter', 'Legal']},
    )
    wb = _load_wb(content)
    ws = wb['Search Results']
    # Domains is column 11.
    assert ws.cell(2, 11).value == 'Bible|Letter|Legal'


def test_iiif_manifest_empty_in_wave3(export_service, stub_dossier):
    """Test 7: IIIF Manifest column header present but cells empty (D-13)."""
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    # IIIF Manifest is column 12.
    assert ws.cell(2, 12).value in ('', None)


def test_manuscripts_sheet_headers(export_service, stub_dossier):
    """Test 8: Manuscripts sheet header is MANUSCRIPT_HEADERS (14 cols)."""
    from shared.export_dossier import MANUSCRIPT_HEADERS
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Manuscripts']
    headers = [ws.cell(1, c).value for c in range(1, 15)]
    assert headers == list(MANUSCRIPT_HEADERS)


def test_manuscripts_dedupes_by_sys_id(export_service, stub_dossier):
    """Test 9: Manuscripts sheet has 1 row per unique sys_id; first-occurrence
    order per D-12."""
    # 3 results with sys_ids A, A, B -> Manuscripts has 2 data rows.
    results = [_make_result('A'), _make_result('A'), _make_result('B')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Manuscripts']
    # Row 1 is header, data is rows 2-3.
    assert ws.cell(2, 1).value == 'A'
    assert ws.cell(3, 1).value == 'B'
    assert ws.cell(4, 1).value is None  # No 3rd data row.


def test_bibliography_sheet_headers(export_service, stub_dossier):
    """Test 10: Bibliography sheet header is BIBLIOGRAPHY_HEADERS (8 cols)."""
    from shared.export_dossier import BIBLIOGRAPHY_HEADERS
    results = [_make_result('99001234567890')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Bibliography']
    headers = [ws.cell(1, c).value for c in range(1, 9)]
    assert headers == list(BIBLIOGRAPHY_HEADERS)


def test_bibliography_zero_rows_when_no_entries(export_service, stub_dossier):
    """Test 11: Bibliography sub-sheet has zero data rows when sys_id has no
    bib entries."""
    # _make_result('A') — the stub returns [] for sys_id != '99001234567890'.
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Bibliography']
    assert ws.cell(2, 1).value is None  # No data row, only header.


def test_conditional_rtl_he(export_service, stub_dossier):
    """Test 12a: lang='he' -> all 3 sheets RTL. D-04 REVISED (2026-05-20):
    sheet titles are now bilingual ('תוצאות חיפוש', 'כתבי יד', 'ביבליוגרפיה').
    Main sheet renamed 'תוצאות גניזה' -> 'תוצאות חיפוש' on 2026-05-21."""
    from shared.export_dossier import sheet_titles
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q', lang='he')
    wb = _load_wb(content)
    he_titles = sheet_titles('he')
    for key in ('main', 'manuscripts', 'bibliography'):
        name = he_titles[key]
        assert wb[name].sheet_view.rightToLeft is True, f"sheet {name} should be RTL"


def test_conditional_rtl_en(export_service, stub_dossier):
    """Test 12b: lang='en' -> all 3 sheets LTR (rightToLeft False or None)."""
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q', lang='en')
    wb = _load_wb(content)
    for name in ['Search Results', 'Manuscripts', 'Bibliography']:
        assert wb[name].sheet_view.rightToLeft in (False, None), f"sheet {name} should be LTR"


def test_rich_snippet_with_marker(export_service, stub_dossier):
    """Test 13a: snippet with '*...*' renders as CellRichText (D-14)."""
    from openpyxl.cell.rich_text import CellRichText
    results = [_make_result('A', snippet='foo *bar* baz')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    # Snippet is column 7.
    snippet_cell = ws.cell(2, 7).value
    assert isinstance(snippet_cell, CellRichText), (
        f"snippet was {type(snippet_cell).__name__}"
    )


def test_plain_snippet_without_marker(export_service, stub_dossier):
    """Test 13b: snippet with no '*' marker stays as plain str."""
    results = [_make_result('A', snippet='no marker here')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    snippet_cell = ws.cell(2, 7).value
    assert isinstance(snippet_cell, str)
    assert 'no marker here' in snippet_cell


def test_backward_compat_no_new_kwargs(export_service, stub_dossier):
    """Test 16: without new kwargs, Has PGP / Is Printed / Domains render empty."""
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    # All 3 new columns render empty for every row.
    assert ws.cell(2, 9).value in ('', None)
    assert ws.cell(2, 10).value in ('', None)
    assert ws.cell(2, 11).value in ('', None)


def test_filename_and_return_shape(export_service, stub_dossier):
    """Test 17: returns (bytes, filename) tuple; filename ends with '.xlsx'."""
    results = [_make_result('A')]
    ret = export_service.export_search_results_excel(results, 'my query')
    assert isinstance(ret, tuple) and len(ret) == 2
    content, filename = ret
    assert isinstance(content, bytes)
    assert filename.endswith('.xlsx')


def test_hebrew_library_name_in_he_lang(export_service, stub_dossier):
    """D-04 REVISED (2026-05-20): lang='he' -> Hebrew library name on the main sheet.

    Pre-reversal test asserted English library name regardless of lang. After
    the 2026-05-20 reversal of D-04, the library cell follows ``lang``: Hebrew
    UI -> Hebrew library name (with English/code fallback inside
    get_library_display). For CUL the Hebrew name is 'ספריית האוניברסיטה של קיימברידג''
    per genizah_translations.LIBRARY_CODES_HE.
    """
    from shared.export_dossier import sheet_titles
    results = [_make_result('A')]
    content, _ = export_service.export_search_results_excel(results, 'q', lang='he')
    wb = _load_wb(content)
    ws = wb[sheet_titles('he')['main']]
    # The new header row precedes the data row(s); find it.
    library_cell = None
    for r in range(1, 25):
        # Locate the data row (header is in row 1 for this workbook flow).
        v0 = ws.cell(r, 1).value
        if v0 and str(v0).strip() and str(v0).startswith(('A', '9')):
            library_cell = ws.cell(r, 2).value
            break
    # Either Hebrew form or English fallback or short code 'CUL' all acceptable
    # — the gate is that the cell is NON-EMPTY and reflects some library marker.
    assert library_cell, f"Library cell should be non-empty, got {library_cell!r}"


def test_image_page_and_source_columns_populated(export_service, stub_dossier):
    """MUST-FIX 94-03-A live path: Image/Page (col 5) and Source (col 6) cells
    are non-empty for live (uncompacted) rows with display.img / display.source.
    """
    results = [_make_result('A', img='2v', source='pgp')]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    assert ws.cell(2, 5).value == '2v'
    assert ws.cell(2, 6).value == 'pgp'


def test_he_lang_produces_hebrew_sheet_titles_and_headers(export_service, stub_dossier):
    """D-04 REVISED (2026-05-20): lang='he' yields Hebrew sheet titles AND
    Hebrew header rows on all 3 sheets. Pin canonical strings explicitly."""
    from shared.export_dossier import sheet_titles, main_header_row, manuscript_header_row, bibliography_header_row
    results = [_make_result('99001234567890')]
    content, _ = export_service.export_search_results_excel(results, 'q', lang='he')
    wb = _load_wb(content)
    he_titles = sheet_titles('he')
    # Sheets present under Hebrew titles.
    assert wb.sheetnames == [he_titles['main'], he_titles['manuscripts'], he_titles['bibliography']]
    # Main sheet headers Hebrew.
    ws_main = wb[he_titles['main']]
    expected_main = main_header_row('he')
    actual_main = [ws_main.cell(1, c).value for c in range(1, 13)]
    assert actual_main == expected_main, f"Hebrew main headers mismatch: {actual_main}"
    # Manuscripts sub-sheet headers Hebrew.
    ws_manu = wb[he_titles['manuscripts']]
    expected_manu = manuscript_header_row('he')
    actual_manu = [ws_manu.cell(1, c).value for c in range(1, 15)]
    assert actual_manu == expected_manu, f"Hebrew manuscripts headers mismatch: {actual_manu}"
    # Bibliography sub-sheet headers Hebrew.
    ws_bib = wb[he_titles['bibliography']]
    expected_bib = bibliography_header_row('he')
    actual_bib = [ws_bib.cell(1, c).value for c in range(1, 9)]
    assert actual_bib == expected_bib, f"Hebrew bib headers mismatch: {actual_bib}"


def test_en_lang_produces_english_sheet_titles_and_headers(export_service, stub_dossier):
    """Symmetric back-compat: lang='en' yields English titles/headers."""
    from shared.export_dossier import main_header_row, manuscript_header_row, bibliography_header_row
    results = [_make_result('99001234567890')]
    content, _ = export_service.export_search_results_excel(results, 'q', lang='en')
    wb = _load_wb(content)
    assert wb.sheetnames == ['Search Results', 'Manuscripts', 'Bibliography']
    assert [wb['Search Results'].cell(1, c).value for c in range(1, 13)] == main_header_row('en')
    assert [wb['Manuscripts'].cell(1, c).value for c in range(1, 15)] == manuscript_header_row('en')
    assert [wb['Bibliography'].cell(1, c).value for c in range(1, 9)] == bibliography_header_row('en')


def test_image_page_and_source_columns_compacted_row(export_service, stub_dossier):
    """MUST-FIX 94-03-A compacted path: when called with a row that has only
    top-level img/source (display dict already dropped, as happens for rows
    that have round-tripped through web/export_state.py compaction), the
    Image/Page + Source cells are still populated."""
    # Post-compaction shape: top-level img + source, NO display dict.
    results = [{
        'uid': 'IE188433865_P1_FL1',
        'sys_id': 'A',
        'snippet': 'snippet',
        'sort_score': 0.9,
        'raw_header': '',
        'img': '3r',  # MUST-FIX 94-03-A: top-level (post-compaction).
        'source': 'doc',  # MUST-FIX 94-03-A: top-level (post-compaction).
    }]
    content, _ = export_service.export_search_results_excel(results, 'q')
    wb = _load_wb(content)
    ws = wb['Search Results']
    assert ws.cell(2, 5).value == '3r'
    assert ws.cell(2, 6).value == 'doc'

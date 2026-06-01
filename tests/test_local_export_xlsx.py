"""Phase 103 LEXP-03/04/05/07: Local Documents xlsx sheet tests.

Offline tests for the desktop xlsx-write logic — exercise
:func:`genizah_app._build_search_results_xlsx_bytes` directly (no Qt deps)
with fake meta_resolver + identity sanitize + stubbed dossier helpers,
mirroring the pattern in :file:`tests/test_desktop_xlsx_multi_sheet.py`.

Covers:
- Mixed (Genizah + LOCAL) → 5-sheet workbook with Local Documents at position 4
- LOCAL rows excluded from Search Results / Manuscripts / Bibliography
- LOCAL-only → EXACTLY [Local Documents, Credits and Info] (empty Genizah sheets ABSENT)
- Bilingual (he/en) sheet names and headers
- D-14 display.source PRIMARY discriminator over 97-prefix
- Genizah-only → unchanged 4-sheet workbook (non-regression)
- Formula-injection safety on Filepath cells
- Missing filepath → blank cells, no exception
"""
from io import BytesIO

import openpyxl
import pytest

from shared.local_sys_id import is_local_sys_id
from shared_export_utils import sanitize_text_for_excel

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

# A valid LOCAL sys_id: 18 digits, all-digit, '97'-prefixed.
LOCAL_ID = '970000000100000001'
# A Genizah sys_id: 15 digits — not LOCAL.
GEN_ID = '990012345678901'


def test_local_sys_id_guard():
    """Guard: confirm our test IDs have the expected LOCAL/non-LOCAL classification."""
    assert is_local_sys_id(LOCAL_ID) is True
    assert is_local_sys_id(GEN_ID) is False


def _local_result(sid, fname, snippet, locator):
    """Build a minimal LOCAL result dict."""
    return {
        'display': {'id': sid, 'source': 'LOCAL', 'shelfmark': fname},
        'sys_id': sid,
        'raw_file_hl': snippet,
        'chunk_locator': locator,
        'p_num': '3',
    }


def _gen_result(sid):
    """Build a minimal Genizah result dict."""
    return {
        'display': {
            'id': sid,
            'source': '',
            'shelfmark': 'T-S 12.1',
            'title': 'Letter',
            'img': '5',
        },
        'sys_id': sid,
        'raw_file_hl': 'gen *hit*',
    }


# Use a path with a clear parent folder 'MyDocs'.
local_filepath_map = {LOCAL_ID: r'C:\Users\me\MyDocs\letter.pdf'}


def _meta_resolver_fake(sid):
    if not sid:
        return None
    return {
        'shelfmark': f'T-S {sid[-4:]}',
        'title': f'Title {sid}',
        'library_code': 'CUL',
        'library_name': 'Cambridge University Library',
    }


def _identity_sanitize(text):
    return '' if text is None else str(text)


@pytest.fixture
def stub_dossier(monkeypatch):
    from shared import export_dossier
    monkeypatch.setattr(export_dossier, 'pgp_subset_for_sys_id', lambda s, **kw: None)
    monkeypatch.setattr(export_dossier, 'nli_subset_for_sys_id', lambda s, **kw: None)
    monkeypatch.setattr(export_dossier, 'catalog_summary_for_sys_id', lambda s, **kw: None)
    monkeypatch.setattr(export_dossier, 'bibliography_for_sys_id', lambda s, **kw: [])


def _build(results, **kwargs):
    from genizah_app import _build_search_results_xlsx_bytes
    return _build_search_results_xlsx_bytes(
        results=results,
        headers_main=[
            'System ID', 'Library', 'Shelfmark', 'Title',
            'Image/Page', 'Source', 'Snippet', 'Full Text',
            'Has PGP', 'Is Printed', 'Domains', 'Image URL',
        ],
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        **kwargs,
    )


def _wb(results, **kwargs):
    content = _build(results, **kwargs)
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


# ---------------------------------------------------------------------------
# Mixed (Genizah + LOCAL) tests
# ---------------------------------------------------------------------------

class TestMixedExport:

    def test_mixed_has_local_documents_sheet(self, stub_dossier):
        wb = _wb(
            [_gen_result(GEN_ID), _local_result(LOCAL_ID, 'letter.pdf', 'a *b* c', 'p. 3')],
            local_filepath_map=local_filepath_map,
        )
        assert wb.sheetnames == [
            'Search Results', 'Manuscripts', 'Bibliography',
            'Local Documents', 'Credits and Info',
        ]

    def test_mixed_local_row_fields(self, stub_dossier):
        wb = _wb(
            [_gen_result(GEN_ID), _local_result(LOCAL_ID, 'letter.pdf', 'a *b* c', 'p. 3')],
            local_filepath_map=local_filepath_map,
        )
        ws = wb['Local Documents']
        # Row 2: filename / parent folder / filepath / page / matched text
        assert ws.cell(row=2, column=1).value == 'letter.pdf', "Filename mismatch"
        assert ws.cell(row=2, column=2).value == 'MyDocs', "Parent folder mismatch"
        fp_val = ws.cell(row=2, column=3).value
        assert fp_val and 'letter.pdf' in str(fp_val), "Filepath mismatch"
        assert ws.cell(row=2, column=4).value == 'p. 3', "Page mismatch"
        matched = ws.cell(row=2, column=5).value
        assert matched is not None and matched != '', "Matched text should be non-empty"

    def test_mixed_search_results_excludes_local(self, stub_dossier):
        wb = _wb(
            [_gen_result(GEN_ID), _local_result(LOCAL_ID, 'letter.pdf', 'a *b* c', 'p. 3')],
            local_filepath_map=local_filepath_map,
        )
        ws = wb['Search Results']
        # Collect all data rows (row 1 is header)
        data_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert GEN_ID in data_ids, "Genizah result should appear on Search Results"
        assert LOCAL_ID not in data_ids, "LOCAL result must NOT appear on Search Results"
        # Exactly 1 data row
        non_empty = [v for v in data_ids if v]
        assert len(non_empty) == 1

    def test_mixed_active_is_search_results(self, stub_dossier):
        wb = _wb(
            [_gen_result(GEN_ID), _local_result(LOCAL_ID, 'letter.pdf', 'a *b* c', 'p. 3')],
            local_filepath_map=local_filepath_map,
        )
        assert wb.active.title == 'Search Results'

    def test_manuscripts_bibliography_exclude_local(self, stub_dossier):
        wb = _wb(
            [_gen_result(GEN_ID), _local_result(LOCAL_ID, 'letter.pdf', 'a *b* c', 'p. 3')],
            local_filepath_map=local_filepath_map,
        )
        for sheet_name in ('Manuscripts', 'Bibliography'):
            ws = wb[sheet_name]
            col_a_vals = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
            assert LOCAL_ID not in col_a_vals, (
                f"LOCAL_ID must NOT appear in column A of {sheet_name}"
            )


# ---------------------------------------------------------------------------
# LOCAL-only tests
# ---------------------------------------------------------------------------

class TestLocalOnlyExport:

    def test_local_only_workbook_shape_exact(self, stub_dossier):
        """D-05: LOCAL-only export must be EXACTLY [Local Documents, Credits and Info]."""
        wb = _wb(
            [_local_result(LOCAL_ID, 'a.pdf', 'x *y*', 'p. 1')],
            local_filepath_map=local_filepath_map,
        )
        assert wb.sheetnames == ['Local Documents', 'Credits and Info']
        assert 'Search Results' not in wb.sheetnames
        assert 'Manuscripts' not in wb.sheetnames
        assert 'Bibliography' not in wb.sheetnames
        assert wb.active.title == 'Local Documents'

    def test_local_only_workbook_shape_exact_he(self, stub_dossier):
        """D-05 + LEXP-07: LOCAL-only HE export has bilingual sheet names."""
        wb = _wb(
            [_local_result(LOCAL_ID, 'a.pdf', 'x *y*', 'p. 1')],
            local_filepath_map=local_filepath_map,
            lang='he',
        )
        assert wb.sheetnames == ['מסמכים מקומיים', 'קרדיט ומידע']
        assert 'Search Results' not in wb.sheetnames
        assert 'Manuscripts' not in wb.sheetnames
        assert 'Bibliography' not in wb.sheetnames
        assert wb.active.title == 'מסמכים מקומיים'

    def test_local_only_partition_by_display_source(self, stub_dossier):
        """D-14: display.source == 'LOCAL' is PRIMARY over 97-prefix.

        A result with GEN_ID (not a 97-prefix) but display['source']=='LOCAL'
        must be treated as a LOCAL row.
        """
        # Use GEN_ID but mark source as LOCAL
        result = _gen_result(GEN_ID)
        result['display']['source'] = 'LOCAL'
        # Provide filepath map keyed to GEN_ID
        fp_map = {GEN_ID: r'C:\Users\me\MyDocs\file.pdf'}
        wb = _wb([result], local_filepath_map=fp_map)
        # Should be treated as LOCAL-only → [Local Documents, Credits and Info]
        assert wb.sheetnames == ['Local Documents', 'Credits and Info']
        # The row should appear on Local Documents
        ws = wb['Local Documents']
        assert ws.max_row >= 2, "LOCAL row should appear on Local Documents sheet"


# ---------------------------------------------------------------------------
# Genizah-only (non-regression) tests
# ---------------------------------------------------------------------------

class TestGenizahOnlyExport:

    def test_genizah_only_unchanged(self, stub_dossier):
        """LEXP-08 xlsx: Genizah-only export produces the unchanged 4-sheet workbook."""
        wb = _wb([_gen_result(GEN_ID)])
        assert wb.sheetnames == [
            'Search Results', 'Manuscripts', 'Bibliography', 'Credits and Info',
        ]
        assert 'Local Documents' not in wb.sheetnames


# ---------------------------------------------------------------------------
# Bilingual headers
# ---------------------------------------------------------------------------

class TestBilingualHeaders:

    def test_local_he_headers(self, stub_dossier):
        """LEXP-07: HE sheet name and column headers for a mixed export."""
        wb = _wb(
            [_gen_result(GEN_ID), _local_result(LOCAL_ID, 'letter.pdf', 'a *b* c', 'p. 3')],
            local_filepath_map=local_filepath_map,
            lang='he',
        )
        assert 'מסמכים מקומיים' in wb.sheetnames
        ws = wb['מסמכים מקומיים']
        header_row = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert header_row == ['שם קובץ', 'תיקייה', 'נתיב מלא', 'עמוד', 'טקסט תואם']


# ---------------------------------------------------------------------------
# Security: formula injection safety
# ---------------------------------------------------------------------------

class TestFormulaSafety:

    def test_formula_injection_filepath_escaped(self, stub_dossier):
        """T-103-04: a filepath beginning with '=' must be escaped by sanitize_fn."""
        from genizah_app import _build_search_results_xlsx_bytes
        injected_id = LOCAL_ID
        injected_filepath_map = {injected_id: '=cmd|calc'}
        content = _build_search_results_xlsx_bytes(
            results=[_local_result(injected_id, 'evil.pdf', 'snippet', 'p. 1')],
            headers_main=[
                'System ID', 'Library', 'Shelfmark', 'Title',
                'Image/Page', 'Source', 'Snippet', 'Full Text',
                'Has PGP', 'Is Printed', 'Domains', 'Image URL',
            ],
            meta_resolver=_meta_resolver_fake,
            sanitize_fn=sanitize_text_for_excel,
            local_filepath_map=injected_filepath_map,
        )
        wb = openpyxl.load_workbook(BytesIO(content), rich_text=True)
        ws = wb['Local Documents']
        fp_cell = ws.cell(row=2, column=3).value
        # sanitize_text_for_excel prefixes leading '=' with "'" per shared_export_utils:51-55
        assert str(fp_cell).startswith("'"), (
            f"Formula injection not escaped: Filepath cell = {fp_cell!r}"
        )


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:

    def test_missing_filepath_blank_no_error(self, stub_dossier):
        """A LOCAL row whose sys_id is NOT in local_filepath_map: blank cells, no exception."""
        result = _local_result(LOCAL_ID, 'missing.pdf', 'some snippet', 'p. 2')
        # Pass an empty map — LOCAL_ID has no filepath entry
        wb = _wb([result], local_filepath_map={})
        ws = wb['Local Documents']
        fp_val = ws.cell(row=2, column=3).value  # Filepath
        parent_val = ws.cell(row=2, column=2).value  # Parent Folder
        assert fp_val in (None, ''), f"Filepath should be blank, got {fp_val!r}"
        assert parent_val in (None, ''), f"Parent Folder should be blank, got {parent_val!r}"

    def test_no_local_filepath_map_kwarg_unchanged(self, stub_dossier):
        """Calling without local_filepath_map (old 4-arg style) works unchanged."""
        wb = _wb([_gen_result(GEN_ID)])
        assert wb.sheetnames == [
            'Search Results', 'Manuscripts', 'Bibliography', 'Credits and Info',
        ]

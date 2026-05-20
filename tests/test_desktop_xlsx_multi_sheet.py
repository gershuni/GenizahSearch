"""Phase 94 EXPORT-META-09: desktop xlsx 3-sheet structure parity.

Offline tests for the desktop xlsx-write logic. The xlsx-write logic is
factored into a module-level helper :func:`genizah_app._build_search_results_xlsx_bytes`
that returns bytes (no Qt UI dependencies); these tests exercise it directly
with fake meta_resolver + sanitize_fn + stubbed dossier helpers.

Wave 4 Task 1 — see :file:`.planning/phases/94-adding-pgp-to-downloaded-data/94-04-PLAN.md`.
"""
from io import BytesIO

import openpyxl
import pytest


def _meta_resolver_fake(sid):
    if not sid:
        return None
    return {
        'shelfmark': f'T-S {sid[-4:]}',
        'title': f'Title {sid}',
        'library_code': 'CUL',
        'library_name': 'Cambridge University Library',
    }


def _identity_sanitize(text):
    # Tests don't need real sanitization; pass through.
    return '' if text is None else str(text)


def _make_result(sys_id, snippet='foo *bar* baz', img='1r', source='ms',
                 full_text='page text'):
    return {
        'display': {
            'id': sys_id, 'shelfmark': f'T-S {sys_id[-4:]}',
            'title': f'Title {sys_id}', 'library_code': 'CUL',
            'img': img, 'source': source,
        },
        'snippet': snippet,
        'raw_file_hl': snippet,
        'full_text': full_text,
    }


@pytest.fixture
def stub_dossier(monkeypatch):
    from shared import export_dossier
    monkeypatch.setattr(export_dossier, 'pgp_subset_for_sys_id', lambda s, **kw: {
        'pgp_url': f'https://pgp.example/{s}',
        'description': f'PGP desc {s}', 'document_type': 'Letter',
        'date_display': '1100', 'languages': ['Hebrew', 'Aramaic'],
        'tags': ['letter'],
    } if s else None)
    monkeypatch.setattr(export_dossier, 'nli_subset_for_sys_id', lambda s, **kw: {
        'catalog_entry': f'Neubauer {s[-4:]}',
        'library_viewer_url': f'https://cudl.example/{s}',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'catalog_summary_for_sys_id', lambda s, **kw: {
        'title': f'CatTitle {s}', 'author_text': 'Author',
        'copy_date': '1180', 'copy_place': 'Fustat',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'bibliography_for_sys_id', lambda s, **kw: [{
        'running_title': 'Med. Soc.', 'title_year': 1967, 'mention_page': '123',
        'article_name': 'Letter', 'article_author_eng': 'Goitein',
        'catalog_acronym': 'MS',
    }] if s == '99001234567890' else [])


def _build(results, **kwargs):
    from genizah_app import _build_search_results_xlsx_bytes
    return _build_search_results_xlsx_bytes(
        results=results,
        headers_main=['System ID', 'Library', 'Shelfmark', 'Title',
                      'Image/Page', 'Source', 'Snippet', 'Full Text',
                      'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest'],
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        **kwargs,
    )


def _load(content):
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


def _find_header_row(ws):
    for r in range(1, 25):
        if ws.cell(r, 1).value == 'System ID':
            return r
    return None


def test_workbook_has_4_sheets_in_order(stub_dossier):
    # MUST-FIX 94-04-A: sheet name is ENGLISH-LOCKED 'Search Results' on
    # both apps (matches web for EXPORT-META-09 parity).
    # Smoke verification round 2 (2026-05-21): a 4th 'Credits and Info'
    # sheet was added; main sheet is still default-active.
    content = _build([_make_result('99001234567890')])
    wb = _load(content)
    assert wb.sheetnames == ['Search Results', 'Manuscripts', 'Bibliography', 'Credits and Info']
    assert wb.active.title == 'Search Results'


def test_main_sheet_12_columns_unified_order(stub_dossier):
    content = _build([_make_result('A')])
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    assert header_row is not None
    headers = [ws.cell(header_row, c).value for c in range(1, 13)]
    assert headers == [
        'System ID', 'Library', 'Shelfmark', 'Title',
        'Image/Page', 'Source', 'Snippet', 'Full Text',
        'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest',
    ]


def test_has_pgp_and_is_printed_yes_or_empty(stub_dossier):
    content = _build(
        [_make_result('A'), _make_result('B')],
        transcription_sys_ids={'A'},
        printed_ids={'B'},
    )
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    # Row A is header_row + 1, Row B is header_row + 2
    assert ws.cell(header_row + 1, 9).value == 'Yes'
    assert ws.cell(header_row + 1, 10).value in ('', None)
    assert ws.cell(header_row + 2, 9).value in ('', None)
    assert ws.cell(header_row + 2, 10).value == 'Yes'


def test_domains_pipe_joined(stub_dossier):
    content = _build(
        [_make_result('A')],
        result_domains={'A': ['Bible', 'Letter', 'Legal']},
    )
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    assert ws.cell(header_row + 1, 11).value == 'Bible|Letter|Legal'


def test_iiif_manifest_empty(stub_dossier):
    content = _build([_make_result('A')])
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    assert ws.cell(header_row + 1, 12).value in ('', None)


def test_full_text_column(stub_dossier):
    content = _build([_make_result('A', full_text='specific page text')])
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    assert ws.cell(header_row + 1, 8).value == 'specific page text'


def test_full_text_excerpt_fallback(stub_dossier):
    # When 'full_text' missing but 'full_text_excerpt' present, the excerpt is used.
    res = _make_result('A', full_text='')
    res.pop('full_text', None)
    res['full_text_excerpt'] = 'excerpt only'
    content = _build([res])
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    assert ws.cell(header_row + 1, 8).value == 'excerpt only'


def test_manuscripts_dedupe(stub_dossier):
    content = _build([_make_result('A'), _make_result('A'), _make_result('B')])
    wb = _load(content)
    ws = wb['Manuscripts']
    assert ws.cell(2, 1).value == 'A'
    assert ws.cell(3, 1).value == 'B'
    assert ws.cell(4, 1).value is None


def test_bibliography_zero_rows_when_empty(stub_dossier):
    # stub returns [] for sys_id != '99001234567890'
    content = _build([_make_result('A')])
    wb = _load(content)
    ws = wb['Bibliography']
    assert ws.cell(2, 1).value is None


def test_bibliography_has_row_when_entries(stub_dossier):
    content = _build([_make_result('99001234567890')])
    wb = _load(content)
    ws = wb['Bibliography']
    assert ws.cell(2, 1).value == '99001234567890'
    assert ws.cell(2, 3).value == 'Goitein'


def test_conditional_rtl_he(stub_dossier):
    """D-04 REVISED (2026-05-20): sheet titles are bilingual when lang='he'."""
    from shared.export_dossier import sheet_titles
    content = _build([_make_result('A')], lang='he')
    wb = _load(content)
    he_titles = sheet_titles('he')
    for key in ('main', 'manuscripts', 'bibliography'):
        name = he_titles[key]
        assert wb[name].sheet_view.rightToLeft is True


def test_conditional_rtl_en(stub_dossier):
    content = _build([_make_result('A')], lang='en')
    wb = _load(content)
    for name in ['Search Results', 'Manuscripts', 'Bibliography']:
        assert wb[name].sheet_view.rightToLeft in (False, None)


def test_rich_snippet_renders(stub_dossier):
    from openpyxl.cell.rich_text import CellRichText
    content = _build([_make_result('A', snippet='foo *bar* baz')])
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    snippet_cell = ws.cell(header_row + 1, 7).value
    assert isinstance(snippet_cell, CellRichText)


def test_main_headers_english_when_caller_pins_english_list(stub_dossier):
    # D-04 REVISED (2026-05-20): MUST-FIX 94-04-B is SUPERSEDED. The helper
    # accepts an explicit ``headers_main`` kwarg for back-compat; when the
    # caller passes a literal English list (as the parity-test fixture does),
    # the resulting workbook reflects exactly those literals regardless of
    # ``lang``. The new bilingual default behavior is covered by
    # ``test_main_headers_hebrew_when_lang_he_and_no_explicit_headers``.
    content = _build([_make_result('A')], lang='he')
    wb = _load(content)
    # In Hebrew lang the sheet title is now Hebrew.
    from shared.export_dossier import sheet_titles
    ws = wb[sheet_titles('he')['main']]
    header_row = _find_header_row(ws)
    assert header_row is not None
    headers = [ws.cell(header_row, c).value for c in range(1, 13)]
    assert headers == [
        'System ID', 'Library', 'Shelfmark', 'Title',
        'Image/Page', 'Source', 'Snippet', 'Full Text',
        'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest',
    ], f"headers were {headers}"


def test_full_text_fetcher_hydrates_when_missing(stub_dossier):
    # MUST-FIX 94-04-D: when the result row lacks full_text /
    # full_text_excerpt, the helper invokes full_text_fetcher(uid).
    from genizah_app import _build_search_results_xlsx_bytes
    results = [{
        'uid': 'IE_x_P_y_FL_z',
        'display': {'id': 'A', 'shelfmark': 'T-S 1', 'title': 'T'},
        'snippet': '',
        'raw_file_hl': '',
        # No 'full_text' or 'full_text_excerpt' keys (PGP tag row case).
    }]
    content = _build_search_results_xlsx_bytes(
        results=results,
        headers_main=['System ID', 'Library', 'Shelfmark', 'Title',
                      'Image/Page', 'Source', 'Snippet', 'Full Text',
                      'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest'],
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        full_text_fetcher=lambda uid: 'hydrated text for ' + uid,
    )
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    # Full Text column is col 8.
    cell = ws.cell(header_row + 1, 8).value
    assert 'hydrated text' in (cell or ''), f"expected hydrated text, got {cell!r}"


def test_full_text_fetcher_none_means_empty_cell(stub_dossier):
    # MUST-FIX 94-04-D: when fetcher is None and stored text is empty,
    # Full Text cell renders as empty (no exception).
    from genizah_app import _build_search_results_xlsx_bytes
    results = [{
        'uid': 'IE_x_P_y_FL_z',
        'display': {'id': 'A', 'shelfmark': 'T-S 1', 'title': 'T'},
        'snippet': '',
    }]
    content = _build_search_results_xlsx_bytes(
        results=results,
        headers_main=['System ID', 'Library', 'Shelfmark', 'Title',
                      'Image/Page', 'Source', 'Snippet', 'Full Text',
                      'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest'],
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        full_text_fetcher=None,
    )
    wb = _load(content)
    ws = wb['Search Results']
    header_row = _find_header_row(ws)
    assert ws.cell(header_row + 1, 8).value in ('', None)


def test_manuscripts_headers_match_shared_constants(stub_dossier):
    from shared.export_dossier import MANUSCRIPT_HEADERS, BIBLIOGRAPHY_HEADERS
    content = _build([_make_result('A')])
    wb = _load(content)
    manu_headers = [wb['Manuscripts'].cell(1, c).value for c in range(1, 15)]
    bib_headers = [wb['Bibliography'].cell(1, c).value for c in range(1, 9)]
    assert manu_headers == list(MANUSCRIPT_HEADERS)
    assert bib_headers == list(BIBLIOGRAPHY_HEADERS)


def test_he_lang_produces_hebrew_sheet_titles_and_headers(stub_dossier):
    """D-04 REVISED (2026-05-20): when no explicit headers_main is passed
    and lang='he', the desktop helper produces Hebrew sheet titles AND
    Hebrew header rows on all sheets. Smoke verification round 2
    (2026-05-21): 4-sheet expectation (added Credits and Info).
    """
    from genizah_app import _build_search_results_xlsx_bytes
    from shared.export_dossier import sheet_titles, main_header_row, manuscript_header_row, bibliography_header_row
    content = _build_search_results_xlsx_bytes(
        results=[_make_result('99001234567890')],
        # headers_main intentionally omitted -> defaults to main_header_row(lang)
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        lang='he',
    )
    wb = _load(content)
    he_titles = sheet_titles('he')
    assert wb.sheetnames == [
        he_titles['main'], he_titles['manuscripts'],
        he_titles['bibliography'], he_titles['credits_info'],
    ]
    ws_main = wb[he_titles['main']]
    header_row = _find_header_row_he(ws_main)
    expected_main = main_header_row('he')
    actual_main = [ws_main.cell(header_row, c).value for c in range(1, 13)]
    assert actual_main == expected_main, f"Hebrew main headers mismatch: {actual_main}"
    ws_manu = wb[he_titles['manuscripts']]
    assert [ws_manu.cell(1, c).value for c in range(1, 15)] == manuscript_header_row('he')
    ws_bib = wb[he_titles['bibliography']]
    assert [ws_bib.cell(1, c).value for c in range(1, 9)] == bibliography_header_row('he')


def _find_header_row_he(ws):
    # The Hebrew main-sheet header starts with 'מספר מערכת' (System ID).
    for r in range(1, 25):
        if ws.cell(r, 1).value == 'מספר מערכת':
            return r
    return None


def test_en_lang_produces_english_sheet_titles_and_headers_default(stub_dossier):
    """Symmetric back-compat for en. When headers_main is omitted, defaults
    to main_header_row('en'); sheet titles stay English.
    Smoke verification round 2 (2026-05-21): 4-sheet expectation."""
    from genizah_app import _build_search_results_xlsx_bytes
    from shared.export_dossier import main_header_row, manuscript_header_row, bibliography_header_row
    content = _build_search_results_xlsx_bytes(
        results=[_make_result('99001234567890')],
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        lang='en',
    )
    wb = _load(content)
    assert wb.sheetnames == ['Search Results', 'Manuscripts', 'Bibliography', 'Credits and Info']
    ws_main = wb['Search Results']
    header_row = _find_header_row(ws_main)
    assert [ws_main.cell(header_row, c).value for c in range(1, 13)] == main_header_row('en')
    assert [wb['Manuscripts'].cell(1, c).value for c in range(1, 15)] == manuscript_header_row('en')
    assert [wb['Bibliography'].cell(1, c).value for c in range(1, 9)] == bibliography_header_row('en')


def test_credit_text_and_search_info_text_ignored_on_main_sheet(stub_dossier):
    """Smoke verification round 2 (2026-05-21): the desktop helper used to
    render ``credit_text`` and ``search_info_text`` above the main-sheet
    header row. Round-2 moved both to the dedicated 'Credits and Info'
    sheet via the new per-field kwargs. The legacy kwargs are still
    accepted (for back-compat with older test fixtures + the desktop
    call site mid-migration) but are IGNORED on the main sheet — the
    header row is now row 1.
    """
    content = _build(
        [_make_result('A')],
        credit_text='Credit line 1\nCredit line 2',
        search_info_text='Query: foo\nMode: text',
    )
    wb = _load(content)
    ws = wb['Search Results']
    # Header row is now row 1 because the credit_text/search_info_text args
    # are IGNORED. Compare the pre-round-2 assertion `header_row > 1`.
    header_row = _find_header_row(ws)
    assert header_row == 1, f"header row should be row 1; got {header_row}"
    # Row 1 should be the header row; 'Credit line 1' must NOT appear on
    # the main sheet anywhere.
    main_cells = [ws.cell(r, 1).value for r in range(1, 10)]
    assert 'Credit line 1' not in main_cells
    # Sanity check — the credits sheet is the canonical home for the
    # credit/search-meta content now.
    assert 'Credits and Info' in wb.sheetnames



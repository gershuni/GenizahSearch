# -*- coding: utf-8 -*-
"""SEED-024 — desktop Joins-Lab search-mode parity + candidate xlsx export.

Headless, QApplication-free coverage of the two module-level pure helpers
(``core_mode_for_join_mode`` mode mapping + ``build_candidate_export_rows`` row
assembly), a real workbook-build integration check (the assembled rows flow
through the shared 4-sheet dossier builder), and source guards pinning the
UI-level changes (the "Search options ▾" dialog removed, the mode selector +
inline options + export wiring present).

The Qt-widget construction / visibility behaviour lives in
``test_join_workbench_construct.py`` (CI-skipped; runs locally under offscreen Qt).
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest

import desktop.join_workbench as jw

_SRC = Path(jw.__file__).read_text(encoding="utf-8")


# --------------------------------------------------------------------------- #
# core_mode_for_join_mode — builder mode key -> core execute_search() string
# --------------------------------------------------------------------------- #


def test_mode_keys_order_matches_web_parity():
    assert jw.JOINS_LAB_MODE_KEYS == ("responsa", "exact", "variants", "fuzzy", "regex")


@pytest.mark.parametrize("mode_key,expected", [
    ("responsa", "exact"),     # structured builder; responsa_options carries the pipeline
    ("exact", "literal"),      # mirrors main-search idx 0
    ("variants", "variants"),
    ("fuzzy", "fuzzy"),
    ("regex", "Regex"),        # core checks capital-R 'Regex'
])
def test_core_mode_for_join_mode(mode_key, expected):
    assert jw.core_mode_for_join_mode(mode_key) == expected


def test_core_mode_for_join_mode_unknown_falls_back_to_exact():
    assert jw.core_mode_for_join_mode("nonsense") == "exact"
    assert jw.core_mode_for_join_mode("") == "exact"


# --------------------------------------------------------------------------- #
# build_candidate_export_rows — Candidate -> shared-dossier result dict
# --------------------------------------------------------------------------- #


def _cand(**kw):
    base = dict(sys_id="990001", page=3, shelfmark="T-S 12.100",
                title="כותרת", snippet="...*מילה*...", full_text="full text",
                uid="990001|3")
    base.update(kw)
    return SimpleNamespace(**base)


def test_build_candidate_export_rows_shape():
    rows = jw.build_candidate_export_rows([_cand()])
    assert len(rows) == 1
    r = rows[0]
    assert r["sys_id"] == "990001"
    assert r["raw_file_hl"] == "...*מילה*..."
    assert r["full_text"] == "full text"
    assert r["uid"] == "990001|3"
    d = r["display"]
    assert d["id"] == "990001"
    assert d["shelfmark"] == "T-S 12.100"
    assert d["title"] == "כותרת"
    assert d["source"] == "Genizah"
    assert d["img"] == 3


def test_build_candidate_export_rows_defaults_and_none_page():
    # Missing shelfmark falls back to sys_id; page None -> empty img cell.
    rows = jw.build_candidate_export_rows([_cand(shelfmark="", page=None, title=None)])
    d = rows[0]["display"]
    assert d["shelfmark"] == "990001"   # sys_id fallback
    assert d["title"] == ""
    assert d["img"] == ""


def test_build_candidate_export_rows_empty():
    assert jw.build_candidate_export_rows([]) == []


# --------------------------------------------------------------------------- #
# Integration — assembled rows build a real 4-sheet workbook
# --------------------------------------------------------------------------- #


def test_candidate_rows_build_real_xlsx():
    """The assembled rows flow through the shared dossier builder and yield a
    valid 4-sheet workbook (catches sanitize_fn / header-shape regressions)."""
    try:
        from io import BytesIO

        import openpyxl

        from genizah_app import _build_search_results_xlsx_bytes
        from shared.export_dossier import main_header_row
    except Exception as exc:  # pragma: no cover - environment dependent
        pytest.skip(f"export deps unavailable: {exc!r}")

    try:
        from shared_export_utils import sanitize_text_for_excel as _sanitize
    except Exception:  # pragma: no cover
        _sanitize = None

    rows = jw.build_candidate_export_rows([
        _cand(sys_id="990001", page=1, shelfmark="T-S 12.100"),
        _cand(sys_id="990002", page=2, shelfmark="T-S 12.101"),
    ])

    def _meta_resolver(sid):
        return {"shelfmark": f"shelf-{sid}", "title": f"title-{sid}",
                "library_code": "CUL", "library_name": "Cambridge"}

    content = _build_search_results_xlsx_bytes(
        results=rows,
        headers_main=main_header_row("en"),
        meta_resolver=_meta_resolver,
        sanitize_fn=_sanitize,
        lang="en",
        search_query="anchor query",
        search_mode="Joins Lab — Responsa-style (✓1 ?0 ✗0)",
    )
    assert isinstance(content, (bytes, bytearray)) and len(content) > 0

    wb = openpyxl.load_workbook(BytesIO(content))
    # Same 4-sheet structure as the main-search export.
    assert "Search Results" in wb.sheetnames
    assert "Manuscripts" in wb.sheetnames
    assert "Bibliography" in wb.sheetnames
    assert "Credits and Info" in wb.sheetnames


# --------------------------------------------------------------------------- #
# Source guards — UI-level changes (no QApplication needed)
# --------------------------------------------------------------------------- #


def test_search_options_dialog_removed():
    assert "def _open_search_options_dialog" not in _SRC
    assert "_btn_search_opts" not in _SRC


def test_mode_selector_and_inline_options_wired():
    assert "def get_mode" in _SRC
    assert "self.mode_combo" in _SRC
    assert "allow_modes" in _SRC
    # inline option checkboxes replaced the dialog
    for name in ("chk_variants", "chk_ja", "chk_flex", "chk_bidir"):
        assert name in _SRC


def test_anchor_builder_enables_modes_other_side_does_not():
    # The anchor/main builder is constructed with allow_modes=True; the
    # other-side builder keeps the default (Responsa-style only).
    assert "allow_modes=True" in _SRC


def test_export_wiring_present():
    assert "def _export_xlsx" in _SRC
    assert "build_candidate_export_rows" in _SRC
    assert "_build_search_results_xlsx_bytes" in _SRC
    assert 'self.btn_export' in _SRC


def test_do_search_uses_mode_aware_core_mode():
    # The hardcoded "exact" mode string is gone from the SearchThread call site.
    assert "core_mode = core_mode_for_join_mode(mode_key)" in _SRC


# --------------------------------------------------------------------------- #
# SEED-024 follow-up — filename normalization + Stop-search + label/colour
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("shelf,expected", [
    ("T-S NS 324.11", "T-S_NS_324.11"),   # the user's example
    ("T-S 12.100", "T-S_12.100"),
    ("a/b:c", "a_b_c"),                    # slash + colon -> underscore
    ("  x  y  ", "x_y"),                   # collapse + trim
    ("", ""),
    (None, ""),
])
def test_normalize_shelfmark_for_filename(shelf, expected):
    assert jw.normalize_shelfmark_for_filename(shelf) == expected


def test_export_filename_uses_normalized_shelfmark():
    assert "joins_candidates_" in _SRC
    assert "normalize_shelfmark_for_filename" in _SRC
    assert "joins_lab_candidates.xlsx" not in _SRC   # old default name gone


def test_stop_search_wiring_present():
    assert "_on_find_clicked" in _SRC
    assert "def _stop_search" in _SRC
    assert "cancel_flag = True" in _SRC          # cancels the running SearchThread
    assert "_set_find_button_searching" in _SRC


def test_find_button_green_with_red_stop():
    assert "#27ae60" in _SRC   # green idle (matches main-search button)
    assert "#c0392b" in _SRC   # red Stop


def test_cancel_renders_partial_results_and_guards_generation():
    # The core returns PARTIAL results on cancel (catches InterruptedError), emitted
    # normally — so _on_results must render `raw` (not discard it) and a generation
    # guard drops stale results from a superseded search.
    assert "def _on_results(self, raw: list, gen=None)" in _SRC
    assert "self._search_gen" in _SRC
    assert "_search_was_partial" in _SRC
    # Enter routes through the Find/Stop toggle so it can't start an overlapping search.
    assert "self._on_find_clicked," in _SRC


def test_responsa_label_matches_main_search():
    # The Responsa combo item now uses tr("Responsa") (-> "פרויקט השו\"ת"),
    # matching the main search — NOT the prior tr("Responsa-style").
    assert 'self.mode_combo.addItem(tr("Responsa"))' in _SRC
    assert 'addItem(tr("Responsa-style"))' not in _SRC

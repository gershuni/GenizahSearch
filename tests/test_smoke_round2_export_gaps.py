# -*- coding: utf-8 -*-
"""Phase 94 smoke verification round 2 (2026-05-21) regression suite.

Pins the round-2 fixes:

1. Main sheet renamed 'Genizah Results' -> 'Search Results' (EN) /
   'תוצאות גניזה' -> 'תוצאות חיפוש' (HE).
2. 4-sheet workbook: ws[4] = 'Credits and Info' / 'קרדיט ומידע' holding
   the canonical Stoekl Ben Ezra citation chain + per-export search
   metadata + a GenizahSearch.com hyperlink.
3. Main sheet no longer carries inline credits (web: bottom block;
   desktop: pre-header rows).
4. Hebrew domain substitution on the main-sheet Domains column when
   lang='he' and a domain_name_map is supplied.
5. Web search-metadata threading: search_mode / search_gap /
   domain_name_map kwargs flow web/export_state.py ->
   web/api.py:export_excel -> export_search_results_excel.

These tests guard against future regressions of all 5 behaviors. They
build the workbook in-memory via the public APIs and assert specific
cells; no real Supabase / FJMS / NLI sidecar access required.
"""
from io import BytesIO

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Shared fixtures (kept minimal; the wave-3 / wave-4 test files have richer
# fixtures — this file only needs the round-2 behaviors)
# ---------------------------------------------------------------------------


class _FakeMetaMgr:
    def get_meta_for_id(self, sys_id):
        return (f'T-S {sys_id[-4:]}', f'Title for {sys_id}')

    def get_library_for_id(self, sys_id):
        return 'CUL'

    def parse_full_id_components(self, header):
        return {}


def _meta_resolver_fake(sid):
    if not sid:
        return None
    return {
        'shelfmark': f'T-S {sid[-4:]}',
        'title': f'Title {sid}',
        'library_code': 'CUL',
        'library_name': 'Cambridge University Library',
    }


def _identity_sanitize(text):
    return '' if text is None else str(text)


def _make_result(sys_id='99001234567890'):
    return {
        'uid': f'IE_x_P_y_FL_{sys_id}',
        'display': {
            'id': sys_id,
            'shelfmark': f'T-S {sys_id[-4:]}',
            'title': f'Title {sys_id}',
            'library_code': 'CUL',
            'img': '1r',
            'source': 'ms',
        },
        'snippet': 'foo *bar* baz',
        'raw_header': '',
        'sort_score': 1.0,
        'full_text': 'page text',
    }


@pytest.fixture
def stub_dossier(monkeypatch):
    """Same stub pattern as the wave-3 / wave-4 tests."""
    from shared import export_dossier
    monkeypatch.setattr(export_dossier, 'pgp_subset_for_sys_id', lambda s, **kw: {
        'pgp_url': f'https://pgp.example/{s}',
        'description': f'PGP desc {s}', 'document_type': 'Letter',
        'date_display': '1100', 'languages': ['Hebrew', 'Aramaic'],
        'tags': ['letter'],
    } if s else None)
    monkeypatch.setattr(export_dossier, 'nli_subset_for_sys_id', lambda s, **kw: {
        'catalog_entry': f'Neubauer {s[-4:]}',
        'library_viewer_url': f'https://cudl.example/{s}',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'catalog_summary_for_sys_id', lambda s, **kw: {
        'title': f'CatTitle {s}', 'author_text': 'Author',
        'copy_date': '1180', 'copy_place': 'Fustat',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'bibliography_for_sys_id', lambda s, **kw: [{
        'running_title': 'Med. Soc.', 'title_year': 1967, 'mention_page': '123',
        'article_name': 'Letter', 'article_author_eng': 'Goitein',
        'catalog_acronym': 'MS',
    }] if s == '99001234567890' else [])


def _build_web(results, **kwargs):
    from web.export_service import ExportService
    svc = ExportService(_FakeMetaMgr())
    content, _ = svc.export_search_results_excel(results, 'my query', **kwargs)
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


def _build_desktop(results, **kwargs):
    from genizah_app import _build_search_results_xlsx_bytes
    content = _build_search_results_xlsx_bytes(
        results=results,
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
        **kwargs,
    )
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


def _find_credits_sheet(wb, lang='en'):
    from shared.export_dossier import sheet_titles
    return wb[sheet_titles(lang)['credits_info']]


def _all_cells(ws, max_row=100, max_col=5):
    """Return a flat list of all non-empty cell values."""
    cells = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and (not isinstance(v, str) or v.strip()):
                cells.append(v)
    return cells


# ===========================================================================
# Gap A: Sheet rename
# ===========================================================================


def test_main_sheet_named_search_results_en_web(stub_dossier):
    wb = _build_web([_make_result()])
    assert 'Search Results' in wb.sheetnames
    assert 'Genizah Results' not in wb.sheetnames


def test_main_sheet_named_search_results_en_desktop(stub_dossier):
    wb = _build_desktop([_make_result()])
    assert 'Search Results' in wb.sheetnames
    assert 'Genizah Results' not in wb.sheetnames


def test_main_sheet_named_search_results_he_web(stub_dossier):
    wb = _build_web([_make_result()], lang='he')
    assert 'תוצאות חיפוש' in wb.sheetnames
    assert 'תוצאות גניזה' not in wb.sheetnames


def test_main_sheet_named_search_results_he_desktop(stub_dossier):
    wb = _build_desktop([_make_result()], lang='he')
    assert 'תוצאות חיפוש' in wb.sheetnames
    assert 'תוצאות גניזה' not in wb.sheetnames


def test_sheet_titles_module_exports_round2_strings():
    from shared.export_dossier import sheet_titles
    assert sheet_titles('en')['main'] == 'Search Results'
    assert sheet_titles('he')['main'] == 'תוצאות חיפוש'
    assert sheet_titles('en')['credits_info'] == 'Credits and Info'
    assert sheet_titles('he')['credits_info'] == 'קרדיט ומידע'


# ===========================================================================
# Gap B: Dedicated Credits and Info sheet (4-sheet workbook)
# ===========================================================================


def test_workbook_has_4_sheets_web(stub_dossier):
    wb = _build_web([_make_result()])
    assert wb.sheetnames == [
        'Search Results', 'Manuscripts', 'Bibliography', 'Credits and Info',
    ]


def test_workbook_has_4_sheets_desktop(stub_dossier):
    wb = _build_desktop([_make_result()])
    assert wb.sheetnames == [
        'Search Results', 'Manuscripts', 'Bibliography', 'Credits and Info',
    ]


def test_credits_sheet_has_citation_lines_web(stub_dossier):
    wb = _build_web([_make_result()])
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    # Stoekl Ben Ezra citation must appear in some cell on the sheet.
    cited = any('Stoekl Ben Ezra' in str(c) for c in cells)
    assert cited, f"Citation missing from credits sheet; cells={cells}"
    # Dataset URL also present.
    has_doi = any('10.5281/zenodo.17734473' in str(c) for c in cells)
    assert has_doi, f"DOI missing from credits sheet; cells={cells}"


def test_credits_sheet_has_citation_lines_desktop(stub_dossier):
    wb = _build_desktop([_make_result()])
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    cited = any('Stoekl Ben Ezra' in str(c) for c in cells)
    assert cited
    has_doi = any('10.5281/zenodo.17734473' in str(c) for c in cells)
    assert has_doi


def test_credits_sheet_has_hyperlink_to_genizahsearch_web(stub_dossier):
    wb = _build_web([_make_result()])
    ws = _find_credits_sheet(wb)
    # Find the cell whose value is the GenizahSearch URL.
    found_url_cell = None
    for r in range(1, 50):
        for c in range(1, 5):
            v = ws.cell(r, c).value
            if v == 'https://genizahsearch.com':
                found_url_cell = ws.cell(r, c)
                break
        if found_url_cell:
            break
    assert found_url_cell is not None, "GenizahSearch URL cell missing"
    assert found_url_cell.hyperlink is not None, "Hyperlink missing on URL cell"
    assert found_url_cell.hyperlink.target == 'https://genizahsearch.com'


def test_credits_sheet_has_hyperlink_to_genizahsearch_desktop(stub_dossier):
    wb = _build_desktop([_make_result()])
    ws = _find_credits_sheet(wb)
    found_url_cell = None
    for r in range(1, 50):
        for c in range(1, 5):
            v = ws.cell(r, c).value
            if v == 'https://genizahsearch.com':
                found_url_cell = ws.cell(r, c)
                break
        if found_url_cell:
            break
    assert found_url_cell is not None
    assert found_url_cell.hyperlink is not None
    assert found_url_cell.hyperlink.target == 'https://genizahsearch.com'


def test_credits_sheet_title_row_en(stub_dossier):
    wb = _build_web([_make_result()], lang='en')
    ws = _find_credits_sheet(wb, 'en')
    # Row 1 col 1 should be 'Credits and Info' (the sheet section title).
    assert ws.cell(1, 1).value == 'Credits and Info'


def test_credits_sheet_title_row_he(stub_dossier):
    wb = _build_web([_make_result()], lang='he')
    ws = _find_credits_sheet(wb, 'he')
    assert ws.cell(1, 1).value == 'קרדיט ומידע'


def test_main_sheet_no_inline_credits_web(stub_dossier):
    """The web inline 'Credits' block at the bottom of the main sheet is gone."""
    wb = _build_web([_make_result()])
    ws = wb['Search Results']
    # Scan the entire main sheet for the literal 'Credits' header or the
    # canonical Stoekl Ben Ezra citation. Neither should appear.
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            assert 'Stoekl Ben Ezra' not in v, (
                f"Inline credits still present on main sheet at row {r}: {v}"
            )


def test_main_sheet_no_inline_credits_desktop(stub_dossier):
    """The desktop inline credit block above the header row is gone."""
    wb = _build_desktop(
        [_make_result()],
        # Pass the legacy kwargs; they must be IGNORED on the main sheet now.
        credit_text='Generated by Genizah Search Pro\nMiDRASH credit text',
        search_info_text='Query: foo\nMode: text',
    )
    ws = wb['Search Results']
    # Header row should be row 1.
    assert ws.cell(1, 1).value == 'System ID', (
        f"Header row not at row 1; got {ws.cell(1, 1).value!r}"
    )
    # The credit text 'Generated by Genizah Search Pro' must NOT appear.
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            assert 'Generated by Genizah Search Pro' not in v
            assert 'MiDRASH credit text' not in v
            assert 'Query: foo' not in v


# ===========================================================================
# Gap C: Web search metadata cells
# ===========================================================================


def test_credits_sheet_has_search_query_cell_web(stub_dossier):
    wb = _build_web(
        [_make_result()],
        search_mode='text',
        search_gap=5,
    )
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    # Search Query label and the query string.
    assert any('Search Query' in str(c) for c in cells)
    assert any('my query' in str(c) for c in cells)


def test_credits_sheet_has_search_mode_cell_web(stub_dossier):
    wb = _build_web([_make_result()], search_mode='text', search_gap=5)
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    assert any('Search Mode' in str(c) for c in cells)
    assert any(c == 'text' for c in cells)


def test_credits_sheet_has_gap_cell_web(stub_dossier):
    wb = _build_web([_make_result()], search_mode='text', search_gap=5)
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    # Gap label and the value 5.
    assert any(c == 'Gap' for c in cells)
    assert any(c == 5 for c in cells)


def test_credits_sheet_has_result_count_cell_web(stub_dossier):
    wb = _build_web([_make_result('A'), _make_result('B'), _make_result('C')])
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    assert any('Result count' in str(c) for c in cells)
    assert any(c == 3 for c in cells)


def test_credits_sheet_has_date_time_cell_web(stub_dossier):
    """Date/time row should be present and contain an ISO-like date."""
    import re as _re
    wb = _build_web([_make_result()])
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    assert any('Date/time of export' in str(c) for c in cells)
    # ISO datetime YYYY-MM-DD HH:MM:SS pattern.
    has_iso = any(
        isinstance(c, str)
        and _re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$', c)
        for c in cells
    )
    assert has_iso, f"No ISO datetime found in credits sheet; cells={cells}"


def test_credits_sheet_omits_lab_mode_on_web(stub_dossier):
    """Web has no Lab Mode UI; the Lab Mode row must be absent."""
    wb = _build_web([_make_result()])
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    assert not any('Lab Mode' in str(c) for c in cells)
    assert not any('Deep Scan' in str(c) for c in cells)


# ===========================================================================
# Gap C': Desktop search metadata cells (parity with web)
# ===========================================================================


def test_credits_sheet_desktop_includes_lab_mode_row(stub_dossier):
    """Desktop has Lab Mode; the row must be present when lab_mode_on is bool."""
    wb = _build_desktop(
        [_make_result()],
        search_query='my query',
        search_mode='Exact',
        search_gap=5,
        lab_mode_on=False,
        deep_scan_on=None,
    )
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    assert any('Lab Mode' in str(c) for c in cells)


def test_credits_sheet_desktop_deep_scan_only_when_lab_on(stub_dossier):
    # Lab Mode off -> Deep Scan row should NOT appear.
    wb = _build_desktop(
        [_make_result()],
        lab_mode_on=False,
        deep_scan_on=True,
    )
    ws = _find_credits_sheet(wb)
    cells = _all_cells(ws)
    assert not any('Deep Scan' in str(c) for c in cells)
    # Lab Mode on -> Deep Scan row appears.
    wb2 = _build_desktop(
        [_make_result()],
        lab_mode_on=True,
        deep_scan_on=True,
    )
    ws2 = _find_credits_sheet(wb2)
    cells2 = _all_cells(ws2)
    assert any('Deep Scan' in str(c) for c in cells2)


# ===========================================================================
# Gap D: Hebrew domain substitution
# ===========================================================================


def test_hebrew_domain_substitution_web(stub_dossier):
    """When lang='he' AND a domain_name_map is provided, the Domains cell
    on the main sheet shows Hebrew names."""
    domain_name_map = {
        'Bible': 'מקרא',
        'Letter': 'מכתב',
        'Legal': 'משפטים',
    }
    wb = _build_web(
        [_make_result('A')],
        result_domains={'A': ['Bible', 'Letter', 'Legal']},
        domain_name_map=domain_name_map,
        lang='he',
    )
    from shared.export_dossier import sheet_titles
    ws = wb[sheet_titles('he')['main']]
    # Domains is col 11; row 2 is the first data row.
    assert ws.cell(2, 11).value == 'מקרא|מכתב|משפטים'


def test_hebrew_domain_substitution_desktop(stub_dossier):
    domain_name_map = {
        'Bible': 'מקרא',
        'Letter': 'מכתב',
    }
    wb = _build_desktop(
        [_make_result('A')],
        result_domains={'A': ['Bible', 'Letter']},
        domain_name_map=domain_name_map,
        lang='he',
    )
    from shared.export_dossier import sheet_titles
    ws = wb[sheet_titles('he')['main']]
    # Desktop header row is row 1; first data row is row 2.
    assert ws.cell(2, 11).value == 'מקרא|מכתב'


def test_hebrew_domain_unknown_names_pass_through_web(stub_dossier):
    """Names not in the map render with their original (English) text."""
    domain_name_map = {'Bible': 'מקרא'}
    wb = _build_web(
        [_make_result('A')],
        result_domains={'A': ['Bible', 'UnknownDomain']},
        domain_name_map=domain_name_map,
        lang='he',
    )
    from shared.export_dossier import sheet_titles
    ws = wb[sheet_titles('he')['main']]
    # Unknown 'UnknownDomain' should pass through verbatim — never dropped.
    assert ws.cell(2, 11).value == 'מקרא|UnknownDomain'


def test_hebrew_domain_unknown_names_pass_through_desktop(stub_dossier):
    domain_name_map = {'Bible': 'מקרא'}
    wb = _build_desktop(
        [_make_result('A')],
        result_domains={'A': ['Bible', 'UnknownDomain']},
        domain_name_map=domain_name_map,
        lang='he',
    )
    from shared.export_dossier import sheet_titles
    ws = wb[sheet_titles('he')['main']]
    assert ws.cell(2, 11).value == 'מקרא|UnknownDomain'


def test_domain_map_ignored_when_lang_en_web(stub_dossier):
    """lang='en' must NOT substitute even if a domain_name_map is provided."""
    domain_name_map = {'Bible': 'מקרא'}
    wb = _build_web(
        [_make_result('A')],
        result_domains={'A': ['Bible']},
        domain_name_map=domain_name_map,
        lang='en',
    )
    ws = wb['Search Results']
    # English Bible passes through (no substitution under lang='en').
    assert ws.cell(2, 11).value == 'Bible'


def test_domain_map_ignored_when_lang_en_desktop(stub_dossier):
    domain_name_map = {'Bible': 'מקרא'}
    wb = _build_desktop(
        [_make_result('A')],
        result_domains={'A': ['Bible']},
        domain_name_map=domain_name_map,
        lang='en',
    )
    ws = wb['Search Results']
    assert ws.cell(2, 11).value == 'Bible'


def test_domain_map_none_falls_through_to_english_web(stub_dossier):
    """When lang='he' but no domain_name_map -> English names render."""
    wb = _build_web(
        [_make_result('A')],
        result_domains={'A': ['Bible', 'Letter']},
        # No domain_name_map kwarg.
        lang='he',
    )
    from shared.export_dossier import sheet_titles
    ws = wb[sheet_titles('he')['main']]
    assert ws.cell(2, 11).value == 'Bible|Letter'


# ===========================================================================
# Hebrew labels — pin the exact strings from the user's round-2 brief
# ===========================================================================


def test_hebrew_label_search_query():
    from shared.export_dossier import search_meta_labels
    assert search_meta_labels('he')['search_query'] == 'שאילתת חיפוש'


def test_hebrew_label_search_mode():
    from shared.export_dossier import search_meta_labels
    # NOTE: the user-specified Hebrew is 'אופן החיפוש' (NOT 'מצב חיפוש'
    # which is the existing tr() entry in genizah_translations.py).
    assert search_meta_labels('he')['search_mode'] == 'אופן החיפוש'


def test_hebrew_label_search_gap():
    from shared.export_dossier import search_meta_labels
    # NOTE: 'רווח' (NOT 'מרווח' from tr()).
    assert search_meta_labels('he')['search_gap'] == 'רווח'


def test_hebrew_label_lab_mode():
    from shared.export_dossier import search_meta_labels
    assert search_meta_labels('he')['lab_mode'] == 'מצב מעבדה'


def test_hebrew_label_deep_scan():
    from shared.export_dossier import search_meta_labels
    assert search_meta_labels('he')['deep_scan'] == 'סריקה עמוקה'


def test_hebrew_label_date_time():
    from shared.export_dossier import search_meta_labels
    assert search_meta_labels('he')['date_time'] == 'תאריך ושעת הייצוא'


def test_hebrew_label_result_count():
    from shared.export_dossier import search_meta_labels
    assert search_meta_labels('he')['result_count'] == 'מספר תוצאות'


def test_hebrew_label_visit_link():
    from shared.export_dossier import search_meta_labels
    assert search_meta_labels('he')['visit_link_label'] == 'בקרו ב-GenizahSearch.com'


# ===========================================================================
# Web state plumbing: domain_name_map flows through set_search_export
# ===========================================================================


@pytest.fixture
def stub_user_storage(monkeypatch):
    """Mirrors tests/test_export_state_enrichment.py stub pattern."""
    from types import SimpleNamespace
    fake_storage = {}
    stub = SimpleNamespace(storage=SimpleNamespace(user=fake_storage))
    monkeypatch.setattr('web.safe_storage.app', stub)
    return fake_storage


def test_set_search_export_persists_domain_name_map(stub_user_storage):
    """The new domain_name_map kwarg on set_search_export must round-trip
    through the export payload and be readable by get_search_export."""
    from web.export_state import set_search_export, get_search_export
    domain_name_map = {'Bible': 'מקרא', 'Letter': 'מכתב'}
    set_search_export(
        results=[{'uid': 'u1', 'sys_id': 'A', 'snippet': '', 'sort_score': 1.0}],
        query='q',
        domain_name_map=domain_name_map,
    )
    payload = get_search_export()
    assert payload is not None
    assert payload.get('domain_name_map') == domain_name_map


def test_update_search_export_enrichment_patches_domain_name_map(stub_user_storage):
    from web.export_state import (
        set_search_export,
        update_search_export_enrichment,
        get_search_export,
    )
    # Set baseline with empty map.
    set_search_export(
        results=[{'uid': 'u1', 'sys_id': 'A', 'snippet': '', 'sort_score': 1.0}],
        query='q',
    )
    assert (get_search_export() or {}).get('domain_name_map') == {}

    # Patch the map.
    update_search_export_enrichment(
        domain_name_map={'Bible': 'מקרא'},
    )
    payload = get_search_export()
    assert payload['domain_name_map'] == {'Bible': 'מקרא'}

    # Re-patching with None leaves the map untouched (opt-in semantics).
    update_search_export_enrichment(transcription_sys_ids={'A'})
    payload2 = get_search_export()
    assert payload2['domain_name_map'] == {'Bible': 'מקרא'}
    assert sorted(payload2['transcription_sys_ids']) == ['A']

# -*- coding: utf-8 -*-
"""The findings xlsx export (EXPORT-01/02/03, phase 136.2).

WHAT THIS FILE IS FOR, and what it deliberately leaves to the sweep. The
masking sweep owns the question "can restricted text reach a downloaded file";
it scans the workbook by CELL and proves the scan can fail. This file owns the
other half: that the workbook says what the PAGE says, that it says it in
words the honesty gate accepts, and that the paths behind it are bounded and
fail whole rather than short.

The recurring defect this phase's suite is written against is the one recorded
seven times over in Phase 136: a check that passes because it measured nothing.
So the assertions here are on CONTENT -- a label that must be present, a
sentence that must appear, a query count that must be small -- and never merely
on the absence of an exception.
"""

from __future__ import annotations

import io
import sqlite3
from typing import Any, Dict, List, Optional

import pytest

from shared.discovery_service import DiscoveryService
from shared.discovery_surface_projection import STATUS_OK, make_envelope
import shared.discovery_display_strings as ds
import web.discovery_export_service as des
from tests.render_smoke import test_findings_render_smoke as tf
from tests.render_smoke.discovery_honesty_gate import assert_surface_honesty

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")


# ===========================================================================
# Fixtures -- rows built by the SAME factory the page's own suite uses, so a
# projection change reaches this file without anyone remembering to update it.
# ===========================================================================

def _rows() -> List[Dict[str, Any]]:
    return [dict(row) for name, row in tf.corpus_rows()
            if name == "SURFACE_FINDING_FIELDS"]


def _excerpt(*, work_side: bool) -> Dict[str, Any]:
    from shared.discovery_surface_projection import surface_safe_excerpt
    row = {
        "identification_id": "synthetic-1", "evidence_id": 1,
        "a_page_id": "p", "frag_before": "alpha ", "frag_span": "beta",
        "frag_after": " gamma", "frag_clipped": 0,
        "work_before": None, "work_span": None, "work_after": None,
        "work_clipped": None, "work_source": None, "attribution": None,
        "n_spans": 1, "text_layer": "htr",
        "frag_hl": None, "work_hl": None, "work_markup": None,
    }
    if work_side:
        row.update({"work_before": "delta ", "work_span": "epsilon",
                    "work_after": " zeta", "work_clipped": 0,
                    "work_source": "exact", "attribution": "Public edition"})
    return surface_safe_excerpt(row)


def _envelope(items: Optional[List[Dict[str, Any]]] = None, **meta_over) -> Dict[str, Any]:
    items = items if items is not None else _rows()
    meta = {"unit": "identification", "bucket": "main", "sort": "band_rank",
            "row_count": len(items), "reported_total": len(items),
            "walk_complete": True, "sidecar_version": "synthetic"}
    meta.update(meta_over)
    env = make_envelope(STATUS_OK, [dict(r) for r in items], len(items), meta=meta)
    for built, source in zip(env["items"], items):
        if "excerpt" in source:
            built["excerpt"] = source["excerpt"]
    return env


def _cells(data: bytes) -> List[str]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: List[str] = []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(values_only=True):
            out.extend(str(v) for v in row if v is not None)
    wb.close()
    return out


def _headers(data: bytes, sheet_index: int = 0) -> List[str]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet_index]]
    first = next(ws.iter_rows(values_only=True))
    wb.close()
    return [str(v) for v in first if v is not None]


# ===========================================================================
# EXPORT-01 -- the workbook says what the page says
# ===========================================================================

@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_workbook_builds_its_sheets_in_both_languages(lang):
    """TWO sheets, not three. The text matches moved onto the identification
    row (owner instruction, 2026-08-20): a reader judging a match had to find
    its counterpart on another sheet, keyed on a shelfmark that repeats."""
    data = des.build_findings_workbook(_envelope(), lang=lang)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    assert wb.sheetnames == [des._pick(des._SHEET_TITLES[k], lang)
                             for k in des.SHEET_KEYS]
    assert len(wb.sheetnames) == 2, (
        "a separate text sheet is back; the passages belong on the row that "
        "makes the claim")
    wb.close()


@pytest.mark.parametrize("lang", ("en", "he"))
def test_there_is_no_band_column(lang):
    """The identification grain carries `best_band_rank` and NO band label,
    evidence source or confidence band. `_render_row_meta` renders no band
    tooltip for exactly that reason -- "deriving a label from a rank would be a
    second band vocabulary". A spreadsheet column is not an exception: a rank in
    a cell reads as a score, and unlike a tooltip it leaves the building.

    Asserted on the HEADERS (no band column may be declared) and on the CELLS
    (the rank's value must not appear as a cell of its own).
    """
    rows = _rows()
    data = des.build_findings_workbook(_envelope(rows), lang=lang)
    headers = " ".join(_headers(data)).lower()
    for forbidden in ("band", "rank", "confidence", "דירוג", "רמת", "ודאות"):
        assert forbidden not in headers, (
            f"the export declared a {forbidden!r} column; the grain has no band "
            "vocabulary and a rank rendered as a value is a second one")
    ranks = {str(r.get("best_band_rank")) for r in rows
             if r.get("best_band_rank") is not None}
    assert ranks, "the fixture carries no band rank, so this test proves nothing"
    assert not (ranks & set(_cells(data))), (
        "a band rank reached a cell verbatim")


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_honesty_gate_passes_over_every_cell(lang):
    """All six detectors, over the workbook's own text.

    The cells are wrapped in a scoped container because the gate is written for
    rendered markup; what it actually inspects is the text, which is exactly
    what a reader of the spreadsheet sees.
    """
    data = des.build_findings_workbook(
        _envelope(), lang=lang, filters={"unit": "identification"},
        generated_at="2026-01-01T00:00:00Z")
    body = "\n".join(_cells(data))
    assert body.strip(), "no cells to check"
    assert_surface_honesty(
        f'<div class="wb">{body}</div>', scope_selector="wb", lang=lang)


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_relation_label_is_present_not_merely_non_crashing(lang):
    """`relation_chip` RAISES on an unknown vocabulary and the row renderer
    swallows that, which drops the element silently. A cell dropped the same way
    is a blank that reads as "no relation".

    So this asserts the POPULATED case: the chip text for the fixture's own
    relation must be in the workbook. A helper that always returned "" would
    satisfy "did not crash" and fail this.
    """
    rows = _rows()
    data = des.build_findings_workbook(_envelope(rows), lang=lang)
    cells = set(_cells(data))
    expected = {ds.relation_chip(r["rendered_relation"], lang)
                for r in rows if r.get("rendered_relation")}
    assert expected, "no fixture row carries a relation; the test is vacuous"
    assert expected & cells, (
        f"none of the relation labels {sorted(expected)} reached a cell")


@pytest.mark.parametrize("lang", ("en", "he"))
def test_a_masked_work_gets_the_honest_sentence_not_a_blank_cell(lang):
    """The bake writes four `None` work pieces for a masked non-Bible work so
    the UI can say so. On a page an empty area is self-evidently not a claim; in
    a spreadsheet an empty CELL reads as a value, so the sentence has to be
    written into it.
    """
    rows = _rows()
    rows[0]["excerpt"] = _excerpt(work_side=False)
    data = des.build_findings_workbook(_envelope(rows), lang=lang)
    cells = _cells(data)
    assert ds.excerpt_strings(lang)["work_unavailable"] in cells, (
        "a work-less excerpt produced a blank cell instead of the honest state")


def test_the_manuscript_and_edition_passages_both_reach_cells():
    rows = _rows()
    rows[0]["excerpt"] = _excerpt(work_side=True)
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    body = "\n".join(_cells(data))
    assert "alpha beta gamma" in body, "the manuscript passage is missing"
    assert "delta epsilon zeta" in body, "the edition passage is missing"


# ===========================================================================
# EXPORT-01 (2026-08-20) -- the evidence rides on the row that makes the claim
# ===========================================================================

def _rich_rows(data: bytes, lang: str = "en"):
    """Every data row of sheet 1 as {header: cell}, with rich text intact."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), rich_text=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows())
    headers = [c.value for c in rows[0]]
    out = [{h: c.value for h, c in zip(headers, row)} for row in rows[1:]]
    wb.close()
    return headers, out


def _runs(value):
    """(text, is_highlighted) for a passage cell, rich or plain."""
    if isinstance(value, str) or value is None:
        return [(value or "", False)]
    runs = []
    for block in value:
        if isinstance(block, str):
            runs.append((block, False))
        else:
            colour = getattr(block.font, "color", None)
            rgb = getattr(colour, "rgb", None) if colour is not None else None
            runs.append((block.text, str(rgb or "").endswith("FF0000")))
    return runs


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_passages_sit_on_the_finding_s_own_row(lang):
    """THE MERGE, asserted on ONE row rather than on the workbook as a whole.

    "Both texts appear somewhere in the file" was true of the two-sheet layout
    too, and it is the property the owner rejected. What has to hold is that
    the shelfmark, the relation and the two passages are the SAME row -- a
    reader judging a match must not have to join two sheets on a shelfmark
    that repeats across rows.
    """
    rows = _rows()
    rows[0]["excerpt"] = _excerpt(work_side=True)
    data = des.build_findings_workbook(_envelope(rows), lang=lang)
    headers, out = _rich_rows(data)
    strings = ds.excerpt_strings(lang)
    frag_h, work_h = strings["frag_label"], strings["work_label"]
    assert frag_h in headers and work_h in headers, (
        f"the passage columns are missing from {headers}")

    row = next(r for r in out
               if "".join(t for t, _ in _runs(r[frag_h])).strip())
    assert "alpha beta gamma" in "".join(t for t, _ in _runs(row[frag_h]))
    assert "delta epsilon zeta" in "".join(t for t, _ in _runs(row[work_h]))
    # ... on the row that also carries the identification itself.
    assert row[des._pick({"en": "Shelfmark", "he": "סימן מדף"}, lang)], (
        "the passage landed on a row with no shelfmark -- it is not the "
        "finding's own row")


def test_the_matched_words_are_red_and_bold_not_merely_present():
    """The highlight is the point of the change, so it is asserted on the RUN
    STRUCTURE. A cell holding the same characters as plain text would satisfy
    "the passage reached a cell" and fail a reader looking for the match."""
    rows = _rows()
    rows[0]["excerpt"] = _excerpt(work_side=True)
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    _headers_, out = _rich_rows(data)
    row = next(r for r in out if "alpha" in
               "".join(t for t, _ in _runs(r["Manuscript passage"])))
    runs = _runs(row["Manuscript passage"])
    highlighted = [t for t, hl in runs if hl]
    assert highlighted == ["beta"], (
        f"expected the matched span alone to be highlighted, got {runs}")
    assert [t for t, hl in runs if not hl] == ["alpha ", " gamma"]


def test_a_literal_marker_in_the_artifact_cannot_forge_a_highlight():
    """The marker is a character that can occur in the source text. If it
    survived into the string the rich-text splitter reads, artifact text would
    choose its own highlighting -- and, worse, an ODD number of them would
    invert the highlight of everything after it."""
    rows = _rows()
    excerpt = dict(_excerpt(work_side=True))
    excerpt["frag_before"] = "lead *forged* "
    excerpt["frag_span"] = "real"
    excerpt["frag_after"] = " tail*"
    rows[0]["excerpt"] = excerpt
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    _h, out = _rich_rows(data)
    row = next(r for r in out if "real" in
               "".join(t for t, _ in _runs(r["Manuscript passage"])))
    assert [t for t, hl in _runs(row["Manuscript passage"]) if hl] == ["real"], (
        "artifact text was able to mark its own highlight")


@pytest.mark.parametrize("text", ("{alpha} beta gam{mel} delta",
                                  "star * inside {a}*b* tail",
                                  "*", "", "   ", "}orphan{ close"))
@pytest.mark.parametrize("ja", (False, True))
@pytest.mark.parametrize("intervals", (None, [[0, 5]], [[2, 4], [8, 12]],
                                       [[-3, 2]], [[5, 999]], [["x", 1]],
                                       [[4, 2]], [[0, 0]], [[1, 3], [2, 6]]))
def test_the_exported_highlight_agrees_with_the_rendered_pane(intervals, ja, text):
    """ONE HIGHLIGHT VOCABULARY, proven rather than asserted in a comment.

    The pane emits HTML and a cell cannot take HTML, so `des._marked_span`
    MIRRORS `findings_rows._compose_excerpt_piece` instead of calling it. A
    mirror drifts. Both are driven here over the same inputs -- including the
    malformed intervals the sidecar could carry -- and the highlighted TEXT
    must match.

    The ONE intended difference: a run of pure whitespace is not marked in a
    cell (a pair of markers around a space reads as a matched word that is not
    there), so whitespace-only fragments are dropped from both sides.
    """
    import html as _html
    import re as _re
    from web.components import findings_rows as fr

    whole = intervals is None

    markup = fr._compose_excerpt_piece(text, intervals, ja, whole_span=whole)
    pane = [_html.unescape(m) for m in _re.findall(
        r'<span class="[^"]*' + _re.escape(fr.ROW_EXCERPT_CLASS)
        + r'-hl[^"]*">(.*?)</span>', markup)]

    parts = des._marked_span(text, intervals, ja, whole=whole).split(des._MARKER)
    cell = [parts[i] for i in range(1, len(parts), 2)]

    # The pane side is escaped the same way before comparing: a literal marker
    # character in the source becomes a space in a cell, which is a documented
    # difference (the cell has no other way to say "this is text, not markup")
    # and not a drift in WHICH characters are matched.
    keep = lambda seq: [p for p in seq if p.strip()]
    pane = [des._strip_markers(p) for p in pane]
    assert "".join(keep(cell)) == "".join(keep(pane)), (
        f"the cell highlighted {cell!r} where the pane highlighted {pane!r}")


def test_a_matched_run_of_literal_markers_keeps_its_highlight():
    """Escaping must not be able to DELETE a highlight.

    A matched run made only of marker characters escapes to spaces, and keying
    the "do not mark whitespace" rule off the ESCAPED text dropped its
    highlight entirely -- tidiness that silently loses content, and invisible
    because the earlier test only put markers in the context pieces.
    """
    marked = des._marked_span("*", [[0, 1]], False, whole=False)
    # Two markers, not three: the source marker itself escaped to a space, and
    # the pair around it is what keeps the run highlighted.
    assert marked.count(des._MARKER) == 2, (
        f"expected the escaped run to stay marked, got {marked!r}")
    parts = marked.split(des._MARKER)
    assert [parts[i] for i in range(1, len(parts), 2)] == [" "]


def test_a_passage_too_long_for_a_cell_cannot_invert_the_highlight():
    """`sanitize_text_for_excel` truncates at 32,700 BLINDLY. A cut between an
    opening marker and its mate leaves an odd count, and every run after it
    flips. The bake caps spans far below this, so the guard is a backstop --
    which is exactly the kind of code that is never exercised unless a test
    exercises it."""
    long_marked = ("x" * (des._CELL_TEXT_MAX - 2)) + des._MARKER + ("y" * 500)
    fitted = des._fit_cell(long_marked)
    assert fitted.count(des._MARKER) % 2 == 0, (
        "the fitted cell has an odd number of markers; the highlight inverts")
    assert fitted.endswith(des._ELLIPSIS), "the cut is not marked"
    # ... and the guard is a no-op on anything of a realistic size.
    assert des._fit_cell("short *match* here") == "short *match* here"


@pytest.mark.parametrize("tail", (True, False))
def test_the_context_clip_respects_its_own_bound(tail):
    """A bound nobody enforces is a bound nobody can rely on: the first version
    sliced to the budget and THEN appended the ellipsis, returning 162
    characters from a function documented as bounded by 160."""
    for length in (1, 159, 160, 161, 400, 4000):
        out = des._clip_context("q" * length, tail=tail)
        assert len(out) <= des._CONTEXT_CHARS, (
            f"clip returned {len(out)} chars for an input of {length}")
    # ... and the marked case really is marked, on the correct edge.
    text = "start " + ("m" * 400) + " finish"
    out = des._clip_context(text, tail=tail)
    assert des._ELLIPSIS in out
    assert (out.endswith("finish") if tail else out.startswith("start"))


def test_a_work_side_with_context_but_no_span_is_reported_unavailable():
    """`_render_panes` keys the ENTIRE work pane on `work_span`. Treating any
    of the three pieces as proof of a work side would show an edition passage
    in the file where the page says the edition is unavailable -- the export
    inventing a claim the surface does not make. No row in the live artifact is
    in that state, which is why nothing caught it."""
    rows = _rows()
    excerpt = dict(_excerpt(work_side=True))
    excerpt["work_span"] = None
    excerpt["work_before"] = "context that is not a match"
    rows[0]["excerpt"] = excerpt
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    body = "\n".join(_cells(data))
    assert ds.excerpt_strings("en")["work_unavailable"] in body
    assert "context that is not a match" not in body, (
        "the export showed work-side context as an edition passage")


@pytest.mark.parametrize("side", ("frag", "work"))
def test_a_passage_the_bake_abbreviated_says_so(side):
    """The bake caps a long span, joins its head and tail with a visible U+22EF
    and FLAGS the row -- 22.7% of fragment spans and 25.3% of work spans in the
    live artifact. Nothing said so, so the mark sat inside a quoted passage
    meaning nothing to the reader. A downloaded file travels without the page
    that could have explained it."""
    rows = _rows()
    excerpt = dict(_excerpt(work_side=True))
    excerpt["frag_clipped"] = 0
    excerpt["work_clipped"] = 0
    excerpt[f"{side}_clipped"] = 1
    rows[0]["excerpt"] = excerpt
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    note = ds.excerpt_strings("en")["clipped_note"]
    # SUBSTRING, not cell equality: the note shares its cell with the
    # transcription qualifier, exactly as `reprojected_note` shares one with
    # `multi_span`.
    assert note in " ".join(_cells(data)), (
        f"a clipped {side} passage did not say so")

    # ... and an unclipped row does NOT carry it, or the note means nothing.
    excerpt["frag_clipped"] = 0
    excerpt["work_clipped"] = 0
    rows[0]["excerpt"] = excerpt
    assert note not in " ".join(_cells(des.build_findings_workbook(
        _envelope(rows), lang="en")))


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_credit_does_not_speak_for_the_human_transcriptions(lang):
    """2,759 FGP and 87 PGP excerpts in the live artifact are human work by
    other projects. A blanket "these are MiDRASH transcriptions" would
    miscredit 5.9% of the rows to a project that did not make them -- and a
    misattribution is a worse failure than a missing credit."""
    data = des.build_findings_workbook(_envelope(), lang=lang)
    body = "\n".join(_cells(data))
    for forbidden in (
        "The manuscript passages in this file are MiDRASH",
        "קטעי כתב היד בקובץ זה הם תעתיקים אוטומטיים של MiDRASH",
    ):
        assert forbidden not in body, (
            "the credit claims every passage is MiDRASH work")
    # The credit must still be there, and must be conditional on the mark the
    # transcription-note column carries.
    assert "MiDRASH" in body
    assert ds.excerpt_strings(lang)["frag_htr_note"] in body or lang == "en"


def test_the_context_is_abbreviated_but_the_export_never_cuts_the_span():
    """THE EXPORT does not cut a matched span. (The BAKE does, above its span
    cap, with a visible U+22EF and a flag -- see
    `test_a_passage_the_bake_abbreviated_says_so`. The distinction matters:
    the bake's cut is recorded in the artifact and can be reported, while a cut
    invented here would be a clipped claim nobody could see.)"""
    span = "MATCHSTART " + ("x" * 4000) + " MATCHEND"
    rows = _rows()
    excerpt = dict(_excerpt(work_side=True))
    excerpt["frag_before"] = "B" * 4000
    excerpt["frag_span"] = span
    excerpt["frag_after"] = "A" * 4000
    rows[0]["excerpt"] = excerpt
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    _h, out = _rich_rows(data)
    row = next(r for r in out if "MATCHSTART" in
               "".join(t for t, _ in _runs(r["Manuscript passage"])))
    body = "".join(t for t, _ in _runs(row["Manuscript passage"]))

    assert span in body, "the matched span was abbreviated"
    # The span's own letters are discounted -- it is deliberately not clipped,
    # so counting them would be measuring the thing the previous line asserts.
    assert body.count("B") - span.count("B") <= des._CONTEXT_CHARS, (
        "the before context was not clipped")
    assert body.count("A") - span.count("A") <= des._CONTEXT_CHARS, (
        "the after context was not clipped")
    assert body.count(des._ELLIPSIS) == 2, (
        "context was cut without saying so on both sides")


@pytest.mark.parametrize("layer,expected", (("htr", True), ("fgp", False),
                                            (None, False)))
def test_the_automated_qualifier_marks_only_the_automated_layer(layer, expected):
    """`_render_panes` appends the qualifier for `text_layer == 'htr'` only;
    FGP and PGP transcriptions are human work. The first draft stamped it on
    every row, which asserted machine reading over a scholar's transcript."""
    rows = _rows()
    excerpt = dict(_excerpt(work_side=True))
    excerpt["text_layer"] = layer
    rows[0]["excerpt"] = excerpt
    data = des.build_findings_workbook(_envelope(rows), lang="en")
    note = ds.excerpt_strings("en")["frag_htr_note"]
    assert (note in _cells(data)) is expected


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_locus_cell_carries_the_address_without_the_layout_cue(lang):
    """The arrow is layout: on a row it says "this belongs to the title above
    me", which a column header states already. In a cell it breaks sorting and
    copy-paste and reached every locus of the first export."""
    rows = _rows()
    labelled = [r for r in rows if str(r.get("locus_label") or "").strip()]
    assert labelled, "no fixture row carries a locus; the test is vacuous"
    data = des.build_findings_workbook(_envelope(rows), lang=lang)
    cells = _cells(data)
    for row in labelled:
        label = str(row["locus_label"]).strip()
        assert label in cells, f"the locus {label!r} did not reach a cell"
        assert ds.locus_subline(label, lang) not in cells, (
            "the locus cell still carries the row-layout cue")
    assert not any("↳" in c for c in cells)


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_workbook_credits_midrash_from_the_one_canonical_citation(lang):
    """THE MANUSCRIPT SIDE OF EVERY PASSAGE HERE IS SOMEONE ELSE'S DATASET.
    The first export shipped thousands of lines of MiDRASH transcription and
    named nobody.

    Asserted against the SHARED constant, and against what the dossier's
    credits sheet prints, so the three copies cannot drift: a stale citation
    credits the wrong record, which is worse than no citation.
    """
    from shared.export_dossier import credits_lines
    from shared.export_utils import MIDRASH_CITATION_LINE

    data = des.build_findings_workbook(_envelope(), lang=lang)
    body = "\n".join(_cells(data))

    # LITERALS, DELIBERATELY -- one per credit line, spelled out here rather
    # than imported. The first version of this test asserted
    # `MIDRASH_CREDIT_LINES` against cells built FROM `MIDRASH_CREDIT_LINES`,
    # which compares a value with itself: the mutation battery changed the
    # dataset constant to "Dataset: MUTANT" and the test stayed GREEN. A gate
    # has to change with the ARTIFACT, not with the fixture that produced it,
    # and these three strings are what the published record actually says.
    for literal in (
        "MiDRASH Automatic Transcriptions (Stoekl Ben Ezra et al., 2025)",
        "Dataset: https://doi.org/10.5281/zenodo.17734473",
        "Citation: Stoekl Ben Ezra, D., Bambaci, L., Kiessling, B.",
        "Zenodo. https://doi.org/10.5281/zenodo.17734473",
    ):
        assert literal in body, f"the workbook omits {literal[:48]!r}"

    # ONE SOURCE, not three copies: whatever the dossier's credits sheet prints
    # as its citation is what this workbook prints. The literals above pin the
    # content; this pins the agreement.
    assert MIDRASH_CITATION_LINE in body
    assert MIDRASH_CITATION_LINE in credits_lines(lang)
    # ... and it is NEVER translated, so it is the same on both workbooks.
    assert MIDRASH_CITATION_LINE in credits_lines("en" if lang == "he" else "he")


# ===========================================================================
# Owner report 2026-08-21 (1) -- a grouped unit exported an empty file
# ===========================================================================

def _grouped_service(monkeypatch):
    """A service whose only real behaviour is `get_findings_enveloped`.

    STUBBED AT THAT ONE SEAM ON PURPOSE. The property under test is which
    GRAIN the collector walks and in what order it emits, and a stub lets the
    two grains be told apart unmistakably -- the grouped rows carry the NULLs
    the real `_FINDINGS_UNIT_SELECT` writes into them, which is the whole
    defect. Driving real SQL here would prove the SELECT lists work, which the
    service's own suite already does.
    """
    svc = DiscoveryService(path_provider=lambda: ":memory:",
                           availability_callable=lambda: True)
    seen = []

    # Two works, three identifications, in an order the leaf grain does NOT
    # produce on its own -- so a passing sort cannot be a coincidence.
    groups = [
        {"display_work_id": "w-B", "neutral_title": "Work B", "sys_id": None,
         "shelfmark_display": None, "identification_id": None},
        {"display_work_id": "w-A", "neutral_title": "Work A", "sys_id": None,
         "shelfmark_display": None, "identification_id": None},
    ]
    leaves = [
        {"identification_id": "i-1", "display_work_id": "w-A",
         "neutral_title": "Work A", "sys_id": "s-1", "shelfmark_display": "T-S 1"},
        {"identification_id": "i-2", "display_work_id": "w-B",
         "neutral_title": "Work B", "sys_id": "s-2", "shelfmark_display": "T-S 2"},
        {"identification_id": "i-3", "display_work_id": "w-A",
         "neutral_title": "Work A", "sys_id": "s-3", "shelfmark_display": "T-S 3"},
    ]

    def _enveloped(unit="identification", page=1, page_size=200, **kw):
        seen.append(unit)
        rows = groups if unit == "work" else leaves
        return {"status": STATUS_OK, "items": [dict(r) for r in rows],
                "total": len(rows), "meta": {"unit": unit}}

    monkeypatch.setattr(svc, "get_findings_enveloped", _enveloped)
    monkeypatch.setattr(svc, "_export_artifact_identity", lambda: ("p", "v"))
    monkeypatch.setattr(svc, "_query_excerpts_for_identifications",
                        lambda ids: {i: {"frag_span": "text"} for i in ids})
    return svc, seen


def test_a_grouped_unit_exports_the_rows_its_expander_opens_onto(monkeypatch):
    """`unit=work` and `unit=manuscript` are GROUP BY grains whose SELECT lists
    write NULL into `identification_id`, `sys_id` and `shelfmark_display`. On
    the page that is complete because the row carries an EXPANDER; in a
    spreadsheet it was a title and two counts -- "useless xlsx, no expandable
    id ... so no ms id and not text" (owner report, 2026-08-21).

    So the export walks the group grain for its ORDER and the leaf grain for
    its ROWS, which is the expander's own contract.
    """
    svc, seen = _grouped_service(monkeypatch)
    env = svc.collect_findings_for_export(unit="work")

    assert seen == ["work", "identification"], (
        f"expected a group walk then a leaf walk, got {seen}")
    items = env["items"]
    assert len(items) == 3, "the file is the leaves, not the groups"
    assert all(r.get("shelfmark_display") for r in items), (
        "a row reached the file with no manuscript on it -- the defect")
    assert all(r.get("identification_id") for r in items), (
        "a row reached the file with no identification id, so it can carry no "
        "text either")
    assert all(r.get("excerpt") for r in items)

    # ... in the GROUP's order, and stable inside each group.
    assert [r["identification_id"] for r in items] == ["i-2", "i-1", "i-3"], (
        "the leaves are not grouped the way the page grouped them")

    meta = env["meta"]
    assert meta["export_unit_requested"] == "work"
    assert meta["export_grain"] == "identification"
    assert meta["export_group_count"] == 2
    assert meta["walk_complete"] is True, (
        "completeness must be measured against the LEAF total, not the group "
        "total")


def test_the_leaf_grain_is_walked_once_and_unchanged_for_an_ungrouped_unit(
        monkeypatch):
    """No second walk, and no reordering, when the reader is already on the
    leaf grain: the two-walk path must not become the only path."""
    svc, seen = _grouped_service(monkeypatch)
    env = svc.collect_findings_for_export(unit="identification")
    assert seen == ["identification"]
    assert [r["identification_id"] for r in env["items"]] == ["i-1", "i-2", "i-3"]
    assert env["meta"]["export_group_count"] is None


@pytest.mark.parametrize("lang", ("en", "he"))
def test_the_about_sheet_says_what_one_row_is_when_it_is_not_what_was_asked(
        monkeypatch, lang):
    """A file of tens of thousands of rows headed "one row per work" is a
    puzzle unless it explains itself."""
    svc, _seen = _grouped_service(monkeypatch)
    env = svc.collect_findings_for_export(unit="work")
    body = " ".join(_cells(des.build_findings_workbook(env, lang=lang)))
    assert des._pick({"en": "One row is", "he": "כל שורה היא"}, lang) in body
    assert des._pick(des._UNIT_NOUNS["work"], lang) in body
    # ... and it does NOT appear when the grain is what the reader chose.
    plain = " ".join(_cells(des.build_findings_workbook(
        svc.collect_findings_for_export(unit="identification"), lang=lang)))
    assert des._pick({"en": "One row is", "he": "כל שורה היא"}, lang) not in plain


@pytest.mark.parametrize("lang", ("en", "he"))
@pytest.mark.parametrize("groups,shape", ((None, "unknown"), (1, "one"),
                                          (2, "many"), (28635, "many")))
def test_the_grain_sentence_reads_as_a_sentence_at_every_count(
        groups, shape, lang):
    """A provenance line that reads like a mail merge invites a reader to skip
    the rest of the sheet -- and the single-group case really did produce "the
    identifications behind those 1 rows"."""
    out = des._grain_sentence(groups, "work", lang)
    assert "{" not in out and "}" not in out, f"unsubstituted placeholder: {out}"
    assert des._pick(des._UNIT_NOUNS["work"], lang) in out
    if shape == "many":
        assert "{:,}".format(groups) in out, (
            f"the group count is missing from {out!r}")
    else:
        # No bare number at all: neither "1" nor an empty gap where one was.
        assert not any(ch.isdigit() for ch in out), out
    assert "  " not in out, f"a gap where the count was: {out!r}"


def test_a_short_group_walk_costs_the_order_and_says_so_without_lying_about_rows(
        monkeypatch):
    """A truncated GROUP walk leaves some leaves unplaced. That is a
    presentation degradation, not a missing row, and folding it into
    `walk_complete` would mark a complete file incomplete -- the opposite of
    the error this export is written against, but still a false statement."""
    svc, _seen = _grouped_service(monkeypatch)
    real = svc.get_findings_enveloped

    def _partial(unit="identification", **kw):
        env = real(unit=unit, **kw)
        if unit == "work":
            env["items"] = env["items"][:1]     # w-B only; w-A unknown
        return env

    monkeypatch.setattr(svc, "get_findings_enveloped", _partial)
    env = svc.collect_findings_for_export(unit="work")
    assert len(env["items"]) == 3, "rows were lost to an ordering problem"
    assert env["meta"]["walk_complete"] is True
    assert env["meta"]["export_group_order_complete"] is False
    # the placed group leads; the unplaced ones keep their own order at the end
    assert [r["identification_id"] for r in env["items"]] == ["i-2", "i-1", "i-3"]
    body = " ".join(_cells(des.build_findings_workbook(env, lang="en")))
    assert "appear at the end" in body



@pytest.mark.parametrize("lang", ("en", "he"))
def test_an_approximate_group_count_is_named_as_a_ceiling_not_a_count(
        monkeypatch, lang):
    """A capped group count must not be printed as a count -- and must not drag
    the rows-moved sentence along with it.

    `DISCOVERY_FINDINGS_COUNT_MAX` turns a total into a CEILING and flags it.
    The About sheet prints `export_group_count` inside a flat sentence ("one row
    per work, of N works"), so a capped number arrives there stated as fact --
    the defect CLAUDE.md calls "a correctness defect, not a tuning choice", in
    the one place a reader cannot check it against the page.

    An approximate count does NOT shorten the walk: the bound is dropped and it
    still ends on a short page. So every row IS placed, and the sentence about
    rows appearing at the end -- which the umbrella flag used to trigger -- was
    telling the reader to look for something that was not there.
    """
    svc, _seen = _grouped_service(monkeypatch)
    real = svc.get_findings_enveloped

    def _capped(unit="identification", **kw):
        env = real(unit=unit, **kw)
        if unit == "work":
            env["meta"] = dict(env["meta"], approximate_total=True)
        return env

    monkeypatch.setattr(svc, "get_findings_enveloped", _capped)
    env = svc.collect_findings_for_export(unit="work")
    meta = env["meta"]

    # The umbrella flag stays conservative. This is the only assertion in the
    # suite that reaches `not (g_approx or g_guard)`.
    assert meta["export_group_order_complete"] is False
    assert meta["export_group_count_approximate"] is True
    # ... and yet nothing moved: every leaf found its group.
    assert meta["export_group_unplaced"] == 0
    assert [r["identification_id"] for r in env["items"]] == ["i-2", "i-1", "i-3"]
    assert meta["walk_complete"] is True, "the ROWS are complete and must say so"

    body = " ".join(_cells(des.build_findings_workbook(env, lang=lang)))
    ceiling = {"en": "ceiling, not a count", "he": "תקרה ולא ספירה"}[lang]
    at_end = {"en": "appear at the end", "he": "מופיעות בסוף"}[lang]
    assert ceiling in body, "a capped group count is stated as if it were one"
    assert at_end not in body, (
        "the file tells the reader some rows appear at the end when every row "
        "was placed in its group")


def test_a_group_walk_that_trips_its_page_guard_leaves_the_order_uncertified(
        monkeypatch):
    """The runaway guard is the OTHER way a grouping stops being certifiable.

    Groups past the guard were never seen, so leaves in them could not be
    placed. Here they all happen to be placed anyway, which is the point: the
    flag must come from the walk having been cut short, not from noticing
    afterwards that something failed to sort.
    """
    svc, seen = _grouped_service(monkeypatch)
    real = svc.get_findings_enveloped

    def _runaway(unit="identification", page=1, page_size=200, **kw):
        env = real(unit=unit, page=page, page_size=page_size, **kw)
        if unit == "work":
            # A FULL page every time, so the walk can only ever end on the
            # guard. The two real groups lead page 1, so every leaf still
            # places and the guard is the sole cause under test.
            filler = [{"display_work_id": "w-%d-%d" % (page, i),
                       "neutral_title": "Filler", "sys_id": None,
                       "shelfmark_display": None, "identification_id": None}
                      for i in range(page_size - 2)]
            env["items"] = [dict(r) for r in env["items"]] + filler
            env["total"] = 1          # guard = 1 page of slack + 8
        return env

    monkeypatch.setattr(svc, "get_findings_enveloped", _runaway)
    env = svc.collect_findings_for_export(unit="work")
    meta = env["meta"]

    assert seen.count("work") == 9, (
        f"the guard did not stop the group walk (pages={seen.count('work')})")
    assert meta["export_group_order_complete"] is False
    assert meta["export_group_unplaced"] == 0
    assert meta["export_group_count_approximate"] is False
    # THE ROWS ARE UNTOUCHED: a group-side degradation must never be reported
    # as a short file.
    assert meta["walk_complete"] is True
    assert [r["identification_id"] for r in env["items"]] == ["i-2", "i-1", "i-3"]
    body = " ".join(_cells(des.build_findings_workbook(env, lang="en")))
    assert "appear at the end" not in body



def test_a_capped_total_does_not_bound_the_walk(monkeypatch):
    """`DISCOVERY_FINDINGS_COUNT_MAX` must not silently halve an export.

    A capped `total` is a CEILING the counter stopped at, not a measurement.
    `_export_walk` already drops it as the STOPPING CONDITION and says so in a
    comment -- but the runaway page guard was still derived from it, so a cap of
    N stopped the walk about eight pages past N and returned an `ok` envelope
    for a file the reader asked to be whole (Codex review of PR #322, P1).

    Here the cap is 1 and 13 pages match. Under the old bound the walk stopped
    at 9 pages and reported a tripped guard; it must now collect every row.
    """
    svc = DiscoveryService(path_provider=lambda: ":memory:",
                           availability_callable=lambda: True)
    pages = 13

    def _enveloped(unit="identification", page=1, page_size=200, **kw):
        # Full pages until the last, which is short -- the only honest end.
        count = page_size if page < pages else 5
        rows = [{"identification_id": "i-%d-%d" % (page, i)}
                for i in range(count)]
        return {"status": STATUS_OK, "items": rows,
                # The CAP, flagged. Deliberately far below the real count.
                "total": 1,
                "meta": {"unit": unit, "approximate_total": True}}

    monkeypatch.setattr(svc, "get_findings_enveloped", _enveloped)
    monkeypatch.setattr(svc, "_export_artifact_identity", lambda: ("p", "v"))
    monkeypatch.setattr(svc, "_query_excerpts_for_identifications", lambda ids: {})

    env = svc.collect_findings_for_export(unit="identification",
                                          with_excerpts=False)
    expected = 200 * (pages - 1) + 5
    assert len(env["items"]) == expected, (
        f"the walk stopped at {len(env['items'])} of {expected} rows -- the "
        "capped total is bounding it again")
    # `None`, not True: the rows cannot be checked against a number that was
    # never a count. What must NOT appear is False, which would mean the guard
    # tripped and the file is knowingly short.
    assert env["meta"]["walk_complete"] is None, (
        "a capped total must leave completeness UNVERIFIED, never asserted "
        "either way")



def test_a_leaf_walk_that_trips_its_guard_refuses_rather_than_shipping_short(
        monkeypatch):
    """FAILS WHOLE, NEVER SHORT -- including when the runaway guard is what
    stopped it.

    A tripped guard on the LEAF walk means rows are missing. The collector used
    to return `ok` and let the file download with "NO -- this file is
    incomplete" on the About sheet, which is not enough for an endpoint whose
    contract is the whole filtered set: a reader who never opens that sheet
    cannot tell a truncated download from a small result set, and the file
    outlives the request that made it (Codex review of PR #323).
    """
    svc = DiscoveryService(path_provider=lambda: ":memory:",
                           availability_callable=lambda: True)

    def _runaway(unit="identification", page=1, page_size=200, **kw):
        # NEVER a short page: the walk can only ever end on the guard.
        return {"status": STATUS_OK,
                "items": [{"identification_id": "i-%d-%d" % (page, i)}
                          for i in range(page_size)],
                "total": 1, "meta": {"unit": unit}}

    monkeypatch.setattr(svc, "get_findings_enveloped", _runaway)
    monkeypatch.setattr(svc, "_export_artifact_identity", lambda: ("p", "v"))
    monkeypatch.setattr(svc, "_query_excerpts_for_identifications", lambda ids: {})

    env = svc.collect_findings_for_export(unit="identification",
                                          with_excerpts=False)
    assert env["status"] != STATUS_OK, (
        "a knowingly truncated walk produced a downloadable workbook")
    assert env["meta"]["reason"] == "export_walk_incomplete"
    assert not env.get("content"), "a refused export must carry no bytes"
    # The builder is the second line of defence and must also refuse it.
    with pytest.raises(ValueError):
        des.build_findings_workbook(env)


def test_the_page_and_the_export_group_on_one_table():
    """The page opens a grouped row onto its children with
    `EXPANSION_KEY_BY_UNIT`; the export flattens the same grouping with it.
    Two copies on opposite sides of the layering guard would drift silently --
    an export grouping on a column the page no longer expands on still
    produces a plausible file."""
    from shared.discovery_service import (
        EXPANSION_KEY_BY_UNIT, EXPORT_GROUPED_UNITS)
    from web.components import findings_rows as fr

    assert fr.EXPANSION_KEY_BY_UNIT is EXPANSION_KEY_BY_UNIT, (
        "the page and the export are looking at two different tables")
    assert EXPORT_GROUPED_UNITS == {"work", "manuscript"}
    assert EXPANSION_KEY_BY_UNIT["identification"] is None


def test_it_refuses_to_build_from_a_non_ok_envelope():
    """NO PARTIAL WORKBOOK, EVER. A short file is indistinguishable from a small
    result set once it has been downloaded, so the builder raises and the route
    turns the status into an HTTP response instead."""
    for status in ("timeout", "busy", "unavailable"):
        broken = dict(_envelope())
        broken["status"] = status
        with pytest.raises(ValueError):
            des.build_findings_workbook(broken)


def test_the_filename_carries_the_artifact_and_never_reader_text():
    """A findings file is only reproducible against the artifact that produced
    it -- and a filename built from user text is the shape that leaked one
    reader's query into another reader's download on the search export."""
    name = des.workbook_filename({"sidecar_version": "discovery-v1-real"})
    assert name == "computed-identifications-discovery-v1-real.xlsx"
    assert des.workbook_filename({}) == "computed-identifications.xlsx"
    dirty = des.workbook_filename({"sidecar_version": "a/b c;\\d"})
    assert "/" not in dirty and " " not in dirty and ";" not in dirty


# ===========================================================================
# EXPORT-01 -- the URL is the reader's view, and is stateless
# ===========================================================================

def test_the_export_url_carries_the_filters_and_never_a_page():
    import web.pages.findings as fp
    state = {
        "unit": "identification", "bucket": "main", "sort": "band_rank",
        "novelty_view": "all", "domain": "liturgy", "author": "someone",
        "work_id": "w000001", "work_label": "x",
        "locus_from": 3, "locus_to": 9, "page": 7,
    }
    params = fp.export_query_params(state, "he")
    assert params["unit"] == "identification"
    assert params["domain"] == "liturgy"
    assert params["work_id"] == "w000001"
    assert params["locus_from"] == "3" and params["locus_to"] == "9"
    assert params["lang"] == "he"
    # THE POINT OF THE TEST. The file is the whole filtered set; sending the
    # reader's page number is how a download quietly stops at 50 rows.
    assert "page" not in params
    url = fp.export_url(state, "he")
    assert url.startswith("/api/export/computed-identifications?")
    assert "page=" not in url


# ===========================================================================
# EXPORT-03 -- bounded, batched, and whole-or-nothing
# ===========================================================================

def test_the_export_has_its_own_budget_and_its_own_executor():
    """Two semaphores over one pool are two names for one budget -- that already
    cost a fix here. A third budget class has to bring its own executor, sized
    to itself, or a slot stops guaranteeing a worker."""
    svc = DiscoveryService(path_provider=lambda: None)
    assert svc._SLOT_EXPORT in svc._SLOT_SPECS
    assert svc._export_capacity >= 1
    executor = svc._executor_for(svc._SLOT_EXPORT)
    assert executor._max_workers == svc._export_capacity
    assert executor is not svc._executor_for(svc._SLOT_HEAVY)
    assert executor is not svc._executor_for(svc._SLOT_BROWSE)


def test_an_unknown_slot_name_raises_rather_than_falling_back():
    """A typo must not silently land on the browse budget."""
    import asyncio
    svc = DiscoveryService(path_provider=lambda: None)
    with pytest.raises(ValueError):
        asyncio.get_event_loop_policy().new_event_loop().run_until_complete(
            svc._run_off_loop(lambda: None, timeout=1, slot="not-a-budget"))


def _excerpt_service(rows: int) -> DiscoveryService:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute(
        "CREATE TABLE discovery_excerpt (identification_id TEXT, evidence_id INT,"
        " a_page_id TEXT, frag_before TEXT, frag_span TEXT, frag_after TEXT,"
        " frag_clipped INT, work_before TEXT, work_span TEXT, work_after TEXT,"
        " work_clipped INT, work_source TEXT, attribution TEXT, n_spans INT,"
        " text_layer TEXT, frag_hl TEXT, work_hl TEXT, work_markup TEXT)")
    conn.executemany(
        "INSERT INTO discovery_excerpt (identification_id, evidence_id, frag_span,"
        " frag_hl) VALUES (?,?,?,?)",
        [(f"id-{i}", i, "span", "[[0,1]]") for i in range(rows)])
    svc = DiscoveryService(path_provider=lambda: ":memory:",
                           availability_callable=lambda: True)
    svc._get_conn = lambda: conn            # type: ignore[assignment]
    return svc


def test_the_excerpt_read_is_batched_not_one_query_per_row(monkeypatch):
    """Uncapped over the main pool a per-row read is tens of thousands of
    serialized SQLite round trips inside ONE request on a server with a SINGLE
    uvicorn worker. The chunked `IN (...)` shape is the same one the
    citation-range fix landed on."""
    monkeypatch.setenv("DISCOVERY_EXPORT_EXCERPT_CHUNK", "50")
    svc = _excerpt_service(120)
    conn = svc._get_conn()
    calls: List[str] = []

    class _CountingConn:
        """`sqlite3.Connection.execute` is read-only, so the count is taken on a
        proxy rather than by patching the driver."""

        def __init__(self, inner):
            self._inner = inner

        def execute(self, sql, *args, **kwargs):
            calls.append(sql)
            return self._inner.execute(sql, *args, **kwargs)

        def __getattr__(self, name):
            return getattr(self._inner, name)

    proxy = _CountingConn(conn)
    svc._get_conn = lambda: proxy  # type: ignore[assignment]
    found = svc._query_excerpts_for_identifications([f"id-{i}" for i in range(120)])

    assert len(found) == 120, "the batched read lost rows"
    assert len(calls) == 3, (
        f"120 ids at a chunk of 50 must be 3 queries, not {len(calls)}")
    # The decode is shared with the single-row reader, so it must have run.
    assert found["id-0"]["frag_hl"] == [[0, 1]]


def test_an_id_with_no_excerpt_is_absent_rather_than_a_placeholder():
    """The caller has to tell "no excerpt row" apart from "a row whose work side
    is empty"; collapsing them here would hand the export a `None` it could not
    classify."""
    svc = _excerpt_service(3)
    found = svc._query_excerpts_for_identifications(["id-0", "id-missing"])
    assert "id-0" in found
    assert "id-missing" not in found


def test_the_walk_fails_whole_rather_than_returning_a_short_set(monkeypatch):
    """A page that comes back degraded aborts the export and returns THAT
    status. The alternative -- a workbook holding the pages that happened to
    succeed -- is the defect the envelope exists to prevent, in the one place
    where the reader keeps the artifact."""
    svc = DiscoveryService(path_provider=lambda: None, availability_callable=lambda: True)
    pages = [
        make_envelope(STATUS_OK, [{"identification_id": "a"}] * 200, 400, meta={}),
        {"status": "timeout", "items": [], "total": None, "meta": {}},
    ]
    seen: List[int] = []

    def _fake(**kwargs):
        seen.append(kwargs.get("page"))
        return pages[len(seen) - 1]

    monkeypatch.setattr(svc, "get_findings_enveloped", _fake)
    monkeypatch.setattr(svc, "_clamp_page_size", staticmethod(lambda n: 200))
    out = svc.collect_findings_for_export()
    assert out["status"] == "timeout", "a degraded page produced an ok envelope"
    assert not out["items"], "a partial result set escaped the failed walk"
    assert seen == [1, 2]


def test_the_collector_reports_whether_the_walk_matched_the_reported_total(monkeypatch):
    """`walk_complete` is CARRIED rather than asserted, so a file built from a
    disagreeing walk says so on its own About sheet instead of looking whole."""
    svc = DiscoveryService(path_provider=lambda: None, availability_callable=lambda: True)
    monkeypatch.setattr(svc, "_clamp_page_size", staticmethod(lambda n: 200))
    monkeypatch.setattr(
        svc, "get_findings_enveloped",
        lambda **kw: make_envelope(STATUS_OK, [{"identification_id": "a"}], 99, meta={}))
    monkeypatch.setattr(svc, "_query_excerpts_for_identifications", lambda ids: {})
    out = svc.collect_findings_for_export()
    assert out["meta"]["walk_complete"] is False
    assert out["meta"]["row_count"] == 1
    assert out["meta"]["reported_total"] == 99

    data = des.build_findings_workbook(out, lang="en")
    body = "\n".join(_cells(data))
    assert "NO — this file is incomplete" in body, (
        "an incomplete export does not say so on its About sheet")

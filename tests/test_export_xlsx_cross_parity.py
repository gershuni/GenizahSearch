"""Phase 94 MUST-FIX 94-04-C: cross-parity regression test for EXPORT-META-09.

Asserts the web app (via :class:`web.export_service.ExportService`) AND the
desktop app (via :func:`genizah_app._build_search_results_xlsx_bytes`)
produce workbooks with IDENTICAL sheet names and IDENTICAL header rows on
identical input. If this test fails, the identical-structure parity
invariant is broken.

Scope (intentional): this parity test pins SHEET NAMES + HEADER ROWS only.
It does NOT pin cell-value identity for data rows. Rationale: web reads
``Full Text`` via ``web/export_service._resolve_result_full_text(res)``
(Tantivy-indexed page text resolved via meta_mgr), while desktop reads via
the ``full_text_fetcher`` callback (MUST-FIX 94-04-D). These two sources
can legitimately produce different strings for the same sys_id (e.g.,
page-text vs full-document text), so a strict cell-value parity assertion
would either fail spuriously or require fragile mock alignment. Identical
structure (EXPORT-META-09) covers what the user can OBSERVE on tab-name +
column-header — that's what the parity gate enforces. Functional drift in
data cells is caught by the per-app unit tests
(:file:`tests/test_export_service_multi_sheet.py` for web,
:file:`tests/test_desktop_xlsx_multi_sheet.py` for desktop).
"""
from io import BytesIO

import openpyxl
import pytest


# Shared fixtures (intentional code duplication with the wave-3 + wave-4 test
# files so this test stays self-contained — the parity check is the whole
# point, so it must NOT reuse a single fixture that could mask drift).
class _FakeMetaMgr:
    def get_meta_for_id(self, sys_id):
        return (f'T-S {sys_id[-4:]}', f'Title for {sys_id}')

    def get_library_for_id(self, sys_id):
        return 'CUL'

    def parse_full_id_components(self, header):
        return {}


def _meta_resolver_fake(sid):
    if not sid:
        return None
    return {
        'shelfmark': f'T-S {sid[-4:]}',
        'title': f'Title {sid}',
        'library_code': 'CUL',
        'library_name': 'Cambridge University Library',
    }


def _identity_sanitize(text):
    return '' if text is None else str(text)


def _make_result(sys_id='99001234567890'):
    return {
        'uid': f'IE_x_P_y_FL_{sys_id}',
        'display': {
            'id': sys_id,
            'shelfmark': f'T-S {sys_id[-4:]}',
            'title': f'Title {sys_id}',
            'library_code': 'CUL',
            'img': '1r',
            'source': 'ms',
        },
        'snippet': 'foo *bar* baz',
        'raw_header': '',
        'sort_score': 1.0,
        'full_text': 'page text',
    }


@pytest.fixture
def stub_dossier(monkeypatch):
    from shared import export_dossier
    monkeypatch.setattr(export_dossier, 'pgp_subset_for_sys_id', lambda s: {
        'pgp_url': f'https://pgp.example/{s}',
        'description': f'PGP desc {s}', 'document_type': 'Letter',
        'date_display': '1100', 'languages': ['Hebrew', 'Aramaic'],
        'tags': ['letter'],
    } if s else None)
    monkeypatch.setattr(export_dossier, 'nli_subset_for_sys_id', lambda s: {
        'catalog_entry': f'Neubauer {s[-4:]}',
        'library_viewer_url': f'https://cudl.example/{s}',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'catalog_summary_for_sys_id', lambda s: {
        'title': f'CatTitle {s}', 'author_text': 'Author',
        'copy_date': '1180', 'copy_place': 'Fustat',
    } if s else None)
    monkeypatch.setattr(export_dossier, 'bibliography_for_sys_id', lambda s: [{
        'running_title': 'Med. Soc.', 'title_year': 1967, 'mention_page': '123',
        'article_name': 'Letter', 'article_author_eng': 'Goitein',
        'catalog_acronym': 'MS',
    }] if s == '99001234567890' else [])


def _build_web(results):
    from web.export_service import ExportService
    svc = ExportService(_FakeMetaMgr())
    content, _ = svc.export_search_results_excel(results, 'q')
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


def _build_desktop(results):
    from genizah_app import _build_search_results_xlsx_bytes
    content = _build_search_results_xlsx_bytes(
        results=results,
        headers_main=['System ID', 'Library', 'Shelfmark', 'Title',
                      'Image/Page', 'Source', 'Snippet', 'Full Text',
                      'Has PGP', 'Is Printed', 'Domains', 'IIIF Manifest'],
        meta_resolver=_meta_resolver_fake,
        sanitize_fn=_identity_sanitize,
    )
    return openpyxl.load_workbook(BytesIO(content), rich_text=True)


def _find_header_row(ws):
    for r in range(1, 25):
        if ws.cell(r, 1).value == 'System ID':
            return r
    return None


def test_sheet_names_identical(stub_dossier):
    # MUST-FIX 94-04-C: web + desktop both produce identical sheet names
    # in identical order on identical input.
    results = [_make_result('99001234567890')]
    wb_web = _build_web(results)
    wb_desktop = _build_desktop(results)
    assert wb_web.sheetnames == wb_desktop.sheetnames, (
        f"PARITY VIOLATION: web sheetnames={wb_web.sheetnames} "
        f"vs desktop sheetnames={wb_desktop.sheetnames}"
    )
    assert wb_web.active.title == wb_desktop.active.title


def test_main_sheet_headers_byte_identical(stub_dossier):
    results = [_make_result('99001234567890')]
    wb_web = _build_web(results)
    wb_desktop = _build_desktop(results)
    sheet_name = 'Genizah Results'
    hr_web = _find_header_row(wb_web[sheet_name])
    hr_desktop = _find_header_row(wb_desktop[sheet_name])
    assert hr_web and hr_desktop, 'header rows not found in one of the workbooks'
    web_headers = [wb_web[sheet_name].cell(hr_web, c).value for c in range(1, 13)]
    desktop_headers = [wb_desktop[sheet_name].cell(hr_desktop, c).value for c in range(1, 13)]
    assert web_headers == desktop_headers, (
        f"PARITY VIOLATION: web headers={web_headers} "
        f"vs desktop headers={desktop_headers}"
    )


def test_manuscripts_sub_sheet_headers_identical(stub_dossier):
    results = [_make_result('99001234567890')]
    wb_web = _build_web(results)
    wb_desktop = _build_desktop(results)
    web_h = [wb_web['Manuscripts'].cell(1, c).value for c in range(1, 15)]
    desktop_h = [wb_desktop['Manuscripts'].cell(1, c).value for c in range(1, 15)]
    assert web_h == desktop_h, (
        f"PARITY VIOLATION: web manuscripts headers={web_h} "
        f"vs desktop manuscripts headers={desktop_h}"
    )


def test_bibliography_sub_sheet_headers_identical(stub_dossier):
    results = [_make_result('99001234567890')]
    wb_web = _build_web(results)
    wb_desktop = _build_desktop(results)
    web_h = [wb_web['Bibliography'].cell(1, c).value for c in range(1, 9)]
    desktop_h = [wb_desktop['Bibliography'].cell(1, c).value for c in range(1, 9)]
    assert web_h == desktop_h, (
        f"PARITY VIOLATION: web bibliography headers={web_h} "
        f"vs desktop bibliography headers={desktop_h}"
    )

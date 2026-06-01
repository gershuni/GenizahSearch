# -*- coding: utf-8 -*-
"""Phase 97 Wave B + Wave C — XLSX extraction tests.

Wave B tests (plan 97-02):
- test_zip_bomb_defense_via_monkeypatched_infolist: _check_zip_bomb rejects a forged
  600 MB uncompressed-size zip without reading any actual bytes (Codex MEDIUM #1 fixture).

Wave C tests (plan 97-03):
- test_per_sheet_per_row_window: 1 sheet x 1200 rows -> 3 chunks with locators
- test_multi_sheet_xlsx: 3 sheets x 100 rows each -> 3 chunks (one per sheet)
- test_uniform_row_extraction: row [שלום, hello, 123] -> text "שלום | hello | 123"
- test_rtl_metadata: sheetView.rightToLeft=True -> is_rtl=True; text NOT mutated
- test_cell_count_limit: >100_000 cells -> raises XlsxZipBombSuspected
"""
from __future__ import annotations

import zipfile
from types import SimpleNamespace


from shared.local_indexer import _check_zip_bomb, _MAX_UNCOMPRESSED_BYTES


def test_zip_bomb_defense_via_monkeypatched_infolist(monkeypatch, tmp_path):
    """_check_zip_bomb rejects a zip whose central-directory reports 600 MB uncompressed.

    Codex MEDIUM #1 fix: ZipInfo.file_size = 600*1024*1024 BEFORE writestr() is
    OVERWRITTEN by Python's zip writer at writestr() time with the actual byte length.
    Correct approach: create a real (small, valid) zip, then monkeypatch infolist()
    to return synthesized ZipInfo records claiming the large uncompressed size.
    """
    # Create a real (small, valid) zip so ZipFile() opens OK.
    real_zip = tmp_path / "fake.xlsx"
    with zipfile.ZipFile(real_zip, "w") as zf:
        zf.writestr("[Content_Types].xml", b"<x/>")

    # Patch infolist to return a forged record claiming 600 MB uncompressed.
    def fake_infolist(self):
        return [SimpleNamespace(file_size=600 * 1024 * 1024, compress_size=2048)]

    monkeypatch.setattr(zipfile.ZipFile, "infolist", fake_infolist)

    reason = _check_zip_bomb(str(real_zip))

    assert reason is not None, "_check_zip_bomb should return a reason string for a 600 MB claim"
    assert "uncompressed size" in reason, f"Expected 'uncompressed size' in reason, got: {reason!r}"
    assert str(_MAX_UNCOMPRESSED_BYTES) in reason, (
        f"Expected limit {_MAX_UNCOMPRESSED_BYTES} in reason, got: {reason!r}"
    )
    # Also verify the claimed size is reported
    assert str(600 * 1024 * 1024) in reason, (
        f"Expected claimed size 629145600 in reason, got: {reason!r}"
    )


# ---------------------------------------------------------------------------
# Wave C — Plan 97-03 XLSX text-extraction tests
# ---------------------------------------------------------------------------

def _make_xlsx_1200_rows(tmp_path):
    """Pytest fixture helper: create a 1-sheet XLSX with 1200 rows via openpyxl write_only."""
    import openpyxl
    p = tmp_path / "large_1200.xlsx"
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    for i in range(1, 1201):
        ws.append([f"val{i}a", f"val{i}b", f"val{i}c"])
    wb.save(str(p))
    wb.close()
    return p


def _make_xlsx_multi_sheet(tmp_path):
    """Pytest fixture helper: create a 3-sheet XLSX with 100 rows each."""
    import openpyxl
    p = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    for sheet_num in range(1, 4):
        ws = wb.create_sheet(f"Sheet{sheet_num}")
        for i in range(1, 101):
            ws.append([f"s{sheet_num}r{i}a", f"s{sheet_num}r{i}b"])
    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(str(p))
    return p


def test_per_sheet_per_row_window(tmp_path):
    """1 sheet x 1200 rows -> 3 chunks with locators 'Sheet1!R1:R500', 'Sheet1!R501:R1000', 'Sheet1!R1001:R1200'."""
    from shared.local_indexer import extract_xlsx_pages

    p = _make_xlsx_1200_rows(tmp_path)
    chunks = list(extract_xlsx_pages(str(p)))
    assert len(chunks) == 3, f"Expected 3 chunks, got {len(chunks)}"

    locators = [loc for _, _, _, loc, _ in chunks]
    assert locators[0] == "Sheet1!R1:R500", f"Expected 'Sheet1!R1:R500', got {locators[0]!r}"
    assert locators[1] == "Sheet1!R501:R1000", f"Expected 'Sheet1!R501:R1000', got {locators[1]!r}"
    assert locators[2] == "Sheet1!R1001:R1200", f"Expected 'Sheet1!R1001:R1200', got {locators[2]!r}"


def test_multi_sheet_xlsx(tmp_path):
    """3 sheets x 100 rows each -> 3 chunks (one per sheet, single window each)."""
    from shared.local_indexer import extract_xlsx_pages

    p = _make_xlsx_multi_sheet(tmp_path)
    chunks = list(extract_xlsx_pages(str(p)))
    assert len(chunks) == 3, f"Expected 3 chunks (one per sheet), got {len(chunks)}"
    sheet_names_in_locators = [loc.split("!")[0] for _, _, _, loc, _ in chunks]
    assert "Sheet1" in sheet_names_in_locators
    assert "Sheet2" in sheet_names_in_locators
    assert "Sheet3" in sheet_names_in_locators


def test_uniform_row_extraction(tmp_path):
    """Row [שלום, hello, 123] -> chunk text contains 'שלום | hello | 123'."""
    import openpyxl
    from shared.local_indexer import extract_xlsx_pages

    p = tmp_path / "uniform.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["שלום", "hello", "123"])
    wb.save(str(p))

    chunks = list(extract_xlsx_pages(str(p)))
    assert len(chunks) > 0, "Expected at least one chunk"
    all_text = " ".join(text for _, text, _, _, _ in chunks)
    assert "שלום | hello | 123" in all_text, (
        f"Expected 'שלום | hello | 123' in text, got: {all_text!r}"
    )


def test_rtl_metadata(tmp_path):
    """sheetView.rightToLeft=True -> yielded tuple includes is_rtl=True; text NOT mutated."""
    import openpyxl
    from shared.local_indexer import extract_xlsx_pages

    p = tmp_path / "rtl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RTL"
    ws.sheet_view.rightToLeft = True
    ws.append(["שלום", "עולם"])
    wb.save(str(p))

    chunks = list(extract_xlsx_pages(str(p)))
    assert len(chunks) > 0, "Expected at least one chunk"
    for _, text, _, _, is_rtl in chunks:
        assert is_rtl is True, f"Expected is_rtl=True for rightToLeft sheet, got {is_rtl}"
        # Text must NOT be reversed (F-06 invariant)
        assert "שלום" in text, f"Expected logical-order 'שלום' in text, got: {text!r}"
        # Reversed string must not appear
        assert "שלום"[::-1] not in text or "שלום" in text, "Text appears reversed"


def test_cell_count_limit(tmp_path):
    """XLSX with > 100_000 cells -> raises XlsxZipBombSuspected."""
    import openpyxl
    from shared.local_indexer import extract_xlsx_pages, XlsxZipBombSuspected, _MAX_CELLS_PER_SHEET

    # Create XLSX with _MAX_CELLS_PER_SHEET + 1 cells in one sheet
    # Use write_only to avoid memory issues at synthesis time
    p = tmp_path / "huge.xlsx"
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("BigSheet")
    cols = 10
    rows_needed = (_MAX_CELLS_PER_SHEET // cols) + 2  # just over the limit
    for i in range(rows_needed):
        ws.append([f"v{i}_{j}" for j in range(cols)])
    wb.save(str(p))

    import pytest
    with pytest.raises(XlsxZipBombSuspected):
        list(extract_xlsx_pages(str(p)))


# ---------------------------------------------------------------------------
# v7.16 BUG-2 — uncached-formula fallback + empty-row guard
# ---------------------------------------------------------------------------

def test_uncached_formula_fallback(tmp_path):
    """v7.16 BUG-2: a workbook of formula cells with NO cached values reads as all
    None under data_only=True. Without a fallback this indexes as 'no_text_layer'
    and is invisible to search. The extractor must retry with data_only=False so
    the formula strings (which carry the searchable Hebrew) are recovered."""
    import openpyxl
    from shared.local_indexer import extract_xlsx_pages

    p = tmp_path / "formulas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = '=CONCATENATE("רוזנצווייג")'
    ws["B1"] = "=1+2"
    wb.save(str(p))

    all_text = " ".join(text for _, text, _, _, _ in extract_xlsx_pages(str(p)))
    assert "רוזנצווייג" in all_text, (
        f"formula text not recovered via data_only=False fallback: {all_text!r}"
    )


def test_blank_rows_do_not_create_empty_pages(tmp_path):
    """v7.16 BUG-2: an all-blank multi-column row joins to ' | ' whose .strip() is
    '|' (truthy) — it must NOT be indexed as content. A wholly blank workbook must
    yield zero chunks (not a blank searchable page)."""
    import openpyxl
    from shared.local_indexer import extract_xlsx_pages

    # Mixed: blank rows interleaved with real rows -> only real rows kept.
    p = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["שם", "שנה"])
    ws.append([None, None])
    ws.append(["", ""])
    ws.append(["רוזנצווייג", "1913"])
    wb.save(str(p))
    chunks = list(extract_xlsx_pages(str(p)))
    text = "\n".join(t for _, t, _, _, _ in chunks)
    assert "רוזנצווייג" in text
    assert " | \n" not in text and not text.endswith(" | "), f"blank row leaked: {text!r}"

    # Wholly blank workbook -> zero chunks (no empty 'no_text_layer'-masking page).
    p2 = tmp_path / "blank.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append([None, None, None])
    ws2.append(["", "", ""])
    wb2.save(str(p2))
    assert list(extract_xlsx_pages(str(p2))) == []

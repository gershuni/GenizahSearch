# -*- coding: utf-8 -*-
"""Shared dossier helpers for research-grade xlsx export (Phase 94).

Both the web (`web/export_service.py`) and desktop (`genizah_app.py`) export
pipelines consume this module to emit the new `Manuscripts` and `Bibliography`
sub-sheets. Behavior is identical across both apps so the workbook structure
is uniform.

D-04 REVISION (2026-05-20, smoke verification gap fix)
-----------------------------------------------------

The original D-04 prohibition on Hebrew content was REVERSED during smoke
verification: when `lang == 'he'`, row content is now Hebrew-preferred with
graceful English fallback (and vice versa for `lang == 'en'`). Sheet titles
and header rows are now bilingual via :func:`sheet_titles`,
:func:`main_header_row`, :func:`manuscript_header_row`,
:func:`bibliography_header_row`. The view-direction conditional RTL is
unchanged. The D-02 prohibition on transcription text in NEW dossier surfaces
is unchanged. The D-10 parallels-envelope strip is unchanged.

REFINED 2026-05-21 (smoke verification round 2)
-----------------------------------------------

After Hillel's round-2 smoke verification, three further refinements:

1. The main sheet was renamed: ``"Genizah Results"`` → ``"Search Results"``
   (English) and ``"תוצאות גניזה"`` → ``"תוצאות חיפוש"`` (Hebrew). This is
   purely a sheet-title change — column layout, row builders, and the
   parity contract are unaffected.
2. A 4th sheet (``"Credits and Info"`` / ``"קרדיט ומידע"``) holds the
   credits text + per-export search metadata (Search Query / Mode / Gap /
   Lab Mode / Deep Scan / Date+time / Result count) + a hyperlink cell to
   GenizahSearch.com. The main sheet no longer carries inline credits.
   See :func:`build_credits_info_sheet`.
3. Domain names on the main sheet are Hebrew-substituted when ``lang='he'``
   if a ``domain_name_map`` (qualified-EN-name → HE-display-name) is provided
   by the caller. See :func:`substitute_domains_with_map`.

Translation rule: where a source DB has both Hebrew and English variants of
a field, the variant matching the UI language is preferred; the other
variant is used as fallback when the preferred one is absent.

Public API
----------

Four lookup helpers (each exception-resilient: wraps the service call in
try/except, logs a warning via the module logger, returns ``None`` or ``[]``):

- :func:`pgp_subset_for_sys_id` — narrow PGP projection (6 keys; never emits
  the full transcription field — see the strict prohibition below).
- :func:`nli_subset_for_sys_id` — narrow NLI projection (2 keys: ``catalog_entry``
  + ``library_viewer_url``). Column header is **NLI Catalog Entry** (NOT
  "NLI Description"); the underlying ``get_catalog_entry()`` returns
  Neubauer–Cowley reference strings, not descriptions (Codex MUST-FIX 2).
- :func:`catalog_summary_for_sys_id` — narrow FJMS catalog projection (4 fields:
  ``title``, ``author_text``, ``copy_date``, ``copy_place``). Uses the narrow
  ``FjmsService.get_catalog_records()`` query; the detail variant (which loads
  ``full_texts``) is intentionally not invoked, since it would break the
  strict prohibition below — Codex MUST-FIX 3.
- :func:`bibliography_for_sys_id` — list of FJMS bib entries with the REAL
  service field names (``running_title``, ``title_year``, ``mention_page``,
  ``article_name``, ``article_author_eng``, ``catalog_acronym``); the
  superseded plan invented field names that don't exist (Codex MUST-FIX 1).

Two row-emitters (return Python primitives only — no openpyxl objects):

- :func:`build_manuscript_row` — assembles one Manuscripts sub-sheet row by
  calling the PGP + NLI + Catalog helpers (3 helpers, NOT 4 — the bibliography
  helper is intentionally NOT called from here; Codex MUST-FIX 4).
- :func:`build_bibliography_rows` — assembles 0..N Bibliography sub-sheet rows
  by calling only :func:`bibliography_for_sys_id`.

Two header constants (length-pinned by tests so row/header drift is caught):

- :data:`MANUSCRIPT_HEADERS` (14 columns).
- :data:`BIBLIOGRAPHY_HEADERS` (8 columns).

D-02 strict prohibition (no transcription text in NEW dossier surfaces)
----------------------------------------------------------------------

The Manuscripts sub-sheet, Bibliography sub-sheet, and JSON envelope
additions (``has_pgp`` / ``is_printed`` / ``domains``) MUST NOT contain
PGP transcription text or FJMS full-text content. The helpers in this
module project a small whitelist of fields per source; the upstream
service dicts may contain additional content fields that this module
deliberately drops. Specifically the prohibited fields by name are
``page_section_text``, ``transcription``, ``full_text``, and ``full_texts``
(naming them here for clarity does not violate the rule — what matters is
that the helper output dicts never contain those keys; see
``tests/test_export_dossier.py::TestPgpSubset::test_no_transcription_text_leak``
and ``TestBibliography::test_no_transcription_leak``).

The pre-existing main-sheet ``Full Text`` column is GRANDFATHERED per
CONTEXT D-02 amendment — that's downstream callers' concern, not this
module's.

D-04 contract (English-only metadata content)
---------------------------------------------

Row content is always English regardless of any ``lang`` parameter passed to
the row builders. ``lang`` on :func:`build_manuscript_row` exists ONLY for
the CALLER's downstream sheet-view direction decision (Hebrew UI → RTL on
the workbook sheet); this module never translates content. Library name
resolution comes from the caller-supplied ``meta_resolver`` callable, which
is expected to call ``genizah_core.get_library_display(code, short=False,
lang='en')`` to hard-pin English (Codex SHOULD-FIX 9).

MUST-FIX 94-01-A — module-scope factory imports
-----------------------------------------------

The 3 service factory functions are imported at MODULE SCOPE (top of file)
so test monkeypatches at ``'shared.export_dossier.<name>'`` actually
intercept the lookup. Lazy in-function imports would let calls bypass the
patched name. See ``tests/test_export_dossier.py`` for the canonical
fixture pattern.

Codex MUST-FIX disposition
--------------------------

1. **Real FJMS bib field names** — :func:`bibliography_for_sys_id` projects
   {running_title, title_year, mention_page, article_name, article_author_eng,
   catalog_acronym}. NEVER {Author, Publisher, Source Name}.
2. **NLI Catalog Entry, not Description** — :func:`nli_subset_for_sys_id`
   uses :meth:`NliCrossrefService.get_catalog_entry` (returns Neubauer-Cowley
   strings).
3. **Narrow catalog query only** — :func:`catalog_summary_for_sys_id` uses
   the narrow ``get_catalog_records`` query; the detail variant (which reads
   ``full_texts``) is intentionally not invoked, since the dossier surfaces
   would otherwise breach D-02.
4. **build_manuscript_row calls 3 helpers, NOT 4** — bibliography rows live
   on a separate sub-sheet built by :func:`build_bibliography_rows`.
"""

import logging
from typing import Any, Callable, Dict, List, Optional

# MUST-FIX 94-01-A: hoist factory functions to module scope so test
# monkeypatches at 'shared.export_dossier.<name>' targets actually
# intercept. Lazy in-function imports break the patches.
from shared.document_service import get_document_for_fragment
from shared.fjms_service import get_fjms_service
from shared.nli_crossref_service import get_nli_crossref_service

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Module-level header constants (Codex SHOULD-FIX 7)
# ---------------------------------------------------------------------------

MANUSCRIPT_HEADERS: List[str] = [
    "System ID",
    "Shelfmark",
    "Library",
    "Title",
    "PGP URL",
    "PGP Description",
    "PGP Type",
    "PGP Date",
    "PGP Languages",
    "PGP Tags",
    "NLI Catalog Entry",  # Codex MUST-FIX 2 — NOT 'NLI Description'.
    "Catalog Summary",
    "Library Viewer URL",
    "GenizahSearch URL",
]  # 14 columns; matches build_manuscript_row output order.


BIBLIOGRAPHY_HEADERS: List[str] = [
    "System ID",
    "Shelfmark",
    "Article Author",
    "Article Name",
    "Running Title",
    "Title Year",
    "Mention Page",
    "Catalog Acronym",
]  # 8 columns; matches build_bibliography_rows row output order.


# Smoke round 4 (2026-05-21): which Manuscripts-sheet columns hold a URL
# that should be rendered as a clickable hyperlink. 0-based indices into
# the build_manuscript_row output: PGP URL = 4, Library Viewer URL = 12,
# GenizahSearch URL = 13. Used by :func:`apply_manuscript_row_hyperlinks`.
MANUSCRIPT_URL_COLUMN_INDICES: tuple = (4, 12, 13)


def apply_manuscript_row_hyperlinks(ws, excel_row: int, row_values: List[Any]) -> None:
    """Mark URL cells in a Manuscripts-sheet data row as clickable.

    Targets the columns listed in :data:`MANUSCRIPT_URL_COLUMN_INDICES`
    (PGP URL / Library Viewer URL / GenizahSearch URL). Cells whose value
    is empty, non-string, or does not parse as an http(s) URL are left
    untouched — that way unknown / missing IIIF viewer URLs fall through as
    plain empty cells without raising.

    ``excel_row`` is the 1-based worksheet row number where the data row was
    just appended (e.g. 2 for the first data row directly under the header).
    ``row_values`` is the 14-cell list returned by :func:`build_manuscript_row`
    — passed in so callers do not have to re-read the cell value after the
    ``sanitize_text_for_excel`` pass that some call sites apply.
    """
    from openpyxl.styles import Font

    hyperlink_font = Font(color="0563C1", underline="single")
    for idx in MANUSCRIPT_URL_COLUMN_INDICES:
        if idx >= len(row_values):
            continue
        val = row_values[idx]
        if not val or not isinstance(val, str):
            continue
        v = val.strip()
        if not (v.startswith('http://') or v.startswith('https://')):
            continue
        cell = ws.cell(row=excel_row, column=idx + 1)
        cell.hyperlink = v
        cell.font = hyperlink_font


# ---------------------------------------------------------------------------
# Bilingual header rows + sheet titles (D-04 REVISED 2026-05-20)
# ---------------------------------------------------------------------------
#
# Hebrew translations match the canonical strings already present in
# ``genizah_translations.TRANSLATIONS`` (used by the desktop ``tr()`` helper).
# When ``lang == 'en'`` the helpers return the corresponding English constant
# verbatim for back-compat with prior callers that read ``MANUSCRIPT_HEADERS``
# / ``BIBLIOGRAPHY_HEADERS`` directly.

_MAIN_HEADERS_EN: List[str] = [
    "System ID", "Library", "Shelfmark", "Title",
    "Image/Page", "Source",
    "Snippet", "Full Text",
    "Has PGP", "Is Printed", "Domains", "IIIF Manifest",
]

_MAIN_HEADERS_HE: List[str] = [
    "מספר מערכת", "ספרייה", "מספר מדף", "כותרת",
    "תמונה/עמוד", "מקור",
    "קטע", "טקסט מלא",
    "יש PGP", "מודפס", "תחומים", "מניפסט IIIF",
]

_MANUSCRIPT_HEADERS_HE: List[str] = [
    "מספר מערכת",
    "מספר מדף",
    "ספרייה",
    "כותרת",
    "כתובת PGP",
    "תיאור PGP",
    "סוג PGP",
    "תאריך PGP",
    "שפות PGP",
    "תגיות PGP",
    "רשומה בקטלוג הספרייה הלאומית",
    "תקציר קטלוגי",
    "קישור לצפייה בספרייה",
    "קישור ל-GenizahSearch",
]

_BIBLIOGRAPHY_HEADERS_HE: List[str] = [
    "מספר מערכת",
    "מספר מדף",
    "מחבר המאמר",
    "שם המאמר",
    "כותרת רצה",
    "שנת הפרסום",
    "עמוד אזכור",
    "קיצור הקטלוג",
]

_SHEET_TITLES_EN: Dict[str, str] = {
    'main': "Search Results",
    'manuscripts': "Manuscripts",
    'bibliography': "Bibliography",
    # Smoke verification round 2 (2026-05-21): new 4th sheet for credits +
    # search metadata. The main sheet no longer carries the credit/metadata
    # rows that previously rode above (desktop) or below (web) the data rows.
    'credits_info': "Credits and Info",
}

_SHEET_TITLES_HE: Dict[str, str] = {
    'main': "תוצאות חיפוש",
    'manuscripts': "כתבי יד",
    'bibliography': "ביבליוגרפיה",
    'credits_info': "קרדיט ומידע",
}


def main_header_row(lang: str = 'en') -> List[str]:
    """Return the 12 main-sheet column headers in the requested language.

    Phase 94 D-04 REVISED (2026-05-20): Hebrew when ``lang == 'he'``, English
    otherwise. The returned list is a fresh copy so callers cannot mutate the
    module constants.
    """
    if lang == 'he':
        return list(_MAIN_HEADERS_HE)
    return list(_MAIN_HEADERS_EN)


def manuscript_header_row(lang: str = 'en') -> List[str]:
    """Return the 14 Manuscripts sub-sheet column headers in the requested language.

    When ``lang == 'en'`` returns the English row matching :data:`MANUSCRIPT_HEADERS`
    verbatim (back-compat). When ``lang == 'he'`` returns the Hebrew row.
    """
    if lang == 'he':
        return list(_MANUSCRIPT_HEADERS_HE)
    return list(MANUSCRIPT_HEADERS)


def bibliography_header_row(lang: str = 'en') -> List[str]:
    """Return the 8 Bibliography sub-sheet column headers in the requested language."""
    if lang == 'he':
        return list(_BIBLIOGRAPHY_HEADERS_HE)
    return list(BIBLIOGRAPHY_HEADERS)


def sheet_titles(lang: str = 'en') -> Dict[str, str]:
    """Return a dict of localized sheet titles keyed by ``main`` / ``manuscripts`` / ``bibliography`` / ``credits_info``.

    Smoke verification round 2 (2026-05-21): a 4th sheet ``credits_info``
    was added — ``"Credits and Info"`` (English) / ``"קרדיט ומידע"`` (Hebrew).
    """
    if lang == 'he':
        return dict(_SHEET_TITLES_HE)
    return dict(_SHEET_TITLES_EN)


# ---------------------------------------------------------------------------
# Credits + Info sheet (smoke verification round 2, 2026-05-21)
# ---------------------------------------------------------------------------
#
# Credits block — the canonical Stoekl Ben Ezra citation chain. Both apps
# used to render this inline on the main sheet (web: bottom; desktop: above
# the header row). The round-2 smoke pass requested it move to its own
# dedicated sheet alongside the per-export search metadata (Query / Mode /
# Gap / Lab Mode / Deep Scan / date+time / result count) and a hyperlink
# cell back to GenizahSearch.com.
#
# Credits lines mirror ``web/export_service.CREDITS_TEXT`` (the canonical
# web text) and ``GenizahGUI._get_credit_header`` (the desktop variant).
# Kept self-contained in the shared module so both apps render identical
# credit text on the new sheet.

CREDITS_TITLE_EN: str = "Credits and Info"
CREDITS_TITLE_HE: str = "קרדיט ומידע"

_CREDITS_LINES_EN: List[str] = [
    "Generated by Dicta Genizah Search",
    "Creator: Hillel Gershuni, gershuni@gmail.com",
    "Data Source: MiDRASH Automatic Transcriptions (Stoekl Ben Ezra et al., 2025)",
    "Dataset: https://doi.org/10.5281/zenodo.17734473",
    (
        "Citation: Stoekl Ben Ezra, D., Bambaci, L., Kiessling, B., Lapin, H., "
        "Ezer, N., Lolli, E., Rustow, M., Dershowitz, N., Kurar Barakat, B., "
        "Gogawale, S., Shmidman, A., Lavee, M., Siew, T., Raziel Kretzmer, V., "
        "Vasyutinsky Shapira, D., Olszowy-Schlanger, J., & Gila, Y. (2025). "
        "MiDRASH Automatic Transcriptions. Zenodo. "
        "https://doi.org/10.5281/zenodo.17734473"
    ),
]

# Hebrew credits lines — only the "Generated by" header and the "Creator"
# attribution are localized. The academic citation stays in its canonical
# English form regardless of UI language (translating a published citation
# is incorrect — the DOI and dataset URL must remain intact and the cited
# authors' names must appear as published).
_CREDITS_LINES_HE: List[str] = [
    "הופק על ידי Dicta Genizah Search",
    "יוצר: הלל גרשוני, gershuni@gmail.com",
    "Data Source: MiDRASH Automatic Transcriptions (Stoekl Ben Ezra et al., 2025)",
    "Dataset: https://doi.org/10.5281/zenodo.17734473",
    (
        "Citation: Stoekl Ben Ezra, D., Bambaci, L., Kiessling, B., Lapin, H., "
        "Ezer, N., Lolli, E., Rustow, M., Dershowitz, N., Kurar Barakat, B., "
        "Gogawale, S., Shmidman, A., Lavee, M., Siew, T., Raziel Kretzmer, V., "
        "Vasyutinsky Shapira, D., Olszowy-Schlanger, J., & Gila, Y. (2025). "
        "MiDRASH Automatic Transcriptions. Zenodo. "
        "https://doi.org/10.5281/zenodo.17734473"
    ),
]


# Search-metadata label dictionary. The user explicitly specified every
# Hebrew label in the smoke-verification briefs on 2026-05-21. Round 3
# (2026-05-21) realigned 'search_mode' Hebrew to 'מצב חיפוש' (matches the
# existing genizah_translations.py entry) and renamed the link-row label
# from 'בקרו ב-GenizahSearch.com' → 'אתר הגניזה של דיקטה' (EN parallel
# 'Visit GenizahSearch.com' → 'Dicta Genizah Search').
_SEARCH_META_LABELS_EN: Dict[str, str] = {
    'credits_section_title': "Credits",
    'metadata_section_title': "Search Metadata",
    'search_query': "Search Query",
    'search_mode': "Search Mode",
    'search_gap': "Gap",
    'lab_mode': "Lab Mode",
    'deep_scan': "Deep Scan",
    'date_time': "Date/time of export",
    'result_count': "Result count",
    'visit_link_label': "Dicta Genizah Search",
    'on': "On",
    'off': "Off",
}

_SEARCH_META_LABELS_HE: Dict[str, str] = {
    'credits_section_title': "קרדיט",
    'metadata_section_title': "פרטי החיפוש",
    'search_query': "שאילתת חיפוש",
    'search_mode': "מצב חיפוש",
    'search_gap': "רווח",
    'lab_mode': "מצב מעבדה",
    'deep_scan': "סריקה עמוקה",
    'date_time': "תאריך ושעת הייצוא",
    'result_count': "מספר תוצאות",
    'visit_link_label': "אתר הגניזה של דיקטה",
    'on': "כן",
    'off': "לא",
}

GENIZAHSEARCH_URL: str = "https://genizahsearch.com"


def credits_lines(lang: str = 'en') -> List[str]:
    """Return the credits citation lines for the given UI language.

    The "Generated by …" header and the "Creator: …" attribution line are
    localized; the academic citation rows (data source, dataset URL, full
    citation with DOI) stay in their canonical English form on both 'en'
    and 'he' workbooks — translating a published citation is incorrect.
    Returns a fresh list copy so callers cannot mutate the module constant.
    """
    if lang == 'he':
        return list(_CREDITS_LINES_HE)
    return list(_CREDITS_LINES_EN)


def search_meta_labels(lang: str = 'en') -> Dict[str, str]:
    """Return the localized label dictionary for the Credits and Info sheet."""
    if lang == 'he':
        return dict(_SEARCH_META_LABELS_HE)
    return dict(_SEARCH_META_LABELS_EN)


def build_credits_info_sheet(
    ws,
    *,
    lang: str = 'en',
    search_query: Optional[str] = None,
    search_mode: Optional[str] = None,
    search_gap: Optional[Any] = None,
    lab_mode_on: Optional[bool] = None,
    deep_scan_on: Optional[bool] = None,
    export_datetime: Optional[str] = None,
    result_count: Optional[int] = None,
) -> None:
    """Populate ``ws`` (an openpyxl Worksheet) as the Credits and Info sheet.

    Sheet layout (col A = label, col B = value):

        Row 1  : sheet header — single cell "Credits and Info" / "קרדיט ומידע"
        Row 2  : "Credits" / "קרדיט" section title
        Row 3+ : one row per credit line (col A holds the credit text)
        Blank  : section separator
        Then   : "Search Metadata" / "פרטי החיפוש" section title
                 Search Query / Mode / Gap rows
                 Lab Mode row (only if lab_mode_on is not None — desktop sets,
                 web omits since web has no Lab Mode toggle)
                 Deep Scan row (only when lab_mode_on is True)
                 Date/time of export
                 Result count
        Blank
        Then   : "Visit GenizahSearch.com" hyperlink row

    Hyperlink: col B cell holds an openpyxl ``Hyperlink`` to
    ``https://genizahsearch.com``; col A holds the localized label text.

    The caller is responsible for setting the worksheet's
    ``sheet_view.rightToLeft`` from its ``lang`` decision — this function
    does NOT touch view direction (consistent with how the row builders for
    the other sheets work).

    Note on label translation: the round-2 smoke brief explicitly specified
    every Hebrew label; we use those exact literals (see
    :data:`_SEARCH_META_LABELS_HE`). 'Search Mode' / 'Gap' differ from the
    existing UI tr() entries — that's intentional per the round-2 brief.
    """
    from openpyxl.styles import Font
    from openpyxl.worksheet.hyperlink import Hyperlink

    labels = search_meta_labels(lang)
    sheet_label = CREDITS_TITLE_HE if lang == 'he' else CREDITS_TITLE_EN

    # Row 1: workbook section title (mirrors header styling on other sheets).
    ws.cell(row=1, column=1, value=sheet_label).font = Font(bold=True, size=14)

    row = 2

    # Credits section header
    ws.cell(row=row, column=1, value=labels['credits_section_title']).font = Font(bold=True)
    row += 1
    for line in credits_lines(lang):
        ws.cell(row=row, column=1, value=line)
        row += 1

    # Blank separator row
    row += 1

    # Search metadata section header
    ws.cell(row=row, column=1, value=labels['metadata_section_title']).font = Font(bold=True)
    row += 1

    def _put(label_key: str, value: Any) -> None:
        nonlocal row
        if value is None or (isinstance(value, str) and not value.strip()):
            return  # Skip rows with no value (D-06 — empty cells, but here we just omit the row entirely so the sheet is compact).
        ws.cell(row=row, column=1, value=labels[label_key]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    _put('search_query', search_query)
    _put('search_mode', search_mode)
    # Gap may legitimately be 0; only the explicit None case omits the row.
    if search_gap is not None and str(search_gap) != '':
        ws.cell(row=row, column=1, value=labels['search_gap']).font = Font(bold=True)
        ws.cell(row=row, column=2, value=search_gap)
        row += 1
    # Lab Mode is desktop-only — None means "not applicable, skip the row".
    if lab_mode_on is not None:
        ws.cell(row=row, column=1, value=labels['lab_mode']).font = Font(bold=True)
        ws.cell(row=row, column=2, value=labels['on'] if lab_mode_on else labels['off'])
        row += 1
        # Deep Scan only shown when Lab Mode is on (mirrors desktop UX).
        if lab_mode_on and deep_scan_on is not None:
            ws.cell(row=row, column=1, value=labels['deep_scan']).font = Font(bold=True)
            ws.cell(row=row, column=2, value=labels['on'] if deep_scan_on else labels['off'])
            row += 1

    _put('date_time', export_datetime)
    if result_count is not None:
        ws.cell(row=row, column=1, value=labels['result_count']).font = Font(bold=True)
        ws.cell(row=row, column=2, value=result_count)
        row += 1

    # Blank separator row before the visit-link
    row += 1

    # Visit GenizahSearch.com hyperlink row
    ws.cell(row=row, column=1, value=labels['visit_link_label']).font = Font(bold=True)
    link_cell = ws.cell(row=row, column=2, value=GENIZAHSEARCH_URL)
    link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=GENIZAHSEARCH_URL)
    # Style the link cell so users see it as a hyperlink.
    link_cell.font = Font(color="0563C1", underline="single")

    # Column widths: A holds labels, B holds values. Make A wider so labels
    # line up nicely; B comfortably wide for long values (queries, URLs).
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70


# ---------------------------------------------------------------------------
# MetaResolver type alias
# ---------------------------------------------------------------------------

# Codex SHOULD-FIX 8 — callable returning a 4-key primitive dict,
# NOT an opaque meta_mgr object. Prevents silent web/desktop drift.
MetaResolver = Callable[[str], Optional[Dict[str, Any]]]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _split_pgp_languages(value: Any) -> List[str]:
    """Normalize a languages projection to a list of stripped non-empty strings.

    The SUPERSEDED-v2 plan had a latent bug: ``list(doc.get('languages_primary'))``
    iterates a comma-separated STRING as characters when the sidecar value is
    ``'Hebrew, Aramaic'`` instead of ``['Hebrew', 'Aramaic']``. The pgp.db
    sidecar's ``languages_primary`` / ``languages_secondary`` columns are
    declared TEXT (see ``scripts/export_pgp_sidecar.py:95-96``) so the
    comma-split branch is the production path; the list-input branch is
    defensive coverage for synthetic or future callers (Codex Q5).
    """
    if value is None or value == '':
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if v and str(v).strip()]
    if isinstance(value, str):
        return [s.strip() for s in value.split(',') if s.strip()]
    return []


# ---------------------------------------------------------------------------
# Helper 1: PGP subset
# ---------------------------------------------------------------------------


def pgp_subset_for_sys_id(
    sys_id: str,
    lang: str = 'en',
) -> Optional[Dict[str, Any]]:
    """Return a 6-key PGP projection for a manuscript or ``None``.

    Keys (always present when the helper returns a dict; values may be None):
    ``pgp_url``, ``description``, ``document_type``, ``date_display``,
    ``languages``, ``tags``.

    Date uses the fallback chain ``inferred_date_display`` →
    ``doc_date_standard`` → ``doc_date_original`` → ``None``.

    Languages are normalized via :func:`_split_pgp_languages` to handle the
    comma-separated TEXT projection bug; primary + secondary are merged with
    duplicates removed (preserving primary order).

    The helper NEVER emits ``page_section_text``, ``transcription``,
    ``full_text``, or any field outside the 6-key whitelist (D-02 boundary).

    D-04 REVISED (2026-05-20): when ``lang == 'he'`` and the PGP sidecar has
    a Hebrew translation for ``description`` (and/or ``document_type``) in
    the ``pgp_translations`` table (per ``shared/translation_service.py``),
    the Hebrew variant is preferred with English fallback. Type / date /
    languages / tags stay in their canonical English/categorical form (these
    are normalized vocabulary, not free-text). ``document_type_he`` is
    preferred when available because the desktop already surfaces translated
    document types in browse/search UI.
    """
    if not sys_id:
        return None
    try:
        doc = get_document_for_fragment(sys_id)
        if not doc:
            return None
        date_display = (
            doc.get('inferred_date_display')
            or doc.get('doc_date_standard')
            or doc.get('doc_date_original')
        )
        langs_primary = _split_pgp_languages(doc.get('languages_primary'))
        langs_secondary = _split_pgp_languages(doc.get('languages_secondary'))
        # Merge primary + secondary, dedupe while preserving order.
        languages = list(langs_primary)
        for entry_lang in langs_secondary:
            if entry_lang not in languages:
                languages.append(entry_lang)

        description = doc.get('description')
        document_type = doc.get('document_type')

        # D-04 REVISED: prefer Hebrew when lang=='he' AND a translation exists.
        if lang == 'he':
            he_translation = _pgp_translation_he_for_sys_id(sys_id)
            if he_translation:
                desc_he = he_translation.get('description_he')
                if desc_he and str(desc_he).strip():
                    description = desc_he
                type_he = he_translation.get('document_type_he')
                if type_he and str(type_he).strip():
                    document_type = type_he

        return {
            'pgp_url': doc.get('pgp_url'),
            'description': description,
            'document_type': document_type,
            'date_display': date_display,
            'languages': languages,
            'tags': list(doc.get('tags') or []),
        }
    except Exception as e:
        logger.warning("pgp_subset_for_sys_id(%s) failed: %s", sys_id, e)
        return None


def _pgp_translation_he_for_sys_id(sys_id: str) -> Optional[Dict[str, Any]]:
    """Return ``{description_he, document_type_he}`` for a sys_id, or ``None``.

    Thin wrapper around :meth:`TranslationService.get_pgp_translations_by_sys_ids`
    for the single-sys_id case. Exception-resilient: returns ``None`` on any
    failure or when the sidecar lacks ``pgp_translations``. Module-scope
    factory (not lazy) so tests can monkeypatch at
    ``shared.export_dossier._pgp_translation_he_for_sys_id``.
    """
    try:
        from shared.translation_service import TranslationService
        svc = TranslationService(thread_safe=True)
        if not svc.pgp_available():
            return None
        result = svc.get_pgp_translations_by_sys_ids([sys_id])
        return result.get(sys_id)
    except Exception as e:
        logger.warning("_pgp_translation_he_for_sys_id(%s) failed: %s", sys_id, e)
        return None


# ---------------------------------------------------------------------------
# Helper 2: NLI subset
# ---------------------------------------------------------------------------


def nli_subset_for_sys_id(sys_id: str) -> Optional[Dict[str, Any]]:
    """Return a 2-key NLI projection for a manuscript or ``None``.

    Returns ``{'catalog_entry': str|None, 'library_viewer_url': str|None}``
    when the NLI sidecar has at least one of the two values for the sys_id;
    otherwise returns ``None`` (no useful data).

    The column header in the workbook is **NLI Catalog Entry**
    (Codex MUST-FIX 2 — :meth:`NliCrossrefService.get_catalog_entry` returns
    Neubauer-Cowley reference strings such as ``'Neubauer - Cowley 2603.1'``,
    NOT descriptions).

    ``thread_safe=True`` is passed to the factory because exports run from
    background tasks on web (NiceGUI) and could run from any thread on
    desktop. The default is ``False`` (see
    ``shared/nli_crossref_service.py:1019``) — established web/shared
    consumers always pass ``True`` (e.g. ``shared/browse_service.py:215-216``)
    and the export pipeline follows the same convention.
    """
    if not sys_id:
        return None
    try:
        svc = get_nli_crossref_service(thread_safe=True)
        if not svc or not svc.is_available():
            return None
        catalog_entry = svc.get_catalog_entry(sys_id)
        viewer = svc.get_library_viewer_url(sys_id)
        viewer_url = viewer.get('url') if isinstance(viewer, dict) else None
        if not catalog_entry and not viewer_url:
            return None
        return {
            'catalog_entry': catalog_entry,
            'library_viewer_url': viewer_url,
        }
    except Exception as e:
        logger.warning("nli_subset_for_sys_id(%s) failed: %s", sys_id, e)
        return None


# ---------------------------------------------------------------------------
# Helper 3: Catalog summary
# ---------------------------------------------------------------------------


def catalog_summary_for_sys_id(
    sys_id: str,
    lang: str = 'en',
) -> Optional[Dict[str, Any]]:
    """Return a narrow FJMS catalog projection for a manuscript or ``None``.

    Returns 4 fields: ``title``, ``author_text``, ``copy_date`` (sentinel-
    normalized to None by the service at :meth:`FjmsService.get_catalog_records`),
    ``copy_place``.

    **Aggregation strategy:** *first non-empty per field* across all records
    returned by :meth:`FjmsService.get_catalog_records`. Multiple records
    arise when a manuscript has been cataloged by several teams; we pick
    the first scholar-provided value for each field independently rather
    than picking a single record verbatim, because cataloging is often
    partial.

    D-04 REVISED (2026-05-20): title preference depends on ``lang``. When
    ``lang == 'he'``, ``title_heb`` (Hebrew) is preferred with English
    fallback (``title``). When ``lang == 'en'``, English ``title`` is
    preferred with Hebrew fallback. ``author_text`` / ``copy_place`` are
    not Hebrew-translated in the FJMS sidecar (free-text scholar notes),
    so they pass through unchanged. ``copy_date`` is numeric — language-
    independent.

    **Codex MUST-FIX 3 / D-02 boundary:** uses
    :meth:`FjmsService.get_catalog_records` only. The detail variant
    (which loads ``full_texts``) is intentionally not invoked from the
    dossier path.
    """
    if not sys_id:
        return None
    try:
        fjms = get_fjms_service(thread_safe=True)
        if not fjms or not fjms.is_available():
            return None
        records = fjms.get_catalog_records(sys_id)
        if not records:
            return None

        def _pick(field):
            for rec in records:
                v = rec.get(field)
                if v is not None and str(v).strip():
                    return v
            return None

        # D-04 REVISED: language-preferred title selection.
        if lang == 'he':
            title = _pick('title_heb') or _pick('title')
        else:
            title = _pick('title') or _pick('title_heb')
        author_text = _pick('author_text')
        copy_date = _pick('copy_date')
        copy_place = _pick('copy_place')

        if not any((title, author_text, copy_date, copy_place)):
            return None

        return {
            'title': title,
            'author_text': author_text,
            'copy_date': copy_date,
            'copy_place': copy_place,
        }
    except Exception as e:
        logger.warning("catalog_summary_for_sys_id(%s) failed: %s", sys_id, e)
        return None


# ---------------------------------------------------------------------------
# Helper 4: Bibliography
# ---------------------------------------------------------------------------


def bibliography_for_sys_id(
    sys_id: str,
    lang: str = 'en',
) -> List[Dict[str, Any]]:
    """Return 0..N FJMS bibliography entries projected to 6 whitelisted keys.

    Keys per entry: ``running_title``, ``title_year``, ``mention_page``,
    ``article_name``, ``article_author_eng``, ``catalog_acronym`` (Codex
    MUST-FIX 1 — these are the REAL FJMS field names; the SUPERSEDED-v2 plan
    invented Author / Publisher / Source Name which do not exist in
    :meth:`FjmsService.get_bibliography`'s return shape).

    The extended fields ``comment`` / ``note_for_display`` / ``catalog_entry``
    that the service may include are deliberately dropped — they don't fit
    the dossier sub-sheet structure and risk D-02 boundary creep.

    D-04 REVISED (2026-05-20): when ``lang == 'he'`` and the underlying
    bibliography row exposes ``running_title_heb`` / ``article_author_heb``
    fields, the Hebrew variants are preferred with English fallback. The
    other 4 keys (``title_year`` int, ``mention_page`` short string,
    ``article_name`` free-text not translated, ``catalog_acronym`` short
    string) are language-neutral.
    """
    if not sys_id:
        return []
    try:
        fjms = get_fjms_service(thread_safe=True)
        if not fjms or not fjms.is_available():
            return []
        entries = fjms.get_bibliography(sys_id) or []
        prefer_he = (lang == 'he')

        def _pick_he(entry, he_key, en_key):
            if prefer_he:
                v = entry.get(he_key)
                if v and str(v).strip():
                    return v
                return entry.get(en_key)
            v = entry.get(en_key)
            if v and str(v).strip():
                return v
            return entry.get(he_key)

        return [
            {
                'running_title': _pick_he(e, 'running_title_heb', 'running_title'),
                'title_year': e.get('title_year'),
                'mention_page': e.get('mention_page'),
                'article_name': e.get('article_name'),
                'article_author_eng': _pick_he(e, 'article_author_heb', 'article_author_eng'),
                'catalog_acronym': e.get('catalog_acronym'),
            }
            for e in entries
        ]
    except Exception as e:
        logger.warning("bibliography_for_sys_id(%s) failed: %s", sys_id, e)
        return []


# ---------------------------------------------------------------------------
# Row emitter 1: Manuscripts
# ---------------------------------------------------------------------------


def build_manuscript_row(
    sys_id: str,
    meta_resolver: Optional[MetaResolver],
    lang: str = 'en',
) -> List[Any]:
    """Build one Manuscripts sub-sheet row for a sys_id.

    Calls 3 helpers — :func:`pgp_subset_for_sys_id`,
    :func:`nli_subset_for_sys_id`, :func:`catalog_summary_for_sys_id`.
    Does NOT call :func:`bibliography_for_sys_id` (Codex MUST-FIX 4 —
    bibliography rows live on the separate Bibliography sub-sheet via
    :func:`build_bibliography_rows`).

    Returns a list of exactly 14 Python primitives matching
    :data:`MANUSCRIPT_HEADERS` (English) / :func:`manuscript_header_row('he')`
    (Hebrew) order. Missing data renders as empty strings (NOT 'N/A' /
    placeholders — D-06).

    D-04 REVISED (2026-05-20): when ``lang == 'he'`` the row prefers
    Hebrew variants from each source DB (PGP description / type via
    ``pgp_translations`` table, FJMS catalog title via ``title_heb``,
    library name via ``get_library_display(code, lang='he')``) with
    English graceful fallback. The Catalog Summary cell field labels
    (``Title:`` / ``Author:`` / ``Date:`` / ``Place:``) also follow the
    requested language.

    ``library_name`` comes from the caller-supplied ``meta_resolver``;
    callers MUST construct the resolver such that ``library_name`` reflects
    ``lang`` (i.e. call ``genizah_core.get_library_display(code,
    short=False, lang=lang)``).
    """
    if meta_resolver is not None and sys_id:
        meta = meta_resolver(sys_id)
    else:
        meta = None

    if isinstance(meta, dict):
        shelfmark = meta.get('shelfmark') or ''
        title = meta.get('title') or ''
        library_name = meta.get('library_name') or ''
    else:
        shelfmark = ''
        title = ''
        library_name = ''

    pgp = pgp_subset_for_sys_id(sys_id, lang=lang) or {}
    nli = nli_subset_for_sys_id(sys_id) or {}
    catalog = catalog_summary_for_sys_id(sys_id, lang=lang) or {}

    # D-05: pipe-joined, NO surrounding spaces.
    languages_pipe = '|'.join(pgp.get('languages') or [])
    tags_pipe = '|'.join(pgp.get('tags') or [])

    # Catalog Summary cell: 'Title: X | Author: Y | Date: Z | Place: W'
    # with empty fields omitted entirely (CONTEXT D-08 helper 3 strategy).
    # D-04 REVISED: labels follow the row's language.
    if lang == 'he':
        _labels = (
            ('title', 'כותרת'),
            ('author_text', 'מחבר'),
            ('copy_date', 'תאריך'),
            ('copy_place', 'מקום'),
        )
    else:
        _labels = (
            ('title', 'Title'),
            ('author_text', 'Author'),
            ('copy_date', 'Date'),
            ('copy_place', 'Place'),
        )
    cat_parts: List[str] = []
    for key, label in _labels:
        v = catalog.get(key)
        if v is not None and str(v).strip():
            cat_parts.append(f"{label}: {v}")
    catalog_summary_str = ' | '.join(cat_parts)

    genizah_url = (
        f'https://genizahsearch.com/browse?sys_id={sys_id}' if sys_id else ''
    )

    return [
        sys_id or '',
        shelfmark,
        library_name,
        title,
        pgp.get('pgp_url') or '',
        pgp.get('description') or '',
        pgp.get('document_type') or '',
        pgp.get('date_display') or '',
        languages_pipe,
        tags_pipe,
        nli.get('catalog_entry') or '',
        catalog_summary_str,
        nli.get('library_viewer_url') or '',
        genizah_url,
    ]


# ---------------------------------------------------------------------------
# Row emitter 2: Bibliography
# ---------------------------------------------------------------------------


def build_bibliography_rows(
    sys_id: str,
    meta_resolver: Optional[MetaResolver],
    lang: str = 'en',
) -> List[List[Any]]:
    """Build 0..N Bibliography sub-sheet rows for a sys_id.

    Calls :func:`bibliography_for_sys_id` ONLY (Codex MUST-FIX 4).

    Returns a list of row-lists, each row exactly 8 cells matching
    :data:`BIBLIOGRAPHY_HEADERS` (English) /
    :func:`bibliography_header_row('he')` (Hebrew) order. Empty list when
    the sys_id has no bib entries.

    Per D-06, missing string fields render as empty cells.

    D-04 REVISED (2026-05-20): when ``lang == 'he'`` the row prefers Hebrew
    variants from each underlying bib entry (running_title_heb,
    article_author_heb) with English graceful fallback. Threading lang
    happens via :func:`bibliography_for_sys_id`.
    """
    if not sys_id:
        return []
    entries = bibliography_for_sys_id(sys_id, lang=lang) or []
    if not entries:
        return []

    if meta_resolver is not None:
        meta = meta_resolver(sys_id)
    else:
        meta = None
    shelfmark = (
        meta.get('shelfmark') if isinstance(meta, dict) else None
    ) or ''

    rows: List[List[Any]] = []
    for entry in entries:
        # title_year may legitimately be an integer (e.g. 1967); preserve
        # numeric type when present.
        title_year_v = entry.get('title_year')
        if title_year_v in (None, ''):
            title_year_cell: Any = ''
        else:
            title_year_cell = title_year_v
        rows.append([
            sys_id or '',
            shelfmark,
            entry.get('article_author_eng') or '',
            entry.get('article_name') or '',
            entry.get('running_title') or '',
            title_year_cell,
            entry.get('mention_page') or '',
            entry.get('catalog_acronym') or '',
        ])
    return rows

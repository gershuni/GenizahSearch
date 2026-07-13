# -*- coding: utf-8 -*-
"""MAPV2-15e — build a clean CANONICAL TEXT corpus from the local Maagarim
export (Bible + Mishnah/Tosefta/Bavli/Yerushalmi + Midrash), classified by
the Academy's own סוגה column, for the shared-source detector.

Rationale (Hillel): the shared-source detector should test a match against
the actual authoritative canonical TEXTS, not against track1's incidental
canonical identifications (which are Bible-dense but Talmud/Mishnah-thin).
Maagarim already holds the full classical canon locally; we tag each work to
a fine cat and normalize it to the matching stream, ready to index.

CANON (feed the mask): Bible, Mishnah, Tosefta, Bavli, Yerushalmi.
Midrash is ingested but tagged separately (it is the discovery-bearing
class — a midrash quote may be a shared source OR an unrecognized witness —
so it does not auto-mask; the AI layer judges it).

Out: data/canon_corpus_maagarim.pkl  (list of {id,cat,title,author,stream})
     results/canon_corpus_report.md
"""
import os
import pickle
import re
from collections import Counter, defaultdict

from normalize import norm_stream

MG = r'C:\Users\gersh\Dropbox\דיקטה\מאגרים\AllTextsOnlyText'
XLSX = r'C:\Users\gersh\Dropbox\דיקטה\מאגרים\חיבורים במאגרים.xlsx'
PROBE = r'C:\Genizahsearch\same_work_spike\probe'
OUT_PKL = PROBE + r'\data\canon_corpus_maagarim.pkl'
OUT_MD = PROBE + r'\results\canon_corpus_report.md'

CANON_CATS = ('Bible', 'Mishnah', 'Tosefta', 'Bavli', 'Yerushalmi')
_YT = re.compile(r'Ytext(\d+)')


def fine_cat(title, sugya):
    t = title or ''
    if sugya == 'מקרא':
        return 'Bible'
    # within תלמוד ומדרש: order matters (תלמוד ירושלמי before בבלי/תלמוד)
    if 'תוספתא' in t:
        return 'Tosefta'
    if 'ירושלמי' in t:
        return 'Yerushalmi'
    if 'משנה' in t and 'תלמוד' not in t:
        return 'Mishnah'
    if 'בבלי' in t or 'תלמוד' in t or 'גמרא' in t:
        return 'Bavli'
    if any(k in t for k in ('מכילתא', 'ספרא', 'ספרי', 'רבה', 'תנחומא',
                            'פסיקתא', 'מדרש', 'אבות דרבי', 'פרקי דרבי',
                            'ילקוט', 'אגדת')):
        return 'Midrash'
    return 'Midrash'   # default bucket inside תלמוד ומדרש


def main():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)   # header: מחבר, חיבור, תאריך, סוגה, שם קובץ
    rows = []
    for r in it:
        au, ti, date, sugya, fname = (list(r) + [None] * 5)[:5]
        if sugya in ('מקרא', 'תלמוד ומדרש'):
            m = _YT.search(str(fname or ''))
            if m:
                rows.append((m.group(1), str(au or ''), str(ti or ''), sugya))
    print(f"canon-sugya rows in xlsx: {len(rows)}", flush=True)

    # map Ytext id -> file path in the export
    id2path = {}
    for f in os.listdir(MG):
        m = _YT.search(f)
        if m:
            id2path[m.group(1)] = os.path.join(MG, f)

    works = []
    cat_letters = defaultdict(int)
    cat_works = Counter()
    missing = 0
    unmapped_titles = []
    for yid, au, ti, sugya in rows:
        path = id2path.get(yid)
        if not path or not os.path.exists(path):
            missing += 1
            continue
        cat = fine_cat(ti, sugya)
        try:
            raw = open(path, encoding='utf-8', errors='replace').read()
        except OSError:
            missing += 1
            continue
        stream = norm_stream(raw)[0]
        if not stream:
            continue
        works.append({'id': f'M:Ytext{yid}', 'cat': cat, 'title': ti,
                      'author': au, 'sugya': sugya, 'stream': stream})
        cat_letters[cat] += len(stream)
        cat_works[cat] += 1
        if cat == 'Midrash' and sugya == 'תלמוד ומדרש' and len(unmapped_titles) < 25:
            # midrash default bucket — record a few to eyeball
            unmapped_titles.append(ti[:60])

    pickle.dump(works, open(OUT_PKL, 'wb'))
    canon_letters = sum(cat_letters[c] for c in CANON_CATS)
    L = ["# Maagarim canonical corpus (MAPV2-15e)", "",
         f"- xlsx canon-sugya rows: {len(rows)}; files missing: {missing}",
         f"- works ingested: {len(works)}",
         f"- CANON letters (Bible+Mishnah+Tosefta+Bavli+Yerushalmi): "
         f"**{canon_letters:,}**", "",
         "## works + normalized letters by fine cat", "",
         "| cat | works | letters | in-canon-mask |",
         "|---|--|--|--|"]
    for c, n in cat_works.most_common():
        L.append(f"| {c} | {n} | {cat_letters[c]:,} | "
                 f"{'yes' if c in CANON_CATS else 'no (AI judges)'} |")
    L.append("\n## sample Midrash-bucket titles (eyeball the cat mapping)\n")
    L.extend(f"- {t}" for t in unmapped_titles)
    open(OUT_MD, 'w', encoding='utf-8').write('\n'.join(L) + '\n')
    print(f"wrote {OUT_PKL} ({len(works)} works) + {OUT_MD}")
    print("works by cat:", dict(cat_works))
    print("CANON letters:", f"{canon_letters:,}")


if __name__ == '__main__':
    main()

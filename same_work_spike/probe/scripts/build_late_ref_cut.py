# -*- coding: utf-8 -*-
"""MAPV2-15i — the '>17th century' cut (Hillel's directive: cut all later
literature). A Genizah fragment (9th-13th c. literary layer) cannot be an
original witness to a work composed after ~1700; a "match" to Bialik/Mapu/an
18th-c. essay is anachronistic noise (really: that late author quoted the same
classical source the fragment does).

We do NOT rebuild the 180 MB reference pkl. We build a BLOCKLIST of reference
work_ids composed >= CUT_YEAR, straight from the Maagarim classification xlsx
(the authoritative source: מחבר / חיבור / תאריך / סוגה / שם קובץ). The
evidence engine filters candidate matches against this list and LOGS the count
(no silent truncation).

Date field is Hebrew-descriptive; the BCE convention is a trailing dash
(`180-` = 180 BCE). We take the latest CE year mentioned as the composition
upper bound; BCE-only or undatable rows are never cut (default keep).

Out: data/late_ref_blocklist.json  { "M:Ytext<N>": {year, genre, title, author} }
"""
import json
import os
import re
from collections import Counter

XLSX = r'C:\Users\gersh\Dropbox\דיקטה\מאגרים\חיבורים במאגרים.xlsx'
PROBE = r'C:\Genizahsearch\same_work_spike\probe'
OUT = PROBE + r'\data\late_ref_blocklist.json'

CUT_YEAR = 1700   # ">17th century": keep <=1699, cut >=1700
_YT = re.compile(r'Ytext(\d+)')
# canonical genres must never be cut (all classical); assert this holds
CANON_SUGYA = {'מקרא', 'תלמוד ומדרש'}


def latest_ce_year(date_str):
    """Latest CE year mentioned; BCE runs (digit + trailing dash/maqaf) and
    undatable strings -> None (keep)."""
    d = date_str or ''
    ce = []
    for m in re.finditer(r'\d{2,4}', d):
        y = int(m.group())
        if not (100 <= y <= 2100):
            continue
        nxt = d[m.end()] if m.end() < len(d) else ''
        # NB: `nxt in '-...'` would do SUBSTRING matching and `'' in s` is True,
        # so a year at end-of-string (where dates usually put it) would falsely
        # read as BCE. Membership against a tuple avoids that.
        if nxt in ('-', '־', '–'):   # hyphen / maqaf / en-dash -> BCE
            continue
        ce.append(y)
    return max(ce) if ce else None


def main():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)   # header

    block = {}
    n_rows = n_yt = n_dated = 0
    by_genre = Counter()
    canon_late = []
    samples = []
    for r in it:
        au, ti, date, sugya, fname = (list(r) + [None] * 5)[:5]
        n_rows += 1
        m = _YT.search(str(fname or ''))
        if not m:
            continue
        n_yt += 1
        y = latest_ce_year(str(date or ''))
        if y is not None:
            n_dated += 1
        if y is not None and y >= CUT_YEAR:
            wid = f'M:Ytext{m.group(1)}'
            block[wid] = {'year': y, 'genre': str(sugya or ''),
                          'title': str(ti or ''), 'author': str(au or '')}
            by_genre[str(sugya or '')] += 1
            if str(sugya or '') in CANON_SUGYA:
                canon_late.append((wid, y, ti))
            if len(samples) < 30:
                samples.append((y, str(sugya or ''), str(ti or '')))

    json.dump(block, open(OUT, 'w', encoding='utf-8'),
              ensure_ascii=False, indent=0)
    print(f"xlsx rows: {n_rows}; with Ytext id: {n_yt}; "
          f"with parseable year: {n_dated}")
    print(f"\nBLOCKLIST (>= {CUT_YEAR}): {len(block)} reference works")
    print("by genre:")
    for g, n in by_genre.most_common():
        print(f"  {n:4d}  {g}")
    # safety: canonical genres must never be cut
    print(f"\nCANON-genre works in blocklist (must be 0): {len(canon_late)}")
    if canon_late:
        for wid, y, ti in canon_late[:10]:
            print(f"  !! {wid} {y} {ti}")
    print("\nsample cut works (year | genre | title):")
    for y, g, t in sorted(samples, reverse=True):
        print(f"   {y}  [{g}]  {t[:50]}")
    print(f"\nwrote {OUT}")


if __name__ == '__main__':
    main()

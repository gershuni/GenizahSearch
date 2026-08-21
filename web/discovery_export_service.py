# -*- coding: utf-8 -*-
"""The findings-export workbook builder (EXPORT-01/02, phase 136.2).

ONE DATA PATH, NOT TWO. Every value in this workbook is the value the findings
page renders, produced by the SAME helper the row renderer calls -- the work
title through ``findings_rows._work_title`` (ruling R routes every title through
``display_work_title``), the relation through ``ds.relation_chip``, the coverage
clause through ``findings_rows.coverage_clause`` (which is itself derived from
``ds.row_headline`` so the one permitted percentage keeps its qualifier), the
novelty badge and the divergence marker through their own row helpers.

That is why this module imports two private names from ``findings_rows``. The
alternative is not "cleaner", it is a SECOND vocabulary: a spreadsheet that
formatted ``neutral_title`` itself would silently opt out of the title curation
and print a halakhic work's name over pages the owner ruled are mostly liturgy,
and a coverage figure composed here would be the second coverage vocabulary the
honesty gate exists to prevent. Identity with the rendered row is the property
being bought.

**THERE IS NO BAND COLUMN, DELIBERATELY.** The identification grain exposes
``best_band_rank`` and no band label, evidence source or confidence band --
``_render_row_meta`` records that it therefore renders no band tooltip, because
"deriving a label from a rank would be a second band vocabulary, which is
precisely what the panel renderer refuses to do for the same reason". A
spreadsheet column is not an exception to that; a rank in a cell reads as a
score, and it would leave the building in a file.

**NO PRECISION FIGURE REACHES A CELL.** The only permitted number is page
coverage in matched letters, and only where ``coverage_clause`` yields it -- the
direct family, non-propagated, with its qualifier attached.

**THE PASSAGES RIDE ON THE FINDING'S OWN ROW** (owner instruction,
2026-08-20). The first draft put them on a second sheet, which made the file
unusable for the thing a reader actually does with it -- read a match and
judge it -- because judging a row meant finding its counterpart by shelfmark
on another sheet. One row now carries the claim and the evidence for it, the
matched words are RED AND BOLD exactly as they are in the parallels export
(`shared.export_utils.build_rich_snippet_cell`, one implementation for both),
and the surrounding context is abbreviated so a row stays a row. The MATCHED
SPAN ITSELF IS NEVER ABBREVIATED: it is the claim, and a silently clipped
claim in a downloaded file is the defect this whole surface is written
against.

PURE. This module takes an envelope and returns bytes. It performs no I/O, opens
no database and reads no request, so the masking sweep can drive it directly
with a seeded row and read the resulting cells -- which is how it is scanned.
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from shared import discovery_display_strings as ds
from shared.export_utils import (
    MIDRASH_CREDIT_LINES,
    build_rich_snippet_cell,
    sanitize_text_for_excel,
)
from web.components import discovery_links
from web.components.findings_rows import (
    coverage_clause,
    divergence_marker,
    novelty_badge,
    _plural,
    _work_title,
)

logger = logging.getLogger(__name__)

__all__ = ["build_findings_workbook", "workbook_filename", "SHEET_KEYS"]


#: The sheets, in order. Named by a stable KEY rather than by their rendered
#: title so a test (and the masking capture) can address a sheet without
#: depending on the display language.
#:
#: THERE IS NO SEPARATE TEXT SHEET. It was one until the owner read a real
#: export: splitting a finding from its own evidence made every judgement a
#: two-sheet lookup keyed on a shelfmark that repeats across rows. The
#: passages are columns on the identification row now.
SHEET_KEYS: Tuple[str, ...] = ("identifications", "about")

_SHEET_TITLES: Dict[str, Dict[str, str]] = {
    "identifications": {"en": "Identifications", "he": "זיהויים"},
    "about": {"en": "About this export", "he": "אודות הייצוא"},
}

_COLUMNS: Tuple[Tuple[str, Dict[str, str]], ...] = (
    ("shelfmark", {"en": "Shelfmark", "he": "סימן מדף"}),
    ("library", {"en": "Library", "he": "ספרייה"}),
    ("sys_id", {"en": "System number", "he": "מספר מערכת"}),
    ("work", {"en": "Work", "he": "חיבור"}),
    ("author", {"en": "Author", "he": "מחבר"}),
    ("locus", {"en": "Location in work", "he": "מיקום בחיבור"}),
    # THE CATALOGUE'S OWN TITLE for the same manuscript, beside ours. The page
    # renders it (`_render_shelfmark`) and an export that dropped it would be
    # quietly less informative than the screen it claims to reproduce -- and
    # would blur the distinction the page is careful about, between what the
    # library says and what we computed. INJECTED, exactly as the row renderer
    # receives it, so this module still reads nothing.
    ("catalogue_title", {"en": "Catalogued as", "he": "בקטלוג"}),
    ("relation", {"en": "Relation", "he": "סוג הזיקה"}),
    # NO `novelty_source` COLUMN. `novelty_source_label` is a masked label and
    # would be safe to print -- but it is NOT in `SURFACE_FINDING_FIELDS`, so
    # on this grain it is never populated and the column would be empty in
    # every row of every export. A permanently blank column is worse than an
    # absent one: it reads as "no comparison source" rather than "this grain
    # does not carry one".
    ("novelty", {"en": "Compared with catalogues", "he": "בהשוואה לקטלוגים"}),
    ("divergence", {"en": "Catalogue note", "he": "הערת קטלוג"}),
    ("pages", {"en": "Pages", "he": "עמודים"}),
    ("coverage", {"en": "Page coverage", "he": "היקף בעמוד"}),
    # THE EVIDENCE, on the row that makes the claim. `{}` marks a header the
    # SHARED excerpt vocabulary owns -- see `_EXCERPT_HEADERS`.
    ("frag_text", {}),
    ("frag_note", {"en": "Transcription note", "he": "הערת תעתיק"}),
    ("work_text", {}),
    ("work_note", {"en": "Edition note", "he": "הערת מהדורה"}),
    ("attribution", {"en": "Attribution", "he": "ייחוס"}),
    ("link", {"en": "Link", "he": "קישור"}),
)

#: The two passage headers come from `ds.excerpt_strings`, not from the table
#: above. The retired text sheet retyped them and had already drifted on the
#: Hebrew -- "קטע מכתב היד" against the pane's "קטע כתב היד" -- which is the
#: second-vocabulary failure this module's docstring is about, in miniature.
_EXCERPT_HEADERS: Dict[str, str] = {
    "frag_text": "frag_label",
    "work_text": "work_label",
}

#: Columns whose cells carry `*`-marked highlight runs and are therefore
#: written as RICH TEXT rather than plain strings.
_PASSAGE_KEYS = frozenset(_EXCERPT_HEADERS)

#: Column widths, in openpyxl width units. The two passage columns are wide
#: because their content is a paragraph; everything else is sized to its header.
_WIDTHS: Dict[str, int] = {
    "shelfmark": 22, "library": 12, "sys_id": 22, "work": 34, "author": 22,
    "locus": 20, "relation": 26, "novelty": 26, "catalogue_title": 34,
    "divergence": 30, "pages": 9, "coverage": 22, "link": 46,
    "frag_text": 70, "frag_note": 22, "work_text": 70, "work_note": 34,
    "attribution": 30,
}

#: How much CONTEXT survives into a cell, per side of the match. The matched
#: span is never trimmed; only the words around it are. Chosen so a row stays
#: readable at one screen of rows rather than one row per screen -- the whole
#: reason the passages could move onto the finding's row at all.
_CONTEXT_CHARS = 160

#: The marker `build_rich_snippet_cell` splits on. A literal one in artifact
#: text would be read as ours, so every chunk of source text is scrubbed of it
#: before the markers go in.
_MARKER = "*"
_ELLIPSIS = "…"


#: How the About sheet names a grouped unit, in the reader's own language.
#: The page's "Show as" labels live in `tr()` (page chrome), which this pure
#: module must not reach into; these are the same two nouns, declined for the
#: sentence they sit in.
_UNIT_NOUNS: Dict[str, Dict[str, str]] = {
    "work": {"en": "work", "he": "חיבור"},
    "manuscript": {"en": "manuscript", "he": "כתב יד"},
}


#: The "what one row is" sentence, in three shapes. THREE, not one with a
#: number substituted: the single-group case produced "the identifications
#: those 1 rows open onto" in English and a worse construction in Hebrew, and a
#: provenance sentence that reads like a mail merge invites a reader to skip
#: the rest of the sheet. The unknown-count shape exists because
#: `export_group_count` is `None` whenever the group walk could not report one,
#: and a sentence with a hole in it is not an option here.
_GRAIN_SENTENCE: Dict[str, Dict[str, str]] = {
    "many": {
        "en": "One computed identification. You were viewing one row per "
              "{unit}; a spreadsheet has no expander, so this file lists the "
              "identifications behind those {count} rows, in the same order "
              "and grouped by the same columns.",
        "he": "זיהוי מחושב אחד. הצפייה הייתה בשורה אחת לכל {unit}; בגיליון "
              "אין אפשרות הרחבה, ולכן הקובץ מפרט את הזיהויים שמאחורי {count} "
              "השורות הללו, באותו סדר ומקובצים לפי אותן עמודות.",
    },
    "one": {
        "en": "One computed identification. You were viewing one row per "
              "{unit}; a spreadsheet has no expander, so this file lists the "
              "identifications behind that single row.",
        "he": "זיהוי מחושב אחד. הצפייה הייתה בשורה אחת לכל {unit}; בגיליון "
              "אין אפשרות הרחבה, ולכן הקובץ מפרט את הזיהויים שמאחורי אותה "
              "שורה אחת.",
    },
    "unknown": {
        "en": "One computed identification. You were viewing one row per "
              "{unit}; a spreadsheet has no expander, so this file lists the "
              "identifications behind those rows, in the same order and "
              "grouped by the same columns.",
        "he": "זיהוי מחושב אחד. הצפייה הייתה בשורה אחת לכל {unit}; בגיליון "
              "אין אפשרות הרחבה, ולכן הקובץ מפרט את הזיהויים שמאחורי אותן "
              "שורות, באותו סדר ומקובצים לפי אותן עמודות.",
    },
}


def _grain_sentence(groups: Any, requested: str, lang: str) -> str:
    """What one row of this file is, when it is not what the reader asked for."""
    try:
        count = int(groups)
    except (TypeError, ValueError):
        count = None
    shape = "unknown" if count is None else ("one" if count == 1 else "many")
    return _pick(_GRAIN_SENTENCE[shape], lang).format(
        unit=_pick(_UNIT_NOUNS.get(requested, {}), lang) or requested,
        count="{:,}".format(count) if count is not None else "")


def _lang_key(lang: Any) -> str:
    return "he" if str(lang or "en").lower().startswith("he") else "en"


def _pick(table: Mapping[str, str], lang: str) -> str:
    return table.get(lang) or table.get("en") or ""


def _text(value: Any) -> str:
    """Every string that enters a cell passes through here.

    ``sanitize_text_for_excel`` is the shared sanitizer both apps' exports
    already use: an XML-1.0 whitelist, formula-injection prevention (a cell
    opening with ``=`` is data, not a formula) and the 32,700-character cell
    limit. A discovery excerpt is short, but ``frag_before``/``after`` are
    context windows and a malformed artifact is exactly the case that must not
    produce a corrupt workbook.
    """
    if value is None:
        return ""
    return sanitize_text_for_excel(str(value))


def _column_label(key: str, label: Mapping[str, str], lang: str) -> str:
    shared_key = _EXCERPT_HEADERS.get(key)
    if shared_key:
        return ds.excerpt_strings(lang)[shared_key]
    return _pick(label, lang)


# ---------------------------------------------------------------------------
# The passages, and their highlight
# ---------------------------------------------------------------------------

#: Newline -> space, ONE CHARACTER FOR ONE.
_WS_MAP = {ord("\r"): " ", ord("\n"): " ", ord("\t"): " "}


def _flatten(text: str) -> str:
    """Line breaks to spaces, WITHOUT moving a single index.

    `frag_hl` / `work_hl` are char offsets INTO THE SPAN PIECE, so anything
    applied before they are painted must be length-preserving.
    `clean_text_single_line` collapses a CR-LF pair to one space and runs of
    spaces to one, which would slide every later offset left and highlight the
    wrong words -- silently, and only on the rows that happen to contain a line
    break. Mapping each character to one character cannot.

    Flattening at all is not optional: the Excel sanitiser DELETES control
    characters rather than replacing them, so a line break left in place would
    fuse the words on either side of it into one.
    """
    return text.translate(_WS_MAP)


def _strip_markers(text: str) -> str:
    """A literal marker character in artifact text, neutralised before ours
    go in -- `build_rich_snippet_cell` splits on it and could not tell the
    two apart."""
    return text.replace(_MARKER, " ")


def _marked_span(text: str, intervals, ja_braces: bool, *, whole: bool) -> str:
    """One excerpt piece as plain text, with markers around every matched run.

    MIRRORS `findings_rows._compose_excerpt_piece`: same per-character flag
    painting, same ruling on the J-corpus braces (colour the content, remove
    the marks). It has to be a mirror rather than a call because the pane
    emits HTML and a cell cannot take HTML --
    `test_the_exported_highlight_agrees_with_the_rendered_pane` drives both
    over the same inputs and compares the highlighted TEXT, so the two cannot
    drift into a second highlight vocabulary.

    The one deliberate difference: a brace-marked Hebrew word gets NO colour
    of its own here. A cell carries exactly one highlight -- the matched words
    -- because a second colour in a spreadsheet reads as a second kind of
    claim, and no wording has been ratified to explain what it would mean.

    `whole=True` is the fallback the pane uses when the bake carried no
    intervals (no work side to be parallel to, or a pre-round-2 asset): the
    span's own offsets are then the claim.
    """
    n = len(text)
    hl = [whole] * n
    for pair in intervals or ():
        try:
            start, end = int(pair[0]), int(pair[1])
        except (TypeError, ValueError, IndexError):
            continue
        for k in range(max(0, start), min(n, end)):
            hl[k] = True
    drop = [False] * n
    if ja_braces:
        for index, ch in enumerate(text):
            if ch in "{}":
                drop[index] = True
    out = []
    i = 0
    while i < n:
        if drop[i]:
            i += 1
            continue
        j = i
        while j < n and not drop[j] and hl[j] == hl[i]:
            j += 1
        raw = text[i:j]
        chunk = _strip_markers(raw)
        # A run of pure whitespace is never marked: a pair of markers around a
        # space reads in the cell as a matched word that is not there. The test
        # is on the RAW run, not the escaped one -- a matched run consisting of
        # literal marker characters escapes to spaces, and keying off the
        # escaped text dropped its highlight, which is content loss dressed as
        # tidiness (Codex round 4, finding 1).
        out.append(f"{_MARKER}{chunk}{_MARKER}" if hl[i] and raw.strip()
                   else chunk)
        i = j
    return "".join(out)


def _clip_context(text: str, *, tail: bool) -> str:
    """Abbreviate ONE context piece, keeping the edge that touches the match.

    `tail=True` is the BEFORE piece (keep its end), `tail=False` the AFTER
    piece (keep its start). The cut is nudged to the nearest word boundary
    within a short reach so the abbreviation does not invent a half-word a
    reader could take for the manuscript's own text, and it is always marked
    with an ellipsis -- an abbreviation nobody can see is a claim about the
    page that this file is not entitled to make.
    """
    if len(text) <= _CONTEXT_CHARS:
        return text
    # THE ELLIPSIS COMES OUT OF THE BUDGET, not on top of it. Slicing to
    # `_CONTEXT_CHARS` and then appending " …" returned 162 characters from a
    # function documented as bounded by 160 (Codex round 4, finding 4) -- a
    # small number and a real one: a bound nobody enforces is a bound nobody
    # can rely on later.
    budget = _CONTEXT_CHARS - len(_ELLIPSIS) - 1
    if tail:
        cut = text[-budget:]
        space = cut.find(" ")
        if 0 <= space < 32:
            cut = cut[space + 1:]
        return f"{_ELLIPSIS} {cut}"
    cut = text[:budget]
    space = cut.rfind(" ")
    if space > budget - 32:
        cut = cut[:space]
    return f"{cut} {_ELLIPSIS}"


#: The Excel cell ceiling. `sanitize_text_for_excel` truncates at 32,700 --
#: BLINDLY, so a cut falling between an opening marker and its mate would leave
#: an odd number of markers and invert the highlight for everything after it in
#: the cell. The bake caps a span far below this (600 stream letters; the
#: longest passage in the live artifact is 1,936 characters), so this is a
#: backstop and not a live path -- but a backstop that fails silently is not a
#: backstop (Codex round 4, finding 2).
_CELL_TEXT_MAX = 32000


def _fit_cell(marked: str) -> str:
    """Bound a marked passage BEFORE the sanitiser can cut it mid-marker."""
    if len(marked) <= _CELL_TEXT_MAX:
        return marked
    head = marked[:_CELL_TEXT_MAX]
    if head.count(_MARKER) % 2:
        head += _MARKER  # close the run the cut fell inside
    return f"{head} {_ELLIPSIS}"


def _passage(row: Mapping[str, Any], side: str) -> str:
    """One side's {before, span, after}, marked and abbreviated."""
    ja = side == "work" and row.get("work_markup") == "ja_braces"

    def piece(name: str, intervals, whole: bool) -> str:
        return _marked_span(_flatten(str(row.get(f"{side}_{name}") or "")),
                            intervals, ja, whole=whole)

    hl = row.get(f"{side}_hl")
    return "".join((
        _clip_context(piece("before", None, False), tail=True),
        piece("span", hl, hl is None),
        _clip_context(piece("after", None, False), tail=False),
    ))


def _flag(value: Any) -> bool:
    """A sidecar 0/1 flag, tolerant of the None the work side carries."""
    try:
        return bool(int(value))
    except (TypeError, ValueError):
        return False


def _excerpt_values(item: Mapping[str, Any],
                    strings: Mapping[str, str]) -> Dict[str, str]:
    """The five evidence cells for one finding row.

    THE HONEST STATES ARE WRITTEN OUT, NOT LEFT BLANK -- and there are two of
    them, which the retired text sheet could express by omitting a row and a
    merged row cannot:

    * no excerpt at all (evidence that is not excerpt-eligible) carries the
      page's own `none` sentence;
    * an excerpt whose work side is empty is the documented masked-non-Bible /
      mismatched-stream / below-threshold case -- the bake writes `None` for
      all four work pieces precisely so the UI can say so -- and carries
      `work_unavailable`.

    A blank cell would be indistinguishable from an edition that simply had
    nothing to show, and unlike an empty area on a page an empty CELL reads as
    a value.
    """
    empty = {"frag_text": "", "frag_note": "", "work_text": "",
             "work_note": "", "attribution": ""}
    excerpt = item.get("excerpt")
    if not isinstance(excerpt, Mapping):
        return {**empty, "frag_text": strings.get("none", "")}

    frag = _passage(excerpt, "frag")
    # `work_span`, NOT any of the three pieces. `_render_panes` keys the whole
    # work pane on `row.get("work_span")`, so a row carrying only context would
    # be shown as an edition passage here while the page truthfully said the
    # edition is unavailable -- the exact divergence this module exists to
    # prevent (Codex round 4, finding 5; the retired text sheet had the same
    # bug). No row in the live artifact is in that state, which is why nothing
    # caught it.
    has_work = bool(excerpt.get("work_span"))
    if not frag.strip() and not has_work:
        return {**empty, "frag_text": strings.get("none", "")}
    # An EMPTY manuscript side beside a POPULATED work side stays blank, which
    # is what the pane does -- it renders the fragment paragraph whatever is in
    # it. There is no ratified sentence for "the manuscript text is missing",
    # and inventing one here would be a second vocabulary for a state the
    # surface has never had to name. `frag_span` is non-empty in all 48,270
    # rows of the live artifact.

    work_note = ""
    if has_work:
        work = _passage(excerpt, "work")
        if excerpt.get("work_source") == "reprojected":
            work_note = strings.get("reprojected_note", "")
    else:
        work = strings.get("work_unavailable", "")

    # THE BAKE'S OWN ABBREVIATION, named. `pieces()` caps a long span and joins
    # its head and tail with a visible U+22EF, flagging the row -- 22.7% of
    # fragment spans and 25.3% of work spans in the live artifact. Neither
    # surface said so, so the mark sat in the middle of a quoted passage
    # meaning nothing (Codex round 4, finding 9). A file that travels without
    # its page has to carry the explanation with it.
    clipped = strings.get("clipped_note", "")
    if clipped and _flag(excerpt.get("work_clipped")) and has_work:
        work_note = (work_note + " " if work_note else "") + clipped

    n_spans = excerpt.get("n_spans")
    try:
        if int(n_spans or 0) > 1:
            note = strings.get("multi_span", "")
            if note:
                work_note = (work_note + " " if work_note else "") + note.format(
                    count=int(n_spans))
    except (TypeError, ValueError):
        pass

    # ONLY for the automated layer, exactly as `_render_panes` appends it. The
    # first draft stamped the qualifier on every row, which asserted machine
    # reading over the FGP and PGP transcriptions that are human work.
    frag_note = (strings.get("frag_htr_note", "")
                 if (excerpt.get("text_layer") or "") == "htr" else "")
    if clipped and _flag(excerpt.get("frag_clipped")):
        frag_note = (frag_note + " " if frag_note else "") + clipped

    return {
        "frag_text": frag,
        "frag_note": frag_note,
        "work_text": work,
        "work_note": work_note,
        "attribution": str(excerpt.get("attribution") or ""),
    }


def _relation_text(item: Mapping[str, Any], lang: str) -> str:
    """The relation chip, or an explicit blank.

    ``relation_chip`` RAISES on a vocabulary it does not know, and the row
    renderer catches that and omits the chip. Omitting a chip is invisible on a
    web row; omitting it silently in a spreadsheet column leaves a blank cell
    that reads as "no relation" rather than "this build emitted a relation this
    release cannot name". So the failure is caught, LOGGED by relation value,
    and the cell is left empty on purpose -- and a test asserts the populated
    case rather than merely asserting no exception, because a helper that
    always returned "" would pass the second check and fail the reader.
    """
    relation = item.get("rendered_relation")
    if not relation:
        return ""
    try:
        return ds.relation_chip(relation, lang)
    except ValueError:
        logger.warning(
            "discovery export: unknown rendered_relation %r has no chip string",
            relation)
        return ""


def _novelty_text(item: Mapping[str, Any], lang: str) -> str:
    badge = novelty_badge(item, lang)
    return badge[0] if badge else ""


def _divergence_text(item: Mapping[str, Any], lang: str) -> str:
    """The catalogue marker AND the statement its tooltip carries.

    On the page the two-sentence statement rides on the chip's tooltip. A
    spreadsheet has no hover, so dropping it would ship the marker without the
    thing that makes the marker honest -- and ruling F's wording is the whole
    point of the marker. Both go in the cell.
    """
    marker = divergence_marker(item, lang)
    if not marker:
        return ""
    text, tooltip = marker
    if tooltip and tooltip != text:
        return "{} — {}".format(text, tooltip)
    return text


def _row_values(item: Mapping[str, Any], lang: str,
                base_url: str = "", catalogue_title=None,
                strings: Optional[Mapping[str, str]] = None) -> Dict[str, str]:
    sys_id = item.get("sys_id")
    link = ""
    if sys_id:
        try:
            # The folio address comes from the ONE builder -- never composed
            # here; the AST guard exists because a surface that builds its own
            # gets it wrong, and a wrong link in a downloaded file outlives the
            # page that produced it. `base_url` only PREFIXES an origin onto
            # that builder's output, which is not a second address derivation.
            #
            # Absolute on purpose: a relative `/browse?...` is dead the moment
            # the workbook is opened outside the browser, which is the normal
            # way a spreadsheet is read.
            link = base_url + discovery_links.browse_url(
                sys_id,
                page=item.get("first_match_page"),
                volume_ie=item.get("first_match_volume_ie"),
            )
        except Exception as e:  # pragma: no cover - the builder is total
            # NEVER hand-build a `/browse` URL here as a fallback. The AST guard
            # exists because a surface that composes its own folio address gets
            # it wrong, and a wrong link in a downloaded file outlives the page.
            logger.warning("discovery export: browse_url failed (%s)",
                           type(e).__name__)
            link = ""
    locus = item.get("locus_label")
    locus_text = ""
    if isinstance(locus, str) and locus.strip():
        # NO STRUCTURAL CUE. On a row the leading arrow says "this line belongs
        # to the title above it"; a column header says the same thing already,
        # and the glyph carried into every cell breaks sorting and copy-paste
        # for no gain (owner report, 2026-08-20). Still routed through the one
        # helper that owns what a locus looks like.
        locus_text = ds.locus_subline(locus.strip(), lang, cue=False)
    values = {
        "shelfmark": _text(item.get("shelfmark_display")),
        "library": _text(item.get("library_code")),
        "sys_id": _text(sys_id),
        "work": _text(_work_title(item, lang)),
        "author": _text(item.get("author")),
        "locus": _text(locus_text),
        "catalogue_title": _text(catalogue_title(item) if catalogue_title else ""),
        "relation": _text(_relation_text(item, lang)),
        "novelty": _text(_novelty_text(item, lang)),
        "divergence": _text(_divergence_text(item, lang)),
        "pages": _text(_plural("pages", item.get("page_count"), lang)),
        "coverage": _text(coverage_clause(item, lang) or ""),
        "link": _text(link),
    }
    # The passage cells are NOT sanitised here: they carry highlight markers,
    # and `_passage_cell` has to sanitise BEFORE splitting on them so a
    # formula-injection prefix cannot land inside a run instead of in front of
    # the cell.
    values.update(_excerpt_values(item, strings or ds.excerpt_strings(lang)))
    return values


def _header_cell(ws: Any, value: str, rtl: bool) -> Any:
    cell = WriteOnlyCell(ws, value=value)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        horizontal="right" if rtl else "left", vertical="top", wrap_text=True)
    return cell


def _body_cell(ws: Any, value: str, rtl: bool) -> Any:
    cell = WriteOnlyCell(ws, value=value)
    cell.alignment = Alignment(
        horizontal="right" if rtl else "left", vertical="top", wrap_text=True)
    return cell


def _passage_cell(ws: Any, value: str, rtl: bool) -> Any:
    """A passage cell, with the matched words RED AND BOLD.

    `build_rich_snippet_cell` is the helper the search and parallels exports
    already render their snippet column with -- the same colour, the same
    weight, one implementation. It sanitises before splitting on the marker
    (the T-94-01 ordering) and returns a plain string when there is nothing
    marked, so the "no text available" sentences stay unhighlighted.
    """
    cell = WriteOnlyCell(
        ws, value=build_rich_snippet_cell(_fit_cell(value),
                                          sanitize_fn=sanitize_text_for_excel))
    cell.alignment = Alignment(
        horizontal="right" if rtl else "left", vertical="top", wrap_text=True)
    return cell


def _apply_widths(ws: Any, keys: Sequence[str], rtl: bool) -> None:
    for index, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(index)].width = _WIDTHS.get(key, 20)
    # The sheet's own reading direction, so a Hebrew workbook opens right to
    # left rather than needing the reader to flip it. View-level, exactly like
    # the conditional RTL the dossier export already applies.
    ws.sheet_view.rightToLeft = bool(rtl)


def build_findings_workbook(
    envelope: Mapping[str, Any],
    *,
    lang: str = "en",
    filters: Optional[Mapping[str, Any]] = None,
    generated_at: Optional[str] = None,
    page_url: Optional[str] = None,
    base_url: str = "",
    catalogue_title=None,
) -> bytes:
    """The workbook, as bytes, from ONE ``collect_findings_for_export`` envelope.

    ``write_only=True``: rows are streamed to the archive as they are appended
    rather than held as a cell graph. The export is uncapped in rows by owner
    decision, so an in-memory workbook of ~28,600 rows -- each now carrying two
    passages of highlighted rich text -- is exactly the shape that must not be
    built.

    Raises ``ValueError`` on an envelope whose status is not ``ok``. There is no
    partial workbook: a short file that looks complete is the defect this whole
    surface is written against, and the caller turns the status into an HTTP
    response instead.
    """
    status = envelope.get("status")
    if status != "ok":
        raise ValueError(
            "build_findings_workbook: refusing to build from a {!r} envelope"
            .format(status))

    lang = _lang_key(lang)
    rtl = lang == "he"
    items: List[Mapping[str, Any]] = list(envelope.get("items") or ())
    meta: Mapping[str, Any] = envelope.get("meta") or {}
    strings = ds.excerpt_strings(lang)

    wb = Workbook(write_only=True)

    # ---- sheet 1: the identifications, each with its own evidence ----------
    keys = [k for k, _ in _COLUMNS]
    ws = wb.create_sheet(_pick(_SHEET_TITLES["identifications"], lang))
    _apply_widths(ws, keys, rtl)
    ws.append([_header_cell(ws, _column_label(key, label, lang), rtl)
               for key, label in _COLUMNS])
    for item in items:
        values = _row_values(item, lang, base_url=base_url,
                             catalogue_title=catalogue_title, strings=strings)
        ws.append([(_passage_cell if k in _PASSAGE_KEYS else _body_cell)(
            ws, values[k], rtl) for k in keys])

    # ---- sheet 2: what this file is ----------------------------------------
    #
    # NOT decoration. Every honesty obligation the page discharges in its
    # chrome -- that these are computed text matches rather than asserted
    # identifications, that the set is not exhaustive, which artifact produced
    # it, and which filters were applied -- has to travel WITH the file, because
    # the file is the thing that leaves the building and gets mailed to a
    # colleague. A spreadsheet of claims with no provenance sheet is the
    # honesty failure this surface is most exposed to.
    ws3 = wb.create_sheet(_pick(_SHEET_TITLES["about"], lang))
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 100
    ws3.sheet_view.rightToLeft = bool(rtl)

    def _about(label: str, value: Any) -> None:
        ws3.append([_header_cell(ws3, _text(label), rtl),
                    _body_cell(ws3, _text(value), rtl)])

    _about(_pick({"en": "What this file contains", "he": "מה מכיל הקובץ"}, lang),
           _pick({
               "en": "Computed text matches between Genizah manuscript pages and "
                     "known works. These are candidate identifications produced by "
                     "software, not catalogue assertions.",
               "he": "התאמות טקסט מחושבות בין דפי כתבי יד מהגניזה לחיבורים ידועים. "
                     "אלה הצעות זיהוי שהופקו על ידי תוכנה, ולא קביעות קטלוגיות.",
           }, lang))
    _about(_pick({"en": "Coverage", "he": "היקף"}, lang),
           ds.recall_disclaimer(lang))
    # WHY A BLANK IN THAT COLUMN IS NOT A VERDICT. `novelty_badge` returns a
    # badge for exactly two shades -- the candidacy one and the fail-closed
    # `not_checked` -- and nothing for the rest, because "a badge asserts
    # something; its absence asserts nothing" and no reader wording has been
    # ratified for the remaining shades. On the page an empty space is
    # self-evidently not a claim. In a spreadsheet an empty CELL reads as a
    # value, so the file has to say what the emptiness means.
    _about(_pick({"en": "Blank in ‘Compared with catalogues’",
                  "he": "תא ריק ב‘בהשוואה לקטלוגים’"}, lang),
           _pick({
               "en": "Not a verdict. This column is filled only where there is "
                     "agreed wording for what the comparison found; a blank cell "
                     "means nothing is being asserted either way.",
               "he": "אין זו קביעה. העמודה מלאה רק כאשר קיים ניסוח מוסכם לתוצאת "
                     "ההשוואה; תא ריק אינו טוען דבר לכאן או לכאן.",
           }, lang))
    _about(_pick({"en": "How rows are grouped", "he": "אופן חלוקת השורות"}, lang),
           ds.rule_sentence(lang))
    # WHAT WAS DONE TO THE PASSAGES, said on the file's own face. The matched
    # words are coloured and the context around them is cut, and a reader who
    # cannot see that a passage was abbreviated may take the cell for the whole
    # of what the page says.
    _about(_pick({"en": "The passages", "he": "הקטעים"}, lang),
           _pick({
               "en": "Matched words are shown in red. A passage is a window "
                     "around the match, not the whole page, and where it was "
                     "cut the text says so with an ellipsis — including in the "
                     "middle of a long match, which the note column names. "
                     "Follow the row's link for the full page.",
               "he": "המילים התואמות מסומנות באדום. הקטע הוא חלון סביב ההתאמה "
                     "ולא הדף כולו, ובמקום שבו נחתך הטקסט מסומן בשלוש נקודות — "
                     "גם באמצעו של קטע ארוך, ועל כך מציינת עמודת ההערה. "
                     "לצפייה בדף המלא יש לפתוח את הקישור שבשורה.",
           }, lang))
    # WHAT ONE ROW IS, whenever that is not what the reader asked for. A file
    # of 40,000 rows headed "one row per work" is a puzzle unless it says why:
    # the page's grouped rows carry an EXPANDER and a spreadsheet has none, so
    # the export flattens the grouping and keeps it as columns instead.
    requested = meta.get("export_unit_requested")
    grain = meta.get("export_grain")
    if requested and grain and requested != grain:
        _about(_pick({"en": "One row is", "he": "כל שורה היא"}, lang),
               _grain_sentence(meta.get("export_group_count"), requested, lang))
        # A short GROUP walk costs the ORDER, not the rows. Said out loud
        # rather than left to look like the reader's own sort choice --
        # but said ONLY about the cause that actually fired. The umbrella
        # flag `export_group_order_complete` goes False for three different
        # reasons and two of them place every row, so printing the
        # rows-at-the-end sentence off the umbrella told the reader to look
        # for something that was not there.
        if int(meta.get("export_group_unplaced") or 0) > 0:
            _about(_pick({"en": "Grouping", "he": "הקיבוץ"}, lang),
                   _pick({
                       "en": "Some rows could not be placed in their group and "
                             "appear at the end. Every matching row is present; "
                             "only the ordering is affected.",
                       "he": "חלק מהשורות לא שובצו בקבוצתן ומופיעות בסוף. כל "
                             "השורות התואמות נמצאות בקובץ; רק הסדר הושפע.",
                   }, lang))
        # The count in the sentence above is a CAP once the artifact stopped
        # counting. A number presented as a count when it is a ceiling is
        # the one thing `DISCOVERY_FINDINGS_COUNT_MAX` is documented never
        # to be allowed to do, and this file is read away from the page.
        if meta.get("export_group_count_approximate"):
            _about(_pick({"en": "Group count", "he": "מספר הקבוצות"}, lang),
                   _pick({
                       "en": "The number above is a ceiling, not a count -- "
                             "counting stopped there. Every matching row is "
                             "still in this file.",
                       "he": "המספר שלמעלה הוא תקרה ולא ספירה — הספירה נעצרה "
                             "בו. כל השורות התואמות נמצאות בקובץ.",
                   }, lang))
    _about(_pick({"en": "Rows in this file", "he": "שורות בקובץ"}, lang),
           meta.get("row_count", len(items)))
    _about(_pick({"en": "Matching rows reported", "he": "שורות תואמות שדווחו"}, lang),
           meta.get("reported_total"))
    # Stated rather than assumed. On a healthy read these two numbers are equal;
    # printing the flag means a file built from a disagreeing walk says so on
    # its own face instead of looking complete.
    # THREE STATES, not two. `walk_complete` is `None` when the reported total
    # was a CAP (`DISCOVERY_FINDINGS_COUNT_MAX`) rather than a count -- the walk
    # then cannot be checked against it, and printing "yes" there would be the
    # file certifying something nobody measured.
    walk_complete = meta.get("walk_complete", True)
    if walk_complete is None:
        completeness = _pick({
            "en": "Unverified — the matching-row total was capped by "
                  "configuration, so this file cannot be checked against it.",
            "he": "לא אומת — סך השורות התואמות הוגבל בתצורה, ולכן לא ניתן "
                  "לאמת מולו את הקובץ.",
        }, lang)
    elif walk_complete:
        completeness = _pick({"en": "yes", "he": "כן"}, lang)
    else:
        completeness = _pick({"en": "NO — this file is incomplete",
                              "he": "לא — הקובץ חלקי"}, lang)
    _about(_pick({"en": "Complete result set", "he": "מערך תוצאות מלא"}, lang),
           completeness)
    _about(_pick({"en": "Data version", "he": "גרסת הנתונים"}, lang),
           meta.get("sidecar_version"))
    if generated_at:
        _about(_pick({"en": "Generated", "he": "הופק בתאריך"}, lang), generated_at)
    if page_url:
        _about(_pick({"en": "Source page", "he": "עמוד המקור"}, lang), page_url)

    # ---- the credit ---------------------------------------------------------
    #
    # THE MANUSCRIPT SIDE OF EVERY PASSAGE IN THIS FILE IS MiDRASH AUTOMATIC
    # TRANSCRIPTION. The first export shipped without saying so: the workbook
    # carried thousands of lines of someone else's dataset and named nobody
    # (owner instruction, 2026-08-20). The citation rows come from
    # `shared.export_utils.MIDRASH_CREDIT_LINES` -- the same three lines the
    # search export, the research dossier and the desktop app print -- and are
    # NEVER translated: a published citation is cited as published.
    ws3.append([])
    ws3.append([_header_cell(
        ws3, _pick({"en": "Credits", "he": "קרדיט"}, lang), rtl)])
    # NOT "every manuscript passage here". 2,759 of the artifact's 48,270
    # excerpts are FGP transcriptions and 87 are PGP -- human work by other
    # projects -- and the export already distinguishes them in the
    # transcription-note column. A blanket credit would miscredit 5.9% of the
    # rows to a project that did not make them (Codex round 4, finding 8).
    _about(_pick({"en": "Manuscript text", "he": "טקסט כתב היד"}, lang),
           _pick({
               "en": "Manuscript passages marked as automated transcriptions "
                     "come from the MiDRASH dataset. Please cite it in any work "
                     "that draws on them. Rows without that mark carry a "
                     "transcription made by another project.",
               "he": "קטעי כתב היד המסומנים כתעתיק אוטומטי לקוחים ממאגר "
                     "MiDRASH. נא לצטטו בכל עבודה הנשענת עליהם. שורות ללא סימון "
                     "זה נושאות תעתיק שנעשה בפרויקט אחר.",
           }, lang))
    for line in MIDRASH_CREDIT_LINES:
        ws3.append([_body_cell(ws3, "", rtl), _body_cell(ws3, _text(line), rtl)])

    if filters:
        ws3.append([])
        ws3.append([_header_cell(
            ws3, _pick({"en": "Filters applied", "he": "מסננים שהוחלו"}, lang), rtl)])
        for key, value in filters.items():
            if value in (None, "", (), []):
                continue
            ws3.append([_body_cell(ws3, _text(key), rtl),
                        _body_cell(ws3, _text(value), rtl)])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def workbook_filename(meta: Optional[Mapping[str, Any]] = None) -> str:
    """A stable, ASCII, per-artifact filename.

    Carries the sidecar version rather than a query name. Two reasons: a
    downloaded findings file is only reproducible against the artifact that
    produced it, and a filename built from user-supplied text is the shape that
    leaked one reader's query into another reader's download on the search
    export (see the STATE-01 note in ``web/api.py``).
    """
    version = ""
    if meta:
        version = str(meta.get("sidecar_version") or "")
    safe = "".join(ch for ch in version if ch.isalnum() or ch in "-_")
    return "computed-identifications-{}.xlsx".format(safe) if safe \
        else "computed-identifications.xlsx"

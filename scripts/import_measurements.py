#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import FIST measurement data from FIST_Computed_Measurements.xlsx into fjms_enrichment.db.

Canonical build order:
  1. scripts/export_fist_enrichment.py  (builds base sidecar from FIST.db)
  2. scripts/import_measurements.py     (THIS SCRIPT -- adds/replaces measurement tables)

This script is the SOLE owner of all measurement-related tables:
  - extra_info           (~743K rows from Extra_Info sheet)
  - computed_measurements (~434K rows from Computed_Measurements sheet)
  - blank_images          (~165K rows from Blank_Images sheet)
  - catalog_sizes         (REPLACES the one from export_fist_enrichment.py with cm-normalized values)
  - manuscript_measurements (summary table, one row per AlmaId, built from unflagged records only)
  - import_meta           (version tracking)

Data sources:
  - fist_data/FIST_Computed_Measurements.xlsx  (4 sheets)
  - FIST_DB_BACKUP/FIST.db                    (for catalog_sizes AlmaId resolution)
  - fist_data/fjms_enrichment.db               (target sidecar)
"""

import os
import sqlite3
import sys
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    # Fallback: simple progress wrapper
    def tqdm(iterable, desc="", unit="", total=None):
        return iterable

BATCH_SIZE = 5000

# Paths
XLSX_PATH = "fist_data/FIST_Computed_Measurements.xlsx"
FIST_DB_PATH = "FIST_DB_BACKUP/FIST.db"
TARGET_DB_PATH = "fist_data/fjms_enrichment.db"

def build_fgp_to_alma_from_fist(fist_db_path: str) -> dict[str, str]:
    """Build FGP→AlmaId lookup from FIST.db (exact integers, no float precision loss).

    The xlsx Extra_Info sheet stores 18-digit AlmaIds as Excel Number (float64),
    which corrupts the last 2-3 digits due to IEEE 754 double precision (~15.9 sig digits).
    Example: xlsx has 9.900017468002052e+17 → int() gives 990001746800205184 (WRONG)
             FIST.db has AlmaId=990001746800205171 (CORRECT)

    So we ignore the xlsx AlmaId column entirely and build the lookup from FIST.db:
    dbo_ImgDigitalImage.FGPImageNumberId → dbo_InventoryAlma.AlmaId
    """
    import sqlite3 as _sqlite3
    conn = _sqlite3.connect(fist_db_path)
    cursor = conn.execute("""
        SELECT img.FGPImageNumberId, CAST(alma.AlmaId AS TEXT)
        FROM dbo_ImgDigitalImage img
        JOIN dbo_InventoryAlma alma ON img.InventoryId = alma.InventoryId
    """)
    fgp_to_alma = {}
    for fgp_id, alma_id in cursor:
        # FIST.db stores FGPImageNumberId as integer (1, 2, 3...)
        # xlsx FGP column uses "C" prefix (C1, C2, C3...)
        # Store both formats for robust matching
        fgp_to_alma[str(fgp_id)] = alma_id
        fgp_to_alma[f"C{fgp_id}"] = alma_id
    conn.close()
    return fgp_to_alma


def build_shelfmark_to_alma_from_fist(fist_db_path: str) -> dict[str, str]:
    """Build Shelfmark→AlmaId lookup from FIST.db for catalog_sizes resolution.

    Uses MIN(AlmaId) per Shelfmark when multiple exist (matching xlsx About sheet convention).
    """
    import sqlite3 as _sqlite3
    conn = _sqlite3.connect(fist_db_path)
    cursor = conn.execute("""
        SELECT inv.Shelfmark, CAST(MIN(alma.AlmaId) AS TEXT)
        FROM dbo_Inventory inv
        JOIN dbo_InventoryAlma alma ON inv.InventoryId = alma.InventoryId
        GROUP BY inv.Shelfmark
    """)
    sm_to_alma = {}
    for shelfmark, alma_id in cursor:
        if shelfmark:
            sm_to_alma[shelfmark.strip()] = alma_id
    conn.close()
    return sm_to_alma


def read_header_map(ws, expected_columns: list[str]) -> dict[str, int]:
    """Read header row and return column_name -> index mapping.

    Verifies expected columns exist. Does NOT hardcode indices.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {}
    for idx, cell_val in enumerate(header_row):
        if cell_val is not None:
            header_map[str(cell_val).strip()] = idx

    missing = [c for c in expected_columns if c not in header_map]
    if missing:
        print(f"  WARNING: Missing expected columns: {missing}")
        print(f"  Available columns: {list(header_map.keys())}")

    return header_map


def step1_extra_info(wb, conn, fgp_to_alma: dict):
    """Step 1: Import Extra_Info sheet using FIST.db AlmaIds (not xlsx float AlmaIds)."""
    print("\nStep 1: Importing Extra_Info sheet...")
    ws = wb["Extra_Info"]

    # Note: AlmaId not in expected_cols — we get it from fgp_to_alma (FIST.db), not xlsx
    expected_cols = [
        "FGP", "Shelfmark", "Material", "Size_Category",
        "NumFolio", "NumBifolio", "PixelWidth", "PixelHeight",
        "Image_Type", "Rotation_Angle_deg",
    ]
    hdr = read_header_map(ws, expected_cols)

    conn.execute("DROP TABLE IF EXISTS extra_info")
    conn.execute("""
        CREATE TABLE extra_info (
            FGP TEXT NOT NULL PRIMARY KEY,
            AlmaId TEXT,
            Shelfmark TEXT,
            Material TEXT,
            Size_Category TEXT,
            NumFolio INTEGER,
            NumBifolio INTEGER,
            PixelWidth INTEGER,
            PixelHeight INTEGER,
            Image_Type TEXT,
            Rotation_Angle_deg REAL
        )
    """)

    batch = []
    total = 0
    alma_found = 0

    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), desc="  extra_info", unit=" rows"):
        fgp = row[hdr.get("FGP", 0)]
        if fgp is None:
            continue
        fgp = str(fgp).strip()

        # AlmaId from FIST.db lookup (exact integer), NOT from xlsx (float precision loss)
        alma_id = fgp_to_alma.get(fgp)
        if alma_id:
            alma_found += 1

        batch.append((
            fgp,
            alma_id,
            row[hdr.get("Shelfmark", 1)] if hdr.get("Shelfmark") is not None else None,
            row[hdr.get("Material", 5)] if hdr.get("Material") is not None else None,
            row[hdr.get("Size_Category", 6)] if hdr.get("Size_Category") is not None else None,
            row[hdr.get("NumFolio", 7)] if hdr.get("NumFolio") is not None else None,
            row[hdr.get("NumBifolio", 8)] if hdr.get("NumBifolio") is not None else None,
            row[hdr.get("PixelWidth", 9)] if hdr.get("PixelWidth") is not None else None,
            row[hdr.get("PixelHeight", 10)] if hdr.get("PixelHeight") is not None else None,
            row[hdr.get("Image_Type", 11)] if hdr.get("Image_Type") is not None else None,
            row[hdr.get("Rotation_Angle_deg", 12)] if hdr.get("Rotation_Angle_deg") is not None else None,
        ))

        if len(batch) >= BATCH_SIZE:
            conn.executemany(
                "INSERT OR IGNORE INTO extra_info VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                batch,
            )
            total += len(batch)
            batch = []

    if batch:
        conn.executemany(
            "INSERT OR IGNORE INTO extra_info VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            batch,
        )
        total += len(batch)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_ei_alma ON extra_info(AlmaId)")

    distinct_alma = conn.execute("SELECT COUNT(DISTINCT AlmaId) FROM extra_info WHERE AlmaId IS NOT NULL").fetchone()[0]
    print(f"  Imported {total:,} rows ({distinct_alma:,} distinct AlmaIds, {alma_found:,} with AlmaId)")

    return total


def validate_alma_ids(fgp_to_alma: dict, libraries_csv_path: str = "libraries.csv"):
    """Validate FIST.db-sourced AlmaIds against libraries.csv system_numbers."""
    print("\n  AlmaId validation (FIST.db-sourced, no xlsx float conversion):")

    alma_values = set(fgp_to_alma.values())
    print(f"    Total distinct AlmaIds: {len(alma_values):,}")

    # Validate against libraries.csv
    if os.path.exists(libraries_csv_path):
        csv_ids = set()
        try:
            with open(libraries_csv_path, "r", encoding="utf-8-sig") as f:
                for i, line in enumerate(f):
                    if i == 0:
                        continue
                    parts = line.strip().split(",")
                    if parts:
                        csv_ids.add(parts[0])
            overlap = len(alma_values & csv_ids)
            pct = (overlap / len(alma_values) * 100) if alma_values else 0
            print(f"    libraries.csv match: {overlap:,}/{len(alma_values):,} ({pct:.1f}%)")
            if overlap == 0:
                print("    FATAL: 0 AlmaIds matched libraries.csv! FIST.db→AlmaId join may be broken.")
                return False
            if pct < 50:
                print(f"    WARNING: Only {pct:.1f}% matched — some AlmaIds may not be in current libraries.csv")
        except Exception as e:
            print(f"    Could not read libraries.csv for validation: {e}")
    else:
        print(f"    libraries.csv not found at {libraries_csv_path} — skipping validation")

    return True


def step2_computed_measurements(wb, conn, fgp_to_alma: dict):
    """Step 2: Import Computed_Measurements sheet."""
    print("\nStep 2: Importing Computed_Measurements sheet...")
    ws = wb["Computed_Measurements"]

    expected_cols = [
        "FGP", "Image_Side", "Component_Num", "Bifolio_Side",
        "Page_Width_cm", "Page_Height_cm", "Num_Lines",
        "Left_Margin_cm", "Right_Margin_cm", "Top_Margin_cm", "Bottom_Margin_cm",
        "Written_Width_cm", "Written_Height_cm",
        "Avg_Line_Height_Text_mm", "Text_Density_per10cm",
        "DpiGrid", "DisplayDPI",
        "Flag_DPI_High", "Flag_DPI_Low", "Flag_Negative_Margin", "Flag_BifolioLoc_Error",
    ]
    hdr = read_header_map(ws, expected_cols)

    conn.execute("DROP TABLE IF EXISTS computed_measurements")
    conn.execute("""
        CREATE TABLE computed_measurements (
            FGP TEXT NOT NULL,
            AlmaId TEXT,
            Image_Side TEXT,
            Component_Num INTEGER,
            Bifolio_Side TEXT,
            Page_Width_cm REAL,
            Page_Height_cm REAL,
            Num_Lines INTEGER,
            Left_Margin_cm REAL,
            Right_Margin_cm REAL,
            Top_Margin_cm REAL,
            Bottom_Margin_cm REAL,
            Written_Width_cm REAL,
            Written_Height_cm REAL,
            Avg_Line_Height_Text_mm REAL,
            Text_Density_per10cm REAL,
            DpiGrid INTEGER,
            DisplayDPI INTEGER,
            Flag_DPI_High INTEGER DEFAULT 0,
            Flag_DPI_Low INTEGER DEFAULT 0,
            Flag_Negative_Margin INTEGER DEFAULT 0,
            Flag_BifolioLoc_Error INTEGER DEFAULT 0
        )
    """)

    batch = []
    total = 0
    matched_alma = 0

    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), desc="  computed", unit=" rows"):
        fgp = row[hdr.get("FGP", 0)]
        if fgp is None:
            continue
        fgp = str(fgp).strip()
        alma_id = fgp_to_alma.get(fgp)
        if alma_id:
            matched_alma += 1

        def g(col, default_idx=0):
            idx = hdr.get(col, default_idx)
            return row[idx] if idx is not None and idx < len(row) else None

        batch.append((
            fgp, alma_id,
            g("Image_Side", 1), g("Component_Num", 2), g("Bifolio_Side", 3),
            g("Page_Width_cm", 4), g("Page_Height_cm", 5), g("Num_Lines", 6),
            g("Left_Margin_cm", 7), g("Right_Margin_cm", 8),
            g("Top_Margin_cm", 9), g("Bottom_Margin_cm", 10),
            g("Written_Width_cm", 11), g("Written_Height_cm", 12),
            g("Avg_Line_Height_Text_mm", 13), g("Text_Density_per10cm", 14),
            g("DpiGrid", 15), g("DisplayDPI", 16),
            g("Flag_DPI_High", 17) or 0, g("Flag_DPI_Low", 18) or 0,
            g("Flag_Negative_Margin", 19) or 0, g("Flag_BifolioLoc_Error", 20) or 0,
        ))

        if len(batch) >= BATCH_SIZE:
            conn.executemany(
                "INSERT INTO computed_measurements VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                batch,
            )
            total += len(batch)
            batch = []

    if batch:
        conn.executemany(
            "INSERT INTO computed_measurements VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            batch,
        )
        total += len(batch)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_cm_alma ON computed_measurements(AlmaId)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cm_fgp ON computed_measurements(FGP)")

    distinct_alma = conn.execute("SELECT COUNT(DISTINCT AlmaId) FROM computed_measurements WHERE AlmaId IS NOT NULL").fetchone()[0]
    print(f"  Imported {total:,} rows ({distinct_alma:,} distinct AlmaIds, {matched_alma:,} matched via FGP lookup)")
    return total


def step3_blank_images(wb, conn, fgp_to_alma: dict):
    """Step 3: Import Blank_Images sheet."""
    print("\nStep 3: Importing Blank_Images sheet...")
    ws = wb["Blank_Images"]

    expected_cols = [
        "FGP", "Fragment_Width_cm", "Fragment_Height_cm", "IsNotWhole", "PuzzleRatio",
    ]
    hdr = read_header_map(ws, expected_cols)

    conn.execute("DROP TABLE IF EXISTS blank_images")
    conn.execute("""
        CREATE TABLE blank_images (
            FGP TEXT NOT NULL,
            AlmaId TEXT,
            Fragment_Width_cm REAL,
            Fragment_Height_cm REAL,
            IsNotWhole INTEGER,
            PuzzleRatio REAL
        )
    """)

    batch = []
    total = 0

    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), desc="  blank_images", unit=" rows"):
        fgp = row[hdr.get("FGP", 0)]
        if fgp is None:
            continue
        fgp = str(fgp).strip()
        alma_id = fgp_to_alma.get(fgp)

        batch.append((
            fgp, alma_id,
            row[hdr.get("Fragment_Width_cm", 1)] if hdr.get("Fragment_Width_cm") is not None else None,
            row[hdr.get("Fragment_Height_cm", 2)] if hdr.get("Fragment_Height_cm") is not None else None,
            row[hdr.get("IsNotWhole", 3)] if hdr.get("IsNotWhole") is not None else None,
            row[hdr.get("PuzzleRatio", 4)] if hdr.get("PuzzleRatio") is not None else None,
        ))

        if len(batch) >= BATCH_SIZE:
            conn.executemany("INSERT INTO blank_images VALUES (?, ?, ?, ?, ?, ?)", batch)
            total += len(batch)
            batch = []

    if batch:
        conn.executemany("INSERT INTO blank_images VALUES (?, ?, ?, ?, ?, ?)", batch)
        total += len(batch)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_bi_alma ON blank_images(AlmaId)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_bi_fgp ON blank_images(FGP)")

    distinct_alma = conn.execute("SELECT COUNT(DISTINCT AlmaId) FROM blank_images WHERE AlmaId IS NOT NULL").fetchone()[0]
    print(f"  Imported {total:,} rows ({distinct_alma:,} distinct AlmaIds)")
    return total


def step4_catalog_sizes(wb, conn, sm_to_alma: dict):
    """Step 4: Replace catalog_sizes with normalized cm values.

    The xlsx Catalog_Sizes sheet has Shelfmark (not UnitCatalogRecId).
    We resolve Shelfmark→AlmaId using the sm_to_alma lookup built from FIST.db.
    """
    print("\nStep 4: Replacing catalog_sizes with normalized cm values...")

    ws = wb["Catalog_Sizes"]
    expected_cols = [
        "Shelfmark", "SizeX_cm", "SizeY_cm",
        "InnerSizeX_cm", "InnerSizeY_cm", "SizeUnit",
        "Measurement_Scope", "Flag_WH_Swap", "Flag_Unit_Error",
    ]
    hdr = read_header_map(ws, expected_cols)

    conn.execute("DROP TABLE IF EXISTS catalog_sizes")
    conn.execute("""
        CREATE TABLE catalog_sizes (
            AlmaId TEXT NOT NULL,
            Shelfmark TEXT,
            SizeX_cm REAL,
            SizeY_cm REAL,
            InnerSizeX_cm REAL,
            InnerSizeY_cm REAL,
            SizeUnit TEXT,
            Measurement_Scope TEXT,
            Flag_WH_Swap TEXT,
            Flag_Unit_Error TEXT
        )
    """)

    batch = []
    total = 0
    matched = 0
    unmatched = 0
    unmatched_samples = []

    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), desc="  catalog_sizes", unit=" rows"):
        shelfmark = row[hdr.get("Shelfmark", 0)]
        if shelfmark is None:
            continue
        shelfmark = str(shelfmark).strip()

        # Resolve Shelfmark → AlmaId via FIST.db lookup
        alma_id = sm_to_alma.get(shelfmark)
        if alma_id is None:
            unmatched += 1
            if len(unmatched_samples) < 5:
                unmatched_samples.append(shelfmark)
            continue

        def g(col):
            idx = hdr.get(col)
            return row[idx] if idx is not None and idx < len(row) else None

        matched += 1
        batch.append((
            alma_id, shelfmark,
            g("SizeX_cm"), g("SizeY_cm"),
            g("InnerSizeX_cm"), g("InnerSizeY_cm"),
            g("SizeUnit"), g("Measurement_Scope"),
            g("Flag_WH_Swap"), g("Flag_Unit_Error"),
        ))

        if len(batch) >= BATCH_SIZE:
            conn.executemany(
                "INSERT INTO catalog_sizes VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                batch,
            )
            total += len(batch)
            batch = []

    if batch:
        conn.executemany(
            "INSERT INTO catalog_sizes VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            batch,
        )
        total += len(batch)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_catsz_alma ON catalog_sizes(AlmaId)")

    distinct_alma = conn.execute("SELECT COUNT(DISTINCT AlmaId) FROM catalog_sizes").fetchone()[0]

    # Audit report
    print(f"\n  Catalog_Sizes Audit:")
    print(f"    Total xlsx rows with shelfmark: {matched + unmatched:,}")
    print(f"    Matched (Shelfmark found in FIST.db): {matched:,}")
    print(f"    Unmatched (Shelfmark not in FIST.db): {unmatched:,}")
    if unmatched_samples:
        print(f"    Unmatched samples: {unmatched_samples}")
    print(f"    Final inserted rows: {total:,}")
    print(f"    Final distinct AlmaIds: {distinct_alma:,}")

    return total


def step5_manuscript_measurements(conn):
    """Step 5: Build manuscript_measurements summary table.

    Aggregates from catalog_sizes, computed_measurements, extra_info, and blank_images.
    ONLY unflagged records are included in aggregation.
    """
    print("\nStep 5: Building manuscript_measurements summary table...")

    conn.execute("DROP TABLE IF EXISTS manuscript_measurements")
    conn.execute("""
        CREATE TABLE manuscript_measurements (
            AlmaId TEXT NOT NULL PRIMARY KEY,
            -- From catalog_sizes (catalog-reported, most authoritative for physical size)
            -- Only unflagged rows (Flag_WH_Swap IS NULL AND Flag_Unit_Error IS NULL)
            -- MAX across all catalogers for this AlmaId -- may combine width from one
            -- cataloger and height from another. Acceptable as upper bounds for
            -- Phase 55 dimension range filtering.
            catalog_width_cm REAL,
            catalog_height_cm REAL,
            catalog_inner_width_cm REAL,
            catalog_inner_height_cm REAL,
            catalog_count INTEGER,
            -- From computed_measurements (image-derived, ONLY unflagged rows)
            -- All 4 flags must be 0: Flag_DPI_High, Flag_DPI_Low, Flag_Negative_Margin, Flag_BifolioLoc_Error
            min_computed_width_cm REAL,
            max_computed_width_cm REAL,
            min_computed_height_cm REAL,
            max_computed_height_cm REAL,
            avg_num_lines REAL,
            min_num_lines INTEGER,
            max_num_lines INTEGER,
            avg_text_density REAL,
            avg_line_height_mm REAL,
            computed_image_count INTEGER,
            -- From extra_info
            material TEXT,
            size_category TEXT,
            total_image_count INTEGER,
            -- From blank_images
            has_blank_images INTEGER DEFAULT 0,
            blank_image_count INTEGER DEFAULT 0
        )
    """)

    # Use a multi-CTE approach to aggregate from all sources
    conn.execute("""
        INSERT INTO manuscript_measurements
        SELECT
            alma.AlmaId,
            -- catalog
            cat_agg.catalog_width_cm,
            cat_agg.catalog_height_cm,
            cat_agg.catalog_inner_width_cm,
            cat_agg.catalog_inner_height_cm,
            cat_agg.catalog_count,
            -- computed
            comp_agg.min_computed_width_cm,
            comp_agg.max_computed_width_cm,
            comp_agg.min_computed_height_cm,
            comp_agg.max_computed_height_cm,
            comp_agg.avg_num_lines,
            comp_agg.min_num_lines,
            comp_agg.max_num_lines,
            comp_agg.avg_text_density,
            comp_agg.avg_line_height_mm,
            comp_agg.computed_image_count,
            -- extra_info
            ei_agg.material,
            ei_agg.size_category,
            ei_agg.total_image_count,
            -- blank
            COALESCE(bi_agg.has_blank_images, 0),
            COALESCE(bi_agg.blank_image_count, 0)
        FROM (
            -- Union all AlmaIds from all tables
            SELECT DISTINCT AlmaId FROM extra_info WHERE AlmaId IS NOT NULL
            UNION
            SELECT DISTINCT AlmaId FROM computed_measurements WHERE AlmaId IS NOT NULL
            UNION
            SELECT DISTINCT AlmaId FROM catalog_sizes WHERE AlmaId IS NOT NULL
            UNION
            SELECT DISTINCT AlmaId FROM blank_images WHERE AlmaId IS NOT NULL
        ) alma
        LEFT JOIN (
            -- Catalog aggregation: only unflagged rows
            SELECT AlmaId,
                MAX(SizeX_cm) as catalog_width_cm,
                MAX(SizeY_cm) as catalog_height_cm,
                MAX(InnerSizeX_cm) as catalog_inner_width_cm,
                MAX(InnerSizeY_cm) as catalog_inner_height_cm,
                COUNT(*) as catalog_count
            FROM catalog_sizes
            WHERE Flag_WH_Swap IS NULL AND Flag_Unit_Error IS NULL
            GROUP BY AlmaId
        ) cat_agg ON alma.AlmaId = cat_agg.AlmaId
        LEFT JOIN (
            -- Computed aggregation: only unflagged rows (all 4 flags = 0)
            SELECT AlmaId,
                MIN(Page_Width_cm) as min_computed_width_cm,
                MAX(Page_Width_cm) as max_computed_width_cm,
                MIN(Page_Height_cm) as min_computed_height_cm,
                MAX(Page_Height_cm) as max_computed_height_cm,
                AVG(Num_Lines) as avg_num_lines,
                MIN(Num_Lines) as min_num_lines,
                MAX(Num_Lines) as max_num_lines,
                AVG(Text_Density_per10cm) as avg_text_density,
                AVG(Avg_Line_Height_Text_mm) as avg_line_height_mm,
                COUNT(*) as computed_image_count
            FROM computed_measurements
            WHERE Flag_DPI_High = 0 AND Flag_DPI_Low = 0
              AND Flag_Negative_Margin = 0 AND Flag_BifolioLoc_Error = 0
              AND AlmaId IS NOT NULL
            GROUP BY AlmaId
        ) comp_agg ON alma.AlmaId = comp_agg.AlmaId
        LEFT JOIN (
            -- Extra info aggregation: material MODE (most common), size_category MODE
            SELECT AlmaId,
                -- MODE via GROUP BY + ORDER BY COUNT DESC LIMIT 1 subquery
                (SELECT Material FROM extra_info e2
                 WHERE e2.AlmaId = ei.AlmaId AND Material IS NOT NULL
                 GROUP BY Material ORDER BY COUNT(*) DESC LIMIT 1) as material,
                (SELECT Size_Category FROM extra_info e3
                 WHERE e3.AlmaId = ei.AlmaId AND Size_Category IS NOT NULL
                 GROUP BY Size_Category ORDER BY COUNT(*) DESC LIMIT 1) as size_category,
                COUNT(*) as total_image_count
            FROM extra_info ei
            WHERE AlmaId IS NOT NULL
            GROUP BY AlmaId
        ) ei_agg ON alma.AlmaId = ei_agg.AlmaId
        LEFT JOIN (
            -- Blank images
            SELECT AlmaId,
                1 as has_blank_images,
                COUNT(*) as blank_image_count
            FROM blank_images
            WHERE AlmaId IS NOT NULL
            GROUP BY AlmaId
        ) bi_agg ON alma.AlmaId = bi_agg.AlmaId
    """)

    # Create indexes for Phase 55 filtering
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_cat_width ON manuscript_measurements(catalog_width_cm)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_cat_height ON manuscript_measurements(catalog_height_cm)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_comp_max_width ON manuscript_measurements(max_computed_width_cm)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_comp_max_height ON manuscript_measurements(max_computed_height_cm)")
    # Indexes for measurement filtering (review concern #4)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_avg_num_lines ON manuscript_measurements(avg_num_lines)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_line_height ON manuscript_measurements(avg_line_height_mm)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_text_density ON manuscript_measurements(avg_text_density)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_material ON manuscript_measurements(material)")

    row_count = conn.execute("SELECT COUNT(*) FROM manuscript_measurements").fetchone()[0]
    has_catalog = conn.execute("SELECT COUNT(*) FROM manuscript_measurements WHERE catalog_width_cm IS NOT NULL").fetchone()[0]
    has_computed = conn.execute("SELECT COUNT(*) FROM manuscript_measurements WHERE max_computed_width_cm IS NOT NULL").fetchone()[0]
    has_blank = conn.execute("SELECT COUNT(*) FROM manuscript_measurements WHERE has_blank_images = 1").fetchone()[0]

    print(f"  Built manuscript_measurements: {row_count:,} rows")
    print(f"    With catalog dimensions: {has_catalog:,}")
    print(f"    With computed dimensions: {has_computed:,}")
    print(f"    With blank images: {has_blank:,}")

    return row_count


def step6_versioning_and_summary(conn):
    """Step 6: Sidecar versioning and print summary statistics."""
    print("\nStep 6: Sidecar versioning and summary...")

    # Sidecar versioning
    conn.execute("CREATE TABLE IF NOT EXISTS import_meta (key TEXT PRIMARY KEY, value TEXT)")
    conn.execute("INSERT OR REPLACE INTO import_meta VALUES ('measurements_version', '1.0.0')")
    conn.execute("INSERT OR REPLACE INTO import_meta VALUES ('measurements_imported_at', datetime('now'))")

    # Summary statistics
    tables = ["extra_info", "computed_measurements", "blank_images", "catalog_sizes", "manuscript_measurements"]
    print("\n  Table summary:")
    for tbl in tables:
        count = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        try:
            distinct = conn.execute(f"SELECT COUNT(DISTINCT AlmaId) FROM {tbl} WHERE AlmaId IS NOT NULL").fetchone()[0]
        except Exception:
            distinct = "N/A"
        print(f"    {tbl}: {count:,} rows ({distinct} distinct AlmaIds)")


def migrate_add_line_height(conn):
    """Add avg_line_height_mm column if missing (for existing DBs without re-import).

    Idempotent: safe to call multiple times. Populates from computed_measurements
    excluding flagged rows (all 4 flags = 0).
    """
    cols = {row[1] for row in conn.execute("PRAGMA table_info(manuscript_measurements)")}
    if 'avg_line_height_mm' in cols:
        return  # Already present -- idempotent
    conn.execute("ALTER TABLE manuscript_measurements ADD COLUMN avg_line_height_mm REAL")
    conn.execute("""
        UPDATE manuscript_measurements SET avg_line_height_mm = (
            SELECT AVG(cm.Avg_Line_Height_Text_mm)
            FROM computed_measurements cm
            WHERE cm.AlmaId = manuscript_measurements.AlmaId
              AND cm.Flag_DPI_High = 0 AND cm.Flag_DPI_Low = 0
              AND cm.Flag_Negative_Margin = 0 AND cm.Flag_BifolioLoc_Error = 0
              AND cm.Avg_Line_Height_Text_mm IS NOT NULL
              AND cm.Avg_Line_Height_Text_mm > 0
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_line_height ON manuscript_measurements(avg_line_height_mm)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_text_density ON manuscript_measurements(avg_text_density)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ms_material ON manuscript_measurements(material)")
    conn.commit()


def main():
    """Main import function."""
    import argparse
    parser = argparse.ArgumentParser(description="Import FIST measurement data into fjms_enrichment.db")
    parser.add_argument("--target", default=TARGET_DB_PATH,
                        help=f"Target SQLite database (default: {TARGET_DB_PATH})")
    parser.add_argument("--dry-run", action="store_true",
                        help="Run import on a disposable copy, print audit, then discard")
    args = parser.parse_args()

    target_path = args.target

    if args.dry_run:
        import shutil
        dry_copy = target_path + ".dry_run_copy"
        shutil.copy2(target_path, dry_copy)
        target_path = dry_copy
        print("*** DRY RUN MODE — writing to disposable copy ***\n")

    print("=" * 60)
    print("FIST Measurement Data Import")
    print(f"  Source: {XLSX_PATH}")
    print(f"  FIST.db: {FIST_DB_PATH}")
    print(f"  Target: {target_path}")
    if args.dry_run:
        print(f"  Mode: DRY RUN")
    print("=" * 60)

    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: xlsx file not found: {XLSX_PATH}")
        sys.exit(1)

    if not os.path.exists(FIST_DB_PATH):
        print(f"ERROR: FIST.db not found: {FIST_DB_PATH}")
        sys.exit(1)

    if not os.path.exists(target_path):
        print(f"ERROR: Target sidecar not found: {target_path}")
        sys.exit(1)

    start = time.time()

    # Open xlsx in read_only mode for streaming
    print("\nOpening xlsx (read_only=True)...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Build AlmaId lookups from FIST.db (exact integers, no xlsx float precision loss)
    print("\nBuilding FGP→AlmaId lookup from FIST.db...")
    fgp_to_alma = build_fgp_to_alma_from_fist(FIST_DB_PATH)
    print(f"  {len(fgp_to_alma):,} FGP→AlmaId mappings")

    print("Building Shelfmark→AlmaId lookup from FIST.db...")
    sm_to_alma = build_shelfmark_to_alma_from_fist(FIST_DB_PATH)
    print(f"  {len(sm_to_alma):,} Shelfmark→AlmaId mappings")

    # Validate AlmaIds against libraries.csv
    if not validate_alma_ids(fgp_to_alma):
        print("\nFATAL: AlmaId validation failed. Aborting.")
        wb.close()
        sys.exit(1)

    # Connect to target sidecar
    conn = sqlite3.connect(target_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")

    # Wrap entire import in a single transaction for safety
    # If any step fails, the transaction rolls back and old data is preserved
    conn.execute("BEGIN")

    try:
        # Step 1: Extra_Info (uses fgp_to_alma from FIST.db, ignores xlsx AlmaId column)
        step1_extra_info(wb, conn, fgp_to_alma)

        # Step 2: Computed_Measurements
        step2_computed_measurements(wb, conn, fgp_to_alma)

        # Step 3: Blank_Images
        step3_blank_images(wb, conn, fgp_to_alma)

        # Step 4: Replace catalog_sizes (uses sm_to_alma from FIST.db)
        step4_catalog_sizes(wb, conn, sm_to_alma)

        # Step 5: Build manuscript_measurements summary
        step5_manuscript_measurements(conn)

        # Step 5b: Migration for existing DBs (idempotent -- no-op after fresh step5)
        migrate_add_line_height(conn)

        # Step 6: Versioning and summary
        step6_versioning_and_summary(conn)

        # Commit the entire transaction
        conn.execute("COMMIT")
        print("\n  Transaction committed successfully.")

    except Exception as e:
        conn.execute("ROLLBACK")
        print(f"\n  ERROR: Import failed, transaction rolled back: {e}")
        raise
    finally:
        wb.close()
        conn.close()

    elapsed = time.time() - start
    print(f"\nDone in {elapsed:.1f}s")

    if args.dry_run:
        os.remove(target_path)
        print(f"\n*** DRY RUN: disposable copy deleted ***")


if __name__ == "__main__":
    main()

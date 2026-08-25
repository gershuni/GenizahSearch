# -*- coding: utf-8 -*-
"""Score a parallels run against the adjudicated Megillat Antiochus deck.

The deck (eval/antiochus/deck.json) is the measuring instrument behind every
recall figure in docs/specs/passage-matching-algorithm.md sections 8.1 and
10.4. This script is how a new run is put on the same scale, so a claim like
"this policy improves recall" is checkable rather than asserted.

Usage:
  # reproduce the recorded comparison table from the archived runs
  python scripts/score_antiochus_deck.py --all

  # score a new run (GUI xlsx export, delta CSV, or JSON rows)
  python scripts/score_antiochus_deck.py --run my_run.xlsx
  python scripts/score_antiochus_deck.py --run delta.csv --side candidate
  python scripts/score_antiochus_deck.py --run rows.json --show-missed

A delta CSV from compare_passage_policies.py holds TWO runs, tagged by its
`presence` column. `--side` says which one to score; without it the file is
refused rather than silently scored as their union, which would credit the
candidate with manuscripts only the baseline found.

Three numbers matter, and the third is the one that gets forgotten:

  precision  graded positives / manuscripts returned
  recall     graded positives found / 83
  UNGRADED   returned manuscripts absent from the deck

UNGRADED is NOT noise. It is either a genuine new find (the deck is a union
of five runs, so a better method can exceed it) or a BROKEN JOIN. A join
break looks like brilliant novelty and catastrophic recall at the same time;
that combination cost two wrong measurements on 2026-08-24. Any ungraded
count above a handful should be read as a join bug until proven otherwise --
`--show-ungraded` lists them so it can be settled by eye in a minute.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from shelfmark_join import canonical_key, load_aliases  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EVAL_DIR = os.path.join(ROOT, 'eval', 'antiochus')
DECK_PATH = os.path.join(EVAL_DIR, 'deck.json')
RUNS_DIR = os.path.join(EVAL_DIR, 'runs')
POSITIVE = ('WITNESS', 'INDIRECT')

# The recall frontier: positives that ONLY word-chunk matching at chunk size
# 2 has ever returned. Letter-level search has never reached more than one of
# these, and section 10.4 explains why -- a Judeo-Arabic translation shares
# almost nothing with an Aramaic query beyond transliterated names. Progress
# on this list is the point of the exercise; progress on the other 63
# positives is mostly re-finding what widest-40 already finds.
FRONTIER = (
    'ENA 1629.10', 'L-G Ar.II.151', 'L-G Ar.II.152', 'MS heb. d.37/71',
    'MS heb. d.60/25', 'MS heb. e.30/56', 'MS heb. e.45/34',
    'MS heb. f.18/35', 'MS heb. f.40/47', 'Ms. 10808.8', 'Ms. C 24',
    'Ms. EVR ARAB I 4838', 'Ms. EVR II A 1225', 'Ms. EVR II A 922',
    'Ms. G.F. vol. 2', 'Ms. VII C 12', 'T-S AS 171.65', 'T-S AS 67.25',
    'T-S AS 72.94', 'T-S Ar.24.174',
)


def load_deck(path: str = DECK_PATH):
    """Return (by_key, aliases), keyed canonically so callers cannot forget.

    Refuses to load a deck whose canonical keys collide on DIFFERENT verdicts.
    Writing both into a dict would let iteration order decide whether a
    manuscript counts as a positive -- an instrument that silently
    contradicts itself, which is worse than one that will not start. It has
    caught a bad alias once already (see `_rejected` in aliases.json).
    """
    with open(path, encoding='utf-8') as fh:
        raw = json.load(fh)
    aliases = load_aliases()
    by_key, conflicts = {}, []
    for shelfmark, rec in raw['items'].items():
        key = canonical_key(shelfmark, aliases)
        prior = by_key.get(key)
        if prior is not None and prior['verdict'] != rec['verdict']:
            conflicts.append(
                f"  {key}: {prior['shelfmark']}={prior['verdict']} vs "
                f"{shelfmark}={rec['verdict']}")
        by_key[key] = dict(rec, shelfmark=shelfmark)
    if conflicts:
        sys.exit('deck is self-contradictory after alias resolution -- fix '
                 'the alias or the grading before measuring anything:\n'
                 + '\n'.join(conflicts))
    return by_key, aliases


def _rows_from_json(path):
    with open(path, encoding='utf-8') as fh:
        data = json.load(fh)
    if isinstance(data, dict):
        data = data.get('results') or data.get('rows') or []
    return [r for r in data if isinstance(r, dict)]


SIDE_KEEP = {
    # `presence` values each side is entitled to claim as its own result.
    'candidate': {'both', 'candidate_only'},
    'baseline': {'both', 'baseline_only'},
    'union': {'both', 'candidate_only', 'baseline_only'},
}


def _rows_from_csv(path):
    with open(path, encoding='utf-8-sig', newline='') as fh:
        return list(csv.DictReader(fh))


def _select_side(rows, side):
    """Narrow a two-run delta CSV to the side actually being scored.

    A file from `compare_passage_policies.py --csv` holds BOTH policies'
    results in one table, distinguished only by `presence`. Scoring it whole
    counts every manuscript the baseline found as though the candidate had
    found it -- so a candidate that lost ground still reports the baseline's
    recall, and the loss reads as an improvement.

    Refuses rather than defaulting: picking `candidate` silently would change
    what an already-saved `--run delta.csv` command means, and a quiet wrong
    answer from a measuring instrument is worse than a loud one.
    """
    if not rows or 'presence' not in rows[0]:
        return rows        # an ordinary single-run export; nothing to choose
    if side is None:
        counts = {}
        for r in rows:
            key = str(r.get('presence') or '?')
            counts[key] = counts.get(key, 0) + 1
        sys.exit(
            'this file has a `presence` column, so it holds TWO runs '
            f'({", ".join(f"{k}={v}" for k, v in sorted(counts.items()))}).\n'
            'Scoring it whole would credit the candidate with manuscripts '
            'only the baseline found.\n'
            'Pass --side candidate | baseline | union.')
    keep = SIDE_KEEP[side]
    picked = [r for r in rows if str(r.get('presence') or '') in keep]
    if (side == 'baseline' and picked
            and 'baseline_score' not in picked[0]
            and any(str(r.get('presence') or '') == 'both' for r in picked)):
        # This file predates the two-score columns: its single `score` holds
        # the CANDIDATE's hit for every shared manuscript, so the baseline's
        # score for those rows is not in the file at all. Returning `score`
        # anyway would report the candidate's numbers under the baseline's
        # name -- the substitution this whole guard exists to stop.
        #
        # Only when a `both` row is present. A file of purely `baseline_only`
        # rows does carry the baseline's own scores (the old writer had no
        # candidate hit to prefer), and refusing that would be over-strict.
        sys.exit(
            'this delta CSV has no `baseline_score` column, so it predates '
            'the two-score writer and does not contain the baseline\'s score '
            'for its shared manuscripts.\n'
            'Re-run compare_passage_policies.py --csv to get one that does, '
            'or score it with --side candidate | union.\n'
            'NOTE: the figures this tool REPORTS (manuscripts, precision, '
            'recall, frontier) never read the score column, so an old file is '
            'still correct for those on any side.')
    if side == 'baseline' and picked and 'baseline_score' in picked[0]:
        # `score` and `record_id` are the CANDIDATE's by contract (see the
        # writer in compare_passage_policies.py). Reading them for the
        # baseline would report the candidate's numbers under the baseline's
        # name -- the same substitution this whole guard exists to stop, one
        # column over. Repoint rather than teach `load_run` a second column
        # vocabulary.
        picked = [dict(r, score=r.get('baseline_score'),
                       record_id=r.get('baseline_record_id')) for r in picked]
    return picked


def _rows_from_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit('reading .xlsx needs openpyxl (pip install openpyxl); '
                 'or export the run as CSV and pass that instead')
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c or '').strip() for c in next(rows)]
    return [dict(zip(header, r)) for r in rows]


def load_run(path, column=None, side=None):
    ext = os.path.splitext(path)[1].lower()
    rows = ({'.json': _rows_from_json, '.csv': _rows_from_csv,
             '.xlsx': _rows_from_xlsx}.get(ext) or _rows_from_json)(path)
    if not rows:
        sys.exit(f'no rows found in {path}')
    rows = _select_side(rows, side)
    if not rows:
        sys.exit(f'no rows left in {path} after --side {side}')
    if column is None:
        for cand in ('shelfmark', 'Shelfmark', 'call_number', 'Call Number',
                     'shelfmarks', 'סימן', 'Shelf Mark'):
            if cand in rows[0]:
                column = cand
                break
    if column is None or column not in rows[0]:
        sys.exit(f'no shelfmark column in {path}; columns are '
                 f'{list(rows[0])[:12]} -- pass --column')
    score_col = next((c for c in ('score', 'Score', 'best_score')
                      if c in rows[0]), None)
    out = {}
    for r in rows:
        sm = str(r.get(column) or '').strip()
        if not sm:
            continue
        try:
            sc = float(r.get(score_col)) if score_col else None
        except (TypeError, ValueError):
            sc = None
        prev = out.get(sm)
        if prev is None or (sc is not None and (prev is None or sc > prev)):
            out[sm] = sc
    return out


def score(run_shelfmarks, deck, aliases):
    """Collapse to manuscripts and grade. Returns a dict of findings."""
    seen, buckets = {}, {'WITNESS': [], 'INDIRECT': [], 'NOISE': [],
                         'UNGRADED': []}
    for sm, sc in run_shelfmarks.items():
        key = canonical_key(sm, aliases)
        if key in seen:
            continue
        seen[key] = sm
        rec = deck.get(key)
        buckets['UNGRADED' if rec is None else rec['verdict']].append((sm, sc))
    total_pos = sum(1 for r in deck.values() if r['verdict'] in POSITIVE)
    found_pos = len(buckets['WITNESS']) + len(buckets['INDIRECT'])
    frontier_keys = {canonical_key(s, aliases) for s in FRONTIER}
    return {
        'manuscripts': len(seen),
        'buckets': buckets,
        'positives_found': found_pos,
        'positives_total': total_pos,
        'precision': found_pos / len(seen) if seen else 0.0,
        'recall': found_pos / total_pos if total_pos else 0.0,
        'frontier_found': sorted(
            deck[k]['shelfmark'] for k in frontier_keys & set(seen)),
        'frontier_total': len(frontier_keys),
        'missed': sorted(rec['shelfmark'] for k, rec in deck.items()
                         if rec['verdict'] in POSITIVE and k not in seen),
    }


def report(name, res, show_missed=False, show_ungraded=False):
    b = res['buckets']
    print(f"\n=== {name} ===")
    print(f"  manuscripts   {res['manuscripts']}")
    print(f"  WITNESS       {len(b['WITNESS'])}")
    print(f"  INDIRECT      {len(b['INDIRECT'])}")
    print(f"  NOISE         {len(b['NOISE'])}")
    print(f"  UNGRADED      {len(b['UNGRADED'])}"
          + ('   <-- new finds OR a broken join; check them' if b['UNGRADED']
             else ''))
    print(f"  precision     {res['precision']:.0%}")
    print(f"  recall        {res['recall']:.0%}  "
          f"({res['positives_found']}/{res['positives_total']})")
    print(f"  frontier      {len(res['frontier_found'])}/{res['frontier_total']}"
          f"  {', '.join(res['frontier_found']) or '-'}")
    if show_ungraded and b['UNGRADED']:
        print('  ungraded rows:')
        for sm, sc in sorted(b['UNGRADED'], key=lambda x: -(x[1] or 0)):
            print(f"     {sm}  score={sc if sc is not None else '-'}")
    if show_missed and res['missed']:
        print(f"  missed {len(res['missed'])} positives:")
        for sm in res['missed']:
            print(f"     {sm}")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--run', help='xlsx / csv / json run to score')
    ap.add_argument('--all', action='store_true',
                    help='score every archived run and print the table')
    ap.add_argument('--column', help='shelfmark column name (autodetected)')
    ap.add_argument('--side', choices=sorted(SIDE_KEEP),
                    help='for a delta CSV from compare_passage_policies.py, '
                         'which run to score. Required for such a file: '
                         'scoring it whole credits the candidate with '
                         'manuscripts only the baseline found')
    ap.add_argument('--deck', default=DECK_PATH)
    ap.add_argument('--show-missed', action='store_true')
    ap.add_argument('--show-ungraded', action='store_true')
    args = ap.parse_args()
    if not args.run and not args.all:
        ap.error('pass --run FILE or --all')

    deck, aliases = load_deck(args.deck)
    total_pos = sum(1 for r in deck.values() if r['verdict'] in POSITIVE)
    print(f"deck: {len(deck)} manuscripts, {total_pos} positives, "
          f"{len(aliases)} aliases")

    if args.all:
        rows = []
        for fn in sorted(os.listdir(RUNS_DIR)):
            if not fn.endswith('.json'):
                continue
            path = os.path.join(RUNS_DIR, fn)
            res = score(load_run(path), deck, aliases)
            rows.append((fn[:-5], res))
        print(f"\n{'run':<22}{'mss':>6}{'prec':>7}{'recall':>8}{'frontier':>10}")
        for name, res in sorted(rows, key=lambda r: r[1]['manuscripts']):
            print(f"{name:<22}{res['manuscripts']:>6}{res['precision']:>7.0%}"
                  f"{res['recall']:>8.0%}"
                  f"{str(len(res['frontier_found'])) + '/' + str(res['frontier_total']):>10}")
        if args.show_missed or args.show_ungraded:
            for name, res in rows:
                report(name, res, args.show_missed, args.show_ungraded)

    if args.run:
        res = score(load_run(args.run, args.column, args.side), deck, aliases)
        report(os.path.basename(args.run), res,
               args.show_missed or True, args.show_ungraded or True)


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""The `fills_gap` probe -- owner ruling K
(`.planning/phases/136-read-surfaces-connections-panel-work-witnesses/136-GATE1-DECISIONS.md`
section K).

The ruling-I re-measurement (`136-NOVELTY-RUN.md` sections 2-3) scored 47/60 = 78.3% shade
agreement with ZERO errors in the false-novel direction -- but its 60-case pool contained ZERO
true `fills_gap` cases, so the axis decision B actually authorizes spend on (does the pinned
gate correctly say "genuinely unknown" ONLY when it is true) was never exercised. This script
builds and runs the purpose-built probe ruling K requires: it measures the false-novel rate on
the population that would ACTUALLY ship as "Candidates for new finds", covering BOTH paths a
row can take to become a candidate --

1. THE MODEL PATH -- real residual rows (the heuristic funnel could not resolve them
   mechanically) that the REAL pinned model (`shared/discovery_novelty.py`'s
   `LLM_MODEL`/`LLM_REASONING_EFFORT`/`NOVELTY_PROMPT_TEMPLATE`, called for real via
   OpenRouter) classifies `fills_gap`.
2. THE BYPASS PATH -- rows where NO checked source has any text at all. These never reach the
   model at all; the funnel ships them as `fills_gap` automatically, with nothing checked
   against them. Ruling K names this path "arguably the HIGHER risk precisely because nothing
   examines it."

Reuses the REAL, committed funnel (`scripts/discovery_novelty_funnel.py` --
`NoveltyCandidate`/`run_heuristic_pass`/`run_heuristic_funnel`/`run_model_arm`/
`assemble_evidence_bundle`) and the REAL data loaders already committed for the ruling-J
hard-case sampler (`scripts/discovery_gate1_evidence.py` -- `load_works`/`load_claims`/
`load_libraries_csv`/`load_fjms_catalog_text`/`load_bib_rows`/`load_fgp_rows`/
`_combined_catalogue_text`), never a second, hand-copied loader. A real, independent
verification of this reconstruction against the live discovery-v1 asset reproduced the EXACT
same full-corpus split the ruling-I re-measurement session reported (`136-NOVELTY-RUN.md`
section 3.2): 65,200 real shipped `(sys_id, work_id)` pairs -> 1,689 resolved `confirms`
mechanically, 8,327 no-source-text bypass, 55,184 residual.

Produces an owner-labelling instrument (Markdown + RTL XLSX, same house style and
reproducibility discipline as `136-NOVELTY-HARDCASES.xlsx` --
`scripts/discovery_gate1_evidence.py::write_hardcases_xlsx`), containing ONLY rows that would
actually ship as candidates -- from BOTH paths, labelled by path, capped at ~40 total, with a
single plain owner question (`genuinely_novel` / `actually_recorded` / `unsure` / `skip`) and
NO pre-filled verdict.

Real cost is read from each OpenRouter response's own `usage.cost` field -- NEVER estimated --
and logged to a gitignored `discovery_data/` cost log; this script's own summary sums that log,
never an in-process accumulator alone (so a resumed, previously-interrupted run reports the
TRUE total spend across every process invocation, not just the last one).

No dependency on any gitignored research tree. `discovery_data/` outputs (the sample, the
model-call checkpoint, the cost log) are gitignored, exactly like every sibling
`discovery_data/` artifact this project's novelty work already produces -- only the instrument
(written to the phase directory) and this script are committed.

Usage:
    python scripts/discovery_novelty_probe.py --run-model \\
        --model-sample-size 300 --bypass-sample-size 20 --seed 20260803

    # Re-run identically (idempotent, resumable): candidates whose verdict is already in the
    # checkpoint are never re-billed.
    python scripts/discovery_novelty_probe.py --run-model \\
        --model-sample-size 300 --bypass-sample-size 20 --seed 20260803

    # Build the instrument from an already-completed model run without spending anything:
    python scripts/discovery_novelty_probe.py --build-instrument-only \\
        --model-sample-size 300 --bypass-sample-size 20 --seed 20260803
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from collections import defaultdict
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from scripts.discovery_gate1_evidence import (  # noqa: E402
    DEFAULT_FGP_DB,
    DEFAULT_FJMS_DB,
    DEFAULT_LIBRARIES_CSV,
    DEFAULT_PGP_DB,
    _combined_catalogue_text,
    connect_readonly,
    load_bib_rows,
    load_claims,
    load_fgp_rows,
    load_fjms_catalog_text,
    load_libraries_csv,
    load_works,
)
from scripts.discovery_novelty_funnel import (  # noqa: E402
    NO_SOURCE_TEXT_REASON,
    NoveltyCandidate,
    assemble_evidence_bundle,
    run_heuristic_funnel,
    run_model_arm,
)
from shared.discovery_novelty import (  # noqa: E402
    CANDIDATE_STATUS,
    INPUT_NORMALIZATION_SHA256,
    LLM_MODEL,
    LLM_MODEL_VERSION,
    LLM_REASONING_EFFORT,
    NOVELTY_PROMPT_TEMPLATE,
    PROMPT_SHA256,
)

PHASE_DIR = os.path.join(
    REPO_ROOT, ".planning", "phases", "136-read-surfaces-connections-panel-work-witnesses",
)
DEFAULT_ASSET = os.path.join(
    REPO_ROOT, "discovery_data",
    "discovery-v1-33499c5b89f9e635565cd1cc8831c012f5373811c2870ddbda7d303e60d4c5ff.db",
)
DEFAULT_CHECKPOINT = os.path.join(REPO_ROOT, "discovery_data", "novelty_probe_model_checkpoint.jsonl")
DEFAULT_COST_LOG = os.path.join(REPO_ROOT, "discovery_data", "novelty_probe_cost_log.jsonl")
DEFAULT_MD_OUT = os.path.join(PHASE_DIR, "136-NOVELTY-FILLSGAP-PROBE.md")
DEFAULT_XLSX_OUT = os.path.join(PHASE_DIR, "136-NOVELTY-FILLSGAP-PROBE.xlsx")

# The single, plain owner question (owner ruling K's own <probe_design> instruction) -- never
# a pre-filled verdict, and never the ten-value shade vocabulary (this probe asks one question:
# is this fragment genuinely not identified in the finding aids we checked?).
VERDICT_TOKENS: Tuple[str, ...] = ("genuinely_novel", "actually_recorded", "unsure", "skip")
VERDICT_QUESTION = (
    "Is this fragment GENUINELY NOT IDENTIFIED in the finding aids we checked -- catalogue, "
    "bibliography, PGP, FGP, and (where present) an internal reference-corpus shelfmark "
    "attribution? Answer `genuinely_novel` if yes (a real previously-unknown case), "
    "`actually_recorded` if you find it identified somewhere among those sources after all, "
    "`unsure` if you cannot tell, or `skip` to decline. A blank cell means \"not yet answered\" "
    "-- it is NOT a label."
)

MODEL_PATH = "model"
BYPASS_PATH = "bypass"


# ---------------------------------------------------------------------------
# Real data loading -- reuses the SAME committed loaders the ruling-J hard-case
# sampler already uses (scripts/discovery_gate1_evidence.py), never a second,
# hand-copied loader.
# ---------------------------------------------------------------------------

def load_pgp_texts(db_path: str) -> Dict[str, Dict[str, str]]:
    """sys_id -> {"description": ..., "transcription": ...} -- REAL free text
    from pgp_data/pgp.db's `documents` table, joined through
    `document_fragments` on sys_id (concatenated across every PGP document
    linked to this fragment), per the same field selection
    `136-NOVELTY-RUN.md` section 2.3 documents for the ruling-I
    re-measurement. Distinct from `discovery_gate1_evidence.py`'s own
    `load_pgp_signal_index` (which returns only presence/named booleans for
    the SAMPLING-ONLY approximation) -- this probe needs the REAL text to
    feed the REAL committed funnel (`run_heuristic_pass`/
    `assemble_evidence_bundle`), not a boolean signal."""
    import sqlite3

    if not os.path.exists(db_path):
        return {}
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        docinfo: Dict[Any, Tuple[str, str]] = {}
        for pid, desc, tr in con.execute("SELECT pgpid, description, transcription FROM documents"):
            docinfo[pid] = (desc or "", tr or "")
        out: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: {"description": [], "transcription": []})
        for sid, did in con.execute("SELECT sys_id, document_id FROM document_fragments"):
            if not sid:
                continue
            info = docinfo.get(did)
            if info is None:
                continue
            desc, tr = info
            if desc:
                out[str(sid)]["description"].append(desc)
            if tr:
                out[str(sid)]["transcription"].append(tr)
    finally:
        con.close()
    return {
        sid: {"description": " ".join(v["description"]), "transcription": " ".join(v["transcription"])}
        for sid, v in out.items()
    }


def _best_claim_all_types(claims: List[Dict[str, Any]]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """One representative shipped claim per (sys_id, work_id) pair, across
    EVERY claim_type (unlike `discovery_gate1_evidence.py`'s own
    `_best_claims_by_sys_work`, which filters to `direct_witness` only) --
    this is the population definition that reproduces the real, measured
    65,200-pair full-corpus split `136-NOVELTY-RUN.md` section 3.2 reports.
    Same deterministic total order (highest matched_letters, then
    lexicographically smallest page_id) as every other selector in this
    project's novelty tooling."""
    best: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for c in claims:
        if c.get("routing_status") != "shipped":
            continue
        key = (c["sys_id"], c["work_id"])
        cand_key = (-(c["matched_letters"] or 0), c["page_id"])
        prev = best.get(key)
        if prev is None or cand_key < prev["_key"]:
            d = dict(c)
            d["_key"] = cand_key
            best[key] = d
    return best


def load_work_witnesses(
    path: Optional[str], crosswalk_path: Optional[str] = None
) -> Dict[str, Dict]:
    """Load `emit_work_witnesses.py`'s output, keyed by MINTED work id.

    Returns `{minted_work_id: {"witnesses": {sys_id: {...}}, "attestation": str}}`.

    Supersedes `load_work_attributions`: that read ONE of the corpus's three
    witness channels, as free text, and understated "already known" by ~3x
    (1,375 flips versus 4,518 measured on the real artifact). This reads the
    pre-resolved three-channel join, so no shelfmark matching happens here --
    that belongs to the pipeline that owns it, not to the novelty gate.

    Same raw->minted translation, and the same refusal without a crosswalk: the
    file is keyed on raw ids, and a mismatched id space fails SILENTLY as an
    empty source rather than as an error.
    """
    if not path:
        return {}
    with open(path, encoding="utf-8") as fh:
        payload = json.load(fh)
    wit = payload.get("witnesses") or {}
    att = payload.get("attestation") or {}
    if not crosswalk_path or not os.path.isfile(crosswalk_path):
        raise SystemExit(
            "work witnesses supplied without a crosswalk -- the file is keyed on "
            "raw work ids and the candidates are keyed on minted ids, so without "
            "the mapping every lookup would silently miss"
        )
    with open(crosswalk_path, encoding="utf-8") as fh:
        crosswalk = json.load(fh)
    out: Dict[str, Dict] = {}
    for raw_id, per_sys in wit.items():
        minted = crosswalk.get(raw_id)
        if minted:
            out[minted] = {"witnesses": per_sys, "attestation": att.get(raw_id)}
    return out


def load_work_attributions(
    path: Optional[str], crosswalk_path: Optional[str] = None
) -> Dict[str, str]:
    """Load `emit_work_attributions.py`'s output, keyed by MINTED work id.

    The file is keyed on RAW work ids (what the reference corpus carries); the
    novelty candidate is keyed on the minted `w######`. This translates through
    the crosswalk, because keying the lookup on the wrong id space is the exact
    failure this bake has hit repeatedly -- and it fails SILENTLY, as an empty
    source rather than an error.

    Returns `{}` when no path is supplied, which is the pre-2026-08-08 behaviour
    and correct for a fixture run.
    """
    if not path:
        return {}
    with open(path, encoding="utf-8") as fh:
        payload = json.load(fh)
    raw = payload.get("attributions") or {}
    if not crosswalk_path or not os.path.isfile(crosswalk_path):
        # No crosswalk => cannot translate. Returning the raw-keyed map would
        # silently match nothing; say so instead.
        raise SystemExit(
            "work attributions supplied without a crosswalk -- the file is keyed "
            "on raw work ids and the candidates are keyed on minted ids, so "
            "without the mapping every lookup would silently miss"
        )
    with open(crosswalk_path, encoding="utf-8") as fh:
        crosswalk = json.load(fh)
    out: Dict[str, str] = {}
    for raw_id, text in raw.items():
        minted = crosswalk.get(raw_id)
        if minted:
            out[minted] = text
    return out


def build_all_candidates(
    asset_path: str,
    libraries_csv: str,
    fjms_db: str,
    pgp_db: str,
    fgp_db: str,
    work_attributions: Optional[Dict[str, str]] = None,
    work_witnesses: Optional[Dict[str, Dict]] = None,
) -> Tuple[List[NoveltyCandidate], Dict[str, Dict[str, Any]], Dict[str, Dict[str, str]]]:
    """Builds a REAL `NoveltyCandidate` for every shipped (sys_id, work_id)
    pair in the live asset, from the real sidecars. Returns
    (candidates, works_by_id, libraries_by_sys_id) -- the latter two so the
    instrument writer can render shelfmarks/titles without re-querying.

    `work_attributions` maps MINTED work id -> the reference corpus's own witness
    attribution for that work (the neutral `src_attr_note`; see
    `scripts/emit_work_attributions.py`). Supply it via
    `load_work_attributions`. Omitted => the source is empty, which is what this
    function did unconditionally until 2026-08-08: the field existed, was
    fingerprinted, and was hardcoded to `None`, so the gate reported checking a
    source it never read.
    """
    conn = connect_readonly(asset_path)
    works = load_works(conn)
    claims = load_claims(conn)

    best = _best_claim_all_types(claims)
    libraries = load_libraries_csv(libraries_csv)
    fjms_catalog = load_fjms_catalog_text(fjms_db)
    bib_rows_idx = load_bib_rows(fjms_db)
    fgp_rows_idx = load_fgp_rows(fgp_db)
    pgp_texts_idx = load_pgp_texts(pgp_db)

    candidates: List[NoveltyCandidate] = []
    for (sid, wid), c in sorted(best.items()):
        w = works.get(wid)
        if w is None:
            continue
        cat_text = _combined_catalogue_text(sid, libraries, fjms_catalog)
        bib_rows = tuple(
            {"text": " ".join(str(v) for v in row[:8] if v), "transcription_type": row[9]}
            for row in bib_rows_idx.get(sid, [])
        )
        bib_rows = tuple(r for r in bib_rows if r["text"])
        pgp = pgp_texts_idx.get(sid, {})
        fgp_rows = fgp_rows_idx.get(sid, [])
        fgp_texts = tuple(x for tup in fgp_rows for x in tup if x)
        # Recorded-witness lookup is per (manuscript, work), NOT per work: the
        # question is whether the corpus attests THIS manuscript, and a per-work
        # answer would mark every claim on a well-attested work as known.
        wit_entry = (work_witnesses or {}).get(wid)
        witness_conf = None
        attestation = (work_attributions or {}).get(wid) or None
        if wit_entry:
            hit = (wit_entry.get("witnesses") or {}).get(str(sid))
            if hit:
                witness_conf = hit.get("confidence")
            # The attestation summary rides along whenever the work has witness
            # data, so the model sees the context even when this manuscript is
            # not among them -- but it is never itself the decision.
            attestation = wit_entry.get("attestation") or attestation
        candidates.append(
            NoveltyCandidate(
                sys_id=sid,
                ref_work_id=wid,
                claimed_title=w.get("neutral_title") or "",
                claimed_author=w.get("author"),
                catalogue_text=cat_text or None,
                bibliography_rows=bib_rows,
                pgp_description=pgp.get("description") or None,
                pgp_transcription=pgp.get("transcription") or None,
                fgp_texts=fgp_texts,
                m_source_shelfmark_text=attestation,
                known_witness_confidence=witness_conf,
                page_mapped=True,
            )
        )
    return candidates, works, libraries


# ---------------------------------------------------------------------------
# Sampling -- deterministic given (candidates, seed), so a re-run against the
# same asset with the same seed draws the identical sample; no persistence of
# the sample itself is needed.
# ---------------------------------------------------------------------------

def _candidate_sort_key(c: NoveltyCandidate) -> Tuple[str, str]:
    return (c.sys_id, c.ref_work_id)


def sample_candidates(candidates: Sequence[NoveltyCandidate], n: int, seed: int) -> List[NoveltyCandidate]:
    import random

    ordered = sorted(candidates, key=_candidate_sort_key)
    rng = random.Random(seed)
    if n >= len(ordered):
        return ordered
    return rng.sample(ordered, n)


# ---------------------------------------------------------------------------
# The real pinned model call (OpenRouter), reading real usage.cost on every
# call, logged to a gitignored cost log -- never estimated. Retries bounded
# transient failures (network errors, non-200s) WITHOUT checkpointing a
# not_checked verdict for them, so a transient failure is retried on the next
# invocation rather than being permanently, wrongly recorded as a model
# abstention.
# ---------------------------------------------------------------------------

def _render_user_message(candidate: NoveltyCandidate) -> str:
    bundle = assemble_evidence_bundle(candidate)
    lines = [
        f"Claimed work: {candidate.claimed_title} (author: {candidate.claimed_author or 'unknown'})",
        "",
    ]
    for source in ("catalogue", "bibliography", "pgp", "fgp", "m_source_shelfmark"):
        texts = bundle.get(source, ())
        if texts:
            joined = " ||| ".join(t for t in texts if t)
            lines.append(f"{source}: {joined}")
        else:
            lines.append(f"{source}: (none)")
    return "\n".join(lines)


def make_openrouter_model_call(
    api_key: str,
    cost_log_path: str,
    timeout: float = 90.0,
    max_attempts: int = 4,
):
    import requests

    session = requests.Session()

    def _model_call(candidate: NoveltyCandidate) -> Optional[Mapping[str, Any]]:
        payload = {
            "model": f"google/{LLM_MODEL}",
            "messages": [
                {"role": "system", "content": NOVELTY_PROMPT_TEMPLATE},
                {"role": "user", "content": _render_user_message(candidate)},
            ],
            "reasoning": {"effort": LLM_REASONING_EFFORT},
            "usage": {"include": True},
            "response_format": {"type": "json_object"},
        }
        last_exc: Optional[BaseException] = None
        for attempt in range(1, max_attempts + 1):
            try:
                resp = session.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    data=json.dumps(payload),
                    timeout=timeout,
                )
                resp.raise_for_status()
                data = resp.json()
                model_echo = data.get("model") or ""
                if LLM_MODEL not in model_echo:
                    raise RuntimeError(
                        f"provider echoed unexpected model {model_echo!r}, expected to contain "
                        f"{LLM_MODEL!r} -- refusing (never silently accept a downgrade)"
                    )
                usage = data.get("usage") or {}
                cost = usage.get("cost")
                content = data["choices"][0]["message"]["content"]
                raw = json.loads(content)
                with open(cost_log_path, "a", encoding="utf-8") as fh:
                    fh.write(
                        json.dumps(
                            {
                                "sys_id": candidate.sys_id,
                                "ref_work_id": candidate.ref_work_id,
                                "cost": cost,
                                "model_echo": model_echo,
                                "attempt": attempt,
                            },
                            ensure_ascii=False,
                        )
                        + "\n"
                    )
                return raw
            except Exception as exc:  # noqa: BLE001 -- real network call; bounded retries below
                last_exc = exc
                if attempt < max_attempts:
                    time.sleep(min(2**attempt, 20))
                    continue
        with open(cost_log_path, "a", encoding="utf-8") as fh:
            fh.write(
                json.dumps(
                    {
                        "sys_id": candidate.sys_id,
                        "ref_work_id": candidate.ref_work_id,
                        "cost": 0.0,
                        "model_echo": None,
                        "error": str(last_exc),
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )
        return None

    return _model_call


def sum_real_cost(cost_log_path: str) -> Tuple[float, int]:
    """Real total spend and real call count, summed from the cost log --
    NEVER an in-process accumulator alone, so a resumed run (killed and
    restarted, per `run_model_arm`'s own checkpoint discipline) reports the
    TRUE total across every process invocation that ever wrote to this log."""
    if not os.path.exists(cost_log_path):
        return 0.0, 0
    total = 0.0
    n = 0
    with open(cost_log_path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            cost = rec.get("cost")
            if cost:
                total += float(cost)
            n += 1
    return total, n


# ---------------------------------------------------------------------------
# The owner-labelling instrument -- Markdown + RTL XLSX, capped at ~40 rows,
# path-labelled, NO pre-filled verdict. House style mirrors
# scripts/discovery_gate1_evidence.py::write_hardcases_xlsx.
# ---------------------------------------------------------------------------

def _manuscript_str(sys_id: str, libraries: Dict[str, Dict[str, str]]) -> str:
    shelfmark = libraries.get(sys_id, {}).get("shelfmark", "")
    return shelfmark if shelfmark else f"sys_id {sys_id} (no shelfmark on file)"


# The owner cannot adjudicate "is this identified in the aids?" without SEEING
# what each aid actually says -- showing only the selection rationale repeats the
# arm-2 instrument failure (a stratum label where the source text belonged, which
# left 18/25 cases undecidable). Every checked source therefore gets its own
# column carrying its OWN free text, provenance-separated exactly as
# `assemble_evidence_bundle` hands it to the model, so what the owner reads is
# what the model read.
_EVIDENCE_CAP = 3000
_NO_TEXT = "(none -- this source has no text at all for this manuscript)"


def _evidence_cell(texts: Sequence[str]) -> str:
    """Renders one source family's own free text for the instrument. Multiple
    rows stay visibly separate. Over-long text is truncated with an EXPLICIT,
    counted marker -- never silently, since a truncation that hid the naming
    phrase would flip a verdict without the owner ever knowing."""
    joined = "\n---\n".join(t.strip() for t in texts if t and t.strip())
    if not joined:
        return _NO_TEXT
    if len(joined) > _EVIDENCE_CAP:
        return (
            joined[:_EVIDENCE_CAP]
            + f"\n[... TRUNCATED for display: {len(joined)} chars total. "
            "Check the source directly before recording `genuinely_novel` on this row.]"
        )
    return joined


def _row_evidence(c: NoveltyCandidate) -> Dict[str, str]:
    """Pulls each checked source's text off the SAME bundle the model was given.

    Refuses to emit restricted-corpus text. This instrument is a TRACKED file
    under `.planning/`, and D-25 forbids restricted-source content in committed
    files; the probe never populates that field today, so a non-empty value here
    means an upstream change has quietly routed restricted text at a committed
    artifact -- fail closed rather than write it out.
    """
    bundle = assemble_evidence_bundle(c)
    if bundle.get("m_source_shelfmark"):
        raise RuntimeError(
            f"restricted-corpus text present on {c.sys_id}/{c.ref_work_id}; refusing to write it "
            "into a tracked owner instrument (D-25). Mask or drop it upstream first."
        )
    return {
        "ev_catalogue": _evidence_cell(bundle.get("catalogue", ())),
        "ev_bibliography": _evidence_cell(bundle.get("bibliography", ())),
        "ev_pgp": _evidence_cell(bundle.get("pgp", ())),
        "ev_fgp": _evidence_cell(bundle.get("fgp", ())),
    }


def build_instrument_rows(
    model_path_candidates: Sequence[NoveltyCandidate],
    bypass_path_candidates: Sequence[NoveltyCandidate],
    libraries: Dict[str, Dict[str, str]],
) -> List[Dict[str, Any]]:
    """Builds the final, capped, path-labelled case list (model-path rows
    first, then bypass-path rows, each internally sorted by (sys_id,
    ref_work_id) for determinism) -- no verdict, no proposal, no shade
    guess beyond stating WHY the row is in the set (its selection path)."""
    rows: List[Dict[str, Any]] = []
    for c in sorted(model_path_candidates, key=_candidate_sort_key):
        rows.append(
            {
                "path": MODEL_PATH,
                "sys_id": c.sys_id,
                "ref_work_id": c.ref_work_id,
                "claimed_title": c.claimed_title,
                "claimed_author": c.claimed_author,
                "manuscript": _manuscript_str(c.sys_id, libraries),
                "catalogue_text": c.catalogue_text or "",
                **_row_evidence(c),
                "reason": (
                    "MODEL PATH: this row failed the mechanical heuristic name-match (it is part of "
                    "the residual) and the pinned gate (gemini-3.6-flash, effort=low) classified it "
                    f"`{CANDIDATE_STATUS}` -- it WOULD ship as a 'Candidate for new finds' in production."
                ),
            }
        )
    for c in sorted(bypass_path_candidates, key=_candidate_sort_key):
        rows.append(
            {
                "path": BYPASS_PATH,
                "sys_id": c.sys_id,
                "ref_work_id": c.ref_work_id,
                "claimed_title": c.claimed_title,
                "claimed_author": c.claimed_author,
                "manuscript": _manuscript_str(c.sys_id, libraries),
                "catalogue_text": "",
                **_row_evidence(c),
                "reason": (
                    "BYPASS PATH: NONE of the four checked-source families (catalogue, bibliography, "
                    "PGP, FGP) has ANY text at all for this manuscript -- this row ships as a "
                    "'Candidate for new finds' automatically, with NOTHING checked against it."
                ),
            }
        )
    for i, row in enumerate(rows, start=1):
        row["case_num"] = i
    return rows


def write_instrument_markdown(rows: List[Dict[str, Any]], path: str, meta: Dict[str, Any]) -> None:
    lines: List[str] = []
    lines.append("# The `fills_gap` Probe -- Owner-Labelling Instrument (owner ruling K)")
    lines.append("")
    lines.append(
        f"**{len(rows)} cases** — {meta['n_model_path']} from the MODEL PATH (of "
        f"{meta['model_sample_size']} real residual candidates sent to the pinned gate, "
        f"{meta['n_model_fills_gap']} were classified `{CANDIDATE_STATUS}`" + (
            f"; capped down to {meta['n_model_path']}" if meta['n_model_fills_gap'] > meta['n_model_path'] else ""
        ) + f") + {meta['n_bypass_path']} from the BYPASS PATH (a random sample of "
        f"{meta['bypass_sample_size']} of the real, full-corpus {meta['bypass_population_size']}-row "
        "no-checked-source-text population)."
    )
    lines.append("")
    lines.append(
        f"Real spend so far on the model path: **${meta['real_cost_total']:.6f}** across "
        f"{meta['real_call_count']} real OpenRouter calls (pinned `{LLM_MODEL}`, "
        f"`reasoning.effort={LLM_REASONING_EFFORT!r}`, prompt hash `{PROMPT_SHA256[:16]}...`)."
    )
    lines.append("")
    lines.append(
        "**The question, for EVERY row below, regardless of path:** " + VERDICT_QUESTION
    )
    lines.append("")
    lines.append(
        "**What this probe can and cannot support (read before labelling):** with " + str(len(rows)) +
        " cases split across two paths, this gives a COARSE per-path rate, not a tight one -- do "
        "not treat either path's resulting fraction as a precise corpus-wide false-novel rate. The "
        "model-path sample is a random draw from the real residual (not stratified by source family, "
        "unlike the ruling-J hard-case pool), so it says nothing about which SOURCE FAMILY drives any "
        "false novelty found. The bypass-path sample is a plain random draw from the real "
        f"{meta['bypass_population_size']}-row population and every one of its rows is, by "
        "construction, ALREADY a shipped candidate (no filtering has happened to it at all) -- so its "
        "labelled fraction is the more direct, if still small-sample, read on that path's real risk."
    )
    lines.append("")
    lines.append("## Cases")
    lines.append("")
    lines.append(
        "Each case below carries the ACTUAL free text of every checked source, exactly as the "
        "pinned gate received it. A source shown as "
        f"\"{_NO_TEXT}\" genuinely has none on file -- that is a finding, not a rendering gap."
    )
    lines.append("")
    for row in rows:
        work = row["claimed_title"] + (f" ({row['claimed_author']})" if row["claimed_author"] else "")
        lines.append(f"### Case {row['case_num']} — {row['path']} path")
        lines.append("")
        lines.append(f"- **Manuscript:** {row['manuscript']} (`{row['sys_id']}`)")
        lines.append(f"- **Our claim:** {work}")
        lines.append(f"- **Why it is in the set:** {row['reason']}")
        lines.append("")
        for label, key in (
            ("Catalogue says", "ev_catalogue"),
            ("Bibliography says", "ev_bibliography"),
            ("PGP says", "ev_pgp"),
            ("FGP says", "ev_fgp"),
        ):
            lines.append(f"**{label}:**")
            lines.append("")
            lines.append("```")
            lines.append(row[key])
            lines.append("```")
            lines.append("")
    lines.append("")
    with open(path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines) + "\n")


def write_instrument_xlsx(rows: List[Dict[str, Any]], path: str, meta: Dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    from shared.export_utils import sanitize_text_for_excel as _san

    def _s(value: Optional[str]) -> str:
        return _san(value or "")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_top = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)
    wrap_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.sheet_view.rightToLeft = True
    headers = [
        "Case #",
        "Verdict",
        "Path",
        "Manuscript",
        "sys_id",
        "Claimed work",
        # The four checked sources, each carrying its OWN free text -- this is
        # the evidence the verdict is actually about, and the model saw exactly
        # these strings. Without them the owner is grading a rationale, not a claim.
        "Catalogue says",
        "Bibliography says",
        "PGP says",
        "FGP says",
        "Why this row is in the set",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    widths = {
        "A": 9, "B": 18, "C": 10, "D": 32, "E": 22, "F": 40,
        "G": 55, "H": 55, "I": 55, "J": 55, "K": 60,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for row in rows:
        work = row["claimed_title"] + (f" ({row['claimed_author']})" if row["claimed_author"] else "")
        ws.append(
            [
                row["case_num"],
                "",  # Verdict -- left blank; owner fills in. NEVER pre-filled.
                _s(row["path"]),
                _s(row["manuscript"]),
                _s(row["sys_id"]),
                _s(work),
                _s(row["ev_catalogue"]),
                _s(row["ev_bibliography"]),
                _s(row["ev_pgp"]),
                _s(row["ev_fgp"]),
                _s(row["reason"]),
            ]
        )
    last_row = ws.max_row
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=1).alignment = wrap_top_center
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = wrap_top
        # Pin a uniform, readable height. Left to auto-fit, a 3000-char evidence
        # cell would make a single row hundreds of lines tall and the sheet
        # unnavigable; the owner expands any row that needs more.
        ws.row_dimensions[r].height = 130
    # Freeze the identity block (case / verdict / path / manuscript / sys_id /
    # claim) so it stays visible while scrolling the wide evidence columns.
    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VERDICT_TOKENS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid verdict",
        error="Choose 'genuinely_novel', 'actually_recorded', 'unsure' or 'skip'. Free text is rejected.",
        promptTitle="Verdict",
        prompt="Is this fragment genuinely not identified in the finding aids we checked? Blank = not yet answered.",
    )
    ws.add_data_validation(dv)
    if last_row >= 2:
        dv.add(f"B2:B{last_row}")

    ws2 = wb.create_sheet(title="Vocabulary & Instructions")
    ws2.sheet_view.rightToLeft = True
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 90

    def _note(text: str) -> None:
        ws2.append([_s(text)])
        ws2.cell(row=ws2.max_row, column=1).alignment = Alignment(
            horizontal="right", vertical="top", wrap_text=True, readingOrder=2
        )
        ws2.merge_cells(start_row=ws2.max_row, start_column=1, end_row=ws2.max_row, end_column=2)

    _note(
        "This workbook is the owner-labelling instrument for the `fills_gap` probe (owner ruling K, "
        "136-GATE1-DECISIONS.md). It contains ONLY rows that would actually ship as 'Candidates for "
        "new finds' in production -- from BOTH paths a row can take to become a candidate, labelled by "
        "path so the two rates can be reported separately."
    )
    ws2.append([])
    _note("THE QUESTION, for every row regardless of path: " + VERDICT_QUESTION)
    ws2.append([])
    _note(
        f"MODEL PATH ({meta['n_model_path']} rows): a real residual candidate (failed the mechanical "
        "heuristic name-match) that the real pinned gate (gemini-3.6-flash, effort=low, called for "
        "real via OpenRouter) classified `fills_gap`."
    )
    _note(
        f"BYPASS PATH ({meta['n_bypass_path']} rows): NONE of the four checked-source families has ANY "
        "text at all for this manuscript. This path NEVER reaches the model -- the funnel ships it as "
        "a candidate automatically. Arguably the higher-risk path, since nothing at all examines it "
        "before it ships."
    )
    ws2.append([])
    _note(
        "IMPORTANT: a BLANK verdict cell is NOT a label -- it means 'not yet answered'. If you cannot "
        "judge a case, enter `unsure` explicitly. If you choose not to judge a case at all, enter "
        "`skip` explicitly -- it is recorded as skipped, never silently treated as an answer."
    )
    ws2.append([])
    _note(
        f"Real spend on the model path so far: ${meta['real_cost_total']:.6f} across "
        f"{meta['real_call_count']} real OpenRouter calls. Pinned contract: model={LLM_MODEL!r}, "
        f"version={LLM_MODEL_VERSION!r}, effort={LLM_REASONING_EFFORT!r}, "
        f"prompt_sha256={PROMPT_SHA256}, input_normalization_sha256={INPUT_NORMALIZATION_SHA256}."
    )
    ws2.append([])
    _note(
        "COARSE, not tight: this is a ~40-case sample. It gives a per-path RATE, not a precise "
        "corpus-wide figure -- treat the resulting fraction accordingly."
    )

    wb.save(path)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--asset", default=DEFAULT_ASSET)
    parser.add_argument("--libraries-csv", default=DEFAULT_LIBRARIES_CSV)
    parser.add_argument("--fjms-db", default=DEFAULT_FJMS_DB)
    parser.add_argument("--pgp-db", default=DEFAULT_PGP_DB)
    parser.add_argument("--fgp-db", default=DEFAULT_FGP_DB)
    parser.add_argument("--checkpoint", default=DEFAULT_CHECKPOINT)
    parser.add_argument("--cost-log", default=DEFAULT_COST_LOG)
    parser.add_argument("--md-out", default=DEFAULT_MD_OUT)
    parser.add_argument("--xlsx-out", default=DEFAULT_XLSX_OUT)
    parser.add_argument("--model-sample-size", type=int, default=300)
    parser.add_argument("--bypass-sample-size", type=int, default=20)
    parser.add_argument("--model-path-cap", type=int, default=20)
    parser.add_argument("--seed", type=int, default=20260803)
    parser.add_argument(
        "--run-model", action="store_true",
        help="Actually call the real pinned model over the model-path sample (real spend). "
             "Without this flag, only the free heuristic pass + sampling + instrument-from-"
             "existing-checkpoint runs.",
    )
    parser.add_argument("--api-key-env", default="OPENROUTER_API_KEY")
    args = parser.parse_args(argv)

    print(f"Loading real data from {args.asset} + real sidecars...")
    candidates, works, libraries = build_all_candidates(
        args.asset, args.libraries_csv, args.fjms_db, args.pgp_db, args.fgp_db
    )
    print(f"Total real shipped (sys_id, work_id) candidates: {len(candidates)}")

    resolved, residual = run_heuristic_funnel(candidates)
    n_confirms = sum(1 for r in resolved.values() if r.novelty_status == "confirms")
    bypass_all = [
        c for c in candidates
        if resolved.get(f"{c.sys_id}::{c.ref_work_id}") is not None
        and resolved[f"{c.sys_id}::{c.ref_work_id}"].reason == NO_SOURCE_TEXT_REASON
    ]
    print(f"  resolved confirms (mechanical name-match): {n_confirms}")
    print(f"  bypass (no source text at all -> fills_gap automatically): {len(bypass_all)}")
    print(f"  residual (would reach the model): {len(residual)}")
    assert n_confirms + len(bypass_all) + len(residual) == len(candidates)

    model_sample = sample_candidates(residual, args.model_sample_size, args.seed)
    bypass_sample = sample_candidates(bypass_all, args.bypass_sample_size, args.seed + 1)
    print(f"Model-path sample: {len(model_sample)} real residual candidates")
    print(f"Bypass-path sample: {len(bypass_sample)} real no-source-text candidates")

    if args.run_model:
        api_key = os.environ.get(args.api_key_env)
        if not api_key:
            print(f"ERROR: {args.api_key_env} not set in environment -- cannot run the real model.")
            return 1
        model_call = make_openrouter_model_call(api_key, args.cost_log)
        print(f"Running the REAL pinned model over {len(model_sample)} candidates (checkpointed, resumable)...")
        results = run_model_arm(model_sample, model_call=model_call, checkpoint_path=args.checkpoint)
    else:
        results = {}
        if os.path.isfile(args.checkpoint):
            with open(args.checkpoint, "r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    rec = json.loads(line)
                    key = f"{rec['sys_id']}::{rec['ref_work_id']}"
                    results[key] = rec

    fills_gap_candidates = [
        c for c in model_sample
        if results.get(f"{c.sys_id}::{c.ref_work_id}", {}).get("novelty_status") == CANDIDATE_STATUS
    ]
    print(f"Model-path candidates classified `{CANDIDATE_STATUS}` (would ship): {len(fills_gap_candidates)}")

    model_path_final = sample_candidates(fills_gap_candidates, args.model_path_cap, args.seed + 2)

    real_cost_total, real_call_count = sum_real_cost(args.cost_log)
    print(f"Real spend so far (from {args.cost_log}): ${real_cost_total:.6f} across {real_call_count} real calls")

    rows = build_instrument_rows(model_path_final, bypass_sample, libraries)
    meta = {
        "model_sample_size": len(model_sample),
        "n_model_fills_gap": len(fills_gap_candidates),
        "n_model_path": len(model_path_final),
        "bypass_sample_size": len(bypass_sample),
        "n_bypass_path": len(bypass_sample),
        "bypass_population_size": len(bypass_all),
        "real_cost_total": real_cost_total,
        "real_call_count": real_call_count,
    }
    write_instrument_markdown(rows, args.md_out, meta)
    write_instrument_xlsx(rows, args.xlsx_out, meta)
    print(f"Wrote {args.md_out} and {args.xlsx_out} ({len(rows)} total cases).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

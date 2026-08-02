#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gate-1 evidence measurement for Phase 136 plan 03 (PANEL-01, PANEL-02, NOVEL-01).

A READ-ONLY measurement script over the deployed discovery sidecar
(``discovery-v1-33499c5b...``, the LIVE asset — the trimmed rebuild has not
run yet). It computes every number the owner needs to rule on the five
"Still OPEN by design" wave-1 decisions named in
``136-CONTEXT.md`` (D-13e, D-16, D-13c, D-13b, D-13d), plus the zero-cost
novelty hard-case candidate set that feeds plan 136-03 Task 3 — originally
three D-23c classes (near-miss titles, alias pairs, granularity), extended
per an OWNER-AUTHORIZED gate-1 ruling (recorded in ``136-GATE1-DECISIONS.md``
item C) with two additional classes: terse/missing catalogue identification
text, and generic collection works. A SIXTH class -- catalogue divergence --
was added per owner decision E (``136-GATE1-DECISIONS.md`` item E, which also
amends the tri-state novelty flag into a seven-value shade enum): real
shipped claims where an available finding aid ties the SAME fragment to a
DIFFERENT work that is not a D-13d granularity variant -- the shade decision
E calls ``diverges``, with zero representation in Classes 1-5.

**Correction E′** (``136-GATE1-DECISIONS.md`` § E′, a same-day correction to
decision E, not a new ruling): decision E's ``refines_granularity`` shade
conflated two opposite directions of a granularity relationship into one
value. The shade enum now widens from SEVEN to EIGHT values by splitting it:
``refines_granularity`` (OUR claim is the finer one -- informative) versus
``aid_more_specific`` (the AID's identification is the finer one -- we add
nothing). ``aid_more_specific`` is the least-novel shade and is excluded from
the candidate toggle alongside ``confirms`` / ``refines_granularity`` /
``diverges`` / ``extends``; ``fills_gap`` remains the sole candidate
selector.

**Rulings F and G** (``136-GATE1-DECISIONS.md`` §§ F/G, a later continuation
after the owner read the actual 97 candidate cases): (F) ``diverges`` is
RETIRED and replaced by a SCOPE split -- ``diverges_work`` (a genuinely
different work) / ``diverges_part`` (a different or finer part of the SAME
work) -- widening the shade enum from EIGHT to NINE values, plus a NEW
sibling ``divergence_correctness`` question (``catalogue_correct`` /
``claim_correct`` / ``unclear``) recorded separately from the shade, because
the owner's own review of the real cases found the catalogue is often right
when it disagrees with a claim -- a fact this module must surface to the
user (hidden by default, explicit warning) without ever letting the system
itself adjudicate on the catalogue's say-so (the catalogue-never-evidence
discipline, applied to a new axis). (G) A boundary correction: when the
catalogue's STRUCTURED identification is generic but its own FREE TEXT
already states our specific identification in prose, the shade is
``confirms``, not ``refines_granularity`` -- with a systemic consequence
that a novelty check keying only on structured ids manufactures false
novelty at scale; this module's OWN Class-6 selector
(``select_catalogue_divergence_candidates``) demonstrates the failure
(over-fires on roughly half of its candidates per the owner's own
characterization) and is left UNCORRECTED here deliberately, per the
"measure it, do not quietly fix it away" instruction.

**Labelling restructure** (same continuation, owner ruling): Classes 1-3
(near-miss / alias / granularity) compare two of OUR OWN claims -- an A↔B
"is this the same work" identity judgment, not a claim-vs-aid novelty
judgment -- and were found, on reading the actual cases, to bake "same work"
into their own selection (same-author + common-title-stem is the selection
criterion itself), so full labelling of all 52 was REPLACED by an ~8-case
IDENTITY SPOT-CHECK per class (``same_work`` / ``different_works`` /
``unsure`` / ``skip``) that tests the constant-answer assumption rather than
building ground truth. Classes 4-6 (terse catalogue / generic collection /
catalogue divergence) are where the answer genuinely varies and carry the
NOVELTY SHADE question; they are EXPANDED from 45 to ~75 candidates
(20/25/30), proportioned to how genuinely hard/consequential each family is
(Class 6 -- the most consequential and, per ruling G, measurably flawed --
gets the largest expansion; Class 5's collection-level ambiguity is second;
Class 4's terse/absent text is the most mechanical of the three). Each row
now carries the ONE question type it was actually constructed to support
(identity XOR shade), fixing an incoherence in the pre-restructure worksheet
where Decision E's shade vocabulary was applied uniformly to all six classes
even though Classes 1-3 have no claim-vs-aid relationship to judge.

**Rulings H and I** (``136-GATE1-DECISIONS.md`` §§ H/I, a later dispatch still,
prompted by the read-only ``136-NOVELTY-PRIOR-ART.md`` research pass): (H) the
prior-art pass measured a real, non-trivial "witness" shape under the OLDER
five-way title-gate vocabulary -- a catalogue entry names a broader standard-
rite prayer-book/cycle/ceremony whose predictable content includes the
claimed unit, without ever naming that unit itself (1,327 of 20,410 in-scope
rows, 6.5%, in the population that vocabulary scored) -- which the CURRENT
shade enum has no bucket for and would misfile as ``fills_gap`` by
elimination. The owner adopted this as a TENTH shade, under a
NON-COLLIDING name (``container_predicts`` -- "witness" already names five
OTHER distinct concepts in this project, per the prior-art pass's own §8
sweep): excluded from "Candidates for new finds" like every other
non-``fills_gap`` shade, but shown NORMALLY (never hidden by default) --
ruling F's default-hidden posture was specifically about rows the owner has
measured reason to believe are OUR false positives, which does not apply
here (there is no disagreement between the aid and the claim). (I) The
pinned gate's validation (40/40 agreement; 99% vs 103 human grades) covers
the FIVE-way vocabulary and the one-title-string input contract -- it does
NOT cover the widened shade enum or ruling G's free-text input contract, so
the owner ruled the pinned config must be RE-MEASURED against the owner-
labelled evaluation set on the new vocabulary/contract before the
production run, not merely run because it was once validated on a
different question. Consequence for this module: a NEW Class 7
(liturgical-container predictability) is added to the hard-case candidate
set, built with the identical zero-model-call, script-reproducible
selection discipline as Classes 4-6, targeting ~12 cases, so the model's
FIRST encounter with this shape is a graded evaluation, not production.

**Ruling J** (``136-GATE1-DECISIONS.md`` § J, a later continuation still,
after a Codex-flagged accuracy gap: the pre-J pool's Classes 4/5/7 all read
EXACTLY ONE field -- libraries.csv column 7 -- and had ZERO representation
of the bib/PGP/FGP failure modes Codex measured as most damaging in
``gen2_novelty_gate.py`` (3,688 ``published_full`` false-knowns, 2,014 PGP
false-knowns [942 sole-source]). Ruling J (a) adopts the funnel-FIRST
architecture -- the LLM arm runs ONLY on the heuristic funnel's residual,
never on all identifications -- and records the resulting UNRECOVERABLE
false-known risk (the funnel only ever demotes, never promotes, so the model
never sees a heuristically-demoted row); (b) REPLACES Classes 4/5/7 with a
THREE-ARM, SOURCE-STRATIFIED sample built from the REAL bib/PGP/FGP/FJMS-
catalogue sidecars (not just libraries.csv): Arm 1 RESIDUAL (rows that would
reach the model, stratified by which source supplied unmatched text, folding
in the former Classes 4 and 7 as two of its seven strata), Arm 2
HEURISTIC-DEMOTED (rows the funnel marks known before any model call,
oversampling `published_full`-sole and PGP-sole demotions to characterize the
false-known risk directly), and Arm 3 NO-SOURCE-TEXT (rows with no checked-
source text at all, shipped as candidates with NO verdict, by design). Class
6 (catalogue divergence) and Classes 1-3 (identity spot-check) are RETAINED
UNCHANGED; Class 5 (generic collection works) is DROPPED (no owner ruling
exists for any specific Class 5 case). See ``select_novelty_arms`` and its
neighboring functions for the full implementation and the exact kept/folded/
dropped accounting.

This module also emits an XLSX labelling workbook (``write_hardcases_xlsx``)
alongside the Markdown, per the owner's request for something easier to work
with than Hebrew RTL in Markdown; it now carries FIVE sheets -- "Identity
Spot-Check", "Novelty Shades" (Class 6 + Arm 1, with its own Correctness
column), "Heuristic-Demoted" (Arm 2), "No-Source-Text" (Arm 3, no verdict
column), and "Vocabulary & Instructions".

Mirrors the shape of ``scripts/bench_discovery.py``: open the live asset
read-only, drive real queries against it, print a table per measurement, and
NEVER present a silent zero as a finding — every count this script reports
that should be nonzero is asserted nonzero, and the script exits 1 if any
such assertion fails.

Usage:
    python scripts/discovery_gate1_evidence.py <asset.db>
    python scripts/discovery_gate1_evidence.py <asset.db> --research-db <fullcorpus.db> \\
        --libraries-csv libraries.csv --evidence-out 136-GATE1-EVIDENCE.md \\
        --hardcases-out 136-NOVELTY-HARDCASES.md --hardcases-xlsx-out 136-NOVELTY-HARDCASES.xlsx

Determinism: every SQL query below is either naturally total (aggregate
COUNT) or carries an explicit ORDER BY over a stable key, and every Python
container built from query results is either a dict keyed by a stable id or
a list sorted by an explicit key before being sliced/rendered — re-running
this script against the same asset produces byte-identical Markdown output.

Masking discipline: this script renders real ``works.neutral_title`` /
``works.author`` values (already the shipped, DATA-04-cleared neutral
titles — the same strings the live product already displays) and real
public catalogue metadata from ``libraries.csv`` (shelfmark / non-placeholder
title — already public, displayed throughout the product). It never reads or
prints a raw M-source/J-source identifier, a restricted corpus name, or
reference text. The caller is expected to run
``scripts/check_atlas_masking.py --scan-asset`` over both written files
before presenting them to the owner (acceptance criteria, not enforced by
this script itself — this script has no import-time dependency on the
masking-gate module so it stays a lightweight read path, mirroring
``web/discovery_assets.py``'s own comment on why it does not import
``scripts.build_discovery_sidecar``).
"""

from __future__ import annotations

import argparse
import csv
import difflib
import hashlib
import json
import os
import re
import sqlite3
import sys
import unicodedata
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PHASE_DIR = os.path.join(
    REPO_ROOT, ".planning", "phases",
    "136-read-surfaces-connections-panel-work-witnesses",
)
DEFAULT_RESEARCH_DB = os.path.join(REPO_ROOT, "same_work_spike", "probe", "data", "fullcorpus.db")
DEFAULT_LIBRARIES_CSV = os.path.join(REPO_ROOT, "libraries.csv")
DEFAULT_EVIDENCE_OUT = os.path.join(PHASE_DIR, "136-GATE1-EVIDENCE.md")
DEFAULT_HARDCASES_OUT = os.path.join(PHASE_DIR, "136-NOVELTY-HARDCASES.md")
DEFAULT_HARDCASES_XLSX_OUT = os.path.join(PHASE_DIR, "136-NOVELTY-HARDCASES.xlsx")

# Real production sidecars for the ruling-J three-arm stratified sampler
# (136-03 continuation, 2026-08-02) -- the SAME checked-source family NOVEL-01
# names and gen2_novelty_gate.py/discovery_identified_gate.py already read,
# at their standard project-root-relative locations (shared/fjms_service.py
# _SIDECAR_DIR/_SIDECAR_FILENAME, shared/document_service.py, shared/fgp_service.py).
DEFAULT_FJMS_DB = os.path.join(REPO_ROOT, "fist_data", "fjms_enrichment.db")
DEFAULT_PGP_DB = os.path.join(REPO_ROOT, "pgp_data", "pgp.db")
DEFAULT_FGP_DB = os.path.join(REPO_ROOT, "fgp_data", "fgp_transcriptions.db")

# ---------------------------------------------------------------------------
# Verdict vocabulary -- the SINGLE source of truth for every owner-facing
# shade token, shared verbatim by the Markdown worksheet, the XLSX dropdown
# validation and the XLSX vocabulary sheet, so the three surfaces can never
# drift out of sync. Order matches the owner's own E/E′/F/H shade table
# (decision E's six original shades, E′'s aid_more_specific inserted right
# after its sibling refines_granularity, F's diverges_work/diverges_part
# replacing the single diverges token, H's container_predicts inserted right
# before its sibling fills_gap -- the shade it would otherwise be misfiled
# into by elimination -- then not_checked's owner-facing alias `unsure` and
# the `skip` non-answer). ``not_checked`` itself is NOT owner-facing -- only
# its owner-facing alias `unsure` is offered as an answer token. THIS
# vocabulary is used ONLY on Class 4-7 ("shade") rows -- see
# IDENTITY_VOCABULARY below for the separate Class 1-3 question.
# ---------------------------------------------------------------------------
SHADE_VOCABULARY: Tuple[Tuple[str, str], ...] = (
    ("confirms", "an aid already ties this fragment to this work"),
    (
        "refines_granularity",
        "OUR claim is MORE SPECIFIC (finer) than what an aid says -- e.g. the catalogue names the "
        "whole work, our claim names a specific book/chapter of it (the D-13d same-author/related-"
        "title rule); the OPPOSITE direction from `aid_more_specific` -- we ADD precision here. "
        "OWNER RULING G: if the aid's own FREE TEXT already states this identification in ANY form "
        "(even under a coarser structured work-id), the correct shade is `confirms`, not this one -- "
        "reserve `refines_granularity` for information the aid contains in NO form, structured or free.",
    ),
    (
        "aid_more_specific",
        "an AID names a MORE SPECIFIC (finer) variant of this fragment's work than our claim does "
        "-- e.g. the catalogue names a chapter/book, our claim names the whole work (the D-13d "
        "same-author/related-title rule); the OPPOSITE direction from `refines_granularity` -- we "
        "add NOTHING here, the aid already knew more (owner correction E′; the LEAST novel shade)",
    ),
    (
        "diverges_work",
        "an aid ties this fragment to a genuinely DIFFERENT WORK (not a granularity variant of ours) "
        "-- the aid and the claim contradict each other on WHICH WORK this is (owner ruling F, "
        "replacing the single `diverges` token). Owner: reading real cases, USUALLY the catalogue is "
        "right and this is OUR false positive -- but not always; record which side is correct in the "
        "separate Correctness column. Hidden by default on every surface, behind an explicit warned "
        "toggle -- never silently shown, never silently suppressed.",
    ),
    (
        "diverges_part",
        "an aid ties this fragment to a DIFFERENT OR FINER PART of the SAME work (owner ruling F -- "
        "\"more delicate and essentially less important\" than diverges_work) -- e.g. the aid names a "
        "specific chapter/section of the work while we name a different one, or the whole work, or "
        "vice versa. Same Correctness column and same hidden-by-default posture as diverges_work.",
    ),
    (
        "container_predicts",
        "an aid names a BROADER rite/cycle/ceremony/container -- a full standard-rite prayer-book "
        "(siddur/machzor) or a named ceremony/occasion -- whose STANDARD, PREDICTABLE content includes "
        "this specific unit, WITHOUT ever naming the unit itself (owner ruling H; e.g. the catalogue "
        "names 'מחזור מנהג אשכנז לשלש רגלים', the claim is a Yotzer for one of its festivals). Distinct "
        "from `confirms` (the aid never names this specific unit) and from `fills_gap` (the content IS "
        "predictable, so it is not 'previously unknown' -- under the pre-H enum this fell through to "
        "`fills_gap` by elimination). Excluded from the candidate toggle like every other non-`fills_gap` "
        "shade, but -- UNLIKE `diverges_work`/`diverges_part` -- shown NORMALLY, never hidden by "
        "default: there is no disagreement here to warn about, only a container relationship.",
    ),
    (
        "fills_gap",
        "the aids identify this fragment as nothing at all -- the genuine \"previously unknown\" case",
    ),
    (
        "extends",
        "aids tie OTHER folios of the SAME manuscript to this work, but not this specific folio",
    ),
    (
        "alias_merge",
        "the two work_ids shown ARE the same underlying work, not yet canonically merged (Class 2's "
        "situation)",
    ),
    (
        "unsure",
        "you cannot judge this case from the information shown -- maps to `not_checked`, costs "
        "nothing, is a real and useful answer",
    ),
    (
        "skip",
        "you choose not to judge this case at all -- recorded as skipped, NEVER filled from a draft "
        "`PROPOSAL`",
    ),
)
# The real SHADE tokens only (excludes the two non-shade answer tokens
# `unsure` / `skip`) -- this is the XLSX DataValidation list together with
# `unsure` / `skip` appended (eleven tokens total): nine real shades that are
# owner-facing (the tenth stored value, `not_checked`, is a fail-closed
# system default never picked directly -- `unsure` is its owner-facing
# alias) plus the two non-shade answers.
SHADE_TOKENS: Tuple[str, ...] = tuple(tok for tok, _ in SHADE_VOCABULARY)

# ---------------------------------------------------------------------------
# Correctness vocabulary (owner ruling F, 136-GATE1-DECISIONS.md § F) -- a
# SEPARATE axis from the shade, applicable ONLY when the shade verdict is
# `diverges_work` or `diverges_part`. A divergence shade records only THAT
# the aid and the claim disagree (and at what scope); it cannot also record
# WHICH side is right, because the owner's own review of the real cases
# found BOTH directions occur under the identical shade. Blank is allowed
# (meaning "not applicable / not yet answered" -- the same convention as a
# blank Verdict cell) and is the correct answer for every non-divergence row.
# ---------------------------------------------------------------------------
CORRECTNESS_VOCABULARY: Tuple[Tuple[str, str], ...] = (
    (
        "catalogue_correct",
        "the catalogue/aid is right; our claim is the false positive -- owner ruling F: reading the "
        "real cases, this is the COMMON outcome",
    ),
    (
        "claim_correct",
        "our claim is right; the aid is wrong, thinner, or itself mistaken",
    ),
    (
        "unclear",
        "cannot tell which side is correct from the information shown",
    ),
)
CORRECTNESS_TOKENS: Tuple[str, ...] = tuple(tok for tok, _ in CORRECTNESS_VOCABULARY)

# ---------------------------------------------------------------------------
# Identity vocabulary (owner-authorized labelling restructure, same
# continuation as rulings F/G) -- the question Classes 1-3 (near-miss /
# alias / granularity) were ACTUALLY constructed to support: are the two
# claims/works shown (A and B) the same underlying work, or genuinely
# different works? This is NOT a novelty shade -- these rows compare two of
# OUR OWN claims against each other, never a claim against a finding aid.
# ---------------------------------------------------------------------------
IDENTITY_VOCABULARY: Tuple[Tuple[str, str], ...] = (
    (
        "same_work",
        "A and B are the same underlying work (or two parts/granularities of the same work)",
    ),
    (
        "different_works",
        "A and B are genuinely different works",
    ),
    (
        "unsure",
        "you cannot judge this pair from the information shown -- a real and useful answer",
    ),
    (
        "skip",
        "you choose not to judge this pair at all -- recorded as skipped",
    ),
)
IDENTITY_TOKENS: Tuple[str, ...] = tuple(tok for tok, _ in IDENTITY_VOCABULARY)

# ---------------------------------------------------------------------------
# Demotion vocabulary (owner ruling J, 136-GATE1-DECISIONS.md Section J;
# 136-03 continuation, 2026-08-02) -- the question Arm 2 (HEURISTIC-DEMOTED)
# rows were built to support: the funnel-first architecture means these rows
# are marked "already recorded" by a HEURISTIC (a bib/PGP presence test) and
# NEVER reach the model at all -- so this is the ONLY place a false-known can
# ever be caught. Not a shade, not an identity call: a straight correctness
# check on the DEMOTION ITSELF.
# ---------------------------------------------------------------------------
DEMOTION_VOCABULARY: Tuple[Tuple[str, str], ...] = (
    (
        "demotion_correct",
        "the demoting source genuinely already names/records THIS SPECIFIC work on THIS "
        "fragment -- the heuristic's demotion (marking it already known, never reaching the "
        "model) is right",
    ),
    (
        "false_known",
        "the demoting source does NOT actually name this specific work -- only its GENERIC "
        "presence (e.g. a bibliography record, a PGP description/transcription on the "
        "fragment) tripped the heuristic. Per owner ruling J this is an UNRECOVERABLE lost "
        "finding: the funnel only ever demotes (discovery -> known, never the reverse) and "
        "the model never sees a heuristically-demoted row, so a false_known here is "
        "permanent unless this labelling catches it",
    ),
    (
        "unsure",
        "you cannot judge this case from the information shown -- a real and useful answer",
    ),
    (
        "skip",
        "you choose not to judge this case at all -- recorded as skipped",
    ),
)
DEMOTION_TOKENS: Tuple[str, ...] = tuple(tok for tok, _ in DEMOTION_VOCABULARY)

# ---------------------------------------------------------------------------
# D-13c / gate-4 page-coverage normalizer -- a faithful, standalone port of
# scripts/build_discovery_sidecar.py::norm_stream_letter_count /
# compute_page_coverage (itself a port of
# same_work_spike/probe/scripts/normalize.py::norm_stream). Re-implemented
# here rather than imported so this measurement script stays free of
# build_discovery_sidecar's masking-gate + argparse import surface (mirrors
# web/discovery_assets.py's stated reason for not importing that module).
# ---------------------------------------------------------------------------
_HEB_MIN, _HEB_MAX = 0x05D0, 0x05EA
_FINAL_FOLD = {
    0x05DA: chr(0x05DB),  # final kaf   -> kaf
    0x05DD: chr(0x05DE),  # final mem   -> mem
    0x05DF: chr(0x05E0),  # final nun   -> nun
    0x05E3: chr(0x05E4),  # final pe    -> pe
    0x05E5: chr(0x05E6),  # final tsadi -> tsadi
}


def norm_stream_letter_count(text: Optional[str]) -> int:
    """Space-free normalized Hebrew base-letter stream length (page_norm_letters)."""
    if not text:
        return 0
    n = 0
    for ch in unicodedata.normalize("NFC", text):
        folded = _FINAL_FOLD.get(ord(ch))
        code = ord(ch) if folded is None else ord(folded)
        if _HEB_MIN <= code <= _HEB_MAX:
            n += 1
    return n


def compute_page_coverage(matched_letters: Optional[int], page_norm_letters: Optional[int]) -> Optional[float]:
    if matched_letters is None:
        return None
    if not page_norm_letters:
        return 0.0
    return min(1.0, matched_letters / page_norm_letters)


# ---------------------------------------------------------------------------
# Band rank -- mirrors the ORDER of scripts/discovery_ids.py's
# _BAND_RANK_GROUPS (strongest first). Reproduced as a small local table
# (rather than reaching into that module's underscore-prefixed private
# lattice) so this script depends only on the PUBLIC discovery_ids surface
# it actually needs (none, at this granularity) and stays trivially
# auditable against the cited source.
# ---------------------------------------------------------------------------
_BAND_RANK_GROUPS: List[List[Tuple[str, str]]] = [
    [("track1_direct", "expert_verified"), ("track1_direct", "high_confidence_algorithmic")],
    [("track1_direct", "tier_a")],
    [("propagated", "corroborated")],
    [("track1_direct", "screening_rb")],
    [("track1_direct", "screening_canon")],
    [("propagated", "weak")],
    [("propagated", "not_evaluated")],
]
_BAND_RANK_INDEX: Dict[Tuple[str, str], int] = {
    pair: i for i, group in enumerate(_BAND_RANK_GROUPS) for pair in group
}
_UNRANKED_BAND = len(_BAND_RANK_GROUPS)
_SCREENING_CANON_RANK = _BAND_RANK_INDEX[("track1_direct", "screening_canon")]
_WEAK_RANK = _BAND_RANK_INDEX[("propagated", "weak")]


def band_rank(evidence_source: str, confidence_band: str) -> int:
    return _BAND_RANK_INDEX.get((evidence_source, confidence_band), _UNRANKED_BAND)


# ---------------------------------------------------------------------------
# Title-relationship machinery (D-13d granularity separation rule; reused
# verbatim, per the plan's instruction, for the novelty hard-case selection
# in Task 1). Pure string/Unicode operations -- no model call, ever.
# ---------------------------------------------------------------------------
_TITLE_PUNCT_RE = re.compile(r"[\"'׳״‘’“”]")
_WS_RE = re.compile(r"\s+")


def normalize_title(title: Optional[str]) -> str:
    """NFC + strip quote/geresh/gershayim marks + collapse whitespace.

    Deliberately does NOT strip nikud/te'amim (works.neutral_title is plain
    text with no vocalization in the shipped asset) and does NOT touch the
    maqaf ``־`` (a real word-joining character, not punctuation to
    discard) -- mirrors the same discipline documented for
    ``norm_stream_letter_count`` above.
    """
    if not title:
        return ""
    t = unicodedata.normalize("NFC", title)
    t = _TITLE_PUNCT_RE.sub("", t)
    t = _WS_RE.sub(" ", t).strip()
    return t


def titles_share_prefix(title_a: str, title_b: str, min_len: int = 4) -> bool:
    """True when the two (already-normalized) titles share a >= min_len prefix.

    Deliberately crude (a literal leading-character-run match, not tokenized
    morphology) -- this is the "concrete, testable" half of D-13d's proposed
    separation rule: cheap, auditable, and explicitly a DISPLAY heuristic,
    never a data fix.
    """
    if len(title_a) < min_len or len(title_b) < min_len:
        return False
    return title_a[:min_len] == title_b[:min_len]


def works_related_by_title(work_a: Dict[str, Any], work_b: Dict[str, Any]) -> bool:
    """D-13d's proposed separation rule: SAME non-null author AND (identical
    normalized title -- an undetected alias/duplicate -- OR a shared >=4-char
    normalized-title prefix, e.g. a shared "<author> on " commentary marker).

    This is the display-time alias/containment test the plan asks for. It is
    intentionally conservative (author-gated) precisely because an
    ungated title-prefix match alone is corpus-noisy (verified empirically
    against this asset: many M-source responsa collections share a single
    generic collection title -- e.g. "Responsa of the Geonim" -- across
    dozens of genuinely distinct items with NO author recorded; gating on a
    matching author removes that entire noise class).
    """
    author_a = work_a.get("author")
    author_b = work_b.get("author")
    if not author_a or not author_b or author_a != author_b:
        return False
    na, nb = normalize_title(work_a.get("neutral_title")), normalize_title(work_b.get("neutral_title"))
    if not na or not nb:
        return False
    if na == nb:
        return True
    return titles_share_prefix(na, nb, min_len=4)


def title_similarity_ratio(work_a: Dict[str, Any], work_b: Dict[str, Any]) -> float:
    na, nb = normalize_title(work_a.get("neutral_title")), normalize_title(work_b.get("neutral_title"))
    if not na or not nb:
        return 0.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


# ---------------------------------------------------------------------------
# sys_id derivation. page_id = "{sys_id}_IE{ie_id}_P{page_num}_FL{fl_id}"
# (docs/specs/discovery-sidecar-schema-v1.md; canonical_refs "Measured
# facts"). Verified against a live sample before being trusted as a
# structural invariant (see the self-test in __main__ guard below is NOT
# present -- verification instead happens via the nonzero-nonNone assertion
# in load_claims()).
# ---------------------------------------------------------------------------
_PAGE_ID_SYS_RE = re.compile(r"^(\d+)_IE")


def sys_id_from_page_id(page_id: str) -> Optional[str]:
    m = _PAGE_ID_SYS_RE.match(page_id)
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# A tiny nonzero-assertion ledger -- every measurement that SHOULD be
# nonzero registers itself here; main() exits 1 if anything failed, so a
# silent zero can never reach the owner as a finding (per-task acceptance
# criteria + the T-136-03-02 threat mitigation).
# ---------------------------------------------------------------------------
class NonzeroLedger:
    def __init__(self) -> None:
        self.failures: List[str] = []

    def check(self, label: str, value: int) -> int:
        if value == 0:
            self.failures.append(label)
        return value

    def ok(self) -> bool:
        return not self.failures


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def connect_readonly(path: str) -> sqlite3.Connection:
    if not os.path.exists(path):
        raise FileNotFoundError(f"asset not found: {path}")
    uri = f"file:{path}?mode=ro"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def load_works(conn: sqlite3.Connection) -> Dict[str, Dict[str, Any]]:
    rows = conn.execute(
        "SELECT work_id, canonical_work_id, neutral_title, author, genre, source_corpus "
        "FROM works ORDER BY work_id"
    ).fetchall()
    return {r["work_id"]: dict(r) for r in rows}


def load_claims(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
    """One row per claim, carrying its DISPLAY evidence's presentation fields.

    ``ORDER BY dc.claim_id`` makes downstream iteration order (and therefore
    any Python dict built by insertion) reproducible run-to-run.
    """
    rows = conn.execute(
        """
        SELECT dc.page_id, dc.work_id, dc.claim_id, dc.claim_type, dc.display_evidence_id,
               de.evidence_source, de.confidence_band, de.adjudication_status,
               de.routing_status, de.matched_letters, de.span_start, de.span_end,
               de.aligned_len
        FROM discovery_claim dc
        JOIN discovery_evidence de ON de.evidence_id = dc.display_evidence_id
        ORDER BY dc.claim_id
        """
    ).fetchall()
    claims = [dict(r) for r in rows]
    for c in claims:
        sid = sys_id_from_page_id(c["page_id"])
        if sid is None:
            raise ValueError(
                f"page_id does not match the expected '{{sys_id}}_IE...' shape: {c['page_id']!r}"
            )
        c["sys_id"] = sid
    return claims


def load_human_confirmed_claim_ids(conn: sqlite3.Connection) -> set:
    return {
        r[0]
        for r in conn.execute(
            "SELECT DISTINCT claim_id FROM discovery_evidence "
            "WHERE adjudication_status = 'human_confirmed' ORDER BY claim_id"
        )
    }


def load_kept_tie_pages(conn: sqlite3.Connection, works: Dict[str, Dict[str, Any]]) -> set:
    """(page_id, canonical_work_id) pairs where the routing audit recorded an
    unresolved tie retained FOR that work on that page. ``demoted_work_id`` is
    NULL on every kept_tie row in this asset (D-02b's documented flaw), so
    only the kept-side of the tie is reconstructable here -- see the
    Methodology note in the rendered brief."""
    out = set()
    for row in conn.execute(
        "SELECT page_id, kept_work_id FROM discovery_routing_audit "
        "WHERE decision = 'kept_tie' ORDER BY id"
    ):
        w = works.get(row["kept_work_id"])
        if w is not None:
            out.add((row["page_id"], w["canonical_work_id"]))
    return out


def build_near_tie_competition(
    claims: List[Dict[str, Any]], works: Dict[str, Dict[str, Any]]
) -> set:
    """(page_id, canonical_work_id) pairs where at least one track1_direct
    claim for that (page, canonical work) has an overlapping, near-equal-
    length competing span from a DIFFERENT canonical work on the same page
    -- the other half of gate 3 ("an overlapping near-tie span from another
    canonical work"). Overlap >= 70% of the shorter span AND a length ratio
    >= 0.7 -- a stated, reproducible threshold, not a tuned constant."""
    page_spans: Dict[str, List[Tuple[str, int, int]]] = defaultdict(list)
    for c in claims:
        if c["evidence_source"] == "track1_direct" and c["span_start"] is not None and c["span_end"] is not None:
            w = works.get(c["work_id"])
            if w is None:
                continue
            page_spans[c["page_id"]].append((w["canonical_work_id"], c["span_start"], c["span_end"]))

    out = set()
    for c in claims:
        if c["evidence_source"] != "track1_direct" or c["span_start"] is None:
            continue
        w = works.get(c["work_id"])
        if w is None:
            continue
        canon = w["canonical_work_id"]
        s0, s1 = c["span_start"], c["span_end"]
        my_len = s1 - s0
        if my_len <= 0:
            continue
        for other_canon, os0, os1 in page_spans.get(c["page_id"], []):
            if other_canon == canon:
                continue
            olen = os1 - os0
            if olen <= 0:
                continue
            overlap = min(s1, os1) - max(s0, os0)
            if overlap <= 0:
                continue
            ratio = min(my_len, olen) / max(my_len, olen)
            overlap_frac = overlap / min(my_len, olen)
            if ratio >= 0.7 and overlap_frac >= 0.7:
                out.add((c["page_id"], canon))
                break
    return out


# ---------------------------------------------------------------------------
# Main-pool-rule classification (references/main-pool-rule.md). Unit =
# identification = (sys_id, canonical_work_id). Gates evaluated in order;
# ANY gate sends the identification to "show more" (show_more); passing all
# four (or a human_confirmed override) -> "main".
# ---------------------------------------------------------------------------

def classify_identifications(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    human_confirmed_claim_ids: set,
    kept_tie_pages: set,
    near_tie_pages: set,
    page_norm_letters: Dict[str, int],
) -> Tuple[Dict[Tuple[str, str], List[Dict[str, Any]]], Dict[Tuple[str, str], Tuple[str, str]]]:
    """Returns (identifications, classification) where ``identifications``
    maps (sys_id, canonical_work_id) -> its usable claims, and
    ``classification`` maps the same key -> (bucket, reason).

    "Usable" claims (surface-visible, current product behaviour): the claim's
    DISPLAY evidence has ``routing_status == 'shipped'``, OR ANY evidence row
    for the claim is ``adjudication_status == 'human_confirmed'`` (the D-13g
    fix folded in here -- a human-confirmed row must never be invisible to
    this measurement merely because routing demoted it)."""
    identifications: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for c in claims:
        is_hc = c["claim_id"] in human_confirmed_claim_ids
        if c["routing_status"] != "shipped" and not is_hc:
            continue
        w = works.get(c["work_id"])
        if w is None:
            continue
        c = dict(c)
        c["canonical_work_id"] = w["canonical_work_id"]
        c["is_human_confirmed"] = is_hc
        identifications[(c["sys_id"], w["canonical_work_id"])].append(c)

    classification: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for key, cl in sorted(identifications.items()):
        sys_id, canon = key
        if any(x["is_human_confirmed"] for x in cl):
            classification[key] = ("main", "human_confirmed_override")
            continue
        if not any(x["claim_type"] == "direct_witness" for x in cl):
            classification[key] = ("show_more", "gate1_no_same_work_claim")
            continue
        best = min(band_rank(x["evidence_source"], x["confidence_band"]) for x in cl)
        if best in (_SCREENING_CANON_RANK, _WEAK_RANK):
            classification[key] = ("show_more", "gate2_best_band_weak")
            continue
        pages = {x["page_id"] for x in cl}
        if pages and all(
            (p, canon) in kept_tie_pages or (p, canon) in near_tie_pages for p in pages
        ):
            classification[key] = ("show_more", "gate3_unresolved_competition")
            continue
        if len(pages) == 1:
            page = next(iter(pages))
            mls = [x["matched_letters"] for x in cl if x["evidence_source"] == "track1_direct" and x["matched_letters"] is not None]
            if not mls:
                classification[key] = ("show_more", "gate4_no_coverage_data")
                continue
            pnl = page_norm_letters.get(page)
            cov = compute_page_coverage(max(mls), pnl)
            if cov is None or cov < 0.8:
                classification[key] = ("show_more", "gate4_low_single_page_coverage")
            else:
                classification[key] = ("main", "gate4_single_page_full_coverage")
            continue
        classification[key] = ("main", "multi_folio_agreement")

    return identifications, classification


def pages_needing_coverage(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    human_confirmed_claim_ids: set,
    kept_tie_pages: set,
    near_tie_pages: set,
) -> List[str]:
    """Pre-pass identical to the gate1-3 portion of classify_identifications,
    used ONLY to determine which single pages need a page_norm_letters
    lookup before the real classification runs (so we query fullcorpus.db
    for exactly the pages needed, never the whole 667K-row table)."""
    identifications: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for c in claims:
        is_hc = c["claim_id"] in human_confirmed_claim_ids
        if c["routing_status"] != "shipped" and not is_hc:
            continue
        w = works.get(c["work_id"])
        if w is None:
            continue
        c = dict(c)
        c["canonical_work_id"] = w["canonical_work_id"]
        c["is_human_confirmed"] = is_hc
        identifications[(c["sys_id"], w["canonical_work_id"])].append(c)

    needed = set()
    for key, cl in identifications.items():
        sys_id, canon = key
        if any(x["is_human_confirmed"] for x in cl):
            continue
        if not any(x["claim_type"] == "direct_witness" for x in cl):
            continue
        best = min(band_rank(x["evidence_source"], x["confidence_band"]) for x in cl)
        if best in (_SCREENING_CANON_RANK, _WEAK_RANK):
            continue
        pages = {x["page_id"] for x in cl}
        if pages and all(
            (p, canon) in kept_tie_pages or (p, canon) in near_tie_pages for p in pages
        ):
            continue
        if len(pages) == 1:
            page = next(iter(pages))
            mls = [x["matched_letters"] for x in cl if x["evidence_source"] == "track1_direct" and x["matched_letters"] is not None]
            if mls:
                needed.add(page)
    return sorted(needed)


def load_page_norm_letters(research_db_path: str, page_ids: List[str]) -> Dict[str, int]:
    if not page_ids:
        return {}
    conn = sqlite3.connect(f"file:{research_db_path}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    result: Dict[str, int] = {}
    CHUNK = 400
    try:
        for i in range(0, len(page_ids), CHUNK):
            chunk = page_ids[i : i + CHUNK]
            placeholders = ",".join("?" for _ in chunk)
            for row in conn.execute(
                f"SELECT page_id, text FROM pages WHERE page_id IN ({placeholders})",  # noqa: S608 -- fixed table name, params bound
                chunk,
            ):
                result[row["page_id"]] = norm_stream_letter_count(row["text"])
    finally:
        conn.close()
    return result


# ---------------------------------------------------------------------------
# Identical-span groups (D-13b / D-13d). Population = the "shipped direct
# set": evidence_kind='witness', evidence_source='track1_direct',
# routing_status='shipped', grouped by (page_id, span_start, span_end).
# ---------------------------------------------------------------------------

def load_identical_span_groups(
    conn: sqlite3.Connection, works: Dict[str, Dict[str, Any]]
) -> Dict[Tuple[str, int, int], List[Dict[str, Any]]]:
    rows = conn.execute(
        """
        SELECT de.evidence_id, de.claim_id, dc.page_id AS page_id, dc.work_id,
               de.span_start, de.span_end, de.matched_letters,
               de.evidence_source, de.confidence_band
        FROM discovery_evidence de
        JOIN discovery_claim dc ON dc.claim_id = de.claim_id
        WHERE de.evidence_kind = 'witness' AND de.evidence_source = 'track1_direct'
          AND de.routing_status = 'shipped'
        ORDER BY dc.page_id, de.span_start, de.span_end, de.evidence_id
        """
    ).fetchall()
    groups: Dict[Tuple[str, int, int], List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        d = dict(r)
        w = works.get(d["work_id"])
        d["canonical_work_id"] = w["canonical_work_id"] if w else None
        groups[(d["page_id"], d["span_start"], d["span_end"])].append(d)
    return {k: v for k, v in groups.items() if len(v) >= 2}


def load_libraries_csv(path: str) -> Dict[str, Dict[str, str]]:
    """sys_id -> {shelfmark, library_code, catalogue_text}. Public catalogue
    metadata already displayed throughout the product (masking-safe)."""
    out: Dict[str, Dict[str, str]] = {}
    if not os.path.exists(path):
        return out
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader, None)
        for row in reader:
            if len(row) < 8:
                continue
            sys_id = row[0].strip()
            if not sys_id:
                continue
            call_numbers = row[2]
            shelfmark = call_numbers.split("|")[0].strip() if call_numbers else ""
            out[sys_id] = {
                "shelfmark": shelfmark,
                "library_code": row[3].strip(),
                "catalogue_text": row[7].strip(),
            }
    return out


def best_claim_for_work(
    claims: List[Dict[str, Any]], work_id: str
) -> Optional[Dict[str, Any]]:
    """Deterministic representative shipped claim for a work: highest
    matched_letters first, then lexicographically smallest page_id (a total
    order, never "whichever row the query happens to return first")."""
    candidates = [
        c for c in claims
        if c["work_id"] == work_id and c["routing_status"] == "shipped"
    ]
    if not candidates:
        return None
    return min(
        candidates,
        key=lambda c: (-(c["matched_letters"] or 0), c["page_id"]),
    )


# ---------------------------------------------------------------------------
# D-13e: the middle-bucket "not otherwise reachable" count.
# ---------------------------------------------------------------------------

def compute_d13e(
    conn: sqlite3.Connection,
    span_groups: Dict[Tuple[str, int, int], List[Dict[str, Any]]],
    classification: Dict[Tuple[str, str], Tuple[str, str]],
) -> Dict[str, Any]:
    # D-13d population within the span groups: groups with >= 2 DIFFERENT
    # canonical works (a group whose members share ONE canonical id is a
    # D-13a duplicate-rendering case, not a D-13d generic-shared-text case).
    generic_claims: List[Tuple[str, str]] = []  # (sys_id, canonical_work_id) per affected claim
    for key, members in span_groups.items():
        canons = {m["canonical_work_id"] for m in members if m["canonical_work_id"]}
        if len(canons) < 2:
            continue
        for m in members:
            if not m["canonical_work_id"]:
                continue
            sid = sys_id_from_page_id(m["page_id"])
            generic_claims.append((sid, m["canonical_work_id"]))

    not_reachable_generic = 0
    overlap_generic = 0
    for k in generic_claims:
        bucket, _reason = classification.get(k, ("show_more", "not_classified"))
        if bucket == "main":
            not_reachable_generic += 1
        else:
            overlap_generic += 1

    # Related-pages population (D-11a): shipped shared_text evidence never
    # maps onto a work identification at all, so it can NEVER be reached via
    # the per-work "show more matches" toggle -- the entire population is
    # "not otherwise reachable". Counted as DIRECTED (anchor, opposite) page
    # pairs, matching D-11a's own reported figure.
    (related_pairs,) = conn.execute(
        "SELECT COUNT(*) FROM (SELECT DISTINCT a_page_id, other_page_id FROM discovery_evidence "
        "WHERE evidence_kind = 'shared_text' AND routing_status = 'shipped')"
    ).fetchone()

    total_middle = len(generic_claims) + related_pairs
    not_reachable = not_reachable_generic + related_pairs
    overlap = overlap_generic

    return {
        "generic_claims_total": len(generic_claims),
        "generic_not_reachable": not_reachable_generic,
        "generic_overlap": overlap_generic,
        "related_pairs": related_pairs,
        "total_middle_bucket": total_middle,
        "not_reachable_total": not_reachable,
        "overlap_total": overlap,
    }


# ---------------------------------------------------------------------------
# D-16: relation distribution.
# ---------------------------------------------------------------------------

def compute_d16(
    conn: sqlite3.Connection,
    claims: List[Dict[str, Any]],
    classification: Dict[Tuple[str, str], Tuple[str, str]],
) -> Dict[str, Any]:
    (corpus_wide,) = ([dict(r) for r in conn.execute(
        "SELECT claim_type, COUNT(*) AS n FROM discovery_claim GROUP BY claim_type ORDER BY claim_type"
    )],)
    shipped = [
        dict(r)
        for r in conn.execute(
            """
            SELECT dc.claim_type, COUNT(*) AS n
            FROM discovery_claim dc
            JOIN discovery_evidence de ON de.evidence_id = dc.display_evidence_id
            WHERE de.routing_status = 'shipped'
            GROUP BY dc.claim_type ORDER BY dc.claim_type
            """
        )
    ]

    main_pool_relation: Dict[str, int] = defaultdict(int)
    for c in claims:
        sid, wid = c["sys_id"], c.get("canonical_work_id")
        if wid is None:
            continue
        bucket, _ = classification.get((sid, wid), ("show_more", "n/a"))
        if bucket == "main":
            main_pool_relation[c["claim_type"]] += 1

    return {
        "corpus_wide": corpus_wide,
        "shipped": shipped,
        "main_pool": dict(sorted(main_pool_relation.items())),
    }


# ---------------------------------------------------------------------------
# D-13c: short-evidence thresholds.
# ---------------------------------------------------------------------------

def compute_d13c(
    conn: sqlite3.Connection,
    classification: Dict[Tuple[str, str], Tuple[str, str]],
    claims: List[Dict[str, Any]],
) -> Dict[str, Any]:
    thresholds = (50, 100, 150, 200)

    direct_lengths = [
        r[0]
        for r in conn.execute(
            "SELECT matched_letters FROM discovery_evidence "
            "WHERE evidence_source = 'track1_direct' AND routing_status = 'shipped' "
            "AND matched_letters IS NOT NULL"
        )
    ]
    propagated_lengths = [
        r[0]
        for r in conn.execute(
            "SELECT aligned_len FROM discovery_evidence "
            "WHERE evidence_source = 'propagated' AND evidence_kind = 'shared_text' "
            "AND routing_status = 'shipped' AND aligned_len IS NOT NULL"
        )
    ]

    direct_cumulative = {t: sum(1 for v in direct_lengths if v < t) for t in thresholds}
    propagated_cumulative = {t: sum(1 for v in propagated_lengths if v < t) for t in thresholds}

    thinnest_direct = min(direct_lengths) if direct_lengths else None

    # Short direct rows (< 150 matched letters) that are nonetheless part of
    # a MAIN identification via multi-folio agreement specifically (the
    # honest counter-argument the owner already accepted: a short liturgical
    # passage may be exactly the correct identification for a prayer book).
    short_in_main_multi_folio = 0
    for c in claims:
        if c["evidence_source"] != "track1_direct" or c["matched_letters"] is None:
            continue
        if c["matched_letters"] >= 150:
            continue
        sid, wid = c["sys_id"], c.get("canonical_work_id")
        if wid is None:
            continue
        bucket, reason = classification.get((sid, wid), ("show_more", "n/a"))
        if bucket == "main" and reason == "multi_folio_agreement":
            short_in_main_multi_folio += 1

    return {
        "thresholds": thresholds,
        "direct_total": len(direct_lengths),
        "direct_cumulative": direct_cumulative,
        "propagated_total": len(propagated_lengths),
        "propagated_cumulative": propagated_cumulative,
        "thinnest_direct": thinnest_direct,
        "short_in_main_multi_folio": short_in_main_multi_folio,
    }


# ---------------------------------------------------------------------------
# D-13b: lead-attribution tie residual.
# ---------------------------------------------------------------------------

def compute_d13b(span_groups: Dict[Tuple[str, int, int], List[Dict[str, Any]]]) -> Dict[str, Any]:
    total_groups = len(span_groups)
    total_claims = sum(len(v) for v in span_groups.values())
    tied_groups = 0
    tied_claims = 0
    for members in span_groups.values():
        ranks = [band_rank(m["evidence_source"], m["confidence_band"]) for m in members]
        best = min(ranks)
        n_at_best = sum(1 for r in ranks if r == best)
        if n_at_best >= 2:
            tied_groups += 1
            tied_claims += n_at_best
    return {
        "total_groups": total_groups,
        "total_claims": total_claims,
        "tied_after_band_rank_groups": tied_groups,
        "tied_after_band_rank_claims": tied_claims,
    }


# ---------------------------------------------------------------------------
# D-13d: granularity separation quantification + worked example.
# ---------------------------------------------------------------------------

WORKED_EXAMPLE_WORK_IDS = ("w000171", "w001281")
# The SPECIFIC instance named in 136-CONTEXT.md D-13d (T-S Misc. 12.31.14,
# sys_id 990051079570205171) -- verified directly against libraries.csv
# during this plan's research. Preferred over any OTHER co-occurrence of the
# same work pair (this pair also co-occurs across dozens of pages of a
# separate, much larger Rashi-commentary manuscript, sys_id
# 990000852430205171 -- a real finding in its own right, but not the
# documented worked example the owner has already reviewed in the mockup).
WORKED_EXAMPLE_SPAN = (0, 962)


def compute_d13d(
    span_groups: Dict[Tuple[str, int, int], List[Dict[str, Any]]],
    works: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    diff_canon_groups = []
    for key, members in span_groups.items():
        canons = {m["canonical_work_id"] for m in members if m["canonical_work_id"]}
        if len(canons) >= 2:
            diff_canon_groups.append((key, members))

    collapse_candidate_groups = []
    generic_groups = []
    for key, members in diff_canon_groups:
        work_ids = sorted({m["work_id"] for m in members})
        found = False
        for i in range(len(work_ids)):
            for j in range(i + 1, len(work_ids)):
                wa, wb = works.get(work_ids[i]), works.get(work_ids[j])
                if wa and wb and works_related_by_title(wa, wb):
                    found = True
                    break
            if found:
                break
        if found:
            collapse_candidate_groups.append((key, members))
        else:
            generic_groups.append((key, members))

    worked_example = None
    exact_example_entry = None
    fallback_example_entry = None
    for key, members in diff_canon_groups:
        work_ids = {m["work_id"] for m in members}
        if not set(WORKED_EXAMPLE_WORK_IDS).issubset(work_ids):
            continue
        if fallback_example_entry is None:
            fallback_example_entry = (key, work_ids, members)
        _page_id, span_start, span_end = key
        if (span_start, span_end) == WORKED_EXAMPLE_SPAN:
            exact_example_entry = (key, work_ids, members)
            break
    chosen_entry = exact_example_entry or fallback_example_entry

    if chosen_entry is not None:
        key, work_ids, members = chosen_entry
        page_id, span_start, span_end = key
        worked_example = {
            "page_id": page_id,
            "sys_id": sys_id_from_page_id(page_id),
            "span_start": span_start,
            "span_end": span_end,
            "matched_letters": members[0]["matched_letters"],
            "works": [
                {
                    "work_id": wid,
                    "canonical_work_id": works[wid]["canonical_work_id"],
                    "neutral_title": works[wid]["neutral_title"],
                    "author": works[wid]["author"],
                }
                for wid in sorted(work_ids)
                if wid in works
            ],
        }

    return {
        "diff_canon_groups_total": len(diff_canon_groups),
        "diff_canon_claims_total": sum(len(v) for _, v in diff_canon_groups),
        "collapse_candidate_groups": len(collapse_candidate_groups),
        "collapse_candidate_claims": sum(len(v) for _, v in collapse_candidate_groups),
        "generic_groups": len(generic_groups),
        "generic_claims": sum(len(v) for _, v in generic_groups),
        "worked_example": worked_example,
        "collapse_candidate_keys": [k for k, _ in collapse_candidate_groups],
    }


# ---------------------------------------------------------------------------
# Novelty hard-case candidate set (feeds Task 3). Zero model calls -- pure
# string/title comparison, reusing works_related_by_title /
# title_similarity_ratio from the D-13d machinery above.
# ---------------------------------------------------------------------------

def _evenly_spaced_indices(n: int, k: int) -> List[int]:
    """``k`` evenly-spaced 0-based indices across a list of length ``n``,
    inclusive of the first and last positions when ``k >= 2``. Deterministic
    (round-half-to-even via Python's ``round``, but every input here is a
    fixed, already-sorted pool, so re-running against the same asset always
    yields the same indices) -- used to build the Class 1-3 IDENTITY
    spot-check (owner-authorized labelling restructure) by sampling spread
    positions across each class's existing deterministic candidate order,
    rather than hand-picking specific cases -- so the selection stays a
    reproducible function of the data, never a manually pinned case list."""
    if k <= 0 or n <= 0:
        return []
    if k == 1:
        return [0]
    if k >= n:
        return list(range(n))
    seen: List[int] = []
    for i in range(k):
        idx = round(i * (n - 1) / (k - 1))
        if idx not in seen:
            seen.append(idx)
    return seen

def select_alias_pair_candidates(works: Dict[str, Dict[str, Any]]) -> List[Tuple[Dict, Dict]]:
    """Class 2 (alias pairs): cluster works by (author, normalized title);
    a cluster of EXACTLY 2 members with DIFFERENT canonical_work_id is a
    clean alias/duplicate candidate. Clusters of 3+ are excluded: verified
    empirically against this asset that large clusters are generic
    multi-item collection titles (e.g. many distinct M-source responsa
    items sharing one collector's name as both "author" and title stem),
    which is corpus noise for this purpose, not an alias signal."""
    clusters: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for w in sorted(works.values(), key=lambda w: w["work_id"]):
        author = w.get("author")
        nt = normalize_title(w.get("neutral_title"))
        if author and nt:
            clusters[(author, nt)].append(w)
    pairs = []
    for key in sorted(clusters.keys()):
        members = clusters[key]
        if len(members) == 2 and members[0]["canonical_work_id"] != members[1]["canonical_work_id"]:
            pairs.append((members[0], members[1]))
    return pairs


def select_near_miss_candidates(
    works: Dict[str, Dict[str, Any]], min_ratio: float = 0.85, max_ratio: float = 0.999
) -> List[Tuple[Dict, Dict, float]]:
    """Class 1 (near-miss titles): same-author pairs whose normalized titles
    are highly similar (SequenceMatcher ratio in [min_ratio, max_ratio)) but
    NOT identical -- a string-comparison-based check can be fooled either
    way (treating them as the same, or missing that they are related)."""
    items = sorted(works.values(), key=lambda w: w["work_id"])
    out = []
    n = len(items)
    for i in range(n):
        wa = items[i]
        na = normalize_title(wa.get("neutral_title"))
        if not na or not wa.get("author"):
            continue
        for j in range(i + 1, n):
            wb = items[j]
            if wa["canonical_work_id"] == wb["canonical_work_id"]:
                continue
            if wb.get("author") != wa.get("author"):
                continue
            nb = normalize_title(wb.get("neutral_title"))
            if not nb or na == nb:
                continue
            ratio = difflib.SequenceMatcher(None, na, nb).ratio()
            if min_ratio <= ratio < max_ratio:
                out.append((wa, wb, ratio))
    out.sort(key=lambda t: (-t[2], t[0]["work_id"], t[1]["work_id"]))
    return out


# ---------------------------------------------------------------------------
# Classes 4 and 5 -- OWNER-AUTHORIZED scope extension recorded in
# 136-GATE1-DECISIONS.md (item C). Added AFTER the original 52 candidates,
# at the gate, because the original three classes were selected adversarially
# to a STRING heuristic, not to an LLM -- these two classes exist so the
# measured error rate is not flattered by cases an LLM finds easy. Same
# discipline as classes 1-3: zero model calls, pure metadata/string
# comparison, fully deterministic (every selection below is either a total
# order over a stable key or an explicit sort before slicing).
# ---------------------------------------------------------------------------

def select_terse_catalogue_candidates(
    claims: List[Dict[str, Any]],
    libraries: Dict[str, Dict[str, str]],
    cap: int = 15,
    max_len: int = 20,
) -> List[Dict[str, Any]]:
    """Class 4 (terse or missing catalogue identification text) -- SUPERSEDED
    by the ruling-J redesign (136-03 continuation, 2026-08-02;
    136-GATE1-DECISIONS.md § J): folded into Arm 1's ``terse_catalogue``
    stratum (see ``_residual_stratum``/``compute_claim_source_signals``,
    which reuse this function's ``max_len`` threshold directly as
    ``_CATALOGUE_TERSE_MAX_LEN``), now checked against the REAL bib/PGP/FGP
    signal rather than catalogue text alone. Left DEFINED but UNUSED by
    ``build_hardcases`` (not deleted) as a standalone, catalogue-text-only
    reference implementation.

    Shipped ``direct_witness`` claims on a manuscript (sys_id) whose OWN
    ``libraries.csv`` ``catalogue_text`` field is either entirely absent
    (empty string, or no row at all) or so short (<= ``max_len`` characters,
    nonzero) that a title comparison has almost nothing to work with.

    One representative claim per sys_id: the highest ``matched_letters``,
    then lexicographically smallest ``page_id``, then ``work_id`` -- a total
    order, never "whichever row the query happens to return first".
    Candidate sys_ids are then sorted by (catalogue-text length ascending,
    sys_id) so the emptiest cases surface first, and capped at ``cap``.
    """
    best_by_sys: Dict[str, Dict[str, Any]] = {}
    for c in claims:
        if c.get("claim_type") != "direct_witness" or c.get("routing_status") != "shipped":
            continue
        sid = c["sys_id"]
        cat = libraries.get(sid, {}).get("catalogue_text", "")
        if len(cat) > max_len:
            continue
        key = (-(c["matched_letters"] or 0), c["page_id"], c["work_id"])
        prev = best_by_sys.get(sid)
        if prev is None or key < prev["_key"]:
            d = dict(c)
            d["_key"] = key
            best_by_sys[sid] = d

    ordered_sys_ids = sorted(
        best_by_sys.keys(),
        key=lambda sid: (len(libraries.get(sid, {}).get("catalogue_text", "")), sid),
    )
    return [{"sys_id": sid, "claim": best_by_sys[sid]} for sid in ordered_sys_ids[:cap]]


def select_generic_collection_candidates(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    cap: int = 15,
    min_cluster_size: int = 3,
) -> List[Dict[str, Any]]:
    """Class 5 (generic collection works) -- DROPPED from ``build_hardcases``
    by the ruling-J redesign (136-03 continuation, 2026-08-02;
    136-GATE1-DECISIONS.md § J): no owner ruling exists for any specific
    Class 5 case, and its collection-level-identity question does not
    correspond to any of the three source-stratified arms. Left DEFINED but
    UNUSED (not deleted) in case a future session revisits the
    collection-level question on its own terms -- see the "kept, folded,
    dropped" accounting in ``build_hardcases`` and in
    ``136-NOVELTY-HARDCASES.md``'s own intro.

    (author, normalized-title) clusters of >= ``min_cluster_size`` works
    carrying >= 2 distinct ``canonical_work_id``s -- precisely the large
    generic-collection-title clusters ``select_alias_pair_candidates``
    (above) explicitly EXCLUDES as corpus noise (its own docstring: "large
    clusters are generic multi-item collection titles -- e.g. many distinct
    M-source responsa items sharing one collector's name as both author and
    title stem"). Here they ARE the signal, not the noise: for "already
    recorded" to mean anything for a single witness of such a collection is
    genuinely ill-defined, not merely hard to string-match.

    For each such cluster this returns REAL shipped ``direct_witness``
    manuscripts (best claim per sys_id, by matched_letters descending then
    page_id then work_id -- a total order), round-robin across clusters
    (largest cluster first, ties broken by the cluster key) one at a time so
    no single collection crowds out the others, capped at ``cap``.
    """
    clusters: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for w in sorted(works.values(), key=lambda w: w["work_id"]):
        author = w.get("author")
        nt = normalize_title(w.get("neutral_title"))
        if author and nt:
            clusters[(author, nt)].append(w)

    big_clusters: List[Tuple[int, Tuple[str, str], List[Dict[str, Any]]]] = []
    for key, members in clusters.items():
        if len(members) >= min_cluster_size:
            canon_ids = {m["canonical_work_id"] for m in members}
            if len(canon_ids) >= 2:
                big_clusters.append((len(members), key, members))
    big_clusters.sort(key=lambda t: (-t[0], t[1]))

    claims_by_workid: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in claims:
        if c.get("claim_type") == "direct_witness" and c.get("routing_status") == "shipped":
            claims_by_workid[c["work_id"]].append(c)

    per_cluster_reps: List[Tuple[Tuple[str, str], List[Dict[str, Any]], List[Dict[str, Any]]]] = []
    for _size, key, members in big_clusters:
        work_ids = sorted({m["work_id"] for m in members})
        by_sys: Dict[str, Dict[str, Any]] = {}
        for wid in work_ids:
            for c in claims_by_workid.get(wid, []):
                sid = c["sys_id"]
                cand_key = (-(c["matched_letters"] or 0), c["page_id"], c["work_id"])
                prev = by_sys.get(sid)
                if prev is None or cand_key < prev["_key"]:
                    d = dict(c)
                    d["_key"] = cand_key
                    by_sys[sid] = d
        reps = sorted(by_sys.values(), key=lambda d: d["_key"])
        per_cluster_reps.append((key, members, reps))

    out: List[Dict[str, Any]] = []
    round_idx = 0
    while len(out) < cap:
        added_this_round = False
        for key, members, reps in per_cluster_reps:
            if round_idx < len(reps):
                out.append({"cluster_key": key, "members": members, "claim": reps[round_idx]})
                added_this_round = True
                if len(out) >= cap:
                    break
        if not added_this_round:
            break
        round_idx += 1
    return out[:cap]


def select_catalogue_divergence_candidates(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    libraries: Dict[str, Dict[str, str]],
    cap: int = 15,
    min_divergent_title_len: int = 6,
) -> List[Dict[str, Any]]:
    """Class 6 (catalogue divergence, owner decision E -- 136-GATE1-DECISIONS.md
    item E): shipped ``direct_witness`` claims on a manuscript whose OWN
    ``libraries.csv`` catalogue-identification text, once normalized, contains
    the normalized title of a DIFFERENT work than the one this claim
    identifies, where that other work is NOT a D-13d granularity variant of
    the claimed work (``works_related_by_title`` returns False for the pair)
    -- decision E's ``diverges`` shade: "an aid ties F to a different work
    that is NOT a granularity variant". This class has ZERO representation
    across Classes 1-5 and is exactly the shade decision E's rationale names
    as currently inflating the novelty (``not_in_finding_aids``) count.

    Selection is pure deterministic string containment over data already
    loaded for Classes 1-5 -- ZERO model calls, same discipline as every
    other class. A manuscript/work pair is skipped as a candidate when the
    claimed work's OWN normalized title is already found in the catalogue
    text (that is agreement, however partial, not divergence). Only work
    titles at least ``min_divergent_title_len`` normalized characters long
    are searched for containment -- a short title would match too much
    catalogue prose to be a trustworthy divergence signal (the same
    reasoning as D-13d's >=4-char PREFIX bar, raised here because this is
    FULL-title containment against free-text catalogue prose, not a
    title-to-title prefix comparison).

    One candidate per manuscript (sys_id): across all of that manuscript's
    shipped claimed works, the one whose best divergent match has the
    LONGEST normalized divergent title wins (a longer matched title is less
    likely to be a spurious substring hit). Candidates are then GROUPED by
    the divergent work they name and round-robined across groups (largest
    group first, ties broken by work_id -- the identical discipline
    ``select_generic_collection_candidates`` above uses for its clusters):
    a single widely-quoted title (e.g. a popular ethical/halakhic work named
    in passing across many unrelated catalogue entries) would otherwise
    crowd out every other divergence pattern, exactly the diversity problem
    that function's own docstring names. Deterministic, no randomness, no
    sampling; re-running against the same asset reproduces the identical
    candidate list byte-for-byte.
    """
    best_by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for c in claims:
        if c.get("claim_type") != "direct_witness" or c.get("routing_status") != "shipped":
            continue
        key = (c["sys_id"], c["work_id"])
        cand_key = (-(c["matched_letters"] or 0), c["page_id"])
        prev = best_by_key.get(key)
        if prev is None or cand_key < prev["_key"]:
            d = dict(c)
            d["_key"] = cand_key
            best_by_key[key] = d

    title_index: List[Tuple[Dict[str, Any], str]] = []
    for w in sorted(works.values(), key=lambda w: w["work_id"]):
        nt = normalize_title(w.get("neutral_title"))
        if len(nt) >= min_divergent_title_len:
            title_index.append((w, nt))

    sys_ids = sorted({sid for sid, _wid in best_by_key.keys()})
    per_sys_best: List[Dict[str, Any]] = []
    for sid in sys_ids:
        cat_text = libraries.get(sid, {}).get("catalogue_text", "")
        if not cat_text:
            continue
        cat_norm = normalize_title(cat_text)
        if not cat_norm:
            continue
        work_ids_for_sys = sorted({wid for s, wid in best_by_key.keys() if s == sid})
        best_case: Optional[Tuple[int, str, Dict[str, Any], Dict[str, Any], Dict[str, Any]]] = None
        for wid in work_ids_for_sys:
            claim = best_by_key[(sid, wid)]
            w_claimed = works.get(wid)
            if w_claimed is None:
                continue
            claimed_norm = normalize_title(w_claimed.get("neutral_title"))
            if claimed_norm and claimed_norm in cat_norm:
                continue  # catalogue text already seems to name this exact claim -- not a divergence
            for o, nt in title_index:
                if o["canonical_work_id"] == w_claimed["canonical_work_id"]:
                    continue
                if nt not in cat_norm:
                    continue
                if works_related_by_title(w_claimed, o):
                    continue  # D-13d granularity variant, not a genuine divergence
                cand = (len(nt), wid, w_claimed, o, claim)
                if best_case is None or cand[0] > best_case[0]:
                    best_case = cand
        if best_case is not None:
            _mlen, _wid, w_claimed, o, claim = best_case
            per_sys_best.append({
                "sys_id": sid,
                "claim": claim,
                "claimed_work": w_claimed,
                "divergent_work": o,
            })

    # Round-robin across DISTINCT divergent-work groups (largest group first,
    # ties by work_id) -- same discipline as select_generic_collection_candidates
    # above, so one frequently-quoted title cannot crowd out every other
    # divergence pattern in the capped list.
    by_divergent: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for entry in per_sys_best:
        by_divergent[entry["divergent_work"]["work_id"]].append(entry)
    groups: List[Tuple[int, str, List[Dict[str, Any]]]] = []
    for wid, entries in by_divergent.items():
        entries.sort(key=lambda e: (e["sys_id"], e["claim"]["work_id"]))
        groups.append((len(entries), wid, entries))
    groups.sort(key=lambda t: (-t[0], t[1]))

    out: List[Dict[str, Any]] = []
    round_idx = 0
    while len(out) < cap:
        added_this_round = False
        for _size, _wid, entries in groups:
            if round_idx < len(entries):
                out.append(entries[round_idx])
                added_this_round = True
                if len(out) >= cap:
                    break
        if not added_this_round:
            break
        round_idx += 1
    return out[:cap]


# ---------------------------------------------------------------------------
# Liturgical-container detection (owner ruling H, 136-GATE1-DECISIONS.md § H;
# 136-NOVELTY-PRIOR-ART.md § 4's worked examples: `יוצר ח פסח` where the aid
# names `מחזור מנהג אשכנז לשלש רגלים`; `יוצרות לשבתות` where the aid names
# `סדור מנהג אשכנז המזרחי`). Both worked examples share ONE precise shape:
# a container NOUN (מחזור/סדור/סידור) immediately followed by `מנהג`
# ("the custom/rite of ...") -- i.e. the aid names a SPECIFIC, NAMED
# standard-rite prayer-book, not merely a generic liturgical-genre
# CLASSIFICATION TAG. This distinction is load-bearing and was verified
# empirically against the live asset before being pinned here: a bare
# container-noun substring test (no `מנהג` qualifier) fires on catalogue
# TAGS like "תוספות של סידור" ("Siddur additions" -- a structural
# classification label, not a named rite book) at a rate that would dilute
# this class with weaker cases; requiring the `מנהג` collocation isolates
# the genuinely NAMED-rite shape the owner's own worked examples describe.
# Defined HERE independently of same_work_spike/probe/scripts/title_gate.py's
# gitignored GENRE_OF/GENERIC_TOKENS vocabularies, which this project's
# committed code may never depend on (see the module docstring's
# masking/import discipline).
# ---------------------------------------------------------------------------
_LITURGICAL_CONTAINER_RITE_RE = re.compile(r"(מחזור|סדור|סידור)\s*מנהג")


def select_liturgical_container_candidates(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    libraries: Dict[str, Dict[str, str]],
    cap: int = 12,
) -> List[Dict[str, Any]]:
    """Class 7 (liturgical-container predictability, owner rulings H and I --
    136-GATE1-DECISIONS.md §§ H, I; 136-NOVELTY-PRIOR-ART.md § 4) -- SUPERSEDED
    by the ruling-J redesign (136-03 continuation, 2026-08-02;
    136-GATE1-DECISIONS.md § J): folded into Arm 1's ``container_predicts``
    stratum (``_residual_stratum``/``compute_claim_source_signals`` reuse
    ``_LITURGICAL_CONTAINER_RITE_RE`` directly). Left DEFINED but UNUSED by
    ``build_hardcases`` (not deleted) as a standalone, catalogue-text-only
    reference implementation.

    Shipped ``direct_witness`` claims where the manuscript's OWN ``libraries.csv``
    catalogue-identification text names a SPECIFIC, NAMED standard-rite
    prayer-book/cycle (the ``_LITURGICAL_CONTAINER_RITE_RE`` collocation --
    a container noun immediately followed by ``מנהג``) whose standard,
    predictable content plausibly includes the claimed unit, without the
    catalogue text ever naming that unit itself. This is the shape the
    prior-art research pass measured as 1,327 of 20,410 in-scope rows
    (6.5%) under the OLDER five-way title-gate vocabulary's ``witness``
    verdict, which owner ruling H adopts as a shade under the non-colliding
    name ``container_predicts`` (see SHADE_VOCABULARY above -- "witness"
    already names five OTHER distinct concepts in this project).

    Measured against the live asset (verified before this selector was
    pinned): the dominant real-world instance of this shape is a
    Genizah leaf catalogued as a named-rite siddur/machzor carrying a
    Psalm, or another fixed liturgical component, as its claimed
    identification -- a standard prayer-rite predictably includes specific
    Psalms/blessings as part of its fixed liturgy, so a novelty check that
    only reads the container's OWN name (never naming "Psalm 145"
    specifically) would otherwise misfile that claim as `fills_gap`.

    Selection is pure deterministic string containment, the identical
    discipline as ``select_catalogue_divergence_candidates`` above -- ZERO
    model calls, fully reproducible. A candidate requires BOTH:
    (1) the catalogue text (normalized) matches the named-rite container
        collocation;
    (2) the claimed work's own FULL normalized title is NOT already
        contained in the catalogue text (that would be `confirms` -- the
        aid already names the specific unit -- not a container relationship
        at all).

    One candidate per manuscript (sys_id): the claim with the most
    ``matched_letters``, then lexicographically smallest ``page_id`` and
    ``work_id`` -- a total order, never "whichever row the query happens to
    return first". Candidates are then GROUPED by the claimed work and
    round-robined across groups (largest group first, ties by work_id) --
    the identical discipline ``select_generic_collection_candidates`` and
    ``select_catalogue_divergence_candidates`` above use for their own
    clusters: a single dominant liturgical unit (measured: Psalms, by a wide
    margin) would otherwise crowd out every other container-predicted work
    in the capped list. Deterministic, no randomness, no sampling;
    re-running against the same asset reproduces the identical candidate
    list byte-for-byte.
    """
    best_by_sys: Dict[str, Dict[str, Any]] = {}
    container_phrase_by_sys: Dict[str, str] = {}
    for c in claims:
        if c.get("claim_type") != "direct_witness" or c.get("routing_status") != "shipped":
            continue
        sid = c["sys_id"]
        cat_text = libraries.get(sid, {}).get("catalogue_text", "")
        if not cat_text:
            continue
        cat_norm = normalize_title(cat_text)
        match = _LITURGICAL_CONTAINER_RITE_RE.search(cat_norm) if cat_norm else None
        if match is None:
            continue
        w = works.get(c["work_id"])
        if w is None:
            continue
        claimed_norm = normalize_title(w.get("neutral_title"))
        if not claimed_norm:
            continue
        if claimed_norm in cat_norm:
            continue  # the aid already names this exact unit -- confirms, not a container relationship
        key = (-(c["matched_letters"] or 0), c["page_id"], c["work_id"])
        prev = best_by_sys.get(sid)
        if prev is None or key < prev["_key"]:
            d = dict(c)
            d["_key"] = key
            best_by_sys[sid] = d
            container_phrase_by_sys[sid] = match.group(0)

    by_work: Dict[str, List[Tuple[str, Dict[str, Any]]]] = defaultdict(list)
    for sid, entry in best_by_sys.items():
        by_work[entry["work_id"]].append((sid, entry))
    groups: List[Tuple[int, str, List[Tuple[str, Dict[str, Any]]]]] = []
    for wid, entries in by_work.items():
        entries.sort(key=lambda e: e[0])  # sys_id
        groups.append((len(entries), wid, entries))
    groups.sort(key=lambda t: (-t[0], t[1]))

    out: List[Dict[str, Any]] = []
    round_idx = 0
    while len(out) < cap:
        added_this_round = False
        for _size, _wid, entries in groups:
            if round_idx < len(entries):
                sid, entry = entries[round_idx]
                out.append({
                    "sys_id": sid,
                    "claim": entry,
                    "container_phrase": container_phrase_by_sys[sid],
                })
                added_this_round = True
                if len(out) >= cap:
                    break
        if not added_this_round:
            break
        round_idx += 1
    return out[:cap]


# ---------------------------------------------------------------------------
# Ruling J -- three-arm, SOURCE-STRATIFIED novelty sample (136-03 continuation,
# 2026-08-02; 136-GATE1-DECISIONS.md § J; 136-NOVELTY-PRIOR-ART.md §§ 6-7).
#
# The prior-art pass found that every selector above (Classes 4-7) reads
# EXACTLY ONE field -- libraries.csv column 7 -- and therefore has ZERO
# representation of the bib/PGP/FGP failure modes Codex measured as most
# damaging in `gen2_novelty_gate.py` (3,688 `published_full` false-knowns,
# 2,014 PGP false-knowns [942 sole-source], FGP's 1,177-known/9,373-fail
# split). A gate could score perfectly on the pre-J pool and still reproduce
# every one of those defects in production. This section reads the REAL
# checked-source sidecars (fist_data/fjms_enrichment.db `bibliography` +
# `catalog`, pgp_data/pgp.db, fgp_data/fgp_transcriptions.db -- the same
# families NOVEL-01 names and gen2_novelty_gate.py/discovery_identified_gate.py
# already read) joined to the live discovery-v1 asset's shipped claims, and
# builds THREE arms instead of one flat pool:
#
#   Arm 1 -- RESIDUAL: no source's mechanical name-match connects the claim,
#            but >=1 source has non-empty text -- these are the rows that
#            WOULD reach the model under ruling J's funnel-first design.
#            Stratified by which source(s) supplied the unmatched text:
#            bib-sole / pgp-sole / fgp-sole / catalogue-sole / multi-source,
#            PLUS two shape-based strata folding in the former Class 4
#            (terse/absent catalogue text) and Class 7 (a liturgical
#            container collocation) -- per the owner's own instruction, these
#            are STRATA of the residual, not a separate exercise: a
#            container-only machzor title HAS text and FAILS name-match, so
#            it belongs in Arm 1 by construction.
#   Arm 2 -- HEURISTIC-DEMOTED: >=1 source's mechanical test SUCCEEDS, so the
#            funnel marks the claim "already recorded" and it NEVER reaches
#            the model at all (ruling J's own point: the funnel only ever
#            demotes, never promotes -- a false positive here is permanent).
#            Oversamples the two Codex-flagged populations: bib
#            `published_full`-sole and PGP-sole demotions.
#   Arm 3 -- NO-SOURCE-TEXT: not one of the four checked-source families has
#            ANY text at all for this manuscript -- these ship as novelty
#            candidates with NO verdict from any source, by construction (the
#            question this arm answers is whether that bypass is safe, not
#            "what shade is this").
#
# SAMPLING-ONLY APPROXIMATION, stated plainly: the name-match test below is a
# SIMPLIFIED, faithfulness-checked port of
# same_work_spike/probe/scripts/bib_gate.py's / discovery_identified_gate.py's
# token-matching (heb_tokens / prefix-stripped variants / >=2 distinct
# matched tokens with >=1 non-generic-liturgical token) -- committed here
# because that tree is gitignored and this script must not import it (the
# same discipline this module's own D-13c/gate-4 page-coverage port already
# documents for build_discovery_sidecar.py). It intentionally does NOT
# reproduce bib_gate.py's third-tier `known_bib_genre` (the GENRE-EDITION
# heuristic, which needs the full BRIDGE/GENRES vocabulary) -- that tier
# collapses here into "present, non-decisive", which is the conservative
# (never-over-claims-a-match) direction for a SAMPLING instrument. This is a
# labelling-instrument approximation, not the production funnel -- plan
# 136-04 owns the real, fully-faithful implementation.
# ---------------------------------------------------------------------------

_HEB_TOKEN_RE = re.compile(r"[א-ת]+")
# Ported verbatim from same_work_spike/probe/scripts/bib_gate.py (STOP/WEAK,
# 2026-07 vintage) -- short, closed Hebrew stopword/genre-word lists, not
# gitignored business logic; cited here rather than imported per this
# module's own "port, never import same_work_spike" discipline.
_STOP_TOKENS = {
    'בר', 'בן', 'בת', 'רבי', 'ר', 'רב', 'של', 'על', 'אל', 'מן', 'עם',
    'לא', 'ידוע', 'מחבר', 'נוסח', 'קטע', 'קטעים', 'חלק', 'אבו', 'מר',
    'לפי', 'פי', 'מאת', 'ליד', 'בין', 'או', 'גם', 'את', 'זה', 'עוד',
}
_WEAK_TOKENS = {
    'פיוט', 'פיוטים', 'פיוטי', 'יוצר', 'יוצרות', 'קדושתא', 'קדושתאות',
    'קרובות', 'קרובה', 'סליחות', 'סליחה', 'קינות', 'קינה', 'תפילה',
    'תפילות', 'ברכה', 'ברכות', 'ברכת', 'מחזור', 'סידור', 'תרגום',
    'מדרש', 'פירוש', 'הלכות', 'שיר', 'שירי', 'שירים', 'שירת', 'תוספת',
    'סדר', 'עבודה', 'שבת', 'שבתות', 'מועדים', 'מועדי', 'שנה', 'השנה',
    'גניזה', 'הגניזה', 'ספר', 'ספרים', 'כתבי', 'יד', 'גאון', 'הגדול',
    'הימים', 'ימים', 'נוראים', 'כיפורים', 'פסח', 'חול', 'קבע', 'קודש',
}
_PREFIX_CHARS = ('ו', 'ה', 'ב', 'ל', 'מ', 'כ', 'ש', 'ד')
_SUFFIX_STRIP = ('ותיו', 'אות', 'ות', 'ים', 'ין', 'יהם', 'יו', 'י', 'ה', 'ת')


def _stem_token(t: str) -> str:
    for suf in _SUFFIX_STRIP:
        if t.endswith(suf) and len(t) - len(suf) >= 3:
            return t[: -len(suf)]
    return t


def _token_variants(t: str) -> set:
    """Prefix-stripped (up to 2 chars) + stemmed match keys for one token."""
    forms = {t}
    s = t
    for _ in range(2):
        if len(s) > 3 and s[0] in _PREFIX_CHARS:
            s = s[1:]
            forms.add(s)
    return forms | {_stem_token(x) for x in forms if len(_stem_token(x)) >= 3}


def _base_token(t: str) -> str:
    s = t
    for _ in range(2):
        if len(s) > 3 and s[0] in _PREFIX_CHARS:
            s = s[1:]
    return s


def _heb_tokens(s: Optional[str]) -> List[str]:
    s = (s or "").replace("״", "").replace("׳", "").replace("'", "")
    return [t for t in _HEB_TOKEN_RE.findall(s) if len(t) >= 2 and t not in _STOP_TOKENS]


def _sampling_name_match(claim_title: Optional[str], claim_author: Optional[str], texts: List[str]) -> bool:
    """SAMPLING-ONLY port of discovery_identified_gate.py's ``name_match``:
    True when the claimed work's title/author strong-token-matches any of
    ``texts`` (>=2 distinct matched base tokens, >=1 NOT a generic/liturgical
    WEAK token)."""
    claim_tokens = _heb_tokens(f"{claim_author or ''} {claim_title or ''}")
    if not claim_tokens:
        return False
    forms: set = set()
    for t in texts:
        for tok in _heb_tokens(t):
            forms |= _token_variants(tok)
    matched: set = set()
    strong = 0
    for t in claim_tokens:
        b = _base_token(t)
        if _token_variants(t) & forms and b not in matched:
            matched.add(b)
            if b not in _WEAK_TOKENS:
                strong += 1
    return len(matched) >= 2 and strong >= 1


def load_fjms_catalog_text(db_path: str) -> Dict[str, str]:
    """sys_id -> combined FJMS ``catalog`` table text (``TitleHeb`` /
    ``GenizahTitleOrgTitle`` / ``Title`` / ``GenizahTitleEngTitle``, non-null
    fields space-joined across every catalog row for that AlmaId). THIS is
    the field ``same_work_spike/probe/rsource/HANDOFF-TO-135.md`` § 6.1
    identifies as the catalogue's OWN identification -- distinct from
    libraries.csv column 7 (the public NLI title, already read elsewhere in
    this module) and from the WRONG field ``catalog_refs`` Codex measured at
    ZERO matched known pairs (CODEX-REVIEW-17-novelty-2a.log finding 2).
    Returns ``{}`` if the sidecar is absent (graceful degradation, mirroring
    every other sidecar reader in this project)."""
    if not os.path.exists(db_path):
        return {}
    out: Dict[str, List[str]] = defaultdict(list)
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        for aid, title_heb, gen_org, title, gen_eng in con.execute(
            "SELECT AlmaId, TitleHeb, GenizahTitleOrgTitle, Title, GenizahTitleEngTitle FROM catalog"
        ):
            for v in (title_heb, gen_org, title, gen_eng):
                if v:
                    out[str(aid)].append(str(v))
    finally:
        con.close()
    return {sid: " ".join(vals) for sid, vals in out.items()}


def load_bib_rows(db_path: str) -> Dict[str, List[Tuple]]:
    """sys_id (AlmaId) -> list of bibliography rows, the SAME columns and
    SELECT as ``same_work_spike/probe/scripts/bib_gate.py``'s ``BibGate.__init__``
    (``RunningTitle, RunningTitleHeb, TitleAcronymHeb, TitleAcronym,
    ArticleName, ArticleAuthorHeb, ArticleAuthorEng, NoteForDisplay,
    MentionType, TranscriptionType, TitleYear`` -- an 11-tuple per row, index
    9 = ``TranscriptionType``, matching that file's own ``raw[9]`` usage)."""
    if not os.path.exists(db_path):
        return {}
    out: Dict[str, List[Tuple]] = defaultdict(list)
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        for row in con.execute(
            "SELECT AlmaId, RunningTitle, RunningTitleHeb, TitleAcronymHeb, TitleAcronym, "
            "ArticleName, ArticleAuthorHeb, ArticleAuthorEng, NoteForDisplay, MentionType, "
            "TranscriptionType, TitleYear FROM bibliography"
        ):
            out[str(row[0])].append(tuple(row[1:]))
    finally:
        con.close()
    return out


def _bib_classify_simplified(
    rows: List[Tuple], claim_title: Optional[str], claim_author: Optional[str]
) -> Tuple[str, bool]:
    """SAMPLING-ONLY simplified port of ``bib_gate.BibGate.classify`` --
    decides only the two categories Codex flagged as decisive for the
    CURRENT (to-be-hardened) heuristic funnel: ``known_bib`` (a genuine
    token-name-match against the claimed title/author -- trustworthy) and
    ``published_full`` (no name-match, but a ``TranscriptionType='Full'`` row
    exists -- Codex finding 1: this alone affects 3,688 known pairs, 3,060
    sole-source, and is the OVER-BROAD category Arm 2 exists to
    characterize). Returns ``(category, named)`` where ``named`` is True for
    ``known_bib``/``published_full`` (the funnel's current demote-worthy
    categories) and False for ``bib_present_other`` (present, non-decisive)
    or ``bib_absent`` (no rows at all)."""
    if not rows:
        return "bib_absent", False
    for row in rows:
        texts = [str(v) for v in row[:8] if v]
        if _sampling_name_match(claim_title, claim_author, texts):
            return "known_bib", True
    transcription_types = {str(row[9]) for row in rows if row[9]}
    if "Full" in transcription_types:
        return "published_full", True
    return "bib_present_other", False


def load_pgp_signal_index(db_path: str) -> Dict[str, Tuple[bool, bool]]:
    """sys_id -> (present, named). ``present`` = sys_id is linked to >=1 PGP
    document via ``document_fragments``; ``named`` = the strongest such
    document has a non-empty description or transcription -- the CURRENT
    (Codex-flagged over-broad, finding 6) heuristic's own decisive test,
    ported in spirit from ``discovery_identified_gate.py::load_pgp``: "a
    document with a description/transcription NAMES the fragment" --
    regardless of whether that text actually names THIS specific claimed
    work. Kept exactly this over-broad here on purpose: Arm 2's entire
    purpose is to characterize this category's false-known rate."""
    if not os.path.exists(db_path):
        return {}
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        docinfo: Dict[Any, Tuple[bool, bool]] = {}
        for pid, desc, tr, has_tr, _dtype in con.execute(
            "SELECT pgpid, description, transcription, has_transcription, document_type FROM documents"
        ):
            docinfo[pid] = (bool((desc or "").strip()), bool((tr or "").strip()) or bool(has_tr))
        out: Dict[str, Tuple[bool, bool]] = {}
        for sid, did in con.execute("SELECT sys_id, document_id FROM document_fragments"):
            if not sid:
                continue
            info = docinfo.get(did)
            if info is None:
                continue
            sid = str(sid)
            named = info[0] or info[1]
            prev = out.get(sid)
            if prev is None or (named and not prev[1]):
                out[sid] = (True, named)
    finally:
        con.close()
    return out


def load_fgp_rows(db_path: str) -> Dict[str, List[Tuple[str, str, str]]]:
    """sys_id -> list of (title_he, author_he, title_en) -- the SAME fields
    ``discovery_identified_gate.py::load_fgp`` reads for its own name-match
    test (never the ``content`` free-text field, matching that file's own
    scope)."""
    if not os.path.exists(db_path):
        return {}
    out: Dict[str, List[Tuple[str, str, str]]] = defaultdict(list)
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        for sid, th, ah, te in con.execute(
            "SELECT sys_id, title_he, author_he, title_en FROM fgp_transcriptions"
        ):
            if sid:
                out[str(sid)].append((th or "", ah or "", te or ""))
    finally:
        con.close()
    return out


_CATALOGUE_TERSE_MAX_LEN = 20  # mirrors the retired select_terse_catalogue_candidates threshold


def _combined_catalogue_text(sys_id: str, libraries: Dict[str, Dict[str, str]], fjms_catalog: Dict[str, str]) -> str:
    """Union of the two genuinely-implemented catalogue sources NOVEL-01
    names: libraries.csv column 7 (the public NLI title) and the FJMS
    ``catalog`` table's own identification fields (HANDOFF-TO-135.md § 6.1)
    -- NEVER the wrong ``catalog_refs`` field."""
    parts = []
    lib_text = libraries.get(sys_id, {}).get("catalogue_text", "")
    if lib_text:
        parts.append(lib_text)
    fjms_text = fjms_catalog.get(sys_id, "")
    if fjms_text:
        parts.append(fjms_text)
    return " ".join(parts)


def compute_claim_source_signals(
    sid: str,
    claim_title: Optional[str],
    claim_author: Optional[str],
    libraries: Dict[str, Dict[str, str]],
    fjms_catalog: Dict[str, str],
    bib_rows_idx: Dict[str, List[Tuple]],
    pgp_signal_idx: Dict[str, Tuple[bool, bool]],
    fgp_rows_idx: Dict[str, List[Tuple[str, str, str]]],
) -> Dict[str, Any]:
    """Per-(sys_id, claimed-work) source-signal bundle -- the SAME four
    checked-source families NOVEL-01/gen2_novelty_gate.py name (bib,
    catalogue, FGP, PGP), computed against REAL production sidecars joined to
    this manuscript's own claim. A SAMPLING approximation (see module note
    above) -- never a real novelty verdict; that is plan 136-04's job."""
    cat_text = _combined_catalogue_text(sid, libraries, fjms_catalog)
    cat_norm = normalize_title(cat_text)
    claim_norm = normalize_title(claim_title)
    cat_present = bool(cat_text)
    cat_named = bool(claim_norm) and bool(cat_norm) and claim_norm in cat_norm

    bib_rows = bib_rows_idx.get(sid, [])
    bib_category, bib_named = _bib_classify_simplified(bib_rows, claim_title, claim_author)
    bib_present = bib_category != "bib_absent"

    pgp_present, pgp_named = pgp_signal_idx.get(sid, (False, False))

    fgp_rows = fgp_rows_idx.get(sid, [])
    fgp_present = bool(fgp_rows)
    fgp_named = False
    if fgp_rows:
        texts = [x for tup in fgp_rows for x in tup if x]
        fgp_named = _sampling_name_match(claim_title, claim_author, texts)

    is_terse_catalogue = (not cat_present) or (len(cat_text) <= _CATALOGUE_TERSE_MAX_LEN)
    container_match = _LITURGICAL_CONTAINER_RITE_RE.search(cat_norm) if cat_norm else None
    is_container_shape = bool(container_match) and not cat_named

    return {
        "cat_text": cat_text,
        "cat_present": cat_present,
        "cat_named": cat_named,
        "bib_present": bib_present,
        "bib_named": bib_named,
        "bib_category": bib_category,
        "pgp_present": pgp_present,
        "pgp_named": pgp_named,
        "fgp_present": fgp_present,
        "fgp_named": fgp_named,
        "is_terse_catalogue": is_terse_catalogue,
        "is_container_shape": is_container_shape,
        "container_phrase": container_match.group(0) if container_match else None,
    }


def _best_claims_by_sys_work(claims: List[Dict[str, Any]]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """One representative shipped ``direct_witness`` claim per (sys_id,
    work_id) pair -- highest ``matched_letters``, then lexicographically
    smallest ``page_id`` -- the same total-order discipline every selector in
    this module uses. A dedicated helper for the ruling-J arm builder so the
    EXISTING Class 6 selector (``select_catalogue_divergence_candidates``,
    left byte-for-byte unchanged by this continuation) is never touched."""
    best: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for c in claims:
        if c.get("claim_type") != "direct_witness" or c.get("routing_status") != "shipped":
            continue
        key = (c["sys_id"], c["work_id"])
        cand_key = (-(c["matched_letters"] or 0), c["page_id"])
        prev = best.get(key)
        if prev is None or cand_key < prev["_key"]:
            d = dict(c)
            d["_key"] = cand_key
            best[key] = d
    return best


def _residual_stratum(sig: Dict[str, Any]) -> str:
    """Priority order, applied once per residual candidate.

    ``container_predicts`` (a specific, substantial catalogue-text shape)
    takes first priority. For the remaining strata, the catalogue counts as
    a genuinely PRESENT source only when it is SUBSTANTIAL (non-terse) --
    ``effective_cat_present = cat_present and not is_terse_catalogue``. This
    is the fix for a real bug caught during this continuation's dry run: an
    earlier version let a terse/absent catalogue field claim priority over
    the four presence-based strata unconditionally, which made `bib_sole` /
    `pgp_sole` / `fgp_sole` UNREACHABLE by construction (a manuscript can
    only be "sole" for bib/pgp/fgp when the catalogue offers nothing, which
    is EXACTLY the condition the old ordering diverted to `terse_catalogue`
    first). With the fix, `terse_catalogue` now means what the former Class 4
    actually tested: the catalogue offers nothing/next-to-nothing AND no
    other checked source has text either -- the genuinely minimal-information
    residual case. When the catalogue is terse/absent but bib/PGP/FGP DO have
    text, this function now correctly falls through to that source's own
    sole/multi-source stratum instead."""
    if sig["is_container_shape"]:
        return "container_predicts"
    effective_cat_present = sig["cat_present"] and not sig["is_terse_catalogue"]
    present = [
        name for name, p in (
            ("bib", sig["bib_present"]), ("pgp", sig["pgp_present"]),
            ("fgp", sig["fgp_present"]), ("catalogue", effective_cat_present),
        ) if p
    ]
    if len(present) == 0:
        # Nothing SUBSTANTIAL anywhere -- reachable only because the
        # catalogue has SOME (terse) text (any_present was true upstream) or
        # is fully absent while still landing here via the residual/no-text
        # split in select_novelty_arms; either way this is the minimal-
        # information former-Class-4 shape.
        return "terse_catalogue"
    if len(present) >= 2:
        return "multi_source"
    if present == ["bib"]:
        return "bib_sole"
    if present == ["pgp"]:
        return "pgp_sole"
    if present == ["fgp"]:
        return "fgp_sole"
    return "catalogue_sole"  # present == ["catalogue"]


def _demotion_stratum(sig: Dict[str, Any]) -> str:
    """Oversamples the two Codex-flagged false-known populations (Codex
    findings 1 and 6): a bib ``published_full`` demotion with no other
    source agreeing, and a PGP demotion with no other source agreeing."""
    only_pgp = sig["pgp_named"] and not (sig["bib_named"] or sig["fgp_named"] or sig["cat_named"])
    only_bib_pubfull = (
        sig["bib_named"] and sig["bib_category"] == "published_full"
        and not (sig["pgp_named"] or sig["fgp_named"] or sig["cat_named"])
    )
    if only_bib_pubfull:
        return "published_full_sole"
    if only_pgp:
        return "pgp_sole"
    return "other_demotion"


_RESIDUAL_STRATUM_TEXT: Dict[str, str] = {
    "bib_sole": "the Friedberg bibliography has text for this manuscript that does NOT name this "
    "specific claimed work (no other checked source has any text at all for this manuscript)",
    "pgp_sole": "the Princeton Geniza Project has a document description/transcription for this "
    "manuscript that does NOT name this specific claimed work (no other checked source has any "
    "text at all)",
    "fgp_sole": "an FGP transcription's own title/author fields do NOT name this specific claimed "
    "work (no other checked source has any text at all)",
    "catalogue_sole": "this manuscript's own catalogue identification (NLI title and/or the FJMS "
    "catalog table) does NOT name this specific claimed work (no other checked source has any "
    "text at all)",
    "multi_source": ">=2 checked sources have text for this manuscript, but NONE of them names this "
    "specific claimed work",
    "terse_catalogue": "this manuscript's own catalogue identification field is empty or too short "
    "(<=20 characters) to compare against -- folds in the former Class 4 (terse/missing catalogue "
    "text)",
    "container_predicts": "this manuscript's own catalogue text names a specific, NAMED standard-rite "
    "container (a container noun immediately followed by מנהג) whose standard, "
    "predictable content plausibly includes this claimed unit, without the catalogue ever naming the "
    "unit itself -- folds in the former Class 7 (owner rulings H/I)",
}
_DEMOTION_STRATUM_TEXT: Dict[str, str] = {
    "published_full_sole": "the CURRENT heuristic treats ANY bibliography row with "
    "TranscriptionType='Full' as naming this claim, regardless of whether that row's own title/author "
    "actually matches (Codex finding 1) -- no OTHER checked source agrees",
    "pgp_sole": "the CURRENT heuristic treats ANY PGP document with a non-empty description or "
    "transcription as naming this claim, regardless of whether that text actually names this specific "
    "work (Codex finding 6) -- no OTHER checked source agrees",
    "other_demotion": "a genuine token-name-match (catalogue, FGP, or bib known_bib) demotes this "
    "claim -- included for comparison against the two oversampled over-broad categories above",
}


def select_novelty_arms(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    libraries: Dict[str, Dict[str, str]],
    fjms_catalog: Dict[str, str],
    bib_rows_idx: Dict[str, List[Tuple]],
    pgp_signal_idx: Dict[str, Tuple[bool, bool]],
    fgp_rows_idx: Dict[str, List[Tuple[str, str, str]]],
    exclude_sys_ids: set,
    cap_residual_per_stratum: int = 5,
    cap_demoted_published_full: int = 10,
    cap_demoted_pgp: int = 10,
    cap_demoted_other: int = 5,
    cap_no_source_text: int = 8,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Owner ruling J's three-arm, source-stratified novelty sample. Returns
    ``(residual_cases, demoted_cases, no_source_text_cases)``, each a list of
    case dicts in this module's standard shape (class/question_type/sys_id/
    shelfmark/catalogue_text/work_titles/reason/proposal[/stratum]).

    Deterministic: candidates are the ``_best_claims_by_sys_work`` total
    order, sorted by (sys_id, work_id) within each stratum before capping --
    no round-robin, no sampling, no randomness. ``exclude_sys_ids`` removes
    every manuscript already selected as a Class 6 (catalogue divergence)
    candidate, so a single manuscript is never presented twice under two
    different question types.
    """
    best = _best_claims_by_sys_work(claims)

    residual_by_stratum: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    demoted_by_stratum: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    no_source_text: List[Dict[str, Any]] = []

    for (sid, wid), c in sorted(best.items()):
        if sid in exclude_sys_ids:
            continue
        w = works.get(wid)
        if w is None:
            continue
        sig = compute_claim_source_signals(
            sid, w.get("neutral_title"), w.get("author"),
            libraries, fjms_catalog, bib_rows_idx, pgp_signal_idx, fgp_rows_idx,
        )
        any_named = sig["bib_named"] or sig["pgp_named"] or sig["fgp_named"] or sig["cat_named"]
        any_present = sig["bib_present"] or sig["pgp_present"] or sig["fgp_present"] or sig["cat_present"]
        entry = {"sys_id": sid, "work_id": wid, "claim": c, "work": w, "sig": sig}
        if any_named:
            demoted_by_stratum[_demotion_stratum(sig)].append(entry)
        elif any_present:
            residual_by_stratum[_residual_stratum(sig)].append(entry)
        else:
            no_source_text.append(entry)

    def _case_from_entry(entry: Dict[str, Any], cls: str, question_type: str, reason: str, proposal: Optional[str], stratum: Optional[str] = None) -> Dict[str, Any]:
        sid, w, c = entry["sys_id"], entry["work"], entry["claim"]
        cat = libraries.get(sid, {})
        title = f"{w['neutral_title']} ({entry['work_id']})"
        case: Dict[str, Any] = {
            "class": cls,
            "question_type": question_type,
            "sys_id": sid,
            "shelfmark": cat.get("shelfmark", ""),
            "catalogue_text": entry["sig"]["cat_text"] or cat.get("catalogue_text", ""),
            "work_titles": [title],
            "reason": reason,
            "proposal": proposal,
        }
        if stratum is not None:
            case["stratum"] = stratum
        return case

    # --- Arm 1: residual, 7 strata, capped independently ---
    residual_cases: List[Dict[str, Any]] = []
    residual_caps = {
        "container_predicts": cap_residual_per_stratum,
        "terse_catalogue": cap_residual_per_stratum,
        "bib_sole": cap_residual_per_stratum,
        "pgp_sole": cap_residual_per_stratum,
        "fgp_sole": cap_residual_per_stratum,
        "catalogue_sole": cap_residual_per_stratum,
        "multi_source": cap_residual_per_stratum,
    }
    for stratum in ("container_predicts", "terse_catalogue", "bib_sole", "pgp_sole", "fgp_sole", "catalogue_sole", "multi_source"):
        entries = sorted(residual_by_stratum.get(stratum, []), key=lambda e: (e["sys_id"], e["work_id"]))
        for entry in entries[: residual_caps[stratum]]:
            proposal_shade = "container_predicts" if stratum == "container_predicts" else "fills_gap"
            reason = (
                f"Residual stratum `{stratum}`: {_RESIDUAL_STRATUM_TEXT[stratum]}. Under owner ruling J's "
                "funnel-first architecture, this row WOULD reach the pinned model gate."
            )
            proposal = (
                f"PROPOSAL (draft, not a label; owner ruling J -- see 136-GATE1-DECISIONS.md § J): "
                f"plausibly `{proposal_shade}` -- confirm or correct."
            )
            residual_cases.append(_case_from_entry(entry, "residual", "shade", reason, proposal, stratum))

    # --- Arm 2: heuristic-demoted, 3 strata, oversampling the two Codex-flagged populations ---
    demoted_cases: List[Dict[str, Any]] = []
    demoted_caps = {
        "published_full_sole": cap_demoted_published_full,
        "pgp_sole": cap_demoted_pgp,
        "other_demotion": cap_demoted_other,
    }
    for stratum in ("published_full_sole", "pgp_sole", "other_demotion"):
        entries = sorted(demoted_by_stratum.get(stratum, []), key=lambda e: (e["sys_id"], e["work_id"]))
        for entry in entries[: demoted_caps[stratum]]:
            reason = f"Demotion stratum `{stratum}`: {_DEMOTION_STRATUM_TEXT[stratum]}."
            proposal = (
                "PROPOSAL (draft, not a label; owner ruling J -- see 136-GATE1-DECISIONS.md § J): "
                "plausibly `demotion_correct` -- confirm or correct to `false_known` if the demoting "
                "source does not actually name this specific work."
            )
            demoted_cases.append(_case_from_entry(entry, "heuristic_demoted", "demotion", reason, proposal, stratum))

    # --- Arm 3: no-source-text, single bucket, NO verdict collected (design decision, not an omission) ---
    no_source_text_cases: List[Dict[str, Any]] = []
    for entry in sorted(no_source_text, key=lambda e: (e["sys_id"], e["work_id"]))[:cap_no_source_text]:
        reason = (
            "None of the four checked-source families (bib, PGP, FGP, catalogue) has ANY text at all "
            "for this manuscript -- this row ships as a novelty candidate automatically, with no "
            "source to check it against."
        )
        no_source_text_cases.append(
            _case_from_entry(entry, "no_source_text", "no_verdict", reason, None)
        )

    return residual_cases, demoted_cases, no_source_text_cases


# ---------------------------------------------------------------------------
# Class-6 owner scope/shade annotations (rulings F/G, 136-GATE1-DECISIONS.md
# §§ F/G). The owner read all 15 of the ORIGINAL Class-6 candidates directly
# (not merely the shade's abstract definition) and issued explicit verdicts
# on 12 of them, keyed here by sys_id (stable across a case-number
# renumbering, unlike the "Case N" labels in the pre-restructure worksheet
# those verdicts were originally phrased against). This dict does NOT decide
# the owner's Task-3 verdict for these cases -- it only upgrades their
# PROPOSAL text (still explicitly marked PROPOSAL, still not a label) to
# reflect what the owner has ALREADY told this project about that specific
# manuscript/work pair, so the regenerated worksheet does not silently
# regress to a stale generic "plausibly diverges" proposal for cases the
# owner has already spoken to directly. The 3 original candidates NOT named
# in either ruling (sys_ids not present here) are left with a generic,
# undetermined-scope proposal -- genuinely open, not silently resolved.
# ---------------------------------------------------------------------------
_CLASS6_OWNER_SCOPE: Dict[str, Dict[str, str]] = {
    # --- ruling F: diverges_work ("usually our claim is wrong") ---
    "990001004230205171": {"shade": "diverges_work", "ruling": "F"},  # case 92: ילקוט שמעוני / תנחומא
    "990000413480205171": {"shade": "diverges_work", "ruling": "F"},  # case 84: משנה תורה ספר זמנים / הגדה של פסח
    "990001398690205171": {"shade": "diverges_work", "ruling": "F"},  # case 86: משנה תורה ספר אהבה / ברכת המזון
    "990000621960205171": {"shade": "diverges_work", "ruling": "F"},  # case 95: תנ"ך בראשית / שאילתות
    "990051080280205171": {"shade": "diverges_work", "ruling": "F"},  # case 97: ויקרא רבה / חובות הלבבות
    "990000905560205171": {"shade": "diverges_work", "ruling": "F"},  # case 91: מכילתא דרשב"י / מכילתא דרבי ישמעאל
    "990000555810205171": {"shade": "diverges_work", "ruling": "F"},  # case 85: הלכות פסוקות / הלכות גדולות
    # --- ruling F: diverges_part ("more delicate and essentially less important") ---
    "990001935160205171": {"shade": "diverges_part", "ruling": "F"},  # case 90: הלכות פסוקות...קידושין / הלכות פסוקות
    "990051173260205171": {"shade": "diverges_part", "ruling": "F"},  # case 94: משנה תורה הקדמה ומניין המצוות / הלכות ציצית
    "990051150540205171": {"shade": "diverges_part", "ruling": "F"},  # case 96: בראשית רבה צה-צו תוספת / בראשית רבה
    # --- ruling G: NOT a divergence at all -- the aid's own FREE TEXT
    # already states the claimed identification; only the structured
    # work-id keying differed (the selector's own over-fire failure mode) ---
    "990000555880205171": {
        "shade": "confirms",
        "ruling": "G",
        "free_text_quote": "שאלות ותשובות מאת האי בן שרירא גאון",
    },  # case 83: תשובות האיי גאון (structured match: תשובות)
    "990001394270205171": {
        "shade": "confirms",
        "ruling": "G",
        "free_text_quote": "יוסיפון בערבית",
    },  # case 87: ספר יוסיפון (ערבי) (structured match: יוסיפון)
}


def build_hardcases(
    claims: List[Dict[str, Any]],
    works: Dict[str, Dict[str, Any]],
    d13d: Dict[str, Any],
    libraries: Dict[str, Dict[str, str]],
    fjms_catalog: Optional[Dict[str, str]] = None,
    bib_rows_idx: Optional[Dict[str, List[Tuple]]] = None,
    pgp_signal_idx: Optional[Dict[str, Tuple[bool, bool]]] = None,
    fgp_rows_idx: Optional[Dict[str, List[Tuple[str, str, str]]]] = None,
    class1_spotcheck_cap: int = 3,
    class2_spotcheck_cap: int = 2,
    class3_spotcheck_cap: int = 3,
    cap_per_class: int = 20,
    class6_cap: int = 30,
    cap_residual_per_stratum: int = 5,
    cap_demoted_published_full: int = 10,
    cap_demoted_pgp: int = 10,
    cap_demoted_other: int = 5,
    cap_no_source_text: int = 8,
) -> List[Dict[str, Any]]:
    fjms_catalog = fjms_catalog or {}
    bib_rows_idx = bib_rows_idx or {}
    pgp_signal_idx = pgp_signal_idx or {}
    fgp_rows_idx = fgp_rows_idx or {}
    cases: List[Dict[str, Any]] = []

    # --- Class 3: granularity -- IDENTITY spot-check (owner-authorized
    # labelling restructure, same continuation as rulings F/G). Full
    # labelling of the class's candidate pool was REPLACED by an
    # ~class3_spotcheck_cap-case spread, testing the constant-answer
    # assumption ("A and B are the same work at two granularities") rather
    # than building ground truth over all of it -- see the module docstring
    # and 136-GATE1-DECISIONS.md's labelling-restructure note. ---
    # Dedupe to ONE representative group per manuscript (sys_id) first --
    # a single large manuscript can contribute dozens of near-duplicate
    # span-groups (verified: sys_id 990000852430205171 alone supplies 11 of
    # the 276 collapse candidates), which would otherwise crowd out
    # diversity in the candidate pool. Keep each manuscript's LARGEST (most
    # matched_letters) group as its representative.
    by_sys_id: Dict[str, Tuple[Tuple[str, int, int], int]] = {}
    for key in d13d["collapse_candidate_keys"]:
        page_id, s0, s1 = key
        sid = sys_id_from_page_id(page_id)
        matched = s1 - s0
        prev = by_sys_id.get(sid)
        if prev is None or matched > prev[1]:
            by_sys_id[sid] = (key, matched)
    granularity_keys = [v[0] for v in by_sys_id.values()]
    # Worked example first if present, then the rest in stable key order.
    worked_page_id = (d13d["worked_example"] or {}).get("page_id")
    granularity_keys.sort(key=lambda k: (k[0] != worked_page_id, k))
    granularity_pool = granularity_keys[:cap_per_class]
    for pool_idx in _evenly_spaced_indices(len(granularity_pool), class3_spotcheck_cap):
        page_id, s0, s1 = granularity_pool[pool_idx]
        sid = sys_id_from_page_id(page_id)
        cat = libraries.get(sid, {})
        # find the two related work_ids on this exact span
        span_group_works = sorted({
            c["work_id"]
            for c in claims
            if c["page_id"] == page_id and c["span_start"] == s0 and c["span_end"] == s1
        })
        titles = [
            f"{works[w]['neutral_title']} ({w})"
            for w in span_group_works
            if w in works
        ]
        cases.append({
            "class": "granularity",
            "question_type": "identity",
            "sys_id": sid,
            "shelfmark": cat.get("shelfmark", ""),
            "catalogue_text": cat.get("catalogue_text", ""),
            "work_titles": titles,
            "reason": (
                f"Byte-identical span {s0}-{s1} on this page is claimed by {len(span_group_works)} "
                "works sharing the same author and a common title stem -- a title-containment/alias "
                "relationship a plain string comparison cannot resolve on its own (same underlying "
                "commentary at two catalogued granularities, or two genuinely distinct works?)."
            ),
            "proposal": (
                "PROPOSAL (draft, not a label): plausibly `same_work` (the SAME underlying work at two "
                "granularities) -- confirm or correct."
            ),
        })

    # --- Class 2: alias pairs -- IDENTITY spot-check (same restructure) ---
    alias_pairs = select_alias_pair_candidates(works)
    alias_pool = alias_pairs[:cap_per_class]
    for pool_idx in _evenly_spaced_indices(len(alias_pool), class2_spotcheck_cap):
        wa, wb = alias_pool[pool_idx]
        rep = best_claim_for_work(claims, wa["work_id"]) or best_claim_for_work(claims, wb["work_id"])
        sid = rep["sys_id"] if rep else None
        cat = libraries.get(sid, {}) if sid else {}
        cases.append({
            "class": "alias",
            "question_type": "identity",
            "sys_id": sid,
            "shelfmark": cat.get("shelfmark", ""),
            "catalogue_text": cat.get("catalogue_text", ""),
            "work_titles": [
                f"{wa['neutral_title']} ({wa['work_id']}, {wa['source_corpus']})",
                f"{wb['neutral_title']} ({wb['work_id']}, {wb['source_corpus']})",
            ],
            "reason": (
                "Two DIFFERENT work_ids share both the same author and an identical normalized title "
                "(a two-member cluster, not a large generic-collection cluster) -- a likely un-merged "
                "cross-corpus alias/duplicate."
            ),
            "proposal": (
                "PROPOSAL (draft, not a label): plausibly `same_work` (an alias pair, not yet "
                "canonically merged) -- confirm or correct."
            ),
        })

    # --- Class 1: near-miss titles -- IDENTITY spot-check (same restructure) ---
    near_miss = select_near_miss_candidates(works)
    near_miss_pool = near_miss[:cap_per_class]
    for pool_idx in _evenly_spaced_indices(len(near_miss_pool), class1_spotcheck_cap):
        wa, wb, ratio = near_miss_pool[pool_idx]
        rep = best_claim_for_work(claims, wa["work_id"]) or best_claim_for_work(claims, wb["work_id"])
        sid = rep["sys_id"] if rep else None
        cat = libraries.get(sid, {}) if sid else {}
        cases.append({
            "class": "near_miss",
            "question_type": "identity",
            "sys_id": sid,
            "shelfmark": cat.get("shelfmark", ""),
            "catalogue_text": cat.get("catalogue_text", ""),
            "work_titles": [
                f"{wa['neutral_title']} ({wa['work_id']})",
                f"{wb['neutral_title']} ({wb['work_id']})",
            ],
            "reason": (
                f"Same author; normalized titles are {ratio:.1%} similar (SequenceMatcher) but NOT "
                "identical -- genuinely different works (e.g. different books/chapters/parts) whose "
                "titles a string comparison could easily conflate in EITHER direction."
            ),
            "proposal": None,
        })

    # --- Classes 4 (terse/missing catalogue text) and 5 (generic collection
    # works) -- SUPERSEDED by the ruling-J three-arm redesign (136-03
    # continuation, 2026-08-02; 136-GATE1-DECISIONS.md § J). Per this
    # continuation's own accounting (136-NOVELTY-HARDCASES.md intro, and this
    # plan's Task 5): Class 4's phenomenon (a terse/absent catalogue field) is
    # CARRIED FORWARD, not dropped -- it is re-selected as Arm 1's
    # `terse_catalogue` stratum below, now against the REAL bib/PGP/FGP
    # signal rather than catalogue-text alone. Class 5 (generic collection
    # works) is DROPPED outright: no owner ruling exists for any specific
    # Class-5 case (only generic PROPOSALS, unlike Class 6's owner-annotated
    # 12 of 15), and its phenomenon -- whether a SINGLE witness of a
    # same-author/same-title-stem collection is "already recorded" at the
    # collection level -- is a work-IDENTITY ambiguity, not a
    # source-coverage gap; it does not correspond to any of the three arms
    # and folding it into one would blur what that arm measures. This drops
    # no owner-authorized work: no owner verdict was ever recorded against a
    # specific Class 5 case. `select_generic_collection_candidates` is left
    # defined above (unused by this function as of this continuation) rather
    # than deleted, in case a future session revisits the collection-level
    # question on its own terms. ---

    # --- Class 6: catalogue divergence (owner decision E, 136-GATE1-DECISIONS.md
    # item E; SHADE SPLIT BY SCOPE + a separate correctness axis per owner
    # ruling F, 136-GATE1-DECISIONS.md item F; EXPANDED 15->class6_cap in the
    # labelling restructure). The underlying selector's structured-id-vs-
    # free-text conflation is left DELIBERATELY UNCORRECTED (owner ruling G,
    # § G -- "measure it, do not quietly fix it away"): it over-fires on
    # roughly half of the original 15 candidates, and expanding the pool
    # with the SAME heuristic is expected to keep surfacing that same
    # failure mode, which is itself the measured signal a future hardening
    # pass should be built against. NOVELTY SHADE question, PLUS a
    # Correctness sub-question (see CORRECTNESS_VOCABULARY) -- neither
    # forced by this script; _CLASS6_OWNER_SCOPE only upgrades the PROPOSAL
    # text for the 12 of 15 original candidates the owner has already ruled
    # on directly (still a PROPOSAL, never a label). ---
    divergence = select_catalogue_divergence_candidates(claims, works, libraries, cap=class6_cap)
    for entry in divergence:
        sid = entry["sys_id"]
        cat = libraries.get(sid, {})
        w_claimed = entry["claimed_work"]
        w_divergent = entry["divergent_work"]
        c = entry["claim"]
        claimed_title = f"{w_claimed['neutral_title']} ({c['work_id']})"
        divergent_title = f"{w_divergent['neutral_title']} ({w_divergent['work_id']})"
        owner_scope = _CLASS6_OWNER_SCOPE.get(sid)
        if owner_scope is None:
            proposal = (
                "PROPOSAL (draft, not a label): plausibly `diverges_work` or `diverges_part` (scope "
                "not yet distinguished by owner rulings F/G for this specific case) -- the catalogue "
                "and this claim name different works or parts; per ruling G, first check whether the "
                "catalogue's own FREE TEXT already states this claim's identification under a "
                "different spelling/phrasing before confirming a divergence -- confirm the shade, the "
                "scope, and (if divergent) the Correctness call, or correct."
            )
        elif owner_scope["shade"] == "confirms":
            proposal = (
                "PROPOSAL (draft, not a label; owner ruling G -- see 136-GATE1-DECISIONS.md § G): "
                "plausibly `confirms`, NOT a divergence -- the catalogue's own free text "
                f"({owner_scope['free_text_quote']!r}) already states this claim's identification; "
                "only the structured work-id keying differed (this class's own selector over-fired "
                "here) -- confirm or correct."
            )
        else:
            proposal = (
                f"PROPOSAL (draft, not a label; owner ruling F -- see 136-GATE1-DECISIONS.md § F): "
                f"plausibly `{owner_scope['shade']}` -- confirm the shade AND supply the separate "
                "Correctness call (catalogue_correct / claim_correct / unclear), or correct."
            )
        cases.append({
            "class": "catalogue_divergence",
            "question_type": "shade",
            "correctness_applicable": True,
            "sys_id": sid,
            "shelfmark": cat.get("shelfmark", ""),
            "catalogue_text": cat.get("catalogue_text", ""),
            "work_titles": [
                f"CLAIMED (this identification): {claimed_title}",
                f"CATALOGUE NAMES (found in the identification text): {divergent_title}",
            ],
            "reason": (
                "This manuscript's own catalogue identification text names a DIFFERENT work "
                f"({w_divergent['neutral_title']!r}) than the one this claim identifies "
                f"({w_claimed['neutral_title']!r}); the two are NOT a granularity variant under the "
                "D-13d author-gated rule (different author, or an unrelated title) -- a genuine "
                "catalogue/claim divergence CANDIDATE (owner ruling F splits the shade into "
                "`diverges_work` / `diverges_part` by scope; owner ruling G warns this selector can "
                "over-fire when the catalogue's free text actually already agrees with the claim under "
                "a different structured key -- see the Vocabulary sheet/table)."
            ),
            "proposal": proposal,
        })

    # --- Class 7 (liturgical-container predictability, owner rulings H/I) --
    # SUPERSEDED: folded into Arm 1's `container_predicts` stratum below,
    # per owner ruling J's explicit instruction ("a container-only machzor
    # title has text, fails name-match, and therefore lands in the
    # residual"). `select_liturgical_container_candidates` is left defined
    # above (its regex/threshold logic is reused conceptually by
    # `_residual_stratum`/`compute_claim_source_signals`) but is no longer
    # invoked directly here.

    # --- Ruling J: the three-arm, SOURCE-STRATIFIED novelty sample
    # (136-GATE1-DECISIONS.md § J) -- REPLACES the former Classes 4/5/7 as
    # the accuracy instrument. Excludes every sys_id already selected for
    # Class 6 (catalogue divergence) so no manuscript is shown twice under
    # two different question types. ---
    exclude_sys_ids = {entry["sys_id"] for entry in divergence}
    residual_cases, demoted_cases, no_source_text_cases = select_novelty_arms(
        claims, works, libraries, fjms_catalog, bib_rows_idx, pgp_signal_idx, fgp_rows_idx,
        exclude_sys_ids=exclude_sys_ids,
        cap_residual_per_stratum=cap_residual_per_stratum,
        cap_demoted_published_full=cap_demoted_published_full,
        cap_demoted_pgp=cap_demoted_pgp,
        cap_demoted_other=cap_demoted_other,
        cap_no_source_text=cap_no_source_text,
    )
    cases.extend(residual_cases)
    cases.extend(demoted_cases)
    cases.extend(no_source_text_cases)

    return cases


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def _fmt_pct(numerator: int, denominator: int) -> str:
    if not denominator:
        return "n/a"
    return f"{100.0 * numerator / denominator:.1f}%"


def render_evidence_brief(
    *,
    asset_path: str,
    identifications_total: int,
    main_total: int,
    show_more_total: int,
    reason_counts: Dict[str, int],
    d13e: Dict[str, Any],
    d16: Dict[str, Any],
    d13c: Dict[str, Any],
    d13b: Dict[str, Any],
    d13d: Dict[str, Any],
) -> str:
    lines: List[str] = []
    a = lines.append
    a("# Phase 136 Plan 03 -- Gate-1 Decision Evidence")
    a("")
    a(f"Measured read-only against the deployed asset `{os.path.basename(asset_path)}` "
      "(the LIVE v2 sidecar -- the trimmed rebuild has not run). Re-running "
      "`scripts/discovery_gate1_evidence.py` against the same file reproduces every number below "
      "exactly (no sampling, no randomness).")
    a("")
    a("**Population note.** The main-pool-rule classification below is computed over \"usable\" claims: "
      "a claim's display evidence has `routing_status='shipped'`, OR any evidence row for the claim is "
      "`adjudication_status='human_confirmed'` (the D-13g fix folded in, so a human-confirmed row is "
      "never invisible to this measurement merely because routing demoted it). This reproduces "
      f"**{identifications_total:,}** identifications, main **{main_total:,}** / show-more "
      f"**{show_more_total:,}** -- within ~2% of the design-pass figures already on record in "
      "`main-pool-rule.md` (36,152 / 28,357), which is the expected order of agreement given this pass "
      "implements gate 3 slightly differently (see the D-13e section below) and a slightly different "
      "human-confirmed rule.")
    a("")
    a("Gate breakdown (identification count by classification reason):")
    a("")
    a("| Reason | Bucket | Count |")
    a("|---|---|---|")
    reason_bucket = {
        "human_confirmed_override": "main",
        "multi_folio_agreement": "main",
        "gate4_single_page_full_coverage": "main",
        "gate1_no_same_work_claim": "show_more",
        "gate2_best_band_weak": "show_more",
        "gate3_unresolved_competition": "show_more",
        "gate4_low_single_page_coverage": "show_more",
        "gate4_no_coverage_data": "show_more",
    }
    for reason in sorted(reason_counts):
        a(f"| {reason} | {reason_bucket.get(reason, '?')} | {reason_counts[reason]:,} |")
    a("")

    # --- D-13e ---
    a("## D-13e -- does the middle \"Also shares text with\" bucket survive as a THIRD level?")
    a("")
    a(f"- D-13d generic identical-span-group claims (>=2 DIFFERENT canonical works on one byte-identical "
      f"span): **{d13e['generic_claims_total']:,}**.")
    a(f"  - Of those, **{d13e['generic_not_reachable']:,}** belong to an identification that this pass "
      "classifies MAIN under the main-pool rule -- i.e. NOT otherwise reachable via \"show more matches\" "
      "(that toggle only ever renders show-more-classified identifications).")
    a(f"  - The remaining **{d13e['generic_overlap']:,}** belong to an identification ALREADY classified "
      "show-more -- these rows are already reachable via \"show more matches\"; giving them a second, "
      "separate middle-bucket home would be a duplicate view.")
    a(f"- Related-pages population (D-11a, shared-text page relations -- these never map onto a WORK "
      "identification at all, so they can never be reached via the per-work \"show more matches\" toggle): "
      f"**{d13e['related_pairs']:,}** directed (anchor, opposite) page pairs. All of it is, by "
      "construction, not otherwise reachable.")
    a("")
    a(f"**Total middle-bucket population: {d13e['total_middle_bucket']:,}.** "
      f"**Not otherwise reachable: {d13e['not_reachable_total']:,}** "
      f"({_fmt_pct(d13e['not_reachable_total'], d13e['total_middle_bucket'])} of the middle bucket). "
      f"**Overlap with \"show more matches\": {d13e['overlap_total']:,}** "
      f"({_fmt_pct(d13e['overlap_total'], d13e['total_middle_bucket'])}).")
    a("")
    a("**Methodology note (gate 3):** this pass implements \"unresolved competition\" as EITHER a "
      "`discovery_routing_audit` `kept_tie` page (direct, non-heuristic) OR an overlapping, near-equal-"
      "length competing span from another canonical work on the same page (overlap >= 70% of the shorter "
      "span AND a length ratio >= 0.7 -- a stated, reproducible threshold). `demoted_work_id` is NULL on "
      "every `kept_tie` row in this asset (a known, already-documented flaw -- D-02b), so the audit table "
      "alone cannot fully reconstruct every tie; the near-tie span test is this pass's way of recovering "
      "most of the gap, and is the main source of any remaining difference from the design-pass figures.")
    a("")
    a("**Question for the owner:** given the numbers above, does the panel keep a distinct THIRD "
      "disclosure level (\"Also shares text with\"), or does it collapse into \"more matches\"? No "
      "recommendation is made here -- D-13e is open by design.")
    a("")

    # --- D-16 ---
    a("## D-16 / PANEL-01 -- does the findings page also get the relation filter?")
    a("")
    a("Relation distribution (claim_type), corpus-wide (all claims, any routing status):")
    a("")
    a("| Relation | Count |")
    a("|---|---|")
    for row in d16["corpus_wide"]:
        a(f"| {row['claim_type']} | {row['n']:,} |")
    a("")
    a("Relation distribution restricted to SHIPPED display claims:")
    a("")
    a("| Relation | Count |")
    a("|---|---|")
    for row in d16["shipped"]:
        a(f"| {row['claim_type']} | {row['n']:,} |")
    a("")
    a("Relation distribution restricted to this pass's MAIN pool (identifications classified `main`):")
    a("")
    a("| Relation | Count |")
    a("|---|---|")
    for relation, n in d16["main_pool"].items():
        a(f"| {relation} | {n:,} |")
    a("")
    a("**Question for the owner:** does a relation filter on the main pool meaningfully narrow the "
      "default view, or does it mostly restate the bucket (since the main pool is already "
      "overwhelmingly `direct_witness`)? No recommendation is made here -- D-16 is open by design.")
    a("")

    # --- D-13c ---
    a("## D-13c -- the short-evidence threshold")
    a("")
    a(f"Thinnest shipped direct match in the whole asset: **{d13c['thinnest_direct']} matched letters**.")
    a("")
    a("Cumulative row counts below each candidate threshold:")
    a("")
    a("| Threshold (matched letters) | Direct family (of {:,}) | Propagated family (of {:,}) |".format(
        d13c["direct_total"], d13c["propagated_total"]))
    a("|---|---|---|")
    for t in d13c["thresholds"]:
        dn = d13c["direct_cumulative"][t]
        pn = d13c["propagated_cumulative"][t]
        a(f"| < {t} | {dn:,} ({_fmt_pct(dn, d13c['direct_total'])}) | "
          f"{pn:,} ({_fmt_pct(pn, d13c['propagated_total'])}) |")
    a("")
    a("**Methodology note (propagated family):** the propagated family's length metric is "
      "`aligned_len` on shipped `shared_text` evidence rows (propagated `witness`-kind rows -- "
      "`corroborated`/`weak` -- carry no length field in this asset at all). This is a slightly "
      "different population than an earlier design-pass count (which counted DISPLAY claims rather "
      "than evidence rows); the counting unit is stated here so the two are never silently conflated.")
    a("")
    a(f"Short direct rows (< 150 matched letters) that are nonetheless part of a MAIN identification "
      f"via multi-folio agreement (the honest counter-argument the owner already accepted -- for a "
      f"prayer book, a short liturgical passage may be exactly the correct identification): "
      f"**{d13c['short_in_main_multi_folio']:,}**.")
    a("")
    a("**Question for the owner:** what is the short-evidence threshold, in matched letters? A "
      "defensible default exists: **150** (the figure the owner has already reviewed counts against, "
      "per `main-pool-rule.md` / `136-CONTEXT.md` D-13c) -- kept as the recommended default unless the "
      "table above changes the owner's mind.")
    a("")

    # --- D-13b ---
    a("## D-13b -- the lead-attribution tie-break")
    a("")
    a(f"Identical-span groups (>=2 shipped direct claims on one byte-identical span): "
      f"**{d13b['total_groups']:,}** groups / **{d13b['total_claims']:,}** claims.")
    a(f"Of those, **{d13b['tied_after_band_rank_groups']:,}** groups "
      f"({_fmt_pct(d13b['tied_after_band_rank_groups'], d13b['total_groups'])}) are STILL tied after "
      f"ordering by band rank alone (**{d13b['tied_after_band_rank_claims']:,}** claims involved) -- "
      "band rank alone cannot pick a lead attribution for the overwhelming majority of these groups.")
    a("")
    a("**Question for the owner:** what breaks a tie after band rank when several works claim one "
      "passage? A defensible default exists: fall back to the existing TOTAL claim ordering already "
      "used elsewhere in the build (`discovery_ids.py`'s `evidence_id` lexicographic tie-break) -- "
      "deterministic, already in the codebase, and requires no new concept.")
    a("")

    # --- D-13d ---
    a("## D-13d -- the granularity separation rule (KNOWN FLAW)")
    a("")
    a(f"Identical-span groups with >=2 DIFFERENT canonical works: "
      f"**{d13d['diff_canon_groups_total']:,}** groups / **{d13d['diff_canon_claims_total']:,}** claims.")
    a("")
    we = d13d["worked_example"]
    if we:
        a("**Worked example** (T-S Misc. 12.31.14):")
        a("")
        a(f"- Page: `{we['page_id']}` (sys_id `{we['sys_id']}`)")
        a(f"- Span: offsets {we['span_start']}-{we['span_end']} ({we['matched_letters']} matched letters)")
        for w in we["works"]:
            a(f"  - `{w['work_id']}` (canonical `{w['canonical_work_id']}`): **{w['neutral_title']}** "
              f"-- author: {w['author']}")
        a("")
        a("Both works share the same author and a common title prefix (\"<author> on ...\") -- the same "
          "underlying commentary recorded at two catalogued granularities (a general work covering the "
          "whole Torah, and a specific work covering only Genesis), carrying DIFFERENT `canonical_work_id`s. "
          "Under the CURRENT rule (D-13d as originally stated) this whole pair is swept into the generic "
          "\"also shares text with\" bucket and neither title renders as a stand-alone identification for "
          "this page, even though the two titles denote a real, nameable commentary.")
        a("")
    a("**Proposed separation rule** (display-time only, not a data fix): treat two works in an identical-"
      "span group as the SAME work at different granularity (collapse like a duplicate, per D-13a) when "
      "they share a non-null `author` field AND EITHER their normalized titles are identical (an "
      "undetected alias) OR share a >= 4-character normalized-title prefix (e.g. a common "
      "\"<author> on ...\" commentary marker). Groups where no such pair exists remain genuinely generic "
      "shared text.")
    a("")
    a(f"**Measured effect:** of the {d13d['diff_canon_groups_total']:,} different-canonical-work identical-"
      f"span groups, **{d13d['collapse_candidate_groups']:,}** groups "
      f"({_fmt_pct(d13d['collapse_candidate_groups'], d13d['diff_canon_groups_total'])}, "
      f"{d13d['collapse_candidate_claims']:,} claims) contain a same-author/related-title pair and are "
      f"candidates for the collapse rule; **{d13d['generic_groups']:,}** groups "
      f"({_fmt_pct(d13d['generic_groups'], d13d['diff_canon_groups_total'])}, "
      f"{d13d['generic_claims']:,} claims) contain no such pair and remain genuinely generic shared text "
      "under the proposed rule.")
    a("")
    a("**Question for the owner:** does this separation rule (same author + identical/prefix-shared "
      "normalized title) correctly draw the line? A defensible default exists: adopt it as stated, since "
      "it is conservative (author-gated, so it never over-collapses the large generic-collection-title "
      "clusters measured separately in the novelty hard-case selection below) and directly resolves the "
      "worked example.")
    a("")

    return "\n".join(lines) + "\n"


# Classes 1-3 ask the IDENTITY question (same_work / different_works /
# unsure / skip) -- an owner-authorized labelling restructure (same
# continuation as rulings F/G): these rows compare two of OUR OWN claims and
# have no claim-vs-aid relationship to judge, so decision E's shade
# vocabulary never applied to them coherently in the first place.
#
# Class 6 (catalogue divergence) and the ruling-J "residual" arm both ask the
# NOVELTY SHADE question. The ruling-J "heuristic_demoted" arm asks a
# DEMOTION-CORRECTNESS question (DEMOTION_VOCABULARY). The ruling-J
# "no_source_text" arm asks NO question at all -- ships with no verdict, by
# design (see the module's ruling-J section and 136-GATE1-DECISIONS.md § J).
# Former Classes 4/5/7 are superseded -- see build_hardcases's own comments
# for the kept/folded/dropped accounting.
_IDENTITY_CLASSES: frozenset = frozenset({"granularity", "alias", "near_miss"})
_DEMOTION_CLASSES: frozenset = frozenset({"heuristic_demoted"})
_NO_VERDICT_CLASSES: frozenset = frozenset({"no_source_text"})

# Per-class guidance for which of the shade vocabulary's tokens are actually
# plausible answers for that class's kind of hard case (136-03 Task 4 --
# "say plainly, per class, which shades are plausible answers for that
# class"). ONLY meaningful for the SHADE classes (Class 6 + ruling J's
# "residual" arm) -- every case can still receive ANY shade, `unsure`, or
# `skip`; this is a reading aid for the owner, never a constraint enforced by
# this script.
_PLAUSIBLE_SHADES_BY_CLASS: Dict[str, Tuple[str, ...]] = {
    # F replaces `diverges` with the scope-split pair; G adds `confirms` as
    # a live possibility here (the class's own selector can over-fire a
    # genuine confirms case into a divergence candidate -- see ruling G).
    "catalogue_divergence": (
        "diverges_work", "diverges_part", "aid_more_specific", "refines_granularity", "confirms",
    ),
    # Ruling J's residual arm folds the former Classes 4 (terse catalogue)
    # and 7 (container_predicts) in as strata (see the `stratum` field on
    # each case) -- the plausible-shade hint spans the union of what those
    # two classes' own hints named, since a single "residual" section now
    # covers both shapes plus the four source-presence strata.
    "residual": ("fills_gap", "confirms", "container_predicts", "refines_granularity", "aid_more_specific"),
}

# Per-class guidance for the IDENTITY classes (1-3) -- always the same two
# substantive answers (same_work / different_works) plus unsure/skip; kept
# as a dict (rather than a single shared tuple) so a future class-specific
# note can be added without touching the render/xlsx code that reads it.
_PLAUSIBLE_IDENTITY_BY_CLASS: Dict[str, Tuple[str, ...]] = {
    "granularity": ("same_work", "different_works"),
    "alias": ("same_work", "different_works"),
    "near_miss": ("same_work", "different_works"),
}


_CLASS_TITLES: Dict[str, str] = {
    "granularity": "Class 3 -- catalogue entry naming a different granularity of the same work (IDENTITY spot-check)",
    "alias": "Class 2 -- alias pairs (IDENTITY spot-check)",
    "near_miss": "Class 1 -- near-miss titles (IDENTITY spot-check)",
    "catalogue_divergence": "Class 6 -- catalogue divergence (NOVELTY SHADE, owner rulings E/E′/F/G -- RETAINED UNCHANGED by the ruling-J redesign)",
    "residual": "Arm 1 -- RESIDUAL: rows that would reach the model (NOVELTY SHADE, source-stratified, owner ruling J -- folds in the former Classes 4 and 7)",
    "heuristic_demoted": "Arm 2 -- HEURISTIC-DEMOTED: rows the funnel marks known before any model call (DEMOTION CORRECTNESS, owner ruling J)",
    "no_source_text": "Arm 3 -- NO-SOURCE-TEXT: rows shipped as candidates with no verdict (owner ruling J)",
}
_CLASS_ORDER: Tuple[str, ...] = (
    "granularity", "alias", "near_miss", "catalogue_divergence",
    "residual", "heuristic_demoted", "no_source_text",
)
# Reverse lookup (rendered "Class" column text -> internal class code) --
# used by the Task 4 read-back below so the JSON label file stores the SAME
# short codes the rest of this module already uses internally, without a
# second, hand-maintained copy of the class-title strings drifting out of
# sync with `_CLASS_TITLES` above.
_CLASS_TITLE_TO_CODE: Dict[str, str] = {v: k for k, v in _CLASS_TITLES.items()}


def assign_case_numbers(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Assigns a stable, sequential ``case_num`` to each case in the SAME
    ``_CLASS_ORDER`` / within-class order the Markdown worksheet renders in,
    so 136-NOVELTY-HARDCASES.md and 136-NOVELTY-HARDCASES.xlsx agree
    case-for-case -- both render from this ONE pre-numbered list, never two
    independent iterations that could silently drift apart."""
    by_class: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in cases:
        by_class[c["class"]].append(c)
    numbered: List[Dict[str, Any]] = []
    n = 0
    for cls in _CLASS_ORDER:
        for item in by_class.get(cls, []):
            n += 1
            item = dict(item)
            item["case_num"] = n
            numbered.append(item)
    return numbered


def _render_case_common(a, item: Dict[str, Any], work_label: str) -> None:
    if item.get("shelfmark"):
        a(f"- **Manuscript:** {item['shelfmark']} (sys_id `{item['sys_id']}`)")
    elif item.get("sys_id"):
        a(f"- **Manuscript:** sys_id `{item['sys_id']}` (no shelfmark on file)")
    else:
        a("- **Manuscript:** (no shipped claim instance found for either work)")
    a(f"- **{work_label}:** {' / '.join(item['work_titles'])}")
    if item.get("catalogue_text"):
        a(f"- **Catalogue's own identification text:** {item['catalogue_text']}")


def render_hardcases_brief(cases: List[Dict[str, Any]]) -> str:
    """``cases`` must already carry ``case_num`` (see ``assign_case_numbers``)."""
    lines: List[str] = []
    a = lines.append
    identity_total = sum(1 for c in cases if c["class"] in _IDENTITY_CLASSES)
    class6_total = sum(1 for c in cases if c["class"] == "catalogue_divergence")
    residual_total = sum(1 for c in cases if c["class"] == "residual")
    demoted_total = sum(1 for c in cases if c["class"] in _DEMOTION_CLASSES)
    no_verdict_total = sum(1 for c in cases if c["class"] in _NO_VERDICT_CLASSES)
    novelty_total = class6_total + residual_total + demoted_total + no_verdict_total

    by_class: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in cases:
        by_class[c["class"]].append(c)

    a("# Phase 136 Plan 03 -- Novelty Hard-Case Candidates")
    a("")
    a(f"**{len(cases)} candidates total: {identity_total} IDENTITY spot-check cases (Classes 1-3, "
      f"UNCHANGED) + {novelty_total} novelty-evaluation cases** (Class 6 catalogue divergence, "
      "RETAINED unchanged, plus owner ruling J's new three-arm, SOURCE-STRATIFIED sample). "
      "**Redesigned per owner ruling J** (`136-GATE1-DECISIONS.md` § J, 2026-08-02) after a read-only "
      "prior-art pass (`136-NOVELTY-PRIOR-ART.md` §§ 6-7) found that the FORMER Classes 4/5/7 all read "
      "EXACTLY ONE field (libraries.csv column 7) and therefore had ZERO representation of the "
      "bib/PGP/FGP failure modes Codex measured as most damaging in the prior heuristic funnel "
      "(`gen2_novelty_gate.py`): 3,688 `published_full` false-knowns, 2,014 PGP false-knowns "
      "(942 sole-source), FGP's 1,177-known/9,373-fail split. A gate could score perfectly on the "
      "pre-J pool and still reproduce every one of those defects in production.")
    a("")
    a("**What changed, explicitly (per this continuation's own accounting instruction -- nothing is "
      "silently dropped):**")
    a("")
    a("| Former pool item | Disposition | Why |")
    a("|---|---|---|")
    a("| Classes 1-3 (identity spot-check, 8 cases) | **KEPT, unchanged** | Tests a different "
      "assumption (A↔B same-work identity), already correctly sized |")
    a("| Class 6 (catalogue divergence, 30 cases incl. the owner's F/G annotations on 12 of the "
      "original 15) | **KEPT, unchanged** | Owner rulings F/G are substantive, dated characterizations "
      "of SPECIFIC real manuscripts (worked cases 83-97) -- dropping or reselecting this class would "
      "silently discard that owner-authorized work before its Task-3 confirmation ever ran. It also "
      "tests a different, already-owner-engaged axis (divergence SHAPE/correctness, deliberately left "
      "an uncorrected heuristic per ruling G) orthogonal to source coverage |")
    a("| Class 4 (terse/missing catalogue text) | **FOLDED IN** as Arm 1's `terse_catalogue` stratum | "
      "Ruling J's own instruction: fold in as a stratum, not a separate exercise; now checked against "
      "the REAL bib/PGP/FGP signal, not catalogue text alone |")
    a("| Class 7 (liturgical-container predictability, owner rulings H/I) | **FOLDED IN** as Arm 1's "
      "`container_predicts` stratum | Same instruction: \"a container-only machzor title has text, "
      "fails name-match, and therefore lands in the residual\" |")
    a("| Class 5 (generic collection works) | **DROPPED** | No owner ruling exists for any specific "
      "Class 5 case (only generic PROPOSALS, unlike Class 6). Its phenomenon -- whether a single "
      "witness of a same-author/same-title-stem collection is \"already recorded\" -- is a "
      "WORK-IDENTITY ambiguity, not a source-coverage gap; it does not correspond to any of the three "
      "arms and forcing it into one would blur what that arm measures. The underlying question remains "
      "valid and is flagged for a future, separately-scoped pass -- not silently discarded, just not "
      "carried by THIS redesign |")
    a("")
    a("Every case in every group is still selected entirely by deterministic string/metadata/source-"
      "presence comparison over the works, manuscripts and finding-aid sidecars already on this "
      "machine (fist_data/fjms_enrichment.db, pgp_data/pgp.db, fgp_data/fgp_transcriptions.db, "
      "libraries.csv) -- **zero model calls, measured cost $0.00**. Any attached draft verdict is "
      "explicitly marked `PROPOSAL` and is a reading aid only, never a label. This worksheet is also "
      "emitted as `136-NOVELTY-HARDCASES.xlsx` (same phase directory, FIVE sheets: \"Identity "
      "Spot-Check\", \"Novelty Shades\", \"Heuristic-Demoted\", \"No-Source-Text\", \"Vocabulary & "
      "Instructions\") for owners who find Hebrew RTL easier to work with in a spreadsheet; both files "
      "render the SAME cases in the SAME order, from the same pre-numbered case list, so the two agree "
      "case-for-case.")
    a("")
    a("## Sizing -- what each arm can and cannot answer, and why this size")
    a("")
    a(f"**Total: {len(cases)}** ({identity_total} identity + {novelty_total} novelty-evaluation, of "
      "which owner ruling J's own sizing instruction covers the "
      f"{novelty_total - class6_total} non-Class-6 cases -- kept under the ~100 novelty-case guidance "
      "with Class 6 counted separately as pre-existing, unchanged owner-engaged work).")
    a("")
    a(f"- **Class 6 (catalogue divergence, {class6_total} cases, unchanged):** answers \"does the "
      "owner confirm the shade + correctness proposals already characterized on 12 of the original 15 "
      "real cases, and how does the selector's own measured over-fire rate (~50%) hold up across the "
      "expanded pool.\" It does NOT test source coverage -- a case here is selected purely on "
      "catalogue-text containment, same as before.")
    a(f"- **Arm 1 -- RESIDUAL, {residual_total} cases across 7 strata (a FIXED per-stratum cap, not a "
      "proportional sample, where the population supports it):** answers \"of "
      "the rows that WOULD reach the pinned model gate, does the model correctly classify a "
      "representative case from EACH source family and from the two folded-in shapes.\" It does NOT "
      "establish a base rate for how COMMON each stratum is in the full corpus -- the cap is a fixed "
      "ceiling per stratum, not a proportional sample, so this arm answers a per-stratum ACCURACY "
      "question, never a POPULATION-SIZE question.")
    a(f"- **Arm 2 -- HEURISTIC-DEMOTED, {demoted_total} cases across 3 strata (oversampling "
      "`published_full`-sole and PGP-sole demotions specifically, per Codex findings 1 and 6):** "
      "answers \"of the rows the funnel marks known WITHOUT ever consulting a model, how many are "
      "FALSE-knowns -- lost findings that ruling J's funnel-first architecture can never recover.\" It "
      "does NOT give a project-wide false-known RATE (3,688 `published_full` and 2,014 PGP pairs exist "
      "corpus-wide per Codex's own measurement; this arm samples a small, oversampled slice of each, "
      "never the full population) -- a rate estimate would need a much larger, proportionally-stratified "
      "sample, which is explicitly NOT what this size buys.")
    a(f"- **Arm 3 -- NO-SOURCE-TEXT, {no_verdict_total} cases, NO verdict collected:** answers, "
      "qualitatively, \"does this population look like genuinely untouched fragments, or does "
      "something in it look surprising/wrong.\" It is NOT a labelling exercise -- these rows ship as "
      "candidates automatically regardless of what the owner observes here, per ruling J's own design "
      "(\"these ship as candidates with no verdict\"); this section exists so the owner can eyeball the "
      "bypass rather than trust it blindly, not to produce a graded number.")
    a("")
    a("**What this sizing does NOT cover, stated plainly:** none of the three arms measures a "
      "corpus-wide RATE (what fraction of all claims fall in each stratum) -- only a per-stratum, "
      "per-shape ACCURACY/correctness check on a small, capped, deterministically-selected "
      "representative. A future pass wanting base rates would need to run the real funnel (plan "
      "136-04) over the full corpus and report its own per-stratum counts, not re-derive them from this "
      "labelling sample.")
    a("")
    a("## Part A -- IDENTITY spot-check (Classes 1-3, UNCHANGED)")
    a("")
    a("**Question type: IDENTITY, not a novelty shade.** These rows compare two of OUR OWN claims "
      "(A and B) -- there is no finding aid in this judgment at all. Answer ONE of:")
    a("")
    a("| Answer | Choose this when... |")
    a("|---|---|")
    for token, description in IDENTITY_VOCABULARY:
        a(f"| `{token}` | {description} |")
    a("")
    a("**How to read the result (recorded in `136-GATE1-DECISIONS.md` so the interpretation is fixed "
      "BEFORE the answers come in, not chosen afterward to fit them):** if ALL cases below come back "
      "`same_work`, that is a measured fact and an argument that the D-13d author-gated collapse rule "
      "(currently collapsing only 276 of 1,367 identical-span groups, 20.2% -- see this plan's D-13d "
      "section) is TOO CONSERVATIVE and should collapse more aggressively. If even ONE case comes back "
      "`different_works`, the constant-answer assumption FAILS and the full 52-case pool needs real "
      "labelling after all, not just a spot-check.")
    a("")

    for cls in _CLASS_ORDER:
        if cls not in _IDENTITY_CLASSES:
            continue
        items = by_class.get(cls, [])
        a(f"### {_CLASS_TITLES[cls]} ({len(items)} candidates)")
        a("")
        for item in items:
            a(f"#### Case {item['case_num']}")
            a("")
            _render_case_common(a, item, "A vs B")
            a(f"- **Why this pair is adversarial to a STRING heuristic:** {item['reason']}")
            if item.get("proposal"):
                a(f"- **{item['proposal']}**")
            a("- **Identity verdict:** _(pending Task 3 -- `same_work` / `different_works` / "
              "`unsure` / `skip`)_")
            a("")

    a("## Part B -- NOVELTY SHADE cases (Class 6, unchanged, + Arm 1 residual)")
    a("")
    a("**Question type: NOVELTY SHADE, a claim-vs-finding-aid judgment (never A↔B identity).** For "
      "EACH case below, answer with the shade that best describes what an enumerable finding aid "
      "(the catalogue's own identification field, bibliography, titles, PGP, FGP, M-source shelfmark "
      "attributions) actually says about THIS fragment and THIS work -- or `unsure` / `skip`. "
      "Amended 2026-08-02 by owner decisions E / E′ / F / G / H (`136-GATE1-DECISIONS.md` items E, "
      "E′, F, G, H): the tri-state (`already_recorded` / `not_in_finding_aids` / `unsure`) collapsed "
      "materially different findings into one bucket -- a catalogue CONTRADICTION and a genuine "
      "\"previously unknown\" both scored the same way, a granularity refinement that helps and one "
      "that adds nothing also scored the same way (E′), a flat wrong-work divergence and a "
      "same-work-wrong-part divergence also scored the same way with no way to record WHICH SIDE is "
      "actually correct (F), and a broader-container relationship (a standard-rite prayer-book "
      "predicting a specific unit it never names) had NO shade at all and fell through to `fills_gap` "
      "by elimination (H) -- so the shade enum now carries TEN values.")
    a("")
    a("| Shade | Choose this when... |")
    a("|---|---|")
    for token, description in SHADE_VOCABULARY:
        a(f"| `{token}` | {description} |")
    a("")
    a("`not_checked` (the fail-closed system default for an unrun/failed/abstained check) is not a "
      "verdict the owner picks directly -- `unsure` is its owner-facing equivalent.")
    a("")
    a("### Correctness (Class 6 ONLY -- answer ONLY if your shade verdict is `diverges_work` or "
      "`diverges_part`, owner ruling F)")
    a("")
    a("A divergence shade records only THAT the aid and the claim disagree, never WHICH SIDE is "
      "right -- the owner's own review of the real cases found BOTH directions occur under the "
      "identical shade. Leave blank / not applicable for every non-divergence shade, and for every "
      "Arm 1 residual row (Arm 1 excludes the manuscripts already selected for Class 6, so a residual "
      "row is never ALSO a divergence candidate).")
    a("")
    a("| Correctness | Choose this when... |")
    a("|---|---|")
    for token, description in CORRECTNESS_VOCABULARY:
        a(f"| `{token}` | {description} |")
    a("")

    for cls in ("catalogue_divergence", "residual"):
        items = by_class.get(cls, [])
        a(f"### {_CLASS_TITLES[cls]} ({len(items)} candidates)")
        a("")
        plausible = _PLAUSIBLE_SHADES_BY_CLASS[cls]
        plausible_str = ", ".join(f"`{s}`" for s in plausible)
        a(f"**Plausible shades for this class:** {plausible_str} (any other shade from the "
          "vocabulary table above is still a valid answer if the case warrants it; `unsure` / "
          "`skip` are always available).")
        a("")
        for item in items:
            a(f"#### Case {item['case_num']}")
            a("")
            _render_case_common(a, item, "Work(s)")
            if item.get("stratum"):
                a(f"- **Residual stratum:** `{item['stratum']}`")
            a(f"- **Why it is hard:** {item['reason']}")
            if item.get("proposal"):
                a(f"- **{item['proposal']}**")
            a(f"- **Shade verdict:** _(pending Task 3 -- {plausible_str}, any other shade from the "
              "vocabulary table above, or `unsure` / `skip`)_")
            if item.get("correctness_applicable"):
                a("- **Correctness (only if `diverges_work` / `diverges_part` above):** "
                  "_(pending Task 3 -- `catalogue_correct` / `claim_correct` / `unclear`, or blank if "
                  "not applicable)_")
            a("")

    a("## Part C -- HEURISTIC-DEMOTED cases (Arm 2, owner ruling J)")
    a("")
    a("**Question type: DEMOTION CORRECTNESS.** These rows were marked \"already recorded\" by the "
      "CURRENT heuristic funnel's own decisive test (a bib `published_full` row, or a PGP document with "
      "any non-empty description/transcription) -- they NEVER reach the model at all under ruling J's "
      "funnel-first architecture. For EACH case, judge whether the demoting source genuinely names "
      "THIS specific claimed work, or only tripped the heuristic through generic presence:")
    a("")
    a("| Answer | Choose this when... |")
    a("|---|---|")
    for token, description in DEMOTION_VOCABULARY:
        a(f"| `{token}` | {description} |")
    a("")
    items = by_class.get("heuristic_demoted", [])
    for item in items:
        a(f"#### Case {item['case_num']}")
        a("")
        _render_case_common(a, item, "Claimed work")
        if item.get("stratum"):
            a(f"- **Demotion stratum:** `{item['stratum']}`")
        a(f"- **Why this demotion is being checked:** {item['reason']}")
        if item.get("proposal"):
            a(f"- **{item['proposal']}**")
        a("- **Demotion verdict:** _(pending Task 3 -- `demotion_correct` / `false_known` / `unsure` "
          "/ `skip`)_")
        a("")

    a("## Part D -- NO-SOURCE-TEXT cases (Arm 3, owner ruling J -- NO VERDICT REQUIRED)")
    a("")
    a("**Question type: none -- informational only.** None of the four checked-source families has "
      "ANY text at all for these manuscripts, so they ship as novelty candidates automatically, with "
      "no source to check them against. This section exists ONLY so the owner can eyeball whether that "
      "bypass looks safe -- there is nothing to confirm or correct, and no verdict is collected here.")
    a("")
    items = by_class.get("no_source_text", [])
    for item in items:
        a(f"#### Case {item['case_num']}")
        a("")
        _render_case_common(a, item, "Claimed work")
        a(f"- **Why no verdict is collected:** {item['reason']}")
        a("")

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# XLSX labelling workbook (owner-requested, 2026-08-02: Hebrew RTL is hard to
# work with in Markdown). Same ``cases`` data as the Markdown worksheet above
# (pre-numbered via ``assign_case_numbers`` so the two files agree
# case-for-case), reissued via THIS script -- never hand-edited -- per the
# same discipline as the Markdown. openpyxl is an existing project
# dependency (see ``shared/export_dossier.py`` / ``web/export_service.py``
# for the house style this mirrors: ``sheet_view.rightToLeft``, bold-white-
# on-blue header row, ``sanitize_text_for_excel`` on every cell).
#
# MASKING NOTE (measured during this plan): ``.xlsx`` is a ZIP archive whose
# inner XML parts are DEFLATE-compressed by default -- a raw byte-level scan
# of the OUTER file (e.g. a naive ``check_atlas_masking.py --scan-asset`` on
# the ``.xlsx`` path) cannot see a literal string that is only present in the
# compressed inner XML. The masking scan for this artifact must be run
# against the DECOMPRESSED inner content (e.g. via ``zipfile`` extraction
# into a scratch file passed as a single explicit ``--scan-asset`` path) --
# see ``136-GATE1-DECISIONS.md``'s "Outstanding (pending Task 3)" section for
# the full methodology note. This module does not perform that extraction
# itself (it stays a lightweight read/write path, mirroring this file's own
# stated reason for not importing the masking-gate module directly); the
# caller is expected to do it before presenting the workbook to the owner.
# ---------------------------------------------------------------------------

def write_hardcases_xlsx(cases: List[Dict[str, Any]], path: str) -> None:
    """``cases`` must already carry ``case_num`` (see ``assign_case_numbers``).

    Writes a FIVE-sheet workbook (owner ruling J, 136-03 continuation,
    2026-08-02 -- each sheet carries the ONE question type the class/arm was
    actually built to support):

    - "Identity Spot-Check" (Classes 1-3, the ~8-case A↔B spot-check,
      UNCHANGED): an Identity dropdown (`same_work` / `different_works` /
      `unsure` / `skip`).
    - "Novelty Shades" (Class 6, unchanged, + Arm 1 residual): a Verdict
      dropdown (the full ten-shade vocabulary) PLUS a Correctness dropdown
      (owner ruling F -- `catalogue_correct` / `claim_correct` / `unclear`,
      meaningful only on Class 6 `diverges_work` / `diverges_part` rows;
      blank elsewhere, including every Arm 1 row).
    - "Heuristic-Demoted" (Arm 2, NEW): a Demotion-verdict dropdown
      (`demotion_correct` / `false_known` / `unsure` / `skip`).
    - "No-Source-Text" (Arm 3, NEW): NO verdict column at all -- informational
      only, per ruling J's own "ships as candidates with no verdict" design.
    - "Vocabulary & Instructions" (all vocabularies + per-class/arm hints +
      the "blank is not a label" note + the kept/folded/dropped accounting).

    Case #s are GLOBAL across all four data sheets (assigned once by
    ``assign_case_numbers``, never renumbered per-sheet), so Task 4's
    round-trip can locate a case by number regardless of which sheet it came
    from. Deterministic at the cell-value / validation-list / sheet-structure
    level -- NOT claimed byte-for-byte on the saved ``.xlsx`` file itself,
    since openpyxl embeds a save timestamp in the workbook's
    ``docProps/core.xml`` on every save.
    """
    # Deferred heavy import (mirrors scripts/bench_discovery.py's own
    # documented pattern): keeps this module importable/usable via --help
    # and --no-write without an openpyxl dependency on those paths.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if repo_root not in sys.path:
        sys.path.insert(0, repo_root)
    from shared.export_utils import sanitize_text_for_excel as _san

    def _s(value: Optional[str]) -> str:
        return _san(value or "")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_top = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)
    wrap_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    identity_cases = [c for c in cases if c["class"] in _IDENTITY_CLASSES]
    shade_cases = [c for c in cases if c["class"] in ("catalogue_divergence", "residual")]
    demoted_cases = [c for c in cases if c["class"] in _DEMOTION_CLASSES]
    no_verdict_cases = [c for c in cases if c["class"] in _NO_VERDICT_CLASSES]

    wb = Workbook()

    def _manuscript_str(item: Dict[str, Any]) -> str:
        return (
            item["shelfmark"] if item.get("shelfmark")
            else (f"sys_id {item['sys_id']} (no shelfmark on file)" if item.get("sys_id")
                  else "(no shipped claim instance found for either work)")
        )

    def _apply_header(ws, headers: List[str]) -> None:
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _apply_body_alignment(ws, n_cols: int) -> None:
        last_row = ws.max_row
        for row_idx in range(2, last_row + 1):
            ws.cell(row=row_idx, column=1).alignment = wrap_top_center
            for col_idx in range(2, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap_top

    # --- Sheet 1: Identity Spot-Check (Classes 1-3) ---
    ws1 = wb.active
    ws1.title = "Identity Spot-Check"
    ws1.sheet_view.rightToLeft = True
    headers1 = [
        "Case #", "Identity verdict", "Class",
        "Manuscript", "sys_id", "A vs B (claim pair)",
        "Catalogue's own identification text", "Why adversarial to a string heuristic",
        "PROPOSAL (draft -- NOT a label)",
    ]
    _apply_header(ws1, headers1)
    widths1 = {"A": 9, "B": 20, "C": 40, "D": 32, "E": 22, "F": 48, "G": 46, "H": 55, "I": 46}
    for col_letter, width in widths1.items():
        ws1.column_dimensions[col_letter].width = width
    for item in identity_cases:
        ws1.append([
            item["case_num"],
            "",  # Identity verdict -- left blank; owner fills in
            _s(_CLASS_TITLES[item["class"]]),
            _s(_manuscript_str(item)),
            _s(item.get("sys_id")),
            _s(" / ".join(item["work_titles"])),
            _s(item.get("catalogue_text")),
            _s(item["reason"]),
            _s(item.get("proposal")),
        ])
    _apply_body_alignment(ws1, len(headers1))
    last_row1 = ws1.max_row
    ws1.freeze_panes = "C2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{last_row1}"
    dv_identity = DataValidation(
        type="list", formula1='"' + ",".join(IDENTITY_TOKENS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid identity verdict",
        error="Choose 'same_work', 'different_works', 'unsure' or 'skip'. Free text is rejected.",
        promptTitle="Identity verdict",
        prompt="Are A and B the same underlying work? Blank = not yet answered.",
    )
    ws1.add_data_validation(dv_identity)
    if last_row1 >= 2:
        dv_identity.add(f"B2:B{last_row1}")

    # --- Sheet 2: Novelty Shades (Classes 4-7) ---
    ws2 = wb.create_sheet(title="Novelty Shades")
    ws2.sheet_view.rightToLeft = True
    headers2 = [
        "Case #", "Shade verdict", "Correctness (diverges_work/diverges_part only)", "Class",
        "Plausible shades", "Manuscript", "sys_id", "Claimed work(s)",
        "Catalogue's own identification text", "Why it is hard",
        "PROPOSAL (draft -- NOT a label)",
    ]
    _apply_header(ws2, headers2)
    widths2 = {
        "A": 9, "B": 20, "C": 30, "D": 46, "E": 30, "F": 32, "G": 22, "H": 48, "I": 46, "J": 55, "K": 46,
    }
    for col_letter, width in widths2.items():
        ws2.column_dimensions[col_letter].width = width
    for item in shade_cases:
        ws2.append([
            item["case_num"],
            "",  # Shade verdict -- left blank; owner fills in
            "",  # Correctness -- left blank; only meaningful on divergence rows
            _s(_CLASS_TITLES[item["class"]]),
            _s(", ".join(_PLAUSIBLE_SHADES_BY_CLASS[item["class"]])),
            _s(_manuscript_str(item)),
            _s(item.get("sys_id")),
            _s(" / ".join(item["work_titles"])),
            _s(item.get("catalogue_text")),
            _s(item["reason"]),
            _s(item.get("proposal")),
        ])
    _apply_body_alignment(ws2, len(headers2))
    last_row2 = ws2.max_row
    ws2.freeze_panes = "D2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{last_row2}"

    # Shade Verdict dropdown -- the FULL vocabulary (nine real shades +
    # `unsure` + `skip`, i.e. every token in SHADE_VOCABULARY -- eleven total).
    # Blank is allowed ("not yet answered" is a legitimate transient state);
    # the Vocabulary sheet explains a blank is NOT a label.
    dv_shade = DataValidation(
        type="list", formula1='"' + ",".join(SHADE_TOKENS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid verdict",
        error="Choose one of the listed shades, or 'unsure' / 'skip'. Free text is rejected.",
        promptTitle="Shade verdict",
        prompt="Pick the shade that best fits, or 'unsure' / 'skip'. Blank = not yet answered.",
    )
    ws2.add_data_validation(dv_shade)
    if last_row2 >= 2:
        dv_shade.add(f"B2:B{last_row2}")

    # Correctness dropdown (owner ruling F) -- applied to EVERY row for
    # simplicity (Excel data validation is range-based, not conditional on
    # another cell's value); the Vocabulary sheet + column header explain it
    # is meaningful ONLY on diverges_work/diverges_part rows and should stay
    # blank otherwise.
    dv_correctness = DataValidation(
        type="list", formula1='"' + ",".join(CORRECTNESS_TOKENS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid correctness call",
        error="Choose 'catalogue_correct', 'claim_correct' or 'unclear', or leave blank.",
        promptTitle="Correctness (divergence rows only)",
        prompt="Only answer if the Shade verdict is diverges_work or diverges_part. Otherwise leave blank.",
    )
    ws2.add_data_validation(dv_correctness)
    if last_row2 >= 2:
        dv_correctness.add(f"C2:C{last_row2}")

    # --- Sheet 3: Heuristic-Demoted (Arm 2, owner ruling J) ---
    ws3d = wb.create_sheet(title="Heuristic-Demoted")
    ws3d.sheet_view.rightToLeft = True
    headers3d = [
        "Case #", "Demotion verdict", "Stratum", "Class",
        "Manuscript", "sys_id", "Claimed work",
        "Catalogue's own identification text", "Why this demotion is being checked",
        "PROPOSAL (draft -- NOT a label)",
    ]
    _apply_header(ws3d, headers3d)
    widths3d = {"A": 9, "B": 20, "C": 22, "D": 46, "E": 32, "F": 22, "G": 46, "H": 46, "I": 55, "J": 46}
    for col_letter, width in widths3d.items():
        ws3d.column_dimensions[col_letter].width = width
    for item in demoted_cases:
        ws3d.append([
            item["case_num"],
            "",  # Demotion verdict -- left blank; owner fills in
            _s(item.get("stratum")),
            _s(_CLASS_TITLES[item["class"]]),
            _s(_manuscript_str(item)),
            _s(item.get("sys_id")),
            _s(" / ".join(item["work_titles"])),
            _s(item.get("catalogue_text")),
            _s(item["reason"]),
            _s(item.get("proposal")),
        ])
    _apply_body_alignment(ws3d, len(headers3d))
    last_row3d = ws3d.max_row
    ws3d.freeze_panes = "D2"
    ws3d.auto_filter.ref = f"A1:{get_column_letter(len(headers3d))}{last_row3d}"
    dv_demotion = DataValidation(
        type="list", formula1='"' + ",".join(DEMOTION_TOKENS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid demotion verdict",
        error="Choose 'demotion_correct', 'false_known', 'unsure' or 'skip'. Free text is rejected.",
        promptTitle="Demotion verdict",
        prompt="Does the demoting source genuinely name THIS specific work? Blank = not yet answered.",
    )
    ws3d.add_data_validation(dv_demotion)
    if last_row3d >= 2:
        dv_demotion.add(f"B2:B{last_row3d}")

    # --- Sheet 4: No-Source-Text (Arm 3, owner ruling J -- NO verdict column) ---
    ws4 = wb.create_sheet(title="No-Source-Text")
    ws4.sheet_view.rightToLeft = True
    headers4 = [
        "Case #", "Class", "Manuscript", "sys_id", "Claimed work",
        "Why no verdict is collected",
    ]
    _apply_header(ws4, headers4)
    widths4 = {"A": 9, "B": 46, "C": 32, "D": 22, "E": 46, "F": 70}
    for col_letter, width in widths4.items():
        ws4.column_dimensions[col_letter].width = width
    for item in no_verdict_cases:
        ws4.append([
            item["case_num"],
            _s(_CLASS_TITLES[item["class"]]),
            _s(_manuscript_str(item)),
            _s(item.get("sys_id")),
            _s(" / ".join(item["work_titles"])),
            _s(item["reason"]),
        ])
    _apply_body_alignment(ws4, len(headers4))
    last_row4 = ws4.max_row
    ws4.freeze_panes = "C2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers4))}{last_row4}"

    # --- Sheet 5: Vocabulary & Instructions ---
    ws3 = wb.create_sheet(title="Vocabulary & Instructions")
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 90

    def _note(text: str) -> None:
        ws3.append([_s(text)])
        ws3.cell(row=ws3.max_row, column=1).alignment = Alignment(
            horizontal="right", vertical="top", wrap_text=True, readingOrder=2
        )
        ws3.merge_cells(start_row=ws3.max_row, start_column=1, end_row=ws3.max_row, end_column=2)

    def _table_header(a_text: str, b_text: str) -> None:
        ws3.append([a_text, b_text])
        ws3.cell(row=ws3.max_row, column=1).font = header_font
        ws3.cell(row=ws3.max_row, column=2).font = header_font
        ws3.cell(row=ws3.max_row, column=1).fill = header_fill
        ws3.cell(row=ws3.max_row, column=2).fill = header_fill

    def _table_row(a_text: str, b_text: str) -> None:
        ws3.append([a_text, _s(b_text)])
        ws3.cell(row=ws3.max_row, column=1).alignment = Alignment(
            horizontal="right", vertical="top", wrap_text=True, readingOrder=2
        )
        ws3.cell(row=ws3.max_row, column=2).alignment = Alignment(
            horizontal="right", vertical="top", wrap_text=True, readingOrder=2
        )

    _note(
        "This workbook is a labelling instrument for the novelty hard-case evaluation set "
        "(plan 136-03 Task 3). The Markdown file 136-NOVELTY-HARDCASES.md in the same phase "
        f"directory remains the authoritative human-readable record of these {len(cases)} candidate "
        "cases and the reasoning for why each is hard; this workbook renders the SAME cases, in the "
        "SAME order, from one shared pre-numbered case list, split across four data sheets by the "
        "question type each class/arm was actually built to support."
    )
    ws3.append([])
    _note(
        "\"Identity Spot-Check\" (Classes 1-3, UNCHANGED): these rows compare two of OUR OWN claims to "
        "each other -- an A vs B \"same underlying work?\" judgment, no finding aid involved at all. "
        "\"Novelty Shades\" (Class 6, unchanged, + Arm 1 residual): these rows judge ONE claim against "
        "the finding aids, with the full ten-shade vocabulary plus a Correctness column for Class 6 "
        "divergence rows. \"Heuristic-Demoted\" (Arm 2, NEW): rows the funnel already marked known "
        "BEFORE any model call -- judge whether that demotion is correct. \"No-Source-Text\" (Arm 3, "
        "NEW): informational only, no verdict collected -- see its own note below."
    )
    ws3.append([])
    _note(
        "Owner ruling J (136-GATE1-DECISIONS.md § J, 2026-08-02) replaced the former Classes 4/5/7 "
        "with a three-arm, SOURCE-STRATIFIED sample after a prior-art pass found those selectors read "
        "ONLY libraries.csv column 7 -- zero representation of the bib/PGP/FGP failure modes Codex "
        "measured as most damaging. Accounting: Classes 1-3 (identity) and Class 6 (catalogue "
        "divergence) are KEPT UNCHANGED (Class 6 carries substantive owner rulings F/G on specific "
        "real cases -- dropping it would discard that work). Class 4 (terse/missing catalogue text) "
        "and Class 7 (liturgical-container predictability) are FOLDED IN as strata of the new Arm 1 "
        "(residual), per the owner's own instruction. Class 5 (generic collection works) is DROPPED: "
        "no owner ruling exists for any specific Class 5 case, and its collection-level-identity "
        "question does not correspond to a source-coverage stratum."
    )
    ws3.append([])
    _note(
        "IMPORTANT: a BLANK verdict cell (Identity, Shade, or Demotion) is NOT a label -- it means "
        "\"not yet answered\". If you cannot judge a case, enter `unsure` explicitly (it costs nothing "
        "and is a real, useful answer) rather than leaving the cell blank. If you choose not to judge a "
        "case at all, enter `skip` explicitly -- it is recorded as skipped, never silently filled from "
        "the case's own draft PROPOSAL. The Correctness column (Novelty Shades sheet) is meaningful "
        "ONLY when the Shade verdict is `diverges_work` or `diverges_part` -- leave it blank on every "
        "other row. The No-Source-Text sheet has NO verdict column at all -- these rows ship as "
        "candidates automatically regardless of what you observe there."
    )
    ws3.append([])

    _table_header("Identity answer", "Choose this when...")
    for token, description in IDENTITY_VOCABULARY:
        _table_row(token, description)
    ws3.append([])

    _table_header("Shade", "Choose this when...")
    for token, description in SHADE_VOCABULARY:
        _table_row(token, description)
    ws3.append([])
    _note(
        "`not_checked` (the fail-closed system default for an unrun/failed/abstained check) is not "
        "a verdict you pick directly -- `unsure` is its owner-facing equivalent."
    )
    ws3.append([])

    _table_header("Correctness (Class 6 divergence rows only)", "Choose this when...")
    for token, description in CORRECTNESS_VOCABULARY:
        _table_row(token, description)
    ws3.append([])

    _table_header("Demotion verdict (Heuristic-Demoted sheet)", "Choose this when...")
    for token, description in DEMOTION_VOCABULARY:
        _table_row(token, description)
    ws3.append([])

    _table_header("Class / Arm", "Plausible answers (a reading aid -- any answer above is still valid)")
    for cls in _CLASS_ORDER:
        if cls in _IDENTITY_CLASSES:
            plausible_str = ", ".join(_PLAUSIBLE_IDENTITY_BY_CLASS[cls])
        elif cls in _DEMOTION_CLASSES:
            plausible_str = ", ".join(DEMOTION_TOKENS)
        elif cls in _NO_VERDICT_CLASSES:
            plausible_str = "(no verdict collected -- informational only)"
        else:
            plausible_str = ", ".join(_PLAUSIBLE_SHADES_BY_CLASS[cls])
        _table_row(_CLASS_TITLES[cls], plausible_str)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Task 4 -- reading the owner's verdicts back from the labelled XLSX.
#
# 136-GATE1-DECISIONS.md's own "Outstanding" section spells out the contract
# this implements: read by Case #, never by row position; a truly blank
# verdict cell is an explicit `skip`, NEVER silently filled from its own
# PROPOSAL draft; reject any value outside the sheet's own vocabulary
# (defense in depth behind the workbook's DataValidation, which a real
# spreadsheet application re-save is not guaranteed to preserve); fail
# CLOSED on a missing/renamed sheet, a missing/renamed column, or a Case #
# that does not match one of the 101 emitted cases (T-136-03-06 -- a label
# file must not silently tolerate a tampered or hand-edited workbook).
# ---------------------------------------------------------------------------

DEFAULT_LABELS_OUT = os.path.join(REPO_ROOT, "discovery_data", "novelty_hardcase_labels-v1.json")
# The date the owner returned the filled-in workbook (136-GATE1-DECISIONS.md,
# "Status of this record" / rulings A-J all share this date).
LABEL_PROVENANCE_DATE = "2026-08-02"

_SHADE_DIVERGENCE_TOKENS = frozenset({"diverges_work", "diverges_part"})

# The exact header row `write_hardcases_xlsx` writes for each sheet -- read
# back and compared verbatim so a renamed/reordered column fails closed
# rather than silently misreading a different field.
_EXPECTED_SHEET_HEADERS: Dict[str, Tuple[str, ...]] = {
    "Identity Spot-Check": (
        "Case #", "Identity verdict", "Class", "Manuscript", "sys_id",
        "A vs B (claim pair)", "Catalogue's own identification text",
        "Why adversarial to a string heuristic", "PROPOSAL (draft -- NOT a label)",
    ),
    "Novelty Shades": (
        "Case #", "Shade verdict", "Correctness (diverges_work/diverges_part only)", "Class",
        "Plausible shades", "Manuscript", "sys_id", "Claimed work(s)",
        "Catalogue's own identification text", "Why it is hard",
        "PROPOSAL (draft -- NOT a label)",
    ),
    "Heuristic-Demoted": (
        "Case #", "Demotion verdict", "Stratum", "Class", "Manuscript", "sys_id",
        "Claimed work", "Catalogue's own identification text",
        "Why this demotion is being checked", "PROPOSAL (draft -- NOT a label)",
    ),
    "No-Source-Text": (
        "Case #", "Class", "Manuscript", "sys_id", "Claimed work",
        "Why no verdict is collected",
    ),
}

_RESIDUAL_STRATUM_RE = re.compile(r"Residual stratum `([a-zA-Z0-9_]+)`")
_PROPOSAL_TOKEN_RE = re.compile(r"plausibly `([a-zA-Z0-9_]+)`")


class LabelReadError(Exception):
    """Raised to fail CLOSED on any structural defect in the labelled workbook."""


def _header_row(ws) -> Tuple[Any, ...]:
    return tuple(c.value for c in next(ws.iter_rows(min_row=1, max_row=1)))


def _require_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise LabelReadError(
            f"expected sheet {name!r} not found in workbook (sheets present: {wb.sheetnames})"
        )
    ws = wb[name]
    got = _header_row(ws)
    want = _EXPECTED_SHEET_HEADERS[name]
    if got != want:
        raise LabelReadError(
            f"sheet {name!r} header row does not match the expected contract.\n"
            f"  expected: {want}\n  got:      {got}"
        )
    return ws


def _clean_cell(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _proposal_token(proposal_text: Optional[str]) -> Optional[str]:
    if not proposal_text:
        return None
    m = _PROPOSAL_TOKEN_RE.search(proposal_text)
    return m.group(1) if m else None


def _provenance(
    sheet: str,
    proposal_token: Optional[str],
    verdict: Optional[str],
    extra: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Confirmed vs corrected is derived, never asserted -- a case with no
    draft PROPOSAL simply carries ``proposal_shown: false`` and
    ``confirmed_proposal: None`` (not applicable), matching the plan's own
    "leaving it untouched does not count as confirming" discipline applied
    to whichever token WAS in the draft, if any."""
    confirmed: Optional[bool] = None
    if proposal_token is not None and verdict not in (None, "skip", "unsure"):
        confirmed = verdict == proposal_token
    prov: Dict[str, Any] = {
        "source": "owner_supplied",
        "method": "xlsx_round_trip",
        "workbook": "136-NOVELTY-HARDCASES.xlsx",
        "sheet": sheet,
        "date": LABEL_PROVENANCE_DATE,
        "proposal_shown": proposal_token is not None,
        "proposal_token": proposal_token,
        "confirmed_proposal": confirmed,
    }
    if extra:
        prov.update(extra)
    return prov


def read_owner_labels_from_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """Reads the owner-filled ``136-NOVELTY-HARDCASES.xlsx`` back and returns
    the structure written to ``discovery_data/novelty_hardcase_labels-v1.json``
    (minus the enclosing content-hash, added by ``write_owner_labels_json``
    so this function stays a pure read).

    Fails CLOSED (raises ``LabelReadError``) on: a missing/renamed sheet, a
    header row that does not match the written contract, a Case # outside
    1..101 or seen twice, or a verdict/correctness value outside its sheet's
    own vocabulary. A truly BLANK verdict cell is recorded as an explicit
    skip and is NEVER filled from the row's own PROPOSAL draft.
    """
    import openpyxl  # deferred, mirrors write_hardcases_xlsx's own pattern

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    cases: Dict[int, Dict[str, Any]] = {}

    def _case_no(d: Dict[str, Any], sheet_name: str) -> int:
        case_no = d["Case #"]
        if not isinstance(case_no, int):
            raise LabelReadError(f"{sheet_name}: non-integer Case # {case_no!r}")
        if case_no in cases:
            raise LabelReadError(f"Case # {case_no} appears more than once across the workbook")
        return case_no

    # ---- Identity Spot-Check (Classes 1-3) ----
    ws = _require_sheet(wb, "Identity Spot-Check")
    headers = _header_row(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        case_no = _case_no(d, "Identity Spot-Check")
        verdict = _clean_cell(d["Identity verdict"])
        if verdict is not None and verdict not in IDENTITY_TOKENS:
            raise LabelReadError(f"case {case_no}: identity verdict {verdict!r} outside {IDENTITY_TOKENS}")
        skipped = verdict is None or verdict == "skip"
        proposal_text = _clean_cell(d["PROPOSAL (draft -- NOT a label)"])
        proposal_token = _proposal_token(proposal_text)
        class_title = _clean_cell(d["Class"])
        cases[case_no] = {
            "case_id": case_no,
            "question_type": "identity",
            "class": _CLASS_TITLE_TO_CODE.get(class_title, class_title),
            "stratum": None,
            "sys_id": _clean_cell(d["sys_id"]),
            "manuscript": _clean_cell(d["Manuscript"]),
            "claim": _clean_cell(d["A vs B (claim pair)"]),
            "catalogue_identification_text": _clean_cell(d["Catalogue's own identification text"]),
            "why_hard": _clean_cell(d["Why adversarial to a string heuristic"]),
            "verdict": {
                "type": "identity",
                "value": None if skipped else verdict,
                "correctness": None,
                "skipped": skipped,
            },
            "label_provenance": _provenance("Identity Spot-Check", proposal_token, verdict),
        }

    # ---- Novelty Shades (Class 6 + Arm 1 residual) ----
    ws = _require_sheet(wb, "Novelty Shades")
    headers = _header_row(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        case_no = _case_no(d, "Novelty Shades")
        shade = _clean_cell(d["Shade verdict"])
        if shade is not None and shade not in SHADE_TOKENS:
            raise LabelReadError(f"case {case_no}: shade verdict {shade!r} outside {SHADE_TOKENS}")
        correctness = _clean_cell(d["Correctness (diverges_work/diverges_part only)"])
        if correctness is not None and correctness not in CORRECTNESS_TOKENS:
            raise LabelReadError(f"case {case_no}: correctness {correctness!r} outside {CORRECTNESS_TOKENS}")
        skipped = shade is None or shade == "skip"
        correctness_gap = (not skipped) and shade in _SHADE_DIVERGENCE_TOKENS and correctness is None
        if correctness_gap:
            print(
                f"WARNING: case {case_no} carries shade {shade!r} but no Correctness call -- "
                "flagged (not failed) per the plan's own instruction",
                file=sys.stderr,
            )
        class_title = _clean_cell(d["Class"])
        class_code = _CLASS_TITLE_TO_CODE.get(class_title, class_title)
        why = _clean_cell(d["Why it is hard"]) or ""
        stratum = None
        if class_code == "residual":
            m = _RESIDUAL_STRATUM_RE.search(why)
            stratum = m.group(1) if m else None
        proposal_text = _clean_cell(d["PROPOSAL (draft -- NOT a label)"])
        proposal_token = _proposal_token(proposal_text)
        cases[case_no] = {
            "case_id": case_no,
            "question_type": "shade",
            "class": class_code,
            "stratum": stratum,
            "sys_id": _clean_cell(d["sys_id"]),
            "manuscript": _clean_cell(d["Manuscript"]),
            "claim": _clean_cell(d["Claimed work(s)"]),
            "catalogue_identification_text": _clean_cell(d["Catalogue's own identification text"]),
            "why_hard": why or None,
            "verdict": {
                "type": "shade",
                "value": None if skipped else shade,
                "correctness": correctness,
                "skipped": skipped,
                "correctness_gap": correctness_gap,
            },
            "label_provenance": _provenance("Novelty Shades", proposal_token, shade),
        }

    # ---- Heuristic-Demoted (Arm 2) ----
    ws = _require_sheet(wb, "Heuristic-Demoted")
    headers = _header_row(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        case_no = _case_no(d, "Heuristic-Demoted")
        verdict = _clean_cell(d["Demotion verdict"])
        if verdict is not None and verdict not in DEMOTION_TOKENS:
            raise LabelReadError(f"case {case_no}: demotion verdict {verdict!r} outside {DEMOTION_TOKENS}")
        skipped = verdict is None or verdict == "skip"
        proposal_text = _clean_cell(d["PROPOSAL (draft -- NOT a label)"])
        proposal_token = _proposal_token(proposal_text)
        class_title = _clean_cell(d["Class"])
        cases[case_no] = {
            "case_id": case_no,
            "question_type": "demotion",
            "class": _CLASS_TITLE_TO_CODE.get(class_title, class_title),
            "stratum": _clean_cell(d["Stratum"]),
            "sys_id": _clean_cell(d["sys_id"]),
            "manuscript": _clean_cell(d["Manuscript"]),
            "claim": _clean_cell(d["Claimed work"]),
            "catalogue_identification_text": _clean_cell(d["Catalogue's own identification text"]),
            "why_hard": _clean_cell(d["Why this demotion is being checked"]),
            "verdict": {
                "type": "demotion",
                "value": None if skipped else verdict,
                "correctness": None,
                "skipped": skipped,
            },
            "label_provenance": _provenance(
                "Heuristic-Demoted", proposal_token, verdict,
                extra={"blank_cell": d["Demotion verdict"] in (None, "")},
            ),
        }

    # ---- No-Source-Text (Arm 3 -- NO verdict column, by design) ----
    ws = _require_sheet(wb, "No-Source-Text")
    headers = _header_row(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        case_no = _case_no(d, "No-Source-Text")
        class_title = _clean_cell(d["Class"])
        cases[case_no] = {
            "case_id": case_no,
            "question_type": "no_verdict_by_design",
            "class": _CLASS_TITLE_TO_CODE.get(class_title, class_title),
            "stratum": None,
            "sys_id": _clean_cell(d["sys_id"]),
            "manuscript": _clean_cell(d["Manuscript"]),
            "claim": _clean_cell(d["Claimed work"]),
            "catalogue_identification_text": None,
            "why_hard": _clean_cell(d["Why no verdict is collected"]),
            "verdict": {
                "type": "no_verdict_by_design",
                "value": None,
                "correctness": None,
                "skipped": False,
            },
            "label_provenance": {
                "source": "no_verdict_by_design",
                "method": "xlsx_round_trip",
                "workbook": "136-NOVELTY-HARDCASES.xlsx",
                "sheet": "No-Source-Text",
                "date": LABEL_PROVENANCE_DATE,
                "note": (
                    "Arm 3 ships as a candidate automatically per owner ruling J; no verdict "
                    "column exists on this sheet and none was ever solicited."
                ),
            },
        }

    expected_ids = set(range(1, 102))
    got_ids = set(cases.keys())
    if got_ids != expected_ids:
        missing = sorted(expected_ids - got_ids)
        extra = sorted(got_ids - expected_ids)
        raise LabelReadError(
            f"case numbering does not match the expected 1..101 contiguous set. "
            f"missing={missing} extra={extra}"
        )

    ordered = [cases[i] for i in range(1, 102)]

    labelled_ct = sum(1 for c in ordered if c["verdict"]["value"] is not None)
    skipped_ct = sum(1 for c in ordered if c["verdict"].get("skipped"))
    no_verdict_ct = sum(1 for c in ordered if c["question_type"] == "no_verdict_by_design")
    correctness_gap_ct = sum(1 for c in ordered if c["verdict"].get("correctness_gap"))

    return {
        "schema_version": 1,
        "source_workbook": "136-NOVELTY-HARDCASES.xlsx",
        "generated_at": LABEL_PROVENANCE_DATE,
        "total_cases": len(ordered),
        "labelled_count": labelled_ct,
        "skipped_count": skipped_ct,
        "no_verdict_by_design_count": no_verdict_ct,
        "correctness_gaps_flagged": correctness_gap_ct,
        "cases": ordered,
    }


def write_owner_labels_json(xlsx_path: str, json_out_path: str) -> Dict[str, Any]:
    """Reads ``xlsx_path`` (see ``read_owner_labels_from_xlsx``), computes a
    content hash over the ``cases`` array -- the one thing plan 136-04 must
    re-verify has not been hand-edited post-labelling (T-136-03-06) -- and
    writes the result to ``json_out_path``. Returns the written dict."""
    result = read_owner_labels_from_xlsx(xlsx_path)
    canonical_cases = json.dumps(result["cases"], sort_keys=True, ensure_ascii=False)
    content_hash = "sha256:" + hashlib.sha256(canonical_cases.encode("utf-8")).hexdigest()
    result = dict(result)
    result["hash_method"] = "sha256 over json.dumps(cases, sort_keys=True, ensure_ascii=False)"
    result["content_hash"] = content_hash

    if not all("label_provenance" in c for c in result["cases"]):
        raise LabelReadError("internal error: a case is missing label_provenance")
    for c in result["cases"]:
        if c["verdict"].get("skipped") and c["verdict"]["value"] is not None:
            raise LabelReadError(
                f"internal error: case {c['case_id']} is marked skipped but carries a verdict value"
            )

    os.makedirs(os.path.dirname(json_out_path), exist_ok=True)
    with open(json_out_path, "w", encoding="utf-8") as fh:
        json.dump(result, fh, ensure_ascii=False, indent=2, sort_keys=False)
        fh.write("\n")
    return result


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Phase 136 plan 03 -- gate-1 decision evidence + novelty hard-case candidates "
                    "(read-only measurement over the deployed discovery sidecar)."
    )
    parser.add_argument(
        "asset", nargs="?", default=None,
        help="path to the discovery-v1-*.db sidecar to measure (not required with "
             "--read-labels-from, which only round-trips the labelled XLSX)",
    )
    parser.add_argument(
        "--read-labels-from", default=None, metavar="XLSX_PATH",
        help="Task 4: read the owner-filled 136-NOVELTY-HARDCASES.xlsx back and write "
             "discovery_data/novelty_hardcase_labels-v1.json (see --labels-out). Skips the "
             "asset measurement pipeline entirely -- 'asset' is not required with this flag.",
    )
    parser.add_argument(
        "--labels-out", default=DEFAULT_LABELS_OUT,
        help=f"output path for the owner-labels JSON (default: {DEFAULT_LABELS_OUT})",
    )
    parser.add_argument(
        "--research-db", default=DEFAULT_RESEARCH_DB,
        help="path to the (gitignored, local-only) fullcorpus.db research DB used for the D-13c/gate-4 "
             "page-coverage lookup (default: same_work_spike/probe/data/fullcorpus.db)",
    )
    parser.add_argument(
        "--libraries-csv", default=DEFAULT_LIBRARIES_CSV,
        help="path to libraries.csv (public catalogue metadata for hard-case manuscripts)",
    )
    parser.add_argument(
        "--fjms-db", default=DEFAULT_FJMS_DB,
        help="path to fjms_enrichment.db (bib `bibliography` table + FJMS `catalog` table; owner "
             "ruling J's three-arm sampler; default: fist_data/fjms_enrichment.db)",
    )
    parser.add_argument(
        "--pgp-db", default=DEFAULT_PGP_DB,
        help="path to pgp.db (PGP `documents` + `document_fragments`; owner ruling J's three-arm "
             "sampler; default: pgp_data/pgp.db)",
    )
    parser.add_argument(
        "--fgp-db", default=DEFAULT_FGP_DB,
        help="path to fgp_transcriptions.db; owner ruling J's three-arm sampler; default: "
             "fgp_data/fgp_transcriptions.db)",
    )
    parser.add_argument("--evidence-out", default=DEFAULT_EVIDENCE_OUT)
    parser.add_argument("--hardcases-out", default=DEFAULT_HARDCASES_OUT)
    parser.add_argument("--hardcases-xlsx-out", default=DEFAULT_HARDCASES_XLSX_OUT)
    parser.add_argument(
        "--no-write", action="store_true",
        help="print console tables only; do not write the Markdown/XLSX artifacts",
    )
    args = parser.parse_args(argv)

    # Task 4 mode: read the owner-labelled workbook back and write the JSON
    # ground-truth file. Entirely independent of the asset-measurement
    # pipeline below (no DB is opened), so it exits before the ledger/asset
    # handling that mode does not need.
    if args.read_labels_from:
        try:
            result = write_owner_labels_json(args.read_labels_from, args.labels_out)
        except LabelReadError as exc:
            print(f"FAIL: {exc}", file=sys.stderr)
            return 1
        print(
            f"wrote {args.labels_out}: {result['total_cases']} cases "
            f"({result['labelled_count']} labelled, {result['skipped_count']} skipped, "
            f"{result['no_verdict_by_design_count']} no-verdict-by-design, "
            f"{result['correctness_gaps_flagged']} correctness gaps flagged) "
            f"content_hash={result['content_hash']}"
        )
        return 0

    if not args.asset:
        parser.error("the following arguments are required: asset (unless --read-labels-from is given)")

    ledger = NonzeroLedger()

    conn = connect_readonly(args.asset)
    try:
        works = load_works(conn)
        ledger.check("works", len(works))

        claims = load_claims(conn)
        ledger.check("claims", len(claims))

        human_confirmed_claim_ids = load_human_confirmed_claim_ids(conn)
        ledger.check("human_confirmed_claim_ids", len(human_confirmed_claim_ids))

        kept_tie_pages = load_kept_tie_pages(conn, works)
        ledger.check("kept_tie_pages", len(kept_tie_pages))

        near_tie_pages = build_near_tie_competition(claims, works)
        ledger.check("near_tie_pages", len(near_tie_pages))

        needed_pages = pages_needing_coverage(
            claims, works, human_confirmed_claim_ids, kept_tie_pages, near_tie_pages
        )
        ledger.check("pages_needing_coverage", len(needed_pages))
        page_norm_letters = load_page_norm_letters(args.research_db, needed_pages)
        ledger.check("page_norm_letters_resolved", len(page_norm_letters))

        identifications, classification = classify_identifications(
            claims, works, human_confirmed_claim_ids, kept_tie_pages, near_tie_pages, page_norm_letters
        )
        ledger.check("identifications", len(identifications))

        # Fold canonical_work_id back onto each claim dict for the D-16/D-13c
        # main-pool lookups below (classify_identifications built its own
        # copies internally; re-derive here once, cheaply, via `works`).
        for c in claims:
            w = works.get(c["work_id"])
            c["canonical_work_id"] = w["canonical_work_id"] if w else None

        main_total = sum(1 for b, _ in classification.values() if b == "main")
        show_more_total = sum(1 for b, _ in classification.values() if b == "show_more")
        ledger.check("main_total", main_total)
        ledger.check("show_more_total", show_more_total)
        reason_counts: Dict[str, int] = defaultdict(int)
        for _bucket, reason in classification.values():
            reason_counts[reason] += 1

        span_groups = load_identical_span_groups(conn, works)
        ledger.check("span_groups", len(span_groups))

        d13e = compute_d13e(conn, span_groups, classification)
        ledger.check("d13e_total_middle_bucket", d13e["total_middle_bucket"])

        d16 = compute_d16(conn, claims, classification)
        ledger.check("d16_corpus_wide_rows", len(d16["corpus_wide"]))

        d13c = compute_d13c(conn, classification, claims)
        ledger.check("d13c_direct_total", d13c["direct_total"])
        ledger.check("d13c_propagated_total", d13c["propagated_total"])

        d13b = compute_d13b(span_groups)
        ledger.check("d13b_total_groups", d13b["total_groups"])

        d13d = compute_d13d(span_groups, works)
        ledger.check("d13d_diff_canon_groups_total", d13d["diff_canon_groups_total"])

        libraries = load_libraries_csv(args.libraries_csv)
        ledger.check("libraries_csv_rows", len(libraries))

        # Ruling J (136-GATE1-DECISIONS.md § J) source sidecars -- real
        # bib/PGP/FGP/FJMS-catalogue data for the three-arm stratified
        # sampler. Graceful degradation (empty dict) if a sidecar is
        # missing, matching every other sidecar reader in this project --
        # NOT ledger-checked individually (a genuinely absent sidecar is a
        # real, reportable condition, not a bug in this script), but the
        # resulting arm totals below ARE ledger-checked in aggregate.
        fjms_catalog = load_fjms_catalog_text(args.fjms_db)
        bib_rows_idx = load_bib_rows(args.fjms_db)
        pgp_signal_idx = load_pgp_signal_index(args.pgp_db)
        fgp_rows_idx = load_fgp_rows(args.fgp_db)
        print(
            f"ruling-J sidecars: fjms_catalog {len(fjms_catalog):,} / bib {len(bib_rows_idx):,} / "
            f"pgp {len(pgp_signal_idx):,} / fgp {len(fgp_rows_idx):,}", file=sys.stderr,
        )

        hardcases = build_hardcases(
            claims, works, d13d, libraries,
            fjms_catalog=fjms_catalog, bib_rows_idx=bib_rows_idx,
            pgp_signal_idx=pgp_signal_idx, fgp_rows_idx=fgp_rows_idx,
        )
        ledger.check("hardcases", len(hardcases))
        ledger.check("hardcases_class1_near_miss_spotcheck", sum(1 for c in hardcases if c["class"] == "near_miss"))
        ledger.check("hardcases_class2_alias_spotcheck", sum(1 for c in hardcases if c["class"] == "alias"))
        ledger.check("hardcases_class3_granularity_spotcheck", sum(1 for c in hardcases if c["class"] == "granularity"))
        ledger.check("hardcases_class6_catalogue_divergence", sum(1 for c in hardcases if c["class"] == "catalogue_divergence"))
        ledger.check("hardcases_arm1_residual", sum(1 for c in hardcases if c["class"] == "residual"))
        ledger.check("hardcases_arm2_heuristic_demoted", sum(1 for c in hardcases if c["class"] == "heuristic_demoted"))
        ledger.check("hardcases_arm3_no_source_text", sum(1 for c in hardcases if c["class"] == "no_source_text"))
        ledger.check("hardcases_identity_total", sum(1 for c in hardcases if c["class"] in _IDENTITY_CLASSES))
        ledger.check("hardcases_novelty_total", sum(1 for c in hardcases if c["class"] not in _IDENTITY_CLASSES))
        # Acceptance criteria: any attached draft verdict MUST be explicitly
        # marked PROPOSAL and separable from an owner's answer -- assert it
        # on every case that carries one (a case may also carry none at all,
        # e.g. the near-miss class below, which is a valid choice per the
        # plan's "MAY attach a draft" wording).
        for case in hardcases:
            proposal = case.get("proposal")
            if proposal is not None and not proposal.startswith("PROPOSAL"):
                raise AssertionError(
                    f"hard-case draft verdict is not marked PROPOSAL: {proposal!r}"
                )

        # Assign stable case numbers ONCE, in the same class order both
        # rendered artifacts (Markdown + XLSX) share -- see assign_case_numbers.
        hardcases = assign_case_numbers(hardcases)
        ledger.check("hardcases_numbered", len(hardcases))

    finally:
        conn.close()

    # ---- console summary ----
    print("=" * 78)
    print("Discovery gate-1 evidence (Phase 136 plan 03)")
    print("=" * 78)
    print(f"asset            : {args.asset}")
    print(f"identifications  : {len(identifications):,}  (main {main_total:,} / show-more {show_more_total:,})")
    print(f"span-groups      : {len(span_groups):,}")
    print(f"D-13e middle pop : {d13e['total_middle_bucket']:,} "
          f"(not-reachable {d13e['not_reachable_total']:,} / overlap {d13e['overlap_total']:,})")
    print(f"D-13b tied groups: {d13b['tied_after_band_rank_groups']:,} of {d13b['total_groups']:,}")
    print(f"D-13d collapse   : {d13d['collapse_candidate_groups']:,} of {d13d['diff_canon_groups_total']:,}")
    print(f"D-13c thinnest   : {d13c['thinnest_direct']} matched letters")
    print(f"hard-cases       : {len(hardcases):,}")
    print("=" * 78)

    if not ledger.ok():
        print(
            "FAIL: the following measurements returned an unexpected ZERO "
            f"(never presenting a silent zero as a finding): {ledger.failures}",
            file=sys.stderr,
        )
        return 1

    if not args.no_write:
        evidence_md = render_evidence_brief(
            asset_path=args.asset,
            identifications_total=len(identifications),
            main_total=main_total,
            show_more_total=show_more_total,
            reason_counts=dict(reason_counts),
            d13e=d13e,
            d16=d16,
            d13c=d13c,
            d13b=d13b,
            d13d=d13d,
        )
        os.makedirs(os.path.dirname(args.evidence_out), exist_ok=True)
        with open(args.evidence_out, "w", encoding="utf-8") as fh:
            fh.write(evidence_md)
        print(f"wrote {args.evidence_out}")

        hardcases_md = render_hardcases_brief(hardcases)
        os.makedirs(os.path.dirname(args.hardcases_out), exist_ok=True)
        with open(args.hardcases_out, "w", encoding="utf-8") as fh:
            fh.write(hardcases_md)
        print(f"wrote {args.hardcases_out}")

        write_hardcases_xlsx(hardcases, args.hardcases_xlsx_out)
        print(f"wrote {args.hardcases_xlsx_out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""
סקריפט להשוואת גרסאות HTR ומציאת חילופי אותיות (שתיים→אחת ואחת→שתיים)
משווה בין V0.7 (AllGenizah_OLD.txt) ל-V0.8 (Transcriptions.txt)

שימוש:
    python analyze_char_merges.py --v7 AllGenizah_OLD.txt --v8 Transcriptions.txt --output merges_report.xlsx

עם קורפוס לסינון מילים אמיתיות:
    python analyze_char_merges.py --v7 ... --v8 ... --corpus /path/to/corpus_folder
"""

import argparse
import re
import sys
import os
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set
import unicodedata

# Hebrew character range
HEBREW_RANGE = range(0x0590, 0x05FF + 1)

def is_hebrew(char: str) -> bool:
    """בודק אם תו הוא אות עברית"""
    return len(char) == 1 and ord(char) in HEBREW_RANGE

def normalize_hebrew(text: str) -> str:
    """נרמול טקסט עברי - הסרת ניקוד וסימנים"""
    # הסרת ניקוד
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    # השארת רק אותיות עבריות ורווחים
    return text

def extract_hebrew_words(text: str) -> List[str]:
    """חילוץ מילים עבריות מטקסט"""
    text = normalize_hebrew(text)
    # מוצא רצפים של אותיות עבריות
    words = re.findall(r'[\u0590-\u05FF]+', text)
    return [w for w in words if len(w) >= 2]  # מילים עם לפחות 2 אותיות


def load_corpus(corpus_path: str) -> Set[str]:
    """
    טוען קורפוס של מילים אמיתיות.
    יכול להיות:
    - תיקייה עם קבצי טקסט
    - קובץ עם מילה בכל שורה (פלט של prepare_corpus.py)

    מחזיר set של כל המילים העבריות בקורפוס.
    """
    corpus_words = set()
    corpus_path = Path(corpus_path)

    if not corpus_path.exists():
        print(f"אזהרה: הקורפוס {corpus_path} לא נמצא")
        return corpus_words

    # אם זה קובץ בודד - קורא מילה בכל שורה
    if corpus_path.is_file():
        print(f"  קורא קובץ מילים: {corpus_path}")
        try:
            with open(corpus_path, 'r', encoding='utf-8-sig') as f:
                for line in f:
                    word = line.strip()
                    if word and len(word) >= 2:
                        corpus_words.add(word)
        except Exception as e:
            print(f"  שגיאה בקריאת {corpus_path}: {e}")
        print(f"  נטענו {len(corpus_words)} מילים מהקובץ")
        return corpus_words

    # אם זו תיקייה - קורא את כל קבצי הטקסט
    text_files = list(corpus_path.glob('*.txt')) + list(corpus_path.glob('**/*.txt'))
    print(f"  נמצאו {len(text_files)} קבצי טקסט בקורפוס")

    for txt_file in text_files:
        try:
            with open(txt_file, 'r', encoding='utf-8-sig') as f:
                text = f.read()
                words = extract_hebrew_words(text)
                corpus_words.update(words)
        except Exception as e:
            print(f"  שגיאה בקריאת {txt_file}: {e}")

    print(f"  נטענו {len(corpus_words)} מילים ייחודיות מהקורפוס")
    return corpus_words

def parse_v8_file(filepath: str) -> Dict[str, dict]:
    """
    פורסר לקובץ V0.8 (Transcriptions.txt)
    פורמט: ==> ID <==

    מחזיר dict עם ID מנורמל כמפתח, וערך שכולל את הטקסט והשורה הראשונה
    """
    docs = {}
    current_id = None
    current_raw_id = None
    current_text = []

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        for line in f:
            line = line.strip()
            if line.startswith('==>') and line.endswith('<=='):
                # שומר מסמך קודם
                if current_id and current_text:
                    full_text = '\n'.join(current_text)
                    first_line = current_text[0] if current_text else ''
                    docs[current_id] = {
                        'text': full_text,
                        'first_line': first_line,
                        'raw_id': current_raw_id
                    }
                # מתחיל מסמך חדש
                current_raw_id = line[3:-3].strip()
                current_id = normalize_doc_id(current_raw_id)
                current_text = []
            elif current_id:
                current_text.append(line)

        # מסמך אחרון
        if current_id and current_text:
            full_text = '\n'.join(current_text)
            first_line = current_text[0] if current_text else ''
            docs[current_id] = {
                'text': full_text,
                'first_line': first_line,
                'raw_id': current_raw_id
            }

    return docs

def parse_v7_file(filepath: str) -> Dict[str, dict]:
    """
    פורסר לקובץ V0.7 (AllGenizah_OLD.txt)
    פורמט: ### Q:\ERC\...\IE..._P..._FL...—reco.xml - SHELFMARK - filename.tif

    מחזיר dict עם ID מנורמל כמפתח, וערך שכולל את הטקסט והשורה הראשונה
    """
    docs = {}
    current_id = None
    current_text = []

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        for line in f:
            line = line.strip()
            # V0.7 משתמש ב-### כמפריד
            if line.startswith('###'):
                # שומר מסמך קודם
                if current_id and current_text:
                    full_text = '\n'.join(current_text)
                    first_line = current_text[0] if current_text else ''
                    docs[current_id] = {
                        'text': full_text,
                        'first_line': first_line
                    }
                # מתחיל מסמך חדש - מחלץ ID מהנתיב
                # פורמט: ### Q:\ERC\all_version_0.9\990000571730205171\IE47712399\IE47712399_P000028_FL47712624—reco.xml - SP RNL...
                header = line[3:].strip()
                current_id = extract_id_from_v7_header(header)
                current_text = []
            elif current_id:
                current_text.append(line)

        # מסמך אחרון
        if current_id and current_text:
            full_text = '\n'.join(current_text)
            first_line = current_text[0] if current_text else ''
            docs[current_id] = {
                'text': full_text,
                'first_line': first_line
            }

    return docs


def extract_id_from_v7_header(header: str) -> str:
    """
    מחלץ ID מכותרת V0.7
    דוגמה: Q:\ERC\all_version_0.9\990000571730205171\IE47712399\IE47712399_P000028_FL47712624—reco.xml - SP RNL...
    מחזיר: 990000571730205171_IE47712399_P000028_FL47712624
    """
    # מחפש את הדפוס: IE..._P..._FL... (לפני —reco.xml או .xml)
    match = re.search(r'(IE\d+[_-]P\d+[_-]FL\d+)', header, re.IGNORECASE)
    if match:
        ie_p_fl = match.group(1).replace('-', '_')
        # מחפש גם את ה-sys_id (מספר ארוך לפני IE)
        sys_match = re.search(r'(\d{12,18})', header)
        if sys_match:
            return f"{sys_match.group(1)}_{ie_p_fl}"
        return ie_p_fl

    # fallback - נרמול רגיל
    return normalize_doc_id(header)

def normalize_doc_id(doc_id: str) -> str:
    """
    נרמול מזהה מסמך לצורך התאמה בין גרסאות
    מחלץ את המרכיבים העיקריים: sys_id, IE, P, FL
    """
    # מחלץ מספרים ומזהים
    # דוגמה: 990000412990205171_IE104549337_P000001_FL104549339
    parts = re.findall(r'(\d{10,}|IE\d+|P\d+|FL\d+)', doc_id, re.IGNORECASE)
    if parts:
        return '_'.join(parts)
    return doc_id

def align_words(words1: List[str], words2: List[str]) -> List[Tuple[Optional[str], Optional[str]]]:
    """
    יישור מילים בין שתי רשימות באמצעות SequenceMatcher
    מחזיר רשימה של זוגות (מילה1, מילה2) או (None, מילה) לתוספות/מחיקות
    """
    matcher = SequenceMatcher(None, words1, words2)
    aligned = []

    for op, i1, i2, j1, j2 in matcher.get_opcodes():
        if op == 'equal':
            for k in range(i2 - i1):
                aligned.append((words1[i1 + k], words2[j1 + k]))
        elif op == 'replace':
            # התאמה 1:1 כשאפשר
            len1, len2 = i2 - i1, j2 - j1
            for k in range(max(len1, len2)):
                w1 = words1[i1 + k] if k < len1 else None
                w2 = words2[j1 + k] if k < len2 else None
                aligned.append((w1, w2))
        elif op == 'delete':
            for k in range(i2 - i1):
                aligned.append((words1[i1 + k], None))
        elif op == 'insert':
            for k in range(j2 - j1):
                aligned.append((None, words2[j1 + k]))

    return aligned

def find_char_substitutions(word1: str, word2: str) -> List[Tuple[str, str, str]]:
    """
    מוצא חילופי תווים בין שתי מילים
    מחזיר רשימה של (מקור, יעד, סוג) כאשר סוג הוא 'merge' או 'split'
    """
    if not word1 or not word2:
        return []

    substitutions = []
    matcher = SequenceMatcher(None, word1, word2)

    for op, i1, i2, j1, j2 in matcher.get_opcodes():
        if op == 'replace':
            orig = word1[i1:i2]
            repl = word2[j1:j2]

            # חילופים 1↔1 (אות באות)
            if len(orig) == 1 and len(repl) == 1:
                substitutions.append((orig, repl, 'single'))
            # חילופים 2↔1
            elif len(orig) == 2 and len(repl) == 1:
                substitutions.append((orig, repl, 'merge'))
            elif len(orig) == 1 and len(repl) == 2:
                substitutions.append((orig, repl, 'split'))
            # חילופים 2↔2
            elif len(orig) == 2 and len(repl) == 2:
                substitutions.append((orig, repl, 'swap2'))
            # חילופים 3↔1
            elif len(orig) == 3 and len(repl) == 1:
                substitutions.append((orig, repl, 'merge3'))
            elif len(orig) == 1 and len(repl) == 3:
                substitutions.append((orig, repl, 'split3'))

    return substitutions

def first_line_similarity(line1: str, line2: str) -> float:
    """
    מחשב דמיון בין שתי שורות ראשונות
    מחזיר ערך בין 0 ל-1
    """
    if not line1 or not line2:
        return 0.0

    words1 = extract_hebrew_words(line1)
    words2 = extract_hebrew_words(line2)

    if not words1 or not words2:
        return 0.0

    # חישוב דמיון Jaccard על מילים
    set1 = set(words1)
    set2 = set(words2)
    intersection = len(set1 & set2)
    union = len(set1 | set2)

    if union == 0:
        return 0.0

    return intersection / union


def analyze_documents(v7_docs: Dict[str, dict], v8_docs: Dict[str, dict],
                      progress_callback=None,
                      min_similarity: float = 0.2,
                      corpus_words: Set[str] = None) -> Dict[Tuple[str, str], List[dict]]:
    """
    משווה מסמכים ואוסף סטטיסטיקת חילופים
    מתאים מסמכים לפי ID ומוודא דמיון בשורה הראשונה

    אם corpus_words מסופק, רק זוגות מילים שלפחות אחת מהן בקורפוס ייכללו.
    """
    # מציאת מסמכים משותפים לפי ID
    common_ids = set(v7_docs.keys()) & set(v8_docs.keys())
    print(f"נמצאו {len(common_ids)} מסמכים עם ID תואם מתוך V0.7: {len(v7_docs)}, V0.8: {len(v8_docs)}")

    # אוסף חילופים: (מקור, יעד) -> [רשימת הופעות]
    substitutions = defaultdict(list)
    matched_count = 0
    skipped_low_similarity = 0
    skipped_no_corpus_match = 0
    corpus_filtered_pairs = 0

    for idx, norm_id in enumerate(common_ids):
        if progress_callback and idx % 1000 == 0:
            progress_callback(idx, len(common_ids))

        doc_v7 = v7_docs[norm_id]
        doc_v8 = v8_docs[norm_id]

        text_v7 = doc_v7['text']
        text_v8 = doc_v8['text']
        first_v7 = doc_v7.get('first_line', '')
        first_v8 = doc_v8.get('first_line', '')

        # בדיקת דמיון בשורה הראשונה
        similarity = first_line_similarity(first_v7, first_v8)
        if similarity < min_similarity:
            skipped_low_similarity += 1
            continue

        matched_count += 1

        words_v7 = extract_hebrew_words(text_v7)
        words_v8 = extract_hebrew_words(text_v8)

        aligned = align_words(words_v7, words_v8)

        for w7, w8 in aligned:
            if w7 and w8 and w7 != w8:
                # סינון לפי קורפוס - לפחות אחת מהמילים חייבת להיות אמיתית
                if corpus_words is not None:
                    if w7 not in corpus_words and w8 not in corpus_words:
                        skipped_no_corpus_match += 1
                        continue
                    corpus_filtered_pairs += 1

                subs = find_char_substitutions(w7, w8)
                for orig, repl, sub_type in subs:
                    # רק אותיות עבריות
                    if all(is_hebrew(c) for c in orig) and all(is_hebrew(c) for c in repl):
                        key = (orig, repl)
                        substitutions[key].append({
                            'doc_id': norm_id,
                            'word_v7': w7,
                            'word_v8': w8,
                            'type': sub_type
                        })

    print(f"  מסמכים שעברו בדיקת דמיון: {matched_count}")
    print(f"  מסמכים שדולגו (דמיון נמוך): {skipped_low_similarity}")
    if corpus_words is not None:
        print(f"  זוגות שעברו סינון קורפוס: {corpus_filtered_pairs}")
        print(f"  זוגות שנפסלו (אין מילה אמיתית): {skipped_no_corpus_match}")

    return substitutions

def create_excel_report(substitutions: Dict[Tuple[str, str], List[dict]],
                        output_path: str,
                        min_occurrences: int = 2):
    """
    יוצר דוח אקסל מסודר לפי שכיחות
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("נדרש להתקין openpyxl: pip install openpyxl")
        # יוצר CSV במקום
        create_csv_report(substitutions, output_path.replace('.xlsx', '.csv'), min_occurrences)
        return

    wb = openpyxl.Workbook()

    # --- גיליון 1: סיכום לפי שכיחות ---
    ws_summary = wb.active
    ws_summary.title = "סיכום חילופים"
    ws_summary.sheet_view.rightToLeft = True

    # כותרות
    headers = ['מקור', 'יעד', 'סוג', 'שכיחות', 'דוגמאות מילים']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # מיון לפי שכיחות
    sorted_subs = sorted(substitutions.items(), key=lambda x: len(x[1]), reverse=True)

    row = 2
    for (orig, repl), occurrences in sorted_subs:
        if len(occurrences) < min_occurrences:
            continue

        sub_type = occurrences[0]['type']
        type_label = {
            'merge': 'שתיים→אחת',
            'split': 'אחת→שתיים',
            'merge3': 'שלוש→אחת',
            'split3': 'אחת→שלוש'
        }.get(sub_type, sub_type)

        # דוגמאות מילים (עד 5)
        examples = set()
        for occ in occurrences[:10]:
            examples.add(f"{occ['word_v7']}→{occ['word_v8']}")
            if len(examples) >= 5:
                break

        ws_summary.cell(row=row, column=1, value=orig)
        ws_summary.cell(row=row, column=2, value=repl)
        ws_summary.cell(row=row, column=3, value=type_label)
        ws_summary.cell(row=row, column=4, value=len(occurrences))
        ws_summary.cell(row=row, column=5, value=' | '.join(examples))

        # יישור לימין לעברית
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal='right')

        row += 1

    # רוחב עמודות
    ws_summary.column_dimensions['A'].width = 10
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 50

    # --- גיליון 2: רק שתיים→אחת ---
    ws_merges = wb.create_sheet("שתיים לאחת")
    ws_merges.sheet_view.rightToLeft = True

    headers2 = ['שתי אותיות', 'אות אחת', 'שכיחות', 'דוגמאות']
    for col, header in enumerate(headers2, 1):
        cell = ws_merges.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    row = 2
    for (orig, repl), occurrences in sorted_subs:
        if occurrences[0]['type'] != 'merge' or len(occurrences) < min_occurrences:
            continue

        examples = set()
        for occ in occurrences[:10]:
            examples.add(f"{occ['word_v7']}→{occ['word_v8']}")
            if len(examples) >= 5:
                break

        ws_merges.cell(row=row, column=1, value=orig)
        ws_merges.cell(row=row, column=2, value=repl)
        ws_merges.cell(row=row, column=3, value=len(occurrences))
        ws_merges.cell(row=row, column=4, value=' | '.join(examples))
        row += 1

    ws_merges.column_dimensions['A'].width = 12
    ws_merges.column_dimensions['B'].width = 10
    ws_merges.column_dimensions['C'].width = 10
    ws_merges.column_dimensions['D'].width = 50

    # --- גיליון 3: אחת→שתיים ---
    ws_splits = wb.create_sheet("אחת לשתיים")
    ws_splits.sheet_view.rightToLeft = True

    headers3 = ['אות אחת', 'שתי אותיות', 'שכיחות', 'דוגמאות']
    for col, header in enumerate(headers3, 1):
        cell = ws_splits.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

    row = 2
    for (orig, repl), occurrences in sorted_subs:
        if occurrences[0]['type'] != 'split' or len(occurrences) < min_occurrences:
            continue

        examples = set()
        for occ in occurrences[:10]:
            examples.add(f"{occ['word_v7']}→{occ['word_v8']}")
            if len(examples) >= 5:
                break

        ws_splits.cell(row=row, column=1, value=orig)
        ws_splits.cell(row=row, column=2, value=repl)
        ws_splits.cell(row=row, column=3, value=len(occurrences))
        ws_splits.cell(row=row, column=4, value=' | '.join(examples))
        row += 1

    ws_splits.column_dimensions['A'].width = 10
    ws_splits.column_dimensions['B'].width = 12
    ws_splits.column_dimensions['C'].width = 10
    ws_splits.column_dimensions['D'].width = 50

    # --- גיליון 4: פירוט מלא ---
    ws_detail = wb.create_sheet("פירוט מלא")
    ws_detail.sheet_view.rightToLeft = True

    headers4 = ['מקור', 'יעד', 'סוג', 'מילה V0.7', 'מילה V0.8', 'מזהה מסמך']
    for col, header in enumerate(headers4, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

    row = 2
    for (orig, repl), occurrences in sorted_subs:
        if len(occurrences) < min_occurrences:
            continue
        for occ in occurrences[:100]:  # מגביל ל-100 דוגמאות לכל חילוף
            ws_detail.cell(row=row, column=1, value=orig)
            ws_detail.cell(row=row, column=2, value=repl)
            ws_detail.cell(row=row, column=3, value=occ['type'])
            ws_detail.cell(row=row, column=4, value=occ['word_v7'])
            ws_detail.cell(row=row, column=5, value=occ['word_v8'])
            ws_detail.cell(row=row, column=6, value=occ['doc_id'])
            row += 1

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_detail.column_dimensions[col].width = 15

    # שמירה
    wb.save(output_path)
    print(f"הדוח נשמר ב: {output_path}")

def create_csv_report(substitutions: Dict[Tuple[str, str], List[dict]],
                      output_path: str,
                      min_occurrences: int = 2):
    """
    יוצר דוח CSV כחלופה לאקסל
    """
    import csv

    sorted_subs = sorted(substitutions.items(), key=lambda x: len(x[1]), reverse=True)

    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['מקור', 'יעד', 'סוג', 'שכיחות', 'דוגמאות'])

        for (orig, repl), occurrences in sorted_subs:
            if len(occurrences) < min_occurrences:
                continue

            sub_type = occurrences[0]['type']
            examples = set()
            for occ in occurrences[:5]:
                examples.add(f"{occ['word_v7']}→{occ['word_v8']}")

            writer.writerow([orig, repl, sub_type, len(occurrences), ' | '.join(examples)])

    print(f"הדוח נשמר ב: {output_path}")

def print_summary(substitutions: Dict[Tuple[str, str], List[dict]], top_n: int = 30):
    """
    מדפיס סיכום מהיר לקונסול
    """
    print("\n" + "="*60)
    print("סיכום חילופי אותיות (שתיים↔אחת)")
    print("="*60)

    merges = [(k, v) for k, v in substitutions.items() if v[0]['type'] == 'merge']
    splits = [(k, v) for k, v in substitutions.items() if v[0]['type'] == 'split']

    merges.sort(key=lambda x: len(x[1]), reverse=True)
    splits.sort(key=lambda x: len(x[1]), reverse=True)

    print(f"\n--- שתי אותיות → אחת (Top {top_n}) ---")
    print(f"{'מקור':<8} {'יעד':<6} {'שכיחות':<10} {'דוגמה'}")
    print("-" * 50)
    for (orig, repl), occurrences in merges[:top_n]:
        example = f"{occurrences[0]['word_v7']}→{occurrences[0]['word_v8']}"
        print(f"{orig:<8} {repl:<6} {len(occurrences):<10} {example}")

    print(f"\n--- אות אחת → שתיים (Top {top_n}) ---")
    print(f"{'מקור':<6} {'יעד':<8} {'שכיחות':<10} {'דוגמה'}")
    print("-" * 50)
    for (orig, repl), occurrences in splits[:top_n]:
        example = f"{occurrences[0]['word_v7']}→{occurrences[0]['word_v8']}"
        print(f"{orig:<6} {repl:<8} {len(occurrences):<10} {example}")

def main():
    parser = argparse.ArgumentParser(
        description='השוואת גרסאות HTR ומציאת חילופי אותיות',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('--v7', required=True, help='נתיב לקובץ V0.7 (AllGenizah_OLD.txt)')
    parser.add_argument('--v8', required=True, help='נתיב לקובץ V0.8 (Transcriptions.txt)')
    parser.add_argument('--output', '-o', default='char_merges_report.xlsx',
                        help='נתיב לקובץ הפלט (xlsx או csv)')
    parser.add_argument('--min', type=int, default=2,
                        help='מספר הופעות מינימלי להכללה בדוח (ברירת מחדל: 2)')
    parser.add_argument('--similarity', type=float, default=0.2,
                        help='סף דמיון מינימלי בשורה הראשונה (0-1, ברירת מחדל: 0.2)')
    parser.add_argument('--corpus', type=str, default=None,
                        help='תיקייה עם קבצי טקסט של קורפוס אמיתי לסינון מילים')
    parser.add_argument('--summary-only', action='store_true',
                        help='רק סיכום לקונסול, בלי קובץ אקסל')

    args = parser.parse_args()

    # בדיקת קבצים
    if not Path(args.v7).exists():
        print(f"שגיאה: הקובץ {args.v7} לא נמצא")
        sys.exit(1)
    if not Path(args.v8).exists():
        print(f"שגיאה: הקובץ {args.v8} לא נמצא")
        sys.exit(1)

    # טעינת קורפוס אם צוין
    corpus_words = None
    if args.corpus:
        print(f"\nטוען קורפוס מילים אמיתיות מ: {args.corpus}")
        corpus_words = load_corpus(args.corpus)
        if not corpus_words:
            print("אזהרה: הקורפוס ריק, ממשיך ללא סינון")
            corpus_words = None

    print("\nקורא קבצים...")
    print(f"  V0.7: {args.v7}")
    v7_docs = parse_v7_file(args.v7)
    print(f"    נקראו {len(v7_docs)} מסמכים")

    print(f"  V0.8: {args.v8}")
    v8_docs = parse_v8_file(args.v8)
    print(f"    נקראו {len(v8_docs)} מסמכים")

    print(f"\nמנתח חילופים (סף דמיון: {args.similarity})...")
    if corpus_words:
        print(f"  סינון לפי קורפוס: פעיל ({len(corpus_words)} מילים)")

    def progress(current, total):
        print(f"  התקדמות: {current}/{total} ({100*current//total}%)", end='\r')

    substitutions = analyze_documents(v7_docs, v8_docs,
                                      progress_callback=progress,
                                      min_similarity=args.similarity,
                                      corpus_words=corpus_words)
    print(f"\nנמצאו {len(substitutions)} סוגי חילופים ייחודיים")

    # סיכום לקונסול
    print_summary(substitutions)

    # יצירת דוח
    if not args.summary_only:
        print(f"\nיוצר דוח: {args.output}")
        if args.output.endswith('.xlsx'):
            create_excel_report(substitutions, args.output, args.min)
        else:
            create_csv_report(substitutions, args.output, args.min)

    print("\nסיום!")

if __name__ == '__main__':
    main()
